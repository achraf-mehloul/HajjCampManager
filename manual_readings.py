import datetime
import json
import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from models import db, MDBPanel, ManualReading, MeasurementUnit, DropdownList, DropdownItem, User
from sqlalchemy import desc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

readings_bp = Blueprint('readings', __name__)

# صفحة البحث عن لوحة لإضافة قراءة يدوية
@readings_bp.route('/search-panel-for-reading')
@login_required
def search_panel_for_reading():
    # التحقق من أن المستخدم مصرح له
    if current_user.role == 'contractor' and not current_user.contractor_id:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    # الحصول على اللوحات المتاحة للمستخدم
    if current_user.role == 'admin':
        # المدير يمكنه رؤية جميع اللوحات
        panels = MDBPanel.query.all()
    elif current_user.role in ['contractor', 'regular_contractor']:
        # المقاول العادي أو المقاول المدير يمكنه رؤية اللوحات في مناطق مسؤوليته
        from app import get_contractor_areas
        contractor_areas = get_contractor_areas(current_user.contractor_id, current_user.is_manager if hasattr(current_user, 'is_manager') else False)

        if contractor_areas:
            panels = MDBPanel.query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == current_user.contractor_id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            ).all()
        else:
            panels = MDBPanel.query.filter(MDBPanel.responsible_contractor_id == current_user.contractor_id).all()
    else:
        # للمستخدمين الآخرين (مسجل قراءات، مستخدم عادي)
        assigned_areas = current_user.get_assigned_areas()
        if assigned_areas:  # إذا كانت هناك مناطق مخصصة
            panels = MDBPanel.query.filter(MDBPanel.area_name.in_(assigned_areas)).all()
        else:
            # إذا لم تكن هناك مناطق مخصصة، يمكن رؤية جميع اللوحات
            panels = MDBPanel.query.all()

    return render_template('search_panel_for_reading.html',
                           panels=panels,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة إضافة قراءة يدوية
@readings_bp.route('/panels/<int:panel_id>/add-manual-reading')
@login_required
def add_manual_reading_view(panel_id):
    from app import get_setting

    panel = MDBPanel.query.get_or_404(panel_id)

    # التحقق من الصلاحيات - المدير له صلاحية الوصول لجميع اللوحات
    if current_user.role not in ['admin'] and current_user.role in ['contractor', 'regular_contractor']:
        has_access = False
        contractor_id = current_user.contractor_id

        # التحقق مما إذا كان المقاول مسؤولاً عن اللوحة
        if panel.responsible_contractor_id == contractor_id:
            has_access = True
        else:
            # التحقق مما إذا كانت اللوحة في منطقة مسؤولية المقاول
            from app import get_contractor_areas
            contractor_areas = get_contractor_areas(contractor_id, hasattr(current_user, 'is_manager') and current_user.is_manager)

            if panel.area_name in contractor_areas:
                has_access = True

        if not has_access:
            flash('ليس لديك صلاحية للوصول إلى هذه اللوحة', 'danger')
            return redirect(url_for('contractor_dashboard'))

    # الحصول على وحدات القياس النشطة
    units = MeasurementUnit.query.filter_by(is_active=True).all()

    # الحصول على القوائم المنسدلة المناسبة للمستخدم
    if current_user.role == 'admin':
        visibility_options = ['all', 'admin']
    elif current_user.role == 'contractor':
        visibility_options = ['all', 'contractor']
    else:
        visibility_options = ['all']

    dropdowns = DropdownList.query.filter(
        DropdownList.is_active == True,
        DropdownList.visibility.in_(visibility_options)
    ).all()

    # التحقق من وجود معاملات البدء التلقائي
    auto_start = request.args.get('auto_start') == 'true'
    issue_id = request.args.get('issue_id')
    request_id = request.args.get('request_id')

    # الحصول على القيم الافتراضية من الإعدادات
    default_values = {
        'current': get_setting('default_current', '0'),
        'voltage': get_setting('default_voltage', '220'),
        'power': get_setting('default_power', '0'),
        'energy': get_setting('default_energy', '0'),
        'power_factor': get_setting('default_power_factor', '0.9'),
        'frequency': get_setting('default_frequency', '60'),
        'breaker_capacity': str(panel.breaker_capacity) if panel.breaker_capacity else get_setting('default_breaker_capacity', '0')
    }

    # طباعة سعة القاطع للتأكد من وجودها
    print(f"سعة القاطع للوحة {panel.mdb}: {panel.breaker_capacity}")

    # التحقق من إعداد استخدام النظام الثلاثي
    from app import get_setting
    use_three_phase = get_setting('use_three_phase', 'true') == 'true'

    # اختيار القالب المناسب بناءً على نوع المستخدم وإعداد النظام
    if current_user.role == 'regular_contractor':
        template = 'add_manual_reading_regular_contractor.html'
    elif current_user.role == 'readings_recorder':
        template = 'add_manual_reading_recorder.html'
    else:
        template = 'add_manual_reading_three_phase.html' if use_three_phase else 'add_manual_reading.html'

    return render_template(template,
                           panel=panel,
                           units=units,
                           dropdowns=dropdowns,
                           default_values=default_values,
                           auto_start=auto_start,
                           issue_id=issue_id,
                           request_id=request_id,
                           now=datetime.datetime.now(),
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة قراءة يدوية
@readings_bp.route('/panels/<int:panel_id>/add-manual-reading', methods=['POST'])
@login_required
def add_manual_reading(panel_id):
    from app import get_setting

    panel = MDBPanel.query.get_or_404(panel_id)

    # التحقق من الصلاحيات - المدير له صلاحية الوصول لجميع اللوحات
    if current_user.role not in ['admin'] and current_user.role in ['contractor', 'regular_contractor']:
        has_access = False
        contractor_id = current_user.contractor_id

        # التحقق مما إذا كان المقاول مسؤولاً عن اللوحة
        if panel.responsible_contractor_id == contractor_id:
            has_access = True
        else:
            # التحقق مما إذا كانت اللوحة في منطقة مسؤولية المقاول
            from app import get_contractor_areas
            contractor_areas = get_contractor_areas(contractor_id, current_user.is_manager if hasattr(current_user, 'is_manager') else False)

            if panel.area_name in contractor_areas:
                has_access = True

        if not has_access:
            flash('ليس لديك صلاحية للوصول إلى هذه اللوحة', 'danger')
            return redirect(url_for('contractor_dashboard'))

    # الحصول على بيانات النموذج
    timestamp_str = request.form.get('timestamp')
    reading_type = request.form.get('reading_type', 'manual')

    # تحويل التاريخ والوقت إلى كائن datetime
    try:
        timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%dT%H:%M')
    except ValueError:
        flash('صيغة التاريخ والوقت غير صحيحة', 'danger')
        return redirect(url_for('readings.add_manual_reading_view', panel_id=panel_id))

    # التحقق من إعداد استخدام النظام الثلاثي
    use_three_phase = get_setting('use_three_phase', 'true') == 'true'

    # الحصول على القراءات الأساسية مع استخدام القيم الافتراضية
    if use_three_phase:
        # قراءات التيار
        current_l1 = request.form.get('current_l1', get_setting('default_current', '0'))
        current_l2 = request.form.get('current_l2', get_setting('default_current', '0'))
        current_l3 = request.form.get('current_l3', get_setting('default_current', '0'))
        current_unit_id = request.form.get('current_unit_id')

        # التحقق من نوع المستخدم لتحديد كيفية التعامل مع القيم الأخرى
        if current_user.role == 'readings_recorder' or current_user.role == 'regular_contractor':
            # لمسجل القراءات والمقاول العادي، استخدم قيم فارغة للجهد والقدرة
            voltage_l1_l2 = voltage_l2_l3 = voltage_l3_l1 = voltage_l1_n = voltage_l2_n = voltage_l3_n = None
            voltage_unit_id = request.form.get('voltage_unit_id')

            active_power_l1 = active_power_l2 = active_power_l3 = active_power_total = None
            active_power_unit_id = request.form.get('active_power_unit_id')

            apparent_power_l1 = apparent_power_l2 = apparent_power_l3 = apparent_power_total = None
            apparent_power_unit_id = request.form.get('apparent_power_unit_id')

            reactive_power_l1 = reactive_power_l2 = reactive_power_l3 = reactive_power_total = None
            reactive_power_unit_id = request.form.get('reactive_power_unit_id')

            power_factor_l1 = power_factor_l2 = power_factor_l3 = power_factor_total = None

            # تعيين القيم الأساسية كفارغة أيضًا
            voltage = power = power_factor = None
            power_unit_id = None

            # حساب متوسط التيار للتوافق مع الإصدارات السابقة
            current = request.form.get('current', None)
            if not current:
                try:
                    current = str((float(current_l1) + float(current_l2) + float(current_l3)) / 3)
                except (ValueError, TypeError):
                    current = get_setting('default_current', '0')
        else:
            # للمستخدمين الآخرين، استخدم القيم الافتراضية
            # قراءات الجهد
            voltage_l1_l2 = request.form.get('voltage_l1_l2', get_setting('default_voltage', '220'))
            voltage_l2_l3 = request.form.get('voltage_l2_l3', get_setting('default_voltage', '220'))
            voltage_l3_l1 = request.form.get('voltage_l3_l1', get_setting('default_voltage', '220'))
            voltage_l1_n = request.form.get('voltage_l1_n', get_setting('default_voltage', '220'))
            voltage_l2_n = request.form.get('voltage_l2_n', get_setting('default_voltage', '220'))
            voltage_l3_n = request.form.get('voltage_l3_n', get_setting('default_voltage', '220'))
            voltage_unit_id = request.form.get('voltage_unit_id')

            # قراءات القدرة الفعالة
            active_power_l1 = request.form.get('active_power_l1', get_setting('default_power', '0'))
            active_power_l2 = request.form.get('active_power_l2', get_setting('default_power', '0'))
            active_power_l3 = request.form.get('active_power_l3', get_setting('default_power', '0'))
            active_power_total = request.form.get('active_power_total', get_setting('default_power', '0'))
            active_power_unit_id = request.form.get('active_power_unit_id')

            # قراءات القدرة الظاهرية
            apparent_power_l1 = request.form.get('apparent_power_l1', get_setting('default_power', '0'))
            apparent_power_l2 = request.form.get('apparent_power_l2', get_setting('default_power', '0'))
            apparent_power_l3 = request.form.get('apparent_power_l3', get_setting('default_power', '0'))
            apparent_power_total = request.form.get('apparent_power_total', get_setting('default_power', '0'))
            apparent_power_unit_id = request.form.get('apparent_power_unit_id')

            # قراءات القدرة غير الفعالة
            reactive_power_l1 = request.form.get('reactive_power_l1', get_setting('default_power', '0'))
            reactive_power_l2 = request.form.get('reactive_power_l2', get_setting('default_power', '0'))
            reactive_power_l3 = request.form.get('reactive_power_l3', get_setting('default_power', '0'))
            reactive_power_total = request.form.get('reactive_power_total', get_setting('default_power', '0'))
            reactive_power_unit_id = request.form.get('reactive_power_unit_id')

            # قراءات معامل القدرة
            power_factor_l1 = request.form.get('power_factor_l1', get_setting('default_power_factor', '0.9'))
            power_factor_l2 = request.form.get('power_factor_l2', get_setting('default_power_factor', '0.9'))
            power_factor_l3 = request.form.get('power_factor_l3', get_setting('default_power_factor', '0.9'))
            power_factor_total = request.form.get('power_factor_total', get_setting('default_power_factor', '0.9'))

            # حساب المتوسطات للتوافق مع الإصدارات السابقة
            current = request.form.get('current', None)
            if not current:
                try:
                    current = str((float(current_l1) + float(current_l2) + float(current_l3)) / 3)
                except (ValueError, TypeError):
                    current = get_setting('default_current', '0')

            voltage = request.form.get('voltage', None)
            if not voltage:
                try:
                    voltage = str((float(voltage_l1_l2) + float(voltage_l2_l3) + float(voltage_l3_l1)) / 3)
                except (ValueError, TypeError):
                    voltage = get_setting('default_voltage', '220')

            power = request.form.get('power', None)
            if not power:
                try:
                    power = active_power_total
                except (ValueError, TypeError):
                    power = get_setting('default_power', '0')

            power_factor = request.form.get('power_factor', None)
            if not power_factor:
                try:
                    power_factor = power_factor_total
                except (ValueError, TypeError):
                    power_factor = get_setting('default_power_factor', '0.9')

            power_unit_id = active_power_unit_id
    else:
        # القراءات الأحادية
        current = request.form.get('current', get_setting('default_current', '0'))
        current_unit_id = request.form.get('current_unit_id')

        voltage = request.form.get('voltage', get_setting('default_voltage', '220'))
        voltage_unit_id = request.form.get('voltage_unit_id')

        power = request.form.get('power', get_setting('default_power', '0'))
        power_unit_id = request.form.get('power_unit_id')

        power_factor = request.form.get('power_factor', get_setting('default_power_factor', '0.9'))

        # تعيين قيم فارغة للقراءات الثلاثية
        current_l1 = current_l2 = current_l3 = None
        voltage_l1_l2 = voltage_l2_l3 = voltage_l3_l1 = voltage_l1_n = voltage_l2_n = voltage_l3_n = None
        active_power_l1 = active_power_l2 = active_power_l3 = active_power_total = None
        apparent_power_l1 = apparent_power_l2 = apparent_power_l3 = apparent_power_total = None
        reactive_power_l1 = reactive_power_l2 = reactive_power_l3 = reactive_power_total = None
        power_factor_l1 = power_factor_l2 = power_factor_l3 = power_factor_total = None
        active_power_unit_id = apparent_power_unit_id = reactive_power_unit_id = None

    energy = request.form.get('energy', get_setting('default_energy', '0'))
    energy_unit_id = request.form.get('energy_unit_id')

    frequency = request.form.get('frequency', get_setting('default_frequency', '60'))

    # الحصول على سعة القاطع
    breaker_capacity = request.form.get('breaker_capacity', get_setting('default_breaker_capacity', '0'))

    notes = request.form.get('notes')
    mutawif_name = request.form.get('mutawif_name')

    # تحويل القيم إلى أرقام إذا كانت موجودة
    current = float(current) if current else None
    voltage = float(voltage) if voltage else None
    power = float(power) if power else None
    energy = float(energy) if energy else None
    power_factor = float(power_factor) if power_factor else None
    frequency = float(frequency) if frequency else None
    breaker_capacity = float(breaker_capacity) if breaker_capacity else None

    # تحديد حالة التيار بناءً على نسبة الحمل
    current_status = 'normal'
    load_percentage = None
    warning_threshold = float(get_setting('warning_threshold', '70'))
    danger_threshold = float(get_setting('danger_threshold', '80'))

    # حساب نسبة الحمل لكل فاز على حدة
    if use_three_phase and breaker_capacity is not None and breaker_capacity > 0:
        # تحويل القيم إلى أرقام
        current_l1_val = float(current_l1) if current_l1 else 0
        current_l2_val = float(current_l2) if current_l2 else 0
        current_l3_val = float(current_l3) if current_l3 else 0

        # حساب نسبة الحمل لكل فاز
        load_percentage_l1 = (current_l1_val / breaker_capacity) * 100 if current_l1_val else 0
        load_percentage_l2 = (current_l2_val / breaker_capacity) * 100 if current_l2_val else 0
        load_percentage_l3 = (current_l3_val / breaker_capacity) * 100 if current_l3_val else 0

        # استخدام أعلى نسبة حمل من الفازات الثلاثة
        load_percentage = max(load_percentage_l1, load_percentage_l2, load_percentage_l3)

        # تحديد حالة التيار بناءً على نسبة الحمل
        if load_percentage >= danger_threshold:
            current_status = 'danger'
        elif load_percentage >= warning_threshold:
            current_status = 'warning'

        print(f"نسبة الحمل للفاز 1: {load_percentage_l1}%")
        print(f"نسبة الحمل للفاز 2: {load_percentage_l2}%")
        print(f"نسبة الحمل للفاز 3: {load_percentage_l3}%")
        print(f"أعلى نسبة حمل: {load_percentage}%")
        print(f"حالة التيار: {current_status}")

    elif current is not None and breaker_capacity is not None and breaker_capacity > 0:
        # حساب نسبة الحمل للنظام الأحادي
        load_percentage = (current / breaker_capacity) * 100

        # تحديد حالة التيار بناءً على نسبة الحمل
        if load_percentage >= danger_threshold:
            current_status = 'danger'
        elif load_percentage >= warning_threshold:
            current_status = 'warning'

        print(f"نسبة الحمل: {load_percentage}%")
        print(f"حالة التيار: {current_status}")

    # تحويل القيم الثلاثية إلى أرقام إذا كانت موجودة
    if use_three_phase:
        current_l1 = float(current_l1) if current_l1 else None
        current_l2 = float(current_l2) if current_l2 else None
        current_l3 = float(current_l3) if current_l3 else None

        voltage_l1_l2 = float(voltage_l1_l2) if voltage_l1_l2 else None
        voltage_l2_l3 = float(voltage_l2_l3) if voltage_l2_l3 else None
        voltage_l3_l1 = float(voltage_l3_l1) if voltage_l3_l1 else None
        voltage_l1_n = float(voltage_l1_n) if voltage_l1_n else None
        voltage_l2_n = float(voltage_l2_n) if voltage_l2_n else None
        voltage_l3_n = float(voltage_l3_n) if voltage_l3_n else None

        active_power_l1 = float(active_power_l1) if active_power_l1 else None
        active_power_l2 = float(active_power_l2) if active_power_l2 else None
        active_power_l3 = float(active_power_l3) if active_power_l3 else None
        active_power_total = float(active_power_total) if active_power_total else None

        apparent_power_l1 = float(apparent_power_l1) if apparent_power_l1 else None
        apparent_power_l2 = float(apparent_power_l2) if apparent_power_l2 else None
        apparent_power_l3 = float(apparent_power_l3) if apparent_power_l3 else None
        apparent_power_total = float(apparent_power_total) if apparent_power_total else None

        reactive_power_l1 = float(reactive_power_l1) if reactive_power_l1 else None
        reactive_power_l2 = float(reactive_power_l2) if reactive_power_l2 else None
        reactive_power_l3 = float(reactive_power_l3) if reactive_power_l3 else None
        reactive_power_total = float(reactive_power_total) if reactive_power_total else None

        power_factor_l1 = float(power_factor_l1) if power_factor_l1 else None
        power_factor_l2 = float(power_factor_l2) if power_factor_l2 else None
        power_factor_l3 = float(power_factor_l3) if power_factor_l3 else None
        power_factor_total = float(power_factor_total) if power_factor_total else None

    # الحصول على حالة اللوحة
    panel_status = request.form.get('panel_status')

    # إنشاء قراءة يدوية جديدة
    reading = ManualReading(
        panel_id=panel_id,
        timestamp=timestamp,
        reading_type=reading_type,
        is_three_phase=use_three_phase,
        panel_status=panel_status,  # إضافة حالة اللوحة

        # القيم الأساسية (للتوافق مع الإصدارات السابقة)
        current=current,
        current_unit_id=current_unit_id if current is not None else None,
        voltage=voltage,
        voltage_unit_id=voltage_unit_id if voltage is not None else None,
        power=power,
        power_unit_id=power_unit_id if power is not None else None,
        energy=energy,
        energy_unit_id=energy_unit_id if energy is not None else None,
        power_factor=power_factor,
        frequency=frequency,
        breaker_capacity=breaker_capacity,
        current_status=current_status,  # إضافة حالة التيار
        load_percentage=load_percentage,  # إضافة نسبة الحمل

        # القيم الثلاثية
        current_l1=current_l1,
        current_l2=current_l2,
        current_l3=current_l3,

        voltage_l1_l2=voltage_l1_l2,
        voltage_l2_l3=voltage_l2_l3,
        voltage_l3_l1=voltage_l3_l1,
        voltage_l1_n=voltage_l1_n,
        voltage_l2_n=voltage_l2_n,
        voltage_l3_n=voltage_l3_n,

        active_power_l1=active_power_l1,
        active_power_l2=active_power_l2,
        active_power_l3=active_power_l3,
        active_power_total=active_power_total,
        active_power_unit_id=active_power_unit_id if active_power_total is not None else None,

        apparent_power_l1=apparent_power_l1,
        apparent_power_l2=apparent_power_l2,
        apparent_power_l3=apparent_power_l3,
        apparent_power_total=apparent_power_total,
        apparent_power_unit_id=apparent_power_unit_id if apparent_power_total is not None else None,

        reactive_power_l1=reactive_power_l1,
        reactive_power_l2=reactive_power_l2,
        reactive_power_l3=reactive_power_l3,
        reactive_power_total=reactive_power_total,
        reactive_power_unit_id=reactive_power_unit_id if reactive_power_total is not None else None,

        power_factor_l1=power_factor_l1,
        power_factor_l2=power_factor_l2,
        power_factor_l3=power_factor_l3,
        power_factor_total=power_factor_total,

        notes=notes,
        mutawif_name=mutawif_name,
        created_by=current_user.id,
        created_at=datetime.datetime.now()
    )

    db.session.add(reading)

    # الحصول على القيم من القوائم المنسدلة
    dropdown_values = {}
    for key, value in request.form.items():
        if key.startswith('dropdown_') and value:
            dropdown_id = int(key.replace('dropdown_', ''))
            dropdown_values[dropdown_id] = value

    # حفظ القيم في قاعدة البيانات
    if dropdown_values:
        reading.dropdown_values = json.dumps(dropdown_values)

    # تحديث سعة القاطع في اللوحة إذا كانت غير محددة
    if breaker_capacity and (panel.breaker_capacity is None or panel.breaker_capacity == 0):
        panel.breaker_capacity = breaker_capacity
        db.session.add(panel)

    # التحقق من وجود معاملات البدء التلقائي وتحديث الحالة
    auto_start = request.form.get('auto_start') == 'true'
    issue_id = request.form.get('issue_id')
    request_id = request.form.get('request_id')

    try:
        db.session.commit()

        # تحديث حالة البلاغ أو طلب الفحص إذا كان هناك بدء تلقائي
        if auto_start:
            from models import Issue, InspectionRequest
            current_time = datetime.datetime.now()

            if issue_id:
                issue = Issue.query.get(issue_id)
                if issue and issue.status == 'مفتوح':
                    issue.status = 'قيد المعالجة'
                    issue.started_at = current_time
                    if issue.created_at:
                        issue.processing_time = int((current_time - issue.created_at).total_seconds() / 60)
                    db.session.add(issue)

            elif request_id:
                inspection_request = InspectionRequest.query.get(request_id)
                if inspection_request and inspection_request.status == 'جديد':
                    inspection_request.status = 'قيد التنفيذ'
                    inspection_request.started_at = current_time
                    if inspection_request.created_at:
                        inspection_request.processing_time = int((current_time - inspection_request.created_at).total_seconds() / 60)
                    db.session.add(inspection_request)

            db.session.commit()

        flash('تم إضافة القراءة اليدوية بنجاح', 'success')

        # إضافة رسالة إضافية إذا تم تحديث الحالة
        if auto_start and (issue_id or request_id):
            flash('تم بدء التنفيذ تلقائياً', 'info')

    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حفظ القراءة: {str(e)}', 'danger')

    return redirect(url_for('readings.all_readings'))

# عرض تفاصيل قراءة يدوية
@readings_bp.route('/manual-readings/<int:reading_id>')
@login_required
def view_manual_reading(reading_id):
    reading = ManualReading.query.get_or_404(reading_id)
    panel = MDBPanel.query.get_or_404(reading.panel_id)

    # التحقق من الصلاحيات - المدير له صلاحية الوصول لجميع القراءات
    if current_user.role not in ['admin'] and current_user.role in ['contractor', 'regular_contractor']:
        has_access = False
        contractor_id = current_user.contractor_id

        # التحقق مما إذا كان المقاول مسؤولاً عن اللوحة
        if panel.responsible_contractor_id == contractor_id:
            has_access = True
        else:
            # التحقق مما إذا كانت اللوحة في منطقة مسؤولية المقاول
            from app import get_contractor_areas
            contractor_areas = get_contractor_areas(contractor_id, current_user.is_manager if hasattr(current_user, 'is_manager') else False)

            if panel.area_name in contractor_areas:
                has_access = True

        if not has_access:
            flash('ليس لديك صلاحية للوصول إلى هذه القراءة', 'danger')
            return redirect(url_for('contractor_dashboard'))

    # الحصول على قيم القوائم المنسدلة
    dropdown_values = {}
    if reading.dropdown_values:
        try:
            dropdown_values_dict = json.loads(reading.dropdown_values)
            for dropdown_id, value in dropdown_values_dict.items():
                dropdown = DropdownList.query.get(int(dropdown_id))
                if dropdown:
                    item = DropdownItem.query.filter_by(dropdown_id=dropdown.id, value=value).first()
                    if item:
                        dropdown_values[dropdown.display_name] = item.display_text
        except json.JSONDecodeError:
            pass

    # اختيار القالب المناسب بناءً على نوع المستخدم ونوع القراءة
    if current_user.role == 'regular_contractor':
        template = 'view_manual_reading_regular_contractor.html'
    elif current_user.role == 'readings_recorder':
        template = 'view_manual_reading_recorder.html'
    else:
        template = 'view_manual_reading_three_phase.html' if reading.is_three_phase else 'view_manual_reading.html'

    return render_template(template,
                           reading=reading,
                           panel=panel,
                           dropdown_values=dropdown_values,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# حذف قراءة يدوية
@readings_bp.route('/manual-readings/<int:reading_id>/delete')
@login_required
def delete_manual_reading(reading_id):
    reading = ManualReading.query.get_or_404(reading_id)
    panel_id = reading.panel_id

    # التحقق من أن المستخدم هو من أضاف القراءة أو أنه مدير
    if current_user.role != 'admin' and reading.created_by != current_user.id:
        flash('ليس لديك صلاحية لحذف هذه القراءة', 'danger')
        return redirect(url_for('panel_details', panel_id=panel_id))

    db.session.delete(reading)
    db.session.commit()

    # التحقق من مصدر الطلب (من صفحة التفاصيل أو من صفحة جميع القراءات)
    referrer = request.referrer
    if referrer and 'all-readings' in referrer:
        flash('تم حذف القراءة اليدوية بنجاح', 'success')
        return redirect(url_for('readings.all_readings'))

    flash('تم حذف القراءة اليدوية بنجاح', 'success')
    return redirect(url_for('panel_details', panel_id=panel_id))

# عرض جميع القراءات اليدوية
@readings_bp.route('/all-readings')
@login_required
def all_readings():
    # الحصول على معلمات التصفية
    page = request.args.get('page', 1, type=int)
    panel_id = request.args.get('panel_id')
    area = request.args.get('area')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status_filter')

    # عدد العناصر في الصفحة الواحدة
    per_page = 20

    # بناء الاستعلام الأساسي
    query = ManualReading.query.join(MDBPanel)

    # إذا كان المستخدم مقاول أو مقاول عادي، قم بتصفية القراءات حسب مناطق المسؤولية
    if current_user.role in ['contractor', 'regular_contractor']:
        from app import get_contractor_areas
        contractor_areas = get_contractor_areas(current_user.contractor_id, current_user.is_manager if hasattr(current_user, 'is_manager') else False)

        # تصفية اللوحات حسب المقاول المسؤول أو المنطقة
        if contractor_areas:
            query = query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == current_user.contractor_id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            )
        else:
            # إذا لم تكن هناك مناطق محددة، استخدم فقط المقاول المسؤول
            query = query.filter(MDBPanel.responsible_contractor_id == current_user.contractor_id)
    elif current_user.role != 'admin':
        # فلترة المستخدمين الآخرين حسب المناطق المخصصة
        assigned_areas = current_user.get_assigned_areas()
        if assigned_areas:  # إذا كانت هناك مناطق مخصصة
            query = query.filter(MDBPanel.area_name.in_(assigned_areas))

    # تطبيق التصفية
    if panel_id:
        query = query.filter(ManualReading.panel_id == panel_id)

    if area:
        query = query.filter(MDBPanel.area_name == area)

    if start_date:
        start_datetime = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        query = query.filter(ManualReading.timestamp >= start_datetime)

    if end_date:
        end_datetime = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(ManualReading.timestamp <= end_datetime)

    # تطبيق فلتر الحالة الكهربائية
    if status_filter:
        query = query.filter(ManualReading.current_status == status_filter)

    # ترتيب النتائج حسب التاريخ (الأحدث أولاً)
    query = query.order_by(desc(ManualReading.timestamp))

    # الحصول على إجمالي عدد النتائج
    total_count = query.count()
    total_pages = (total_count + per_page - 1) // per_page

    # الحصول على النتائج للصفحة الحالية
    readings = query.paginate(page=page, per_page=per_page, error_out=False).items

    # الحصول على قائمة اللوحات والمناطق للتصفية
    if current_user.role == 'admin':
        panels = MDBPanel.query.all()
        areas = [area[0] for area in db.session.query(MDBPanel.area_name).distinct() if area[0]]
    else:
        # للمقاول، احصل فقط على اللوحات في مناطق مسؤوليته
        from app import get_contractor_areas
        contractor_areas = get_contractor_areas(current_user.contractor_id, current_user.is_manager)

        if contractor_areas:
            panels = MDBPanel.query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == current_user.contractor_id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            ).all()
            areas = contractor_areas
        else:
            panels = MDBPanel.query.filter(MDBPanel.responsible_contractor_id == current_user.contractor_id).all()
            areas = [panel.area_name for panel in panels]
            # إزالة التكرار
            areas = list(set(areas))

    # إعداد معلومات التصفية الحالية
    current_filters = None
    if panel_id or area or start_date or end_date or status_filter:
        current_filters = {
            'panel_id': panel_id,
            'panel_name': MDBPanel.query.get(panel_id).mdb if panel_id else None,
            'area': area,
            'start_date': start_date,
            'end_date': end_date,
            'status_filter': status_filter
        }

    return render_template('all_manual_readings.html',
                           readings=readings,
                           panels=panels,
                           areas=areas,
                           page=page,
                           per_page=per_page,
                           total_pages=total_pages,
                           current_filters=current_filters,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تصدير القراءات اليدوية
@readings_bp.route('/export-readings')
@login_required
def export_readings():
    # الحصول على معلمات التصفية
    panel_id = request.args.get('panel_id')
    area = request.args.get('area')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    export_format = request.args.get('format', 'excel')

    # بناء الاستعلام الأساسي
    query = ManualReading.query.join(MDBPanel)

    # إذا كان المستخدم مقاول أو مقاول عادي، قم بتصفية القراءات حسب مناطق المسؤولية
    if current_user.role in ['contractor', 'regular_contractor']:
        from app import get_contractor_areas
        contractor_areas = get_contractor_areas(current_user.contractor_id, current_user.is_manager if hasattr(current_user, 'is_manager') else False)

        # تصفية اللوحات حسب المقاول المسؤول أو المنطقة
        if contractor_areas:
            query = query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == current_user.contractor_id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            )
        else:
            # إذا لم تكن هناك مناطق محددة، استخدم فقط المقاول المسؤول
            query = query.filter(MDBPanel.responsible_contractor_id == current_user.contractor_id)

    # تطبيق التصفية
    if panel_id:
        query = query.filter(ManualReading.panel_id == panel_id)

    if area:
        query = query.filter(MDBPanel.area_name == area)

    if start_date:
        start_datetime = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        query = query.filter(ManualReading.timestamp >= start_datetime)

    if end_date:
        end_datetime = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(ManualReading.timestamp <= end_datetime)

    # ترتيب النتائج حسب التاريخ (الأحدث أولاً)
    readings = query.order_by(desc(ManualReading.timestamp)).all()

    # إنشاء DataFrame للتصدير
    data = []
    for reading in readings:
        user = User.query.get(reading.created_by) if reading.created_by else None

        # استخدام مجموع الأطوار الثلاثة
        total_current = None
        if reading.is_three_phase and reading.current_l1 is not None and reading.current_l2 is not None and reading.current_l3 is not None:
            total_current = reading.current_l1 + reading.current_l2 + reading.current_l3
        else:
            total_current = reading.current

        # حساب نسبة الحمل
        load_percentage = None
        if total_current is not None and reading.breaker_capacity is not None and reading.breaker_capacity > 0:
            load_percentage = (total_current / reading.breaker_capacity) * 100

        row = {
            'اللوحة': reading.panel.mdb,
            'رقم تاج ماكسيمو': reading.panel.maximo_tag if reading.panel.maximo_tag else '-',
            'المنطقة': reading.panel.area_name,
            'سعة القاطع (A)': reading.breaker_capacity if reading.breaker_capacity is not None else '-',
            'تاريخ القراءة': reading.timestamp.strftime('%Y-%m-%d %H:%M'),
            'التيار الكلي (A)': f"{total_current:.2f}" if total_current is not None else '-',
            'نسبة الحمل (%)': f"{load_percentage:.2f}" if load_percentage is not None else '-',
            'حالة التيار': reading.current_status if hasattr(reading, 'current_status') and reading.current_status else '-',
        }

        # إضافة قيم التيار للأطوار الثلاثة إذا كانت متوفرة
        if reading.is_three_phase:
            row.update({
                'التيار L1 (A)': f"{reading.current_l1:.2f}" if reading.current_l1 is not None else '-',
                'التيار L2 (A)': f"{reading.current_l2:.2f}" if reading.current_l2 is not None else '-',
                'التيار L3 (A)': f"{reading.current_l3:.2f}" if reading.current_l3 is not None else '-',
            })

        # إضافة باقي المعلومات
        row.update({
            'الجهد': f"{reading.voltage} {reading.voltage_unit.name if reading.voltage_unit else 'V'}" if reading.voltage is not None else '-',
            'القدرة': f"{reading.power} {reading.power_unit.name if reading.power_unit else 'W'}" if reading.power is not None else '-',
            'معامل القدرة': reading.power_factor if reading.power_factor is not None else '-',
            'التردد': f"{reading.frequency} Hz" if reading.frequency is not None else '-',
            'اسم المطوف': reading.mutawif_name if hasattr(reading, 'mutawif_name') and reading.mutawif_name else '-',
            'المستخدم': user.name if user else '-',
            'ملاحظات': reading.notes if reading.notes else '-'
        })

        # إضافة قيم القوائم المنسدلة
        if reading.dropdown_values:
            try:
                dropdown_values = json.loads(reading.dropdown_values)
                for dropdown_id, value in dropdown_values.items():
                    dropdown = DropdownList.query.get(int(dropdown_id))
                    if dropdown:
                        item = DropdownItem.query.filter_by(dropdown_id=dropdown.id, value=value).first()
                        if item:
                            row[dropdown.display_name] = item.display_text
            except json.JSONDecodeError:
                pass

        data.append(row)

    # إنشاء DataFrame
    df = pd.DataFrame(data)

    # تصدير البيانات
    if export_format == 'excel':
        # إنشاء ملف Excel في الذاكرة
        output = io.BytesIO()

        # إنشاء مصنف Excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "القراءات اليدوية"

        # إضافة العناوين
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=column_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # إضافة البيانات
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')

        # تعديل عرض الأعمدة
        for col_idx in range(1, len(df.columns) + 1):
            worksheet.column_dimensions[chr(64 + col_idx)].width = 20

        # حفظ الملف
        workbook.save(output)
        output.seek(0)

        # إرسال الملف للتنزيل
        filename = f"manual_readings_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # إذا وصلنا إلى هنا، فهناك خطأ في تنسيق التصدير
    flash('تنسيق التصدير غير صالح', 'danger')
    return redirect(url_for('readings.all_readings'))

# تصدير قائمة اللوحات ذات الحمل العالي
@readings_bp.route('/export-high-load-panels')
@login_required
def export_high_load_panels():
    # الحصول على معلمات التصدير
    limit = request.args.get('limit', '50')
    export_format = request.args.get('format', 'excel')

    try:
        limit = int(limit)
    except ValueError:
        limit = 50

    # الحصول على عتبات التحذير والخطر
    from app import get_setting
    warning_threshold = float(get_setting('warning_threshold', '70'))
    danger_threshold = float(get_setting('danger_threshold', '80'))

    # بناء الاستعلام الأساسي للحصول على آخر قراءة لكل لوحة
    subquery = db.session.query(
        ManualReading.panel_id,
        db.func.max(ManualReading.timestamp).label('max_timestamp')
    ).group_by(ManualReading.panel_id).subquery()

    query = db.session.query(ManualReading).join(
        subquery,
        db.and_(
            ManualReading.panel_id == subquery.c.panel_id,
            ManualReading.timestamp == subquery.c.max_timestamp
        )
    ).join(MDBPanel)

    # إذا كان المستخدم مقاول أو مقاول عادي، قم بتصفية القراءات حسب مناطق المسؤولية
    if current_user.role in ['contractor', 'regular_contractor']:
        from app import get_contractor_areas
        contractor_areas = get_contractor_areas(current_user.contractor_id, current_user.is_manager if hasattr(current_user, 'is_manager') else False)

        # تصفية اللوحات حسب المقاول المسؤول أو المنطقة
        if contractor_areas:
            query = query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == current_user.contractor_id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            )
        else:
            # إذا لم تكن هناك مناطق محددة، استخدم فقط المقاول المسؤول
            query = query.filter(MDBPanel.responsible_contractor_id == current_user.contractor_id)

    # الحصول على جميع القراءات
    all_readings = query.all()

    # حساب نسبة الحمل لكل قراءة
    readings_with_load = []
    for reading in all_readings:
        # حساب متوسط التيار للأطوار الثلاثة
        avg_current = None
        if reading.is_three_phase and reading.current_l1 is not None and reading.current_l2 is not None and reading.current_l3 is not None:
            avg_current = (reading.current_l1 + reading.current_l2 + reading.current_l3) / 3
        else:
            avg_current = reading.current

        # حساب نسبة الحمل
        load_percentage = None
        if avg_current is not None and reading.breaker_capacity is not None and reading.breaker_capacity > 0:
            load_percentage = (avg_current / reading.breaker_capacity) * 100

            # تحديد حالة التيار
            current_status = 'normal'
            if load_percentage >= danger_threshold:
                current_status = 'danger'
            elif load_percentage >= warning_threshold:
                current_status = 'warning'

            readings_with_load.append({
                'reading': reading,
                'avg_current': avg_current,
                'load_percentage': load_percentage,
                'current_status': current_status
            })

    # ترتيب القراءات حسب نسبة الحمل (من الأعلى إلى الأدنى)
    readings_with_load.sort(key=lambda x: x['load_percentage'], reverse=True)

    # تحديد عدد اللوحات المطلوبة
    readings_with_load = readings_with_load[:limit]

    # إنشاء DataFrame للتصدير
    data = []
    for item in readings_with_load:
        reading = item['reading']
        avg_current = item['avg_current']
        load_percentage = item['load_percentage']
        current_status = item['current_status']

        user = User.query.get(reading.created_by) if reading.created_by else None

        row = {
            'اللوحة': reading.panel.mdb,
            'رقم تاج ماكسيمو': reading.panel.maximo_tag if reading.panel.maximo_tag else '-',
            'المنطقة': reading.panel.area_name,
            'سعة القاطع (A)': reading.breaker_capacity if reading.breaker_capacity is not None else '-',
            'تاريخ القراءة': reading.timestamp.strftime('%Y-%m-%d %H:%M'),
            'التيار Avg (A)': f"{avg_current:.2f}" if avg_current is not None else '-',
            'نسبة الحمل (%)': f"{load_percentage:.2f}" if load_percentage is not None else '-',
            'حالة التيار': current_status,
        }

        # إضافة قيم التيار للأطوار الثلاثة إذا كانت متوفرة
        if reading.is_three_phase:
            row.update({
                'التيار L1 (A)': f"{reading.current_l1:.2f}" if reading.current_l1 is not None else '-',
                'التيار L2 (A)': f"{reading.current_l2:.2f}" if reading.current_l2 is not None else '-',
                'التيار L3 (A)': f"{reading.current_l3:.2f}" if reading.current_l3 is not None else '-',
            })

        # إضافة باقي المعلومات
        row.update({
            'الجهد': f"{reading.voltage} {reading.voltage_unit.name if reading.voltage_unit else 'V'}" if reading.voltage is not None else '-',
            'القدرة': f"{reading.power} {reading.power_unit.name if reading.power_unit else 'W'}" if reading.power is not None else '-',
            'معامل القدرة': reading.power_factor if reading.power_factor is not None else '-',
            'التردد': f"{reading.frequency} Hz" if reading.frequency is not None else '-',
            'اسم المطوف': reading.mutawif_name if hasattr(reading, 'mutawif_name') and reading.mutawif_name else '-',
            'المستخدم': user.name if user else '-',
            'ملاحظات': reading.notes if reading.notes else '-'
        })

        data.append(row)

    # إنشاء DataFrame
    df = pd.DataFrame(data)

    # تصدير البيانات
    if export_format == 'excel':
        # إنشاء ملف Excel في الذاكرة
        output = io.BytesIO()

        # إنشاء مصنف Excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "اللوحات ذات الحمل العالي"

        # إضافة العناوين
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=column_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # إضافة البيانات
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')

        # تعديل عرض الأعمدة
        for col_idx in range(1, len(df.columns) + 1):
            worksheet.column_dimensions[chr(64 + col_idx)].width = 20

        # حفظ الملف
        workbook.save(output)
        output.seek(0)

        # إرسال الملف للتنزيل
        filename = f"high_load_panels_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # إذا وصلنا إلى هنا، فهناك خطأ في تنسيق التصدير
    flash('تنسيق التصدير غير صالح', 'danger')
    return redirect(url_for('readings.all_readings'))

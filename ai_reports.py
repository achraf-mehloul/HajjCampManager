"""
Blueprint لتوليد التقارير بالذكاء الاصطناعي مع Excel وخرائط ذهنية
"""
import io
import json
import datetime
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
try:
    from flask_wtf.csrf import csrf_exempt
except ImportError:
    def csrf_exempt(f): return f

ai_reports_bp = Blueprint('ai_reports', __name__, url_prefix='/ai-reports')


# ==================== صفحة التقارير ====================

@ai_reports_bp.route('/')
@login_required
def index():
    return render_template('ai_reports/index.html')


# ==================== API توليد التقرير ====================

@ai_reports_bp.route('/api/generate', methods=['POST'])
@csrf_exempt
@login_required
def generate_report():
    """توليد تقرير Excel بالذكاء الاصطناعي مع رسوم بيانية"""
    try:
        data     = request.get_json() or {}
        rep_type = data.get('report_type', 'issues')   # issues / readings / assets
        ai_key   = data.get('ai_key', '')               # مفتاح AI اختياري

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return jsonify({'success': False, 'error': 'مكتبة openpyxl غير مثبتة. شغّل: pip install openpyxl'}), 500

        wb = Workbook()

        if rep_type == 'issues':
            _build_issues_sheet(wb)
        elif rep_type == 'readings':
            _build_readings_sheet(wb)
        elif rep_type == 'assets':
            _build_assets_sheet(wb)
        else:
            _build_issues_sheet(wb)

        # ورقة الملخص الذكي
        _build_summary_sheet(wb, rep_type, ai_key)

        buf = io.BytesIO()
        try:
            wb.save(buf)
        except AttributeError as e:
            # في بعض الحالات القديمة لـ openpyxl، الخلايا المدموجة قد تسبب خطأ
            # مثل: 'MergedCell' object has no attribute 'column_letter'
            # كحل آمن، نفك الدمج ثم نحاول الحفظ مرة أخرى.
            if 'MergedCell' in str(e):
                for sheet in wb.worksheets:
                    try:
                        merged_ranges = list(sheet.merged_cells.ranges)
                        for rng in merged_ranges:
                            sheet.unmerge_cells(str(rng))
                    except Exception:
                        continue
                buf = io.BytesIO()
                wb.save(buf)
            else:
                raise
        buf.seek(0)

        filename = f"AI_Report_{rep_type}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== دوال مساعدة ====================

def _header_style(ws, title, color='1F4E79'):
    from openpyxl.styles import Font, Alignment, PatternFill
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=18, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=color)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    ws.row_dimensions[1].height = 35


def _col_headers(ws, headers, row=2):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2E86C1')
        c.alignment = Alignment(horizontal='center', readingOrder=2)
        c.border = border


def _auto_width(ws):
    """
    تعيين عرض الأعمدة تلقائياً كان يسبب مشاكل مع الخلايا المدموجة (MergedCell).
    لتبسيط الأمور وضمان عمل توليد التقارير بدون أخطاء، نترك العرض الافتراضي للأعمدة.
    يمكن تحسين المظهر لاحقاً إذا لزم الأمر، لكن المهم الآن هو استقرار التوليد.
    """
    return



# ==================== ورقة البلاغات ====================

def _build_issues_sheet(wb):
    try:
        from models import Issue
    except ImportError:
        Issue = None

    from openpyxl.styles import PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference

    ws = wb.active
    ws.title = 'البلاغات'
    _header_style(ws, 'تقرير البلاغات - تحليل ذكي')
    headers = ['رقم البلاغ', 'العنوان', 'النوع', 'الحالة', 'الأولوية', 'التاريخ']
    _col_headers(ws, headers)

    issues = []
    if Issue is not None:
        try:
            issues = Issue.query.order_by(Issue.created_at.desc()).limit(500).all()
        except Exception:
            issues = []

    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_counts = {}
    for i, issue in enumerate(issues, 3):
        row_data = [
            issue.id,
            issue.title or '',
            issue.issue_type or '',
            issue.status or '',
            issue.priority or '',
            str(issue.created_at.date()) if issue.created_at else ''
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = Alignment(horizontal='right', readingOrder=2)
            c.border = border
            if i % 2 == 0:
                c.fill = PatternFill('solid', fgColor='EBF5FB')

        status_counts[issue.status or 'غير محدد'] = status_counts.get(issue.status or 'غير محدد', 0) + 1

    if not issues:
        ws.cell(3, 1, 'لا توجد بلاغات مسجلة في النظام حالياً')

    # ورقة إحصائية للرسم البياني
    ws_stat = wb.create_sheet('إحصائيات البلاغات')
    ws_stat.cell(1, 1, 'الحالة')
    ws_stat.cell(1, 2, 'العدد')

    if status_counts:
        for r, (status, count) in enumerate(status_counts.items(), 2):
            ws_stat.cell(r, 1, status)
            ws_stat.cell(r, 2, count)

        # رسم بياني
        try:
            chart = BarChart()
            chart.title = 'توزيع البلاغات حسب الحالة'
            chart.style = 10
            chart.y_axis.title = 'العدد'
            data_ref = Reference(ws_stat, min_col=2, min_row=1, max_row=len(status_counts)+1)
            cats_ref = Reference(ws_stat, min_col=1, min_row=2, max_row=len(status_counts)+1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, 'H3')
        except Exception:
            pass

    _auto_width(ws)


# ==================== ورقة القراءات ====================

def _build_readings_sheet(wb):
    """تقرير القراءات - يستخدم MDBPanel كبديل إذا لم تكن ElectricalReading موجودة"""
    from openpyxl.styles import Alignment, Border, Side, PatternFill

    ws = wb.active
    ws.title = 'القراءات الكهربائية'
    _header_style(ws, 'تقرير اللوحات الكهربائية', '196F3D')

    ElectricalReading = None
    try:
        from models import ElectricalReading as ER
        ElectricalReading = ER
    except ImportError:
        pass

    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if ElectricalReading is not None:
        _col_headers(ws, ['المعرف', 'اللوحة (ID)', 'متوسط الجهد (V)', 'متوسط التيار (A)', 'الحالة', 'التاريخ'])
        try:
            readings = ElectricalReading.query.order_by(ElectricalReading.timestamp.desc()).limit(500).all()
            for i, r in enumerate(readings, 3):
                row_data = [
                    r.id,
                    r.panel_id,
                    round(r.voltage, 2) if r.voltage else '',
                    round(r.current, 2) if r.current else '',
                    r.current_status or 'normal',
                    str(r.timestamp.strftime('%Y-%m-%d %H:%M')) if r.timestamp else ''
                ]
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.alignment = Alignment(horizontal='right', readingOrder=2)
                    c.border = border
                    if i % 2 == 0:
                        c.fill = PatternFill('solid', fgColor='E9F7EF')
            if not readings:
                ws.cell(3, 1, 'لا توجد قراءات كهربائية مسجلة في النظام')
        except Exception as ex:
            ws.cell(3, 1, f'خطأ في جلب القراءات: {ex}')

    else:
        # استخدام بيانات MDBPanel كبديل
        _col_headers(ws, ['المعرف', 'اسم اللوحة', 'الباركود', 'المنطقة', 'الحالة', 'نوع اللوحة'])
        try:
            from models import MDBPanel
            panels = MDBPanel.query.limit(500).all()
            for i, p in enumerate(panels, 3):
                row_data = [p.id, p.mdb or '', p.maximo_tag or '', p.area_name or '', p.status or 'عامل', p.panel_type or '']
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.alignment = Alignment(horizontal='right', readingOrder=2)
                    c.border = border
                    if i % 2 == 0:
                        c.fill = PatternFill('solid', fgColor='E9F7EF')
            if not panels:
                ws.cell(3, 1, 'لا توجد لوحات مسجلة في النظام')
        except Exception:
            ws.cell(3, 1, 'خطأ في جلب البيانات')

    _auto_width(ws)


# ==================== ورقة الأصول ====================

def _build_assets_sheet(wb):
    """تقرير الأصول - يستخدم MDBPanel كبديل إذا لم يكن Asset موجوداً"""
    from openpyxl.styles import Alignment, Border, Side, PatternFill

    ws = wb.active
    ws.title = 'الأصول والمعدات'
    _header_style(ws, 'تقرير الأصول والمعدات', '7D3C98')

    Asset = None
    try:
        from models import Asset as A
        Asset = A
    except ImportError:
        pass

    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if Asset is not None:
        _col_headers(ws, ['رقم الأصل', 'الوصف', 'النوع', 'المنطقة', 'الدولة', 'الحالة'])
        try:
            assets = Asset.query.limit(500).all()
            for i, a in enumerate(assets, 3):
                row_data = [a.asset_number, a.description, a.asset_type, a.area, a.country, a.status or 'نشط']
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.alignment = Alignment(horizontal='right', readingOrder=2)
                    c.border = border
        except Exception:
            ws.cell(3, 1, 'لا توجد أصول مسجلة')
    else:
        # استخدام MDBPanel كبديل
        _col_headers(ws, ['المعرف', 'اسم اللوحة', 'الباركود', 'المنطقة', 'نوع اللوحة', 'الحالة'])
        try:
            from models import MDBPanel
            panels = MDBPanel.query.limit(500).all()
            for i, p in enumerate(panels, 3):
                row_data = [p.id, p.mdb or '', p.maximo_tag or '', p.area_name or '', p.panel_type or '', p.status or 'عامل']
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.alignment = Alignment(horizontal='right', readingOrder=2)
                    c.border = border
                    if i % 2 == 0:
                        c.fill = PatternFill('solid', fgColor='F5EEF8')
            if not panels:
                ws.cell(3, 1, 'لا توجد لوحات مسجلة في النظام')
        except Exception:
            ws.cell(3, 1, 'خطأ في جلب البيانات')

    _auto_width(ws)


# ==================== ورقة الملخص الذكي ====================

def _build_summary_sheet(wb, rep_type, ai_key=''):
    """ورقة ملخص ذكي احترافية"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    ws = wb.create_sheet('الملخص الذكي')

    # عنوان رئيسي احترافي
    ws.merge_cells('A1:H1')
    ws['A1'] = 'تقرير تحليلي ذكي'
    ws['A1'].font = Font(bold=True, size=20, color='FFFFFF', name='Cairo')
    ws['A1'].fill = PatternFill('solid', fgColor='1A237E')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    ws.row_dimensions[1].height = 45

    # سطر معلومات التقرير
    ws.merge_cells('A2:H2')
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    ws['A2'] = f'تاريخ التوليد: {now}  |  نوع التقرير: {rep_type}'
    ws['A2'].font = Font(size=11, color='FFFFFF', name='Cairo')
    ws['A2'].fill = PatternFill('solid', fgColor='283593')
    ws['A2'].alignment = Alignment(horizontal='center', readingOrder=2)
    ws.row_dimensions[2].height = 25

    # عنوان الملخص
    ws.merge_cells('A4:H4')
    ws['A4'] = 'الملخص الذكي'
    ws['A4'].font = Font(bold=True, size=14, color='1A237E', name='Cairo')
    ws['A4'].alignment = Alignment(horizontal='right', readingOrder=2)
    ws.row_dimensions[4].height = 25

    # نص الملخص
    summary_text = _get_ai_summary(rep_type, ai_key)
    ws.merge_cells('A5:H20')
    ws['A5'] = summary_text
    ws['A5'].font = Font(size=12, name='Cairo')
    ws['A5'].alignment = Alignment(wrap_text=True, readingOrder=2, vertical='top')
    thin = Side(style='thin', color='1A237E')
    ws['A5'].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[5].height = 280
    ws.column_dimensions['A'].width = 90


def _get_ai_summary(rep_type, ai_key=''):
    """الحصول على ملخص ذكي - يستخدم النظام المتقدم للمعلومات"""
    type_map = {
        'issues': 'البلاغات والمشاكل',
        'readings': 'القراءات الكهربائية',
        'assets': 'الاصول والمعدات'
    }
    label = type_map.get(rep_type, rep_type)

    prompt = (
        f"انت محلل بيانات خبير في ادارة مخيمات الحج والانظمة الكهربائية. "
        f"اكتب تقريراً تحليلياً احترافياً ومفصلاً باللغة العربية لتقرير '{label}'. "
        f"اذكر اهمية البيانات، المخاطر المحتملة، والتوصيات الاستراتيجية للتحسين. "
        f"اجعل الرد منظماً بنقاط واضحة ومرقمة."
    )

    try:
        from advanced_ai_system import get_ai_response
        # نستخدم model_type 'issue_analyzer' لأنه الأنسب للتقارير
        resp = get_ai_response(prompt, 'issue_analyzer')
        if resp.get('success'):
            ai_text = resp.get('response', '')
            model_used = resp.get('model_used', 'AI')
            return f"ملخص مولد بالذكاء الاصطناعي ({model_used}):\n\n{ai_text}"
        else:
            return _static_summary(label, error=resp.get('error', 'فشل في توليد الرد'))
    except Exception as ex:
        import traceback; traceback.print_exc()
        return _static_summary(label, error=str(ex))


def _static_summary(label, error=None):
    """ملخص احترافي ثابت عند عدم توفر مفتاح AI"""
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    err_note = f"\nملاحظة تقنية: {error}" if error else ""
    return f"""تقرير تحليلي - {label}
تاريخ الانشاء: {now}
{err_note}

نظرة عامة:
يحتوي هذا التقرير على تحليل شامل لبيانات {label} المسجلة في منظومة ادارة مخيمات الحج.

اهمية البيانات:
- توفير رؤية واضحة وشاملة للوضع الحالي
- تسهيل اتخاذ القرارات الادارية والتشغيلية
- رصد الانماط وتحديد نقاط القوة والضعف

التوصيات الاستراتيجية:
1. مراجعة البيانات بشكل دوري للتاكد من دقتها واكتمالها
2. متابعة الحالات المتاخرة واغلاقها في الوقت المناسب
3. تحليل الانماط المتكررة للوقاية من المشكلات المستقبلية
4. تفعيل التنبيهات التلقائية للحالات الحرجة
5. توثيق الاجراءات والحلول لبناء قاعدة معرفية

لتفعيل الملخص الذكي بالذكاء الاصطناعي:
اذهب الى: نماذج الذكاء الاصطناعي - اضف مفتاح Gemini API
او ادخل المفتاح مباشرة في صفحة التقارير قبل التنزيل.
"""


print("AI Reports Blueprint loaded")


import os
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
import datetime
import json
import tempfile
import uuid
import numpy as np
import pdfkit
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, abort, session
from werkzeug.utils import secure_filename
import pandas as pd
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from translations import translate
from flask_wtf.csrf import CSRFProtect

# استيراد Blueprints
from units_management import units_bp
from dropdowns_management import dropdowns_bp
from manual_readings import readings_bp
from trips_ai import trips_ai_bp

# استيراد النماذج من ملف models.py
from models import (
    db, get_setting, MDBPanel, Contractor, ContractorTeamMember, ContractorTeam, Issue, User, MapArea,
    DynamicColumn, SystemSettings, ElectricalReading, Alert, DynamicColumnValue, HotSpot, MeasurementUnit,
    DropdownList, DropdownItem, ManualReading, InspectionRequest, InspectionRequestType, RequestAssignment,
    IssueAssignment, InspectionRequestAssignment, Country, Company, Camp, PanelCampAssignment,
    # النماذج الجديدة للمميزات الذكية
    ParkingArea, TrafficZone, TransportRoute, UrbanPlan, ZoningArea, HousingUnit, HousingInspection,
    SafetyInspection, HazardReport, InfrastructureNetwork, InfrastructureMalfunction, LandParcel,
    SurveyPoint, WasteCollectionPoint, EnvironmentalReport, GovernmentIntegration, ExternalReport,
    PilgrimageLocation, PilgrimageReport, VirtualTourInteraction,
    # نماذج الرحلات والذكاء الاصطناعي
    TripPackage, TripBooking, TripReview, CustomTrip, TripItinerary, AIModel,
    # نماذج البانوراما 360
    Asset, Panorama360, PanoramaAssetHotspot
)
from sqlalchemy.orm import joinedload

# استيراد وحدة الإشعارات
from notifications import notify_contractor, create_google_maps_url

# دالة للحصول على مناطق المسؤولية للمقاول
def get_contractor_areas(contractor_id, is_manager=False):
    """
    الحصول على مناطق المسؤولية للمقاول

    Args:
        contractor_id (int): معرف المقاول
        is_manager (bool): ما إذا كان المقاول مديرًا

    Returns:
        list: قائمة بأسماء المناطق المسؤول عنها المقاول
    """
    contractor = Contractor.query.get(contractor_id)
    if not contractor:
        return []

    areas = []

    # إضافة مناطق المقاول نفسه
    if contractor.area_responsibility:
        try:
            # محاولة تحليل البيانات كـ JSON
            areas = json.loads(contractor.area_responsibility)
        except (json.JSONDecodeError, TypeError):
            # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
            if isinstance(contractor.area_responsibility, str):
                areas = [area.strip() for area in contractor.area_responsibility.split(',')]

    # إذا كان المقاول مديرًا، أضف مناطق المقاولين التابعين له
    if is_manager:
        sub_contractors = Contractor.query.filter_by(parent_contractor_id=contractor_id).all()
        for sub in sub_contractors:
            if sub.area_responsibility:
                try:
                    sub_areas = json.loads(sub.area_responsibility)
                    areas.extend(sub_areas)
                except (json.JSONDecodeError, TypeError):
                    if isinstance(sub.area_responsibility, str):
                        sub_areas = [area.strip() for area in sub.area_responsibility.split(',')]
                        areas.extend(sub_areas)

    # إزالة التكرارات
    return list(set(areas))

def get_users_in_area(area_name):
    """الحصول على جميع المستخدمين المخصصين لمنطقة معينة"""
    users = []

    # البحث في المستخدمين المخصصين مباشرة للمنطقة
    all_users = User.query.filter_by(is_active=True).all()
    for user in all_users:
        if user.has_area_access(area_name):
            users.append(user)

    # البحث في المجموعات المخصصة للمنطقة
    from models import UserGroup, UserGroupMembership
    groups = UserGroup.query.filter_by(is_active=True).all()
    for group in groups:
        group_areas = group.get_assigned_areas()
        if area_name in group_areas:
            # إضافة جميع أعضاء المجموعة
            memberships = UserGroupMembership.query.filter_by(
                group_id=group.id,
                is_active=True
            ).all()
            for membership in memberships:
                if membership.user not in users and membership.user.is_active:
                    users.append(membership.user)

    return users

def auto_assign_to_area_users(request_or_issue, area_name):
    """توزيع الطلب أو البلاغ تلقائياً على جميع المستخدمين في المنطقة"""
    try:
        # الحصول على المستخدمين في المنطقة
        area_users = get_users_in_area(area_name)

        if not area_users:
            print(f"لا يوجد مستخدمون مخصصون للمنطقة: {area_name}")
            return

        # تحديد نوع الطلب
        if isinstance(request_or_issue, InspectionRequest):
            # طلب فحص
            for user in area_users:
                assignment = InspectionRequestAssignment(
                    inspection_request_id=request_or_issue.id,
                    user_id=user.id
                )
                db.session.add(assignment)
            print(f"تم توزيع طلب الفحص {request_or_issue.id} على {len(area_users)} مستخدم في المنطقة {area_name}")

        elif isinstance(request_or_issue, Issue):
            # بلاغ
            for user in area_users:
                assignment = IssueAssignment(
                    issue_id=request_or_issue.id,
                    user_id=user.id
                )
                db.session.add(assignment)
            print(f"تم توزيع البلاغ {request_or_issue.id} على {len(area_users)} مستخدم في المنطقة {area_name}")

        db.session.commit()

    except Exception as e:
        print(f"خطأ في التوزيع التلقائي: {str(e)}")
        db.session.rollback()

# دالة لإنشاء رابط Google Maps من الإحداثيات
def create_google_maps_url(x_coordinate, y_coordinate):
    """إنشاء رابط Google Maps من الإحداثيات"""
    if x_coordinate is not None and y_coordinate is not None:
        return f"https://maps.google.com/?q={y_coordinate},{x_coordinate}"
    return None

# دالة للحصول على الفرق الفرعية بشكل متكرر
def get_sub_teams(team, all_teams):
    """الحصول على جميع الفرق الفرعية بشكل متكرر"""
    sub_teams = ContractorTeam.query.filter_by(parent_team_id=team.id).all()
    for sub_team in sub_teams:
        all_teams.append(sub_team)
        get_sub_teams(sub_team, all_teams)

# دالة للتحقق مما إذا كان الفريق من نسل فريق آخر
def is_descendant(team_id, potential_ancestor_id):
    """التحقق مما إذا كان الفريق من نسل فريق آخر"""
    # الحصول على الفريق المحتمل أنه سلف
    potential_ancestor = ContractorTeam.query.get(potential_ancestor_id)
    if not potential_ancestor:
        return False

    # الحصول على جميع الفرق الفرعية للفريق المحتمل أنه سلف
    all_descendants = []
    get_sub_teams(potential_ancestor, all_descendants)

    # التحقق مما إذا كان الفريق من ضمن النسل
    for descendant in all_descendants:
        if descendant.id == int(team_id):
            return True

    return False

# تهيئة تطبيق Flask
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production-' + os.urandom(24).hex())
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///mdb_dashboard.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls', 'jpg', 'jpeg', 'png', 'gif'}  # إضافة امتدادات الصور
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # الحد الأقصى لحجم الملف (16 ميجابايت)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # تعطيل الكاش للتطوير
app.config['TEMPLATES_AUTO_RELOAD'] = True  # إعادة تحميل القوالب تلقائياً

# تفعيل CSRF Protection
csrf = CSRFProtect(app)

# استثناء بعض الـ routes من CSRF (routes مسجلة مبكراً)
csrf.exempt('panorama360.import_assets')
csrf.exempt('panorama360.submit_report')
# ملاحظة: إعفاء trips_ai وai_reports والمفضلة يتم بعد register_blueprint (انظر أسفل)


# تسجيل الخطوط العربية لملفات PDF
fonts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'fonts')
if not os.path.exists(fonts_dir):
    os.makedirs(fonts_dir)

# تحميل الخطوط العربية إذا لم تكن موجودة
arabic_font_path = os.path.join(fonts_dir, 'NotoSansArabic-Regular.ttf')
if not os.path.exists(arabic_font_path):
    import urllib.request
    try:
        urllib.request.urlretrieve('https://github.com/googlefonts/noto-fonts/raw/main/hinted/ttf/NotoSansArabic/NotoSansArabic-Regular.ttf', arabic_font_path)
    except:
        print("فشل تحميل الخط العربي. سيتم استخدام الخط الافتراضي.")

# تسجيل الخطوط العربية
try:
    pdfmetrics.registerFont(TTFont('NotoSansArabic', arabic_font_path))
except:
    print("فشل تسجيل الخط العربي. سيتم استخدام الخط الافتراضي.")

# تحسين أداء قاعدة البيانات
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,  # للتحقق من اتصال قاعدة البيانات قبل الاستخدام
    'pool_recycle': 280,    # إعادة تدوير الاتصالات كل 280 ثانية
    'pool_size': 20,        # زيادة حجم تجمع الاتصالات لتحسين الأداء
    'max_overflow': 40,     # زيادة الحد الأقصى للاتصالات الإضافية
    'pool_timeout': 30,     # زيادة مهلة الانتظار
    'echo': False,          # إيقاف طباعة استعلامات SQL لتحسين الأداء
    'echo_pool': False      # إيقاف طباعة معلومات تجمع الاتصالات
}

# تحسين الأمان
from datetime import timedelta
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('PRODUCTION', 'False') == 'True'  # تفعيل في الإنتاج فقط
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=12)  # مدة الجلسة 12 ساعة

# تعيين خيار إعادة إنشاء قاعدة البيانات
RECREATE_DB = False  # ❌ معطّل عند التسليم حتى لا تُعاد تهيئة قاعدة البيانات

# تهيئة قاعدة البيانات
db.init_app(app)

# تسجيل Blueprints
app.register_blueprint(units_bp)
app.register_blueprint(dropdowns_bp)
app.register_blueprint(readings_bp)

# تسجيل Blueprint البانوراما 360 (الجديد)
try:
    from blueprints.panorama360_bp import panorama360_bp
    app.register_blueprint(panorama360_bp)
    print("Panorama 360° Blueprint registered successfully")
except ImportError as e:
    print(f"Panorama 360° module not found: {str(e)}")
except Exception as e:
    print(f"Error registering camp management Blueprint: {str(e)}")

# تسجيل Blueprint المميزات الذكية للمدينة
try:
    from smart_city_features import smart_city_bp
    app.register_blueprint(smart_city_bp)
    print("Smart City Features Blueprint registered successfully")
except ImportError as e:
    print(f"Smart City Features module not found: {str(e)}")
except Exception as e:
    print(f"Error registering Smart City Features Blueprint: {str(e)}")

# تسجيل Blueprint تجربة الحجاج والسياح 360°
# تم التعليق مؤقتاً بسبب تعارض في Routes
# try:
#     print("محاولة استيراد pilgrimage_360...")
#     from pilgrimage_360 import pilgrimage_bp
#     print("تم استيراد pilgrimage_bp بنجاح")
#     app.register_blueprint(pilgrimage_bp)
#     print("Pilgrimage 360° Blueprint registered successfully")
#     print(f"Blueprint URL prefix: {pilgrimage_bp.url_prefix}")
# except ImportError as e:
#     print(f"Pilgrimage 360° module not found: {str(e)}")
#     import traceback
#     traceback.print_exc()
# except Exception as e:
#     print(f"Error registering Pilgrimage 360° Blueprint: {str(e)}")
#     import traceback
#     traceback.print_exc()

# تسجيل Blueprint الرحلات والذكاء الاصطناعي
try:
    app.register_blueprint(trips_ai_bp)
    print("Trips AI Blueprint registered successfully")
except ImportError as e:
    print(f"Trips AI module not found: {str(e)}")
except Exception as e:
    print(f"Error registering Trips AI Blueprint: {str(e)}")

# تسجيل Blueprint تقارير الذكاء الاصطناعي (جديد)
try:
    from ai_reports import ai_reports_bp
    app.register_blueprint(ai_reports_bp)
    print("AI Reports Blueprint registered successfully")
except ImportError as e:
    print(f"AI Reports module not found: {str(e)}")
except Exception as e:
    print(f"Error registering AI Reports Blueprint: {str(e)}")

# ===== إعفاء API endpoints من CSRF بعد تسجيل الـ Blueprints =====
# يجب أن يكون بعد register_blueprint حتى تكون الـ view functions موجودة
_csrf_exempt_endpoints = [
    'trips_ai.api_chat',
    'trips_ai.api_plan_trip',
    'trips_ai.api_smart_spatial_guide',
    'trips_ai.api_spatial_guide',
    'trips_ai.test_ai_model',
    'ai_reports.generate_report',
    'api_favorites_add',
    'api_favorites_remove',
    'api_favorites_list',
]
for _ep in _csrf_exempt_endpoints:
    if _ep in app.view_functions:
        csrf.exempt(app.view_functions[_ep])


# إضافة دالة get_setting وترجمة إلى قوالب Jinja
app.jinja_env.globals.update(get_setting=get_setting, translate=translate)

# وظيفة لتعيين قيمة إعداد في قاعدة البيانات
def set_setting(key, value, description=None):
    """تعيين قيمة إعداد في قاعدة البيانات"""
    setting = SystemSettings.query.filter_by(key=key).first()
    if setting:
        setting.value = value
        if description:
            setting.description = description
    else:
        setting = SystemSettings(key=key, value=value, description=description)
        db.session.add(setting)
    db.session.commit()
    return setting

# تهيئة الإعدادات الافتراضية
def init_default_settings():
    """تهيئة الإعدادات الافتراضية للنظام"""
    default_settings = {
        # إعدادات اللغة
        'language': ('ar', 'لغة العرض (ar للعربية، en للإنجليزية)'),

        # إعدادات الحدود الافتراضية
        'default_warning_threshold': ('70', 'نسبة التحذير الافتراضية (%)'),
        'default_danger_threshold': ('80', 'نسبة الخطر الافتراضية (%)'),
        'default_min_voltage': ('210', 'الحد الأدنى الافتراضي للجهد (فولت)'),
        'default_max_voltage': ('250', 'الحد الأقصى الافتراضي للجهد (فولت)'),

        # إعدادات التنبيهات
        'alert_check_interval': ('15', 'الفاصل الزمني للتحقق من التنبيهات (دقيقة)'),
        'alert_time_window': ('60', 'النافذة الزمنية للتنبيهات (دقيقة)'),
        'alert_method': ('peak', 'طريقة التنبيه (peak, average)'),
        'alert_calculation_period': ('60', 'فترة حساب التنبيهات (دقيقة)'),

        # إعدادات الألوان
        'alert_color_normal': ('#28A745', 'لون الحالة الطبيعية'),
        'alert_color_warning': ('#FFC107', 'لون التحذير'),
        'alert_color_danger': ('#DC3545', 'لون الخطر'),

        # إعدادات الإشعارات
        'send_notifications': ('true', 'إرسال إشعارات للمقاولين'),
        'notification_method': ('email', 'طريقة إرسال الإشعارات (email, sms, both)'),

        # إعدادات الخريطة
        'default_map_center': ('21.3583, 39.9719', 'مركز الخريطة الافتراضي (عرفة، مكة المكرمة)'),
        'default_map_zoom': ('14', 'مستوى تكبير الخريطة الافتراضي'),

        # إعدادات القراءات اليدوية الافتراضية
        'default_current': ('0', 'القيمة الافتراضية للتيار (أمبير)'),
        'default_voltage': ('220', 'القيمة الافتراضية للجهد (فولت)'),
        'default_power': ('0', 'القيمة الافتراضية للقدرة (واط)'),
        'default_energy': ('0', 'القيمة الافتراضية للطاقة (كيلوواط ساعة)'),
        'default_power_factor': ('0.9', 'القيمة الافتراضية لمعامل القدرة'),
        'default_frequency': ('60', 'القيمة الافتراضية للتردد (هرتز)'),
        'default_breaker_capacity': ('0', 'القيمة الافتراضية لسعة القاطع (أمبير)'),

        # إعدادات القراءات الثلاثية
        'use_three_phase': ('true', 'استخدام القراءات الثلاثية الطور')
    }

    for key, (value, description) in default_settings.items():
        if get_setting(key) is None:
            set_setting(key, value, description)

# تهيئة وحدات القياس الافتراضية
def init_default_units():
    """تهيئة وحدات القياس الافتراضية"""
    # التحقق من وجود وحدات قياس
    if MeasurementUnit.query.count() > 0:
        return

    # وحدات التيار
    current_units = [
        {'name': 'A', 'display_name': 'أمبير', 'category': 'current', 'conversion_factor': 1.0, 'is_default': True},
        {'name': 'kA', 'display_name': 'كيلو أمبير', 'category': 'current', 'conversion_factor': 1000.0, 'is_default': False},
        {'name': 'mA', 'display_name': 'ملي أمبير', 'category': 'current', 'conversion_factor': 0.001, 'is_default': False}
    ]

    # وحدات الجهد
    voltage_units = [
        {'name': 'V', 'display_name': 'فولت', 'category': 'voltage', 'conversion_factor': 1.0, 'is_default': True},
        {'name': 'kV', 'display_name': 'كيلو فولت', 'category': 'voltage', 'conversion_factor': 1000.0, 'is_default': False},
        {'name': 'mV', 'display_name': 'ملي فولت', 'category': 'voltage', 'conversion_factor': 0.001, 'is_default': False}
    ]

    # وحدات القدرة الفعالة (P)
    power_units = [
        {'name': 'W', 'display_name': 'واط', 'category': 'power', 'conversion_factor': 1.0, 'is_default': False},
        {'name': 'kW', 'display_name': 'كيلو واط', 'category': 'power', 'conversion_factor': 1000.0, 'is_default': True},
        {'name': 'MW', 'display_name': 'ميجا واط', 'category': 'power', 'conversion_factor': 1000000.0, 'is_default': False}
    ]

    # وحدات القدرة الظاهرية (S)
    apparent_power_units = [
        {'name': 'VA', 'display_name': 'فولت أمبير', 'category': 'apparent_power', 'conversion_factor': 1.0, 'is_default': False},
        {'name': 'kVA', 'display_name': 'كيلو فولت أمبير', 'category': 'apparent_power', 'conversion_factor': 1000.0, 'is_default': True},
        {'name': 'MVA', 'display_name': 'ميجا فولت أمبير', 'category': 'apparent_power', 'conversion_factor': 1000000.0, 'is_default': False}
    ]

    # وحدات القدرة غير الفعالة (Q)
    reactive_power_units = [
        {'name': 'VAR', 'display_name': 'فار', 'category': 'reactive_power', 'conversion_factor': 1.0, 'is_default': False},
        {'name': 'kVAR', 'display_name': 'كيلو فار', 'category': 'reactive_power', 'conversion_factor': 1000.0, 'is_default': True},
        {'name': 'MVAR', 'display_name': 'ميجا فار', 'category': 'reactive_power', 'conversion_factor': 1000000.0, 'is_default': False}
    ]

    # وحدات الطاقة
    energy_units = [
        {'name': 'Wh', 'display_name': 'واط ساعة', 'category': 'energy', 'conversion_factor': 1.0, 'is_default': False},
        {'name': 'kWh', 'display_name': 'كيلو واط ساعة', 'category': 'energy', 'conversion_factor': 1000.0, 'is_default': True},
        {'name': 'MWh', 'display_name': 'ميجا واط ساعة', 'category': 'energy', 'conversion_factor': 1000000.0, 'is_default': False}
    ]

    # إضافة جميع الوحدات
    all_units = current_units + voltage_units + power_units + apparent_power_units + reactive_power_units + energy_units

    for unit_data in all_units:
        unit = MeasurementUnit(
            name=unit_data['name'],
            display_name=unit_data['display_name'],
            category=unit_data['category'],
            conversion_factor=unit_data['conversion_factor'],
            is_default=unit_data['is_default'],
            is_active=True,
            created_at=datetime.datetime.now()
        )
        db.session.add(unit)

    db.session.commit()

# تهيئة القوائم المنسدلة الافتراضية
def init_default_dropdowns():
    """تهيئة القوائم المنسدلة الافتراضية"""
    # التحقق من وجود قوائم منسدلة
    if DropdownList.query.count() > 0:
        return

    # قائمة حالة القاطع
    breaker_status = DropdownList(
        name='breaker_status',
        display_name='حالة القاطع',
        field_type='breaker',
        visibility='all',
        description='حالة القاطع الكهربائي',
        is_active=True,
        created_at=datetime.datetime.now()
    )
    db.session.add(breaker_status)

    # قائمة سعة القاطع
    breaker_capacity = DropdownList(
        name='breaker_capacity',
        display_name='سعة القاطع',
        field_type='breaker_capacity',
        visibility='all',
        description='سعة القاطع الكهربائي',
        is_active=True,
        created_at=datetime.datetime.now()
    )
    db.session.add(breaker_capacity)

    db.session.commit()

    # إضافة عناصر قائمة حالة القاطع
    breaker_status_items = [
        {'value': 'on', 'display_text': 'مغلق (يعمل)', 'order': 1},
        {'value': 'off', 'display_text': 'مفتوح (لا يعمل)', 'order': 2},
        {'value': 'tripped', 'display_text': 'مفصول (تم فصله تلقائياً)', 'order': 3},
        {'value': 'maintenance', 'display_text': 'قيد الصيانة', 'order': 4}
    ]

    for item_data in breaker_status_items:
        item = DropdownItem(
            dropdown_id=breaker_status.id,
            value=item_data['value'],
            display_text=item_data['display_text'],
            order=item_data['order'],
            is_active=True,
            created_at=datetime.datetime.now()
        )
        db.session.add(item)

    # إضافة عناصر قائمة سعة القاطع
    breaker_capacity_items = [
        {'value': '10', 'display_text': '10 أمبير', 'order': 1},
        {'value': '16', 'display_text': '16 أمبير', 'order': 2},
        {'value': '20', 'display_text': '20 أمبير', 'order': 3},
        {'value': '25', 'display_text': '25 أمبير', 'order': 4},
        {'value': '32', 'display_text': '32 أمبير', 'order': 5},
        {'value': '40', 'display_text': '40 أمبير', 'order': 6},
        {'value': '50', 'display_text': '50 أمبير', 'order': 7},
        {'value': '63', 'display_text': '63 أمبير', 'order': 8},
        {'value': '80', 'display_text': '80 أمبير', 'order': 9},
        {'value': '100', 'display_text': '100 أمبير', 'order': 10},
        {'value': '125', 'display_text': '125 أمبير', 'order': 11},
        {'value': '160', 'display_text': '160 أمبير', 'order': 12},
        {'value': '200', 'display_text': '200 أمبير', 'order': 13},
        {'value': '250', 'display_text': '250 أمبير', 'order': 14},
        {'value': '315', 'display_text': '315 أمبير', 'order': 15},
        {'value': '400', 'display_text': '400 أمبير', 'order': 16},
        {'value': '500', 'display_text': '500 أمبير', 'order': 17},
        {'value': '630', 'display_text': '630 أمبير', 'order': 18},
        {'value': '800', 'display_text': '800 أمبير', 'order': 19},
        {'value': '1000', 'display_text': '1000 أمبير', 'order': 20},
        {'value': '1250', 'display_text': '1250 أمبير', 'order': 21},
        {'value': '1600', 'display_text': '1600 أمبير', 'order': 22},
        {'value': '2000', 'display_text': '2000 أمبير', 'order': 23},
        {'value': '2500', 'display_text': '2500 أمبير', 'order': 24},
        {'value': '3200', 'display_text': '3200 أمبير', 'order': 25}
    ]

    for item_data in breaker_capacity_items:
        item = DropdownItem(
            dropdown_id=breaker_capacity.id,
            value=item_data['value'],
            display_text=item_data['display_text'],
            order=item_data['order'],
            is_active=True,
            created_at=datetime.datetime.now()
        )
        db.session.add(item)

    db.session.commit()

def add_trips_ai_data():
    """إضافة بيانات الرحلات والذكاء الاصطناعي"""
    try:
        # إضافة حزم الرحلات التجريبية
        if TripPackage.query.count() == 0:
            packages = [
                TripPackage(
                    package_name="حج اقتصادي - 14 يوم",
                    package_name_en="Economy Hajj - 14 Days",
                    package_type="hajj",
                    duration_days=14,
                    price_per_person=8000.0,
                    max_participants=50,
                    current_bookings=0,
                    description="حزمة حج اقتصادية تشمل الإقامة والنقل والوجبات الأساسية",
                    description_en="Economy Hajj package including accommodation, transport and basic meals",
                    included_services='["إقامة في فنادق 3 نجوم", "النقل", "وجبتين يومياً", "مرشد سياحي"]',
                    excluded_services='["تذاكر الطيران", "التأمين الصحي", "المصروف الشخصي"]',
                    accommodation_type="hotel",
                    accommodation_rating=3,
                    meals_included="half_board",
                    transport_type="bus",
                    departure_city="الرياض",
                    start_date=datetime.datetime(2024, 6, 1),
                    end_date=datetime.datetime(2024, 6, 14),
                    booking_deadline=datetime.datetime(2024, 5, 1),
                    is_active=True,
                    is_featured=True,
                    rating=4.2,
                    reviews_count=25,
                    difficulty_level="easy",
                    age_restrictions="18+ سنة",
                    health_requirements="شهادة صحية مطلوبة",
                    required_documents='["جواز سفر", "تأشيرة حج", "شهادة تطعيم"]'
                ),
                TripPackage(
                    package_name="عمرة مميزة - 7 أيام",
                    package_name_en="Premium Umrah - 7 Days",
                    package_type="umrah",
                    duration_days=7,
                    price_per_person=4500.0,
                    max_participants=30,
                    current_bookings=0,
                    description="حزمة عمرة مميزة مع إقامة فاخرة وخدمات متكاملة",
                    description_en="Premium Umrah package with luxury accommodation and comprehensive services",
                    included_services='["إقامة في فنادق 5 نجوم", "النقل الفاخر", "جميع الوجبات", "مرشد متخصص", "زيارة المعالم"]',
                    excluded_services='["تذاكر الطيران", "التسوق الشخصي"]',
                    accommodation_type="hotel",
                    accommodation_rating=5,
                    meals_included="full_board",
                    transport_type="bus",
                    departure_city="جدة",
                    start_date=datetime.datetime(2024, 4, 15),
                    end_date=datetime.datetime(2024, 4, 21),
                    booking_deadline=datetime.datetime(2024, 3, 15),
                    is_active=True,
                    is_featured=True,
                    rating=4.8,
                    reviews_count=42,
                    difficulty_level="easy",
                    age_restrictions="جميع الأعمار",
                    health_requirements="فحص طبي أساسي",
                    required_documents='["جواز سفر", "تأشيرة عمرة"]'
                ),
                TripPackage(
                    package_name="جولة سياحية دينية - 5 أيام",
                    package_name_en="Religious Tourism - 5 Days",
                    package_type="tourism",
                    duration_days=5,
                    price_per_person=2800.0,
                    max_participants=40,
                    current_bookings=0,
                    description="جولة سياحية لزيارة المعالم الدينية والتاريخية في مكة والمدينة",
                    description_en="Religious tourism to visit religious and historical landmarks in Mecca and Medina",
                    included_services='["إقامة في فنادق 4 نجوم", "النقل", "الإفطار", "مرشد سياحي", "دخول المعالم"]',
                    excluded_services='["الغداء والعشاء", "التسوق", "الأنشطة الإضافية"]',
                    accommodation_type="hotel",
                    accommodation_rating=4,
                    meals_included="breakfast",
                    transport_type="bus",
                    departure_city="الدمام",
                    start_date=datetime.datetime(2024, 3, 10),
                    end_date=datetime.datetime(2024, 3, 14),
                    booking_deadline=datetime.datetime(2024, 2, 10),
                    is_active=True,
                    is_featured=False,
                    rating=4.5,
                    reviews_count=18,
                    difficulty_level="moderate",
                    age_restrictions="12+ سنة",
                    health_requirements="لا توجد متطلبات خاصة",
                    required_documents='["هوية وطنية أو إقامة"]'
                )
            ]

            for package in packages:
                db.session.add(package)

            db.session.commit()
            print("تم إضافة حزم الرحلات التجريبية بنجاح!")

        # إضافة نماذج الذكاء الاصطناعي التجريبية
        if AIModel.query.count() == 0:
            ai_models = [
                AIModel(
                    model_name="مساعد الرحلات الذكي",
                    model_type="chatbot",
                    api_provider="demo",
                    api_key="demo-key",
                    model_version="v1.0",
                    temperature=0.7,
                    max_tokens=1000,
                    system_prompt="أنت مساعد ذكي لموقع الحج والعمرة. ساعد المستخدمين بمعلومات مفيدة ودقيقة.",
                    is_active=True,
                    is_default=True
                ),
                AIModel(
                    model_name="مخطط الرحلات الذكي",
                    model_type="trip_planner",
                    api_provider="demo",
                    api_key="demo-key",
                    model_version="v1.0",
                    temperature=0.5,
                    max_tokens=2000,
                    system_prompt="أنت خبير في تخطيط رحلات الحج والعمرة. قدم اقتراحات مفصلة ومفيدة للمسافرين.",
                    is_active=True,
                    is_default=True
                )
            ]

            for model in ai_models:
                db.session.add(model)

            db.session.commit()
            print("تم إضافة نماذج الذكاء الاصطناعي التجريبية بنجاح!")

    except Exception as e:
        print(f"خطأ في إضافة بيانات الرحلات والذكاء الاصطناعي: {e}")
        db.session.rollback()

def add_comprehensive_demo_data():
    """إضافة بيانات تجريبية شاملة لجميع المميزات"""
    try:
        # إضافة مواقع الحج والعمرة
        if PilgrimageLocation.query.count() == 0:
            pilgrimage_locations = [
                PilgrimageLocation(
                    location_name="المسجد الحرام",
                    location_name_en="Masjid al-Haram",
                    location_type="holy_site",
                    description="أقدس مكان في الإسلام، يحتوي على الكعبة المشرفة",
                    description_en="The holiest place in Islam, containing the Kaaba",
                    center_lat=21.4225,
                    center_lng=39.8262,
                    operating_hours="24/7",
                    capacity=2000000,
                    current_occupancy=150000,
                    accessibility_features='["مداخل لذوي الاحتياجات الخاصة", "مصاعد", "كراسي متحركة"]',
                    virtual_tour_url="https://www.youtube.com/embed/kZhF8hWz8qY",
                    amenities='["مصاعد", "مكيفات", "مياه زمزم", "مصليات منفصلة"]',
                    # النماذج ثلاثية الأبعاد
                    model_3d_glb="/static/3d_models/kaaba.glb",
                    model_3d_gltf="/static/3d_models/kaaba.gltf",
                    model_3d_usdz="/static/3d_models/kaaba.usdz",
                    model_3d_fbx="/static/3d_models/kaaba.fbx",
                    model_3d_preview="/static/images/kaaba-3d-preview.jpg",
                    model_3d_settings='{"autoRotate": true, "cameraControls": true, "ar": true}',
                    # فيديوهات 360°
                    video_360_urls='["https://www.youtube.com/embed/kZhF8hWz8qY"]',
                    video_360_thumbnails='["https://img.youtube.com/vi/kZhF8hWz8qY/maxresdefault.jpg"]',
                    is_featured=True,
                    crowd_level="high",
                    safety_rating=5,
                    cleanliness_rating=5,
                    display_order=1
                ),
                PilgrimageLocation(
                    location_name="المسجد النبوي",
                    location_name_en="Prophet's Mosque",
                    location_type="holy_site",
                    description="ثاني أقدس مسجد في الإسلام، يحتوي على قبر النبي محمد صلى الله عليه وسلم",
                    description_en="The second holiest mosque in Islam, containing the tomb of Prophet Muhammad",
                    center_lat=24.4672,
                    center_lng=39.6117,
                    operating_hours="24/7",
                    capacity=1000000,
                    current_occupancy=80000,
                    accessibility_features='["مداخل لذوي الاحتياجات الخاصة", "مصاعد", "كراسي متحركة"]',
                    virtual_tour_url="https://www.youtube.com/embed/7d16CpWp-ok",
                    amenities='["مكتبة", "مكيفات", "مياه باردة", "مصليات منفصلة"]',
                    # النماذج ثلاثية الأبعاد
                    model_3d_glb="/static/3d_models/masjid-al-haram.glb",
                    model_3d_gltf="/static/3d_models/masjid-al-haram.gltf",
                    model_3d_usdz="/static/3d_models/masjid-al-haram.usdz",
                    model_3d_fbx="/static/3d_models/masjid-al-haram.fbx",
                    model_3d_preview="/static/images/masjid-preview.jpg",
                    model_3d_settings='{"autoRotate": true, "cameraControls": true, "ar": true}',
                    # فيديوهات 360°
                    video_360_urls='["https://www.youtube.com/embed/7d16CpWp-ok"]',
                    video_360_thumbnails='["https://img.youtube.com/vi/7d16CpWp-ok/maxresdefault.jpg"]',
                    is_featured=True,
                    crowd_level="medium",
                    safety_rating=5,
                    cleanliness_rating=5,
                    display_order=2
                ),
                PilgrimageLocation(
                    location_name="جبل النور - غار حراء",
                    location_name_en="Mount of Light - Cave of Hira",
                    location_type="historical_site",
                    description="الجبل الذي يحتوي على غار حراء حيث نزل الوحي على النبي محمد",
                    description_en="The mountain containing Cave Hira where Prophet Muhammad received his first revelation",
                    center_lat=21.4594,
                    center_lng=39.8578,
                    operating_hours="طوال اليوم",
                    capacity=50,
                    current_occupancy=15,
                    accessibility_features='["مسار صعب، غير مناسب لذوي الاحتياجات الخاصة"]',
                    virtual_tour_url="https://www.youtube.com/embed/Jhi6nRXNzFE",
                    amenities='["مسارات مشي", "لوحات إرشادية"]',
                    is_featured=True,
                    crowd_level="low",
                    safety_rating=3,
                    cleanliness_rating=4,
                    display_order=3
                ),
                PilgrimageLocation(
                    location_name="جبل أحد",
                    location_name_en="Mount Uhud",
                    location_type="historical_site",
                    description="موقع غزوة أحد التاريخية، يحتوي على مقبرة شهداء أحد",
                    description_en="Site of the historic Battle of Uhud, contains the cemetery of Uhud martyrs",
                    center_lat=24.4951,
                    center_lng=39.6189,
                    operating_hours="طوال اليوم",
                    capacity=1000,
                    current_occupancy=50,
                    accessibility_features='["مداخل سهلة", "مسارات معبدة جزئياً"]',
                    virtual_tour_url="https://www.youtube.com/embed/example4",
                    amenities='["مقبرة الشهداء", "مسجد", "مواقف سيارات"]',
                    is_featured=False,
                    crowd_level="low",
                    safety_rating=4,
                    cleanliness_rating=4,
                    display_order=4
                )
            ]

            for location in pilgrimage_locations:
                db.session.add(location)

            db.session.commit()
            print("تم إضافة مواقع الحج والعمرة التجريبية بنجاح!")

        # إضافة مناطق المواقف
        if ParkingArea.query.count() == 0:
            parking_areas = [
                ParkingArea(
                    name="موقف الحرم الشريف - الشمالي",
                    area_type="public",
                    center_lat=21.4235,
                    center_lng=39.8255,
                    capacity=2000,
                    current_occupancy=1200,
                    hourly_rate=5.0,
                    operating_hours="24/7",
                    accessibility_features='["مواقف لذوي الاحتياجات الخاصة", "مصاعد"]',
                    contact_info="هاتف: 920000001",
                    notes="موقف مغطى مع خدمة نقل مجانية للحرم",
                    is_active=True
                ),
                ParkingArea(
                    name="موقف الحرم الشريف - الجنوبي",
                    area_type="public",
                    center_lat=21.4215,
                    center_lng=39.8270,
                    capacity=1500,
                    current_occupancy=800,
                    hourly_rate=5.0,
                    operating_hours="24/7",
                    accessibility_features='["مواقف لذوي الاحتياجات الخاصة"]',
                    contact_info="هاتف: 920000002",
                    notes="موقف مغطى مع محطات شحن السيارات الكهربائية",
                    is_active=True
                ),
                ParkingArea(
                    name="موقف المسجد النبوي - الشرقي",
                    area_type="public",
                    center_lat=24.4680,
                    center_lng=39.6125,
                    capacity=1000,
                    current_occupancy=400,
                    hourly_rate=3.0,
                    operating_hours="24/7",
                    accessibility_features='["مواقف لذوي الاحتياجات الخاصة", "مسارات سهلة"]',
                    contact_info="هاتف: 920000003",
                    notes="موقف مكشوف قريب من المسجد النبوي",
                    is_active=True
                )
            ]

            for parking in parking_areas:
                db.session.add(parking)

            db.session.commit()
            print("تم إضافة مناطق المواقف التجريبية بنجاح!")

        # إضافة وحدات السكن
        if HousingUnit.query.count() == 0:
            housing_units = [
                HousingUnit(
                    unit_number="H001",
                    building_number="B001",
                    unit_type="residential",
                    housing_category="apartment",
                    center_lat=21.4188,
                    center_lng=39.8258,
                    condition_status="good",
                    occupancy_status="vacant",
                    occupancy_level=0,
                    max_capacity=6,
                    floor_area=120.0,
                    number_of_rooms=2,
                    number_of_bathrooms=2,
                    has_parking=True,
                    accessibility_features='["مصعد", "مداخل واسعة"]',
                    owner_name="شركة أبراج البيت",
                    owner_contact="920000100",
                    rental_status="rented",
                    monthly_rent=8000.0,
                    utilities_included='["كهرباء", "ماء", "إنترنت"]',
                    notes="شقة مفروشة مع إطلالة على الحرم"
                ),
                HousingUnit(
                    unit_number="H002",
                    building_number="V001",
                    unit_type="residential",
                    housing_category="villa",
                    center_lat=21.4300,
                    center_lng=39.8100,
                    condition_status="excellent",
                    occupancy_status="vacant",
                    occupancy_level=0,
                    max_capacity=12,
                    floor_area=300.0,
                    number_of_rooms=4,
                    number_of_bathrooms=3,
                    has_parking=True,
                    accessibility_features='["حديقة", "مسبح خاص"]',
                    owner_name="مؤسسة الحرمين",
                    owner_contact="920000101",
                    rental_status="vacant",
                    monthly_rent=15000.0,
                    utilities_included='["كهرباء", "ماء"]',
                    notes="فيلا فاخرة مع حديقة ومسبح خاص"
                ),
                HousingUnit(
                    unit_number="M001",
                    building_number="T001",
                    unit_type="residential",
                    housing_category="apartment",
                    center_lat=24.4700,
                    center_lng=39.6100,
                    condition_status="good",
                    occupancy_status="occupied",
                    occupancy_level=5,
                    max_capacity=8,
                    floor_area=150.0,
                    number_of_rooms=3,
                    number_of_bathrooms=2,
                    has_parking=True,
                    accessibility_features='["مصعد", "بلكونة"]',
                    owner_name="شركة أبراج طيبة",
                    owner_contact="920000102",
                    rental_status="rented",
                    monthly_rent=6000.0,
                    utilities_included='["كهرباء", "ماء", "إنترنت"]',
                    notes="شقة مفروشة في وسط المدينة المنورة"
                )
            ]

            for unit in housing_units:
                db.session.add(unit)

            db.session.commit()
            print("تم إضافة وحدات السكن التجريبية بنجاح!")

    except Exception as e:
        print(f"خطأ في إضافة البيانات التجريبية الشاملة: {e}")
        db.session.rollback()

# تهيئة قاعدة البيانات عند بدء التطبيق
def initialize_database():
    """تهيئة قاعدة البيانات عند بدء التطبيق"""
    with app.app_context():
        # إنشاء جميع الجداول
        db.create_all()
        print("Database tables created")

        # تنفيذ ملفات الترحيل
        try:
            from migrations.add_scada_connected_field import run_migration
            run_migration()
        except Exception as e:
            print(f"حدث خطأ أثناء تنفيذ ملف الترحيل add_scada_connected_field: {str(e)}")

        try:
            from migrations.add_mutawif_name_field import run_migration
            run_migration()
        except Exception as e:
            print(f"حدث خطأ أثناء تنفيذ ملف الترحيل add_mutawif_name_field: {str(e)}")

        try:
            from migrations.add_inspection_request_model import run_migration
            run_migration()
        except Exception as e:
            print(f"حدث خطأ أثناء تنفيذ ملف الترحيل add_inspection_request_model: {str(e)}")

        try:
            from migrations.add_inspection_request_types import run_migration
            run_migration()
        except Exception as e:
            print(f"حدث خطأ أثناء تنفيذ ملف الترحيل add_inspection_request_types: {str(e)}")

        try:
            from migrations.add_responsible_person_to_inspection_request import run_migration
            run_migration()
        except Exception as e:
            print(f"حدث خطأ أثناء تنفيذ ملف الترحيل add_responsible_person_to_inspection_request: {str(e)}")

        # إضافة بيانات الرحلات والذكاء الاصطناعي
        add_trips_ai_data()

        # إضافة البيانات التجريبية الشاملة
        add_comprehensive_demo_data()

# استدعاء دالة تهيئة قاعدة البيانات قبل تشغيل التطبيق
initialize_database()

# تفعيل نظام تسجيل الدخول
from flask_login import LoginManager, login_user, logout_user, login_required, current_user

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'يرجى تسجيل الدخول للوصول إلى هذه الصفحة'
login_manager.login_message_category = 'warning'

# إضافة وظيفة set_password للمستخدم
def set_password(user, password):
    # تخزين كلمة المرور كما هي مؤقتًا (لأغراض التطوير فقط)
    user.password_hash = password
    return user

# دالة تحميل المستخدم لنظام تسجيل الدخول
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# إضافة دالة للحصول على الإعدادات والتاريخ في القوالب
@app.context_processor
def inject_settings():
    return dict(
        get_setting=get_setting,
        current_date=datetime.datetime.now().strftime("%Y-%m-%d")
    )

# إضافة مرشحات JSON مخصصة
@app.template_filter('fromjson')
def fromjson_filter(value):
    """تحويل نص JSON إلى كائن Python"""
    try:
        import json
        return json.loads(value) if value else []
    except:
        return []

# التحقق من امتدادات الملفات المسموح بها
def allowed_file(filename, allowed_extensions=None):
    if allowed_extensions is None:
        allowed_extensions = app.config['ALLOWED_EXTENSIONS']
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

# الصفحة الرئيسية
@app.route('/')
def index():
    # توجيه المستخدم إلى صفحة تسجيل الدخول إذا لم يكن مسجل الدخول
    if not current_user.is_authenticated:
        return redirect(url_for('login'))

    # التحقق مما إذا كان المستخدم مسجل الدخول ودوره مقاول
    if current_user.role == 'contractor':
        return redirect(url_for('contractor_dashboard'))

    # إذا كان المستخدم مقاول عادي أو مسجل قراءات، توجيهه إلى صفحة القراءات اليدوية
    if current_user.role in ['regular_contractor', 'readings_recorder']:
        return redirect(url_for('readings.all_readings'))

    # الحصول على البيانات للعرض في الجدول
    panels = MDBPanel.query.all()

    # إحصائيات للوحة المعلومات
    total_panels = MDBPanel.query.count()
    panel_types = db.session.query(MDBPanel.panel_type, db.func.count(MDBPanel.id)).group_by(MDBPanel.panel_type).all()
    years = db.session.query(MDBPanel.implementation_year, db.func.count(MDBPanel.id)).group_by(MDBPanel.implementation_year).all()
    areas = db.session.query(MDBPanel.area_name, db.func.count(MDBPanel.id)).group_by(MDBPanel.area_name).all()

    # إحصائيات الحالة
    status_counts = db.session.query(MDBPanel.status, db.func.count(MDBPanel.id)).group_by(MDBPanel.status).all()

    # الحصول على القيم الفريدة للفلاتر
    unique_areas = [area[0] for area in db.session.query(MDBPanel.area_name).distinct() if area[0]]
    unique_panel_types = [panel_type[0] for panel_type in db.session.query(MDBPanel.panel_type).distinct() if panel_type[0]]
    unique_years = [year[0] for year in db.session.query(MDBPanel.implementation_year).distinct() if year[0]]

    # الحصول على الأعمدة الديناميكية النشطة
    dynamic_columns = DynamicColumn.query.filter_by(is_active=True).all()

    return render_template('index.html',
                           panels=panels,
                           total_panels=total_panels,
                           panel_types=panel_types,
                           years=years,
                           areas=areas,
                           status_counts=status_counts,
                           unique_areas=unique_areas,
                           unique_panel_types=unique_panel_types,
                           unique_years=unique_years,
                           dynamic_columns=dynamic_columns,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحات تسجيل الدخول والخروج
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        # التحقق من وجود البيانات
        if not username or not password:
            flash('يرجى إدخال اسم المستخدم وكلمة المرور', 'danger')
            return render_template('login.html', current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

        try:
            user = User.query.filter_by(username=username).first()

            # استخدام الدالة الجديدة check_password للتحقق من كلمة المرور المشفرة
            if user and user.check_password(password):
                if not user.is_active:
                    flash('حسابك غير نشط. يرجى التواصل مع المسؤول', 'danger')
                    return render_template('login.html', current_date=datetime.datetime.now().strftime("%Y-%m-%d"))
                
                login_user(user)
                user.last_login = datetime.datetime.now()
                db.session.commit()

                next_page = request.args.get('next')
                return redirect(next_page) if next_page else redirect(url_for('index'))
            else:
                # رسالة خطأ باللغتين العربية والإنجليزية
                if request.headers.get('Accept-Language', '').startswith('en'):
                    flash('Invalid username or password', 'danger')
                else:
                    flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'danger')
        except Exception as e:
            flash('حدث خطأ أثناء تسجيل الدخول', 'danger')
            print(f"Login error: {e}")

    return render_template('login.html', current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('تم تسجيل الخروج بنجاح', 'success')
    return redirect(url_for('login'))

# صفحة إعدادات اللغة
@app.route('/language-settings')
@login_required
def language_settings():
    return render_template('language_settings.html',
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تعيين لغة العرض
@app.route('/set-language', methods=['POST'])
@login_required
def set_language():
    language = request.form.get('language', 'ar')
    set_setting('language', language)

    # تخزين اللغة في الجلسة
    session['language'] = language

    # رسالة نجاح باللغة المختارة
    if language == 'ar':
        flash('تم تغيير لغة العرض بنجاح', 'success')
    else:
        flash('Display language changed successfully', 'success')

    # العودة إلى الصفحة السابقة
    referrer = request.referrer
    if referrer:
        return redirect(referrer)
    return redirect(url_for('index'))

# صفحة إدارة المستخدمين
@app.route('/users')
@login_required
def users():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    users_list = User.query.all()
    contractors = Contractor.query.all()

    # الحصول على المناطق الفريدة
    unique_areas = [area[0] for area in db.session.query(MDBPanel.area_name).distinct() if area[0]]

    # إضافة مناطق افتراضية إذا لم توجد مناطق
    if not unique_areas:
        unique_areas = ['عرفة', 'منى', 'مزدلفة', 'العزيزية', 'الشرائع', 'العوالي']

    # الحصول على المقاولين المديرين
    manager_contractors = Contractor.query.filter_by(is_manager=True).all()

    # الحصول على المجموعات
    from models import UserGroup
    user_groups = UserGroup.query.filter_by(is_active=True).all()

    return render_template('users.html',
                           users=users_list,
                           contractors=contractors,
                           unique_areas=unique_areas,
                           manager_contractors=manager_contractors,
                           user_groups=user_groups,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة مستخدم جديد
@app.route('/users/add', methods=['POST'])
def add_user():
    # تعليق مؤقت للتحقق من الصلاحيات
    # if current_user.role != 'admin':
    #     flash('ليس لديك صلاحية للقيام بهذه العملية', 'danger')
    #     return redirect(url_for('index'))

    username = request.form.get('username')
    name = request.form.get('name')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role')


    # الحصول على المناطق المخصصة
    selected_areas = request.form.getlist('assigned_areas')



    # تحديث حقل is_manager من النموذج المخفي إذا كان موجودًا
    if 'is_manager' in request.form:
        is_manager_value = request.form.get('is_manager')
        is_manager = (is_manager_value == '1')





    # التحقق من عدم وجود مستخدم بنفس اسم المستخدم
    existing_user = User.query.filter_by(username=username).first()
    if existing_user:
        flash('اسم المستخدم موجود بالفعل', 'danger')
        return redirect(url_for('users'))

    # إنشاء المستخدم الجديد
    user = User(
        username=username,
        name=name,
        email=email,
        role=role,
        contractor_id=None,
        is_active=True,
        is_manager=False,
        created_at=datetime.datetime.now()
    )
    user.set_password(password)

    # تحديد المناطق المخصصة
    if selected_areas:
        user.set_assigned_areas(selected_areas)

    db.session.add(user)
    db.session.commit()



    flash('تم إضافة المستخدم بنجاح', 'success')
    return redirect(url_for('users'))

# تحرير مستخدم
@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
def edit_user(user_id):
    user = User.query.get_or_404(user_id)

    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        role = request.form.get('role')
        contractor_id = request.form.get('contractor_id') if role in ['contractor', 'regular_contractor'] else None
        is_manager = 'is_manager' in request.form  # التحقق مما إذا كان المستخدم مقاول مدير

        # الحصول على المناطق المخصصة
        selected_areas = request.form.getlist('assigned_areas')

        user.name = name
        user.email = email
        user.role = role
        user.contractor_id = contractor_id
        user.is_manager = is_manager if role == 'contractor' else False  # تعيين حقل المدير فقط إذا كان المستخدم مقاول

        # تحديث المناطق المخصصة
        user.set_assigned_areas(selected_areas)

        db.session.commit()

        flash('تم تحديث المستخدم بنجاح', 'success')
        return redirect(url_for('users'))

    contractors = Contractor.query.all()

    # الحصول على المناطق الفريدة
    unique_areas = [area[0] for area in db.session.query(MDBPanel.area_name).distinct() if area[0]]

    # إضافة مناطق افتراضية إذا لم توجد مناطق
    if not unique_areas:
        unique_areas = ['عرفة', 'منى', 'مزدلفة', 'العزيزية', 'الشرائع', 'العوالي']

    return render_template('edit_user.html',
                           user=user,
                           contractors=contractors,
                           unique_areas=unique_areas,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إعادة تعيين كلمة المرور
@app.route('/users/<int:user_id>/reset-password', methods=['GET', 'POST'])
def reset_password(user_id):

    user = User.query.get_or_404(user_id)

    if request.method == 'POST':
        password = request.form.get('password')

        user.set_password(password)
        db.session.commit()

        flash('تم إعادة تعيين كلمة المرور بنجاح', 'success')
        return redirect(url_for('users'))

    return render_template('reset_password.html',
                           user=user,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تفعيل/تعطيل مستخدم
@app.route('/users/<int:user_id>/toggle')
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)

    user.is_active = not user.is_active
    db.session.commit()

    status = 'تفعيل' if user.is_active else 'تعطيل'
    flash(f'تم {status} المستخدم بنجاح', 'success')
    return redirect(url_for('users'))

# صفحة الخارطة
@app.route('/map')
def map_view():
    # الحصول على البيانات للعرض على الخارطة مع تحميل العلاقات
    panels_query = MDBPanel.query.options(
        db.joinedload(MDBPanel.camp).joinedload(Camp.company).joinedload(Company.country),
        db.joinedload(MDBPanel.company).joinedload(Company.country),
        db.joinedload(MDBPanel.country)
    )

    # تطبيق فلترة المناطق حسب نوع المستخدم
    if current_user.is_authenticated:
        if current_user.role in ['contractor', 'regular_contractor']:
            # فلترة المقاولين
            contractor = Contractor.query.get(current_user.contractor_id)
            if contractor:
                contractor_areas = get_contractor_areas(contractor.id, current_user.is_manager)
                panels_query = panels_query.filter(
                    db.or_(
                        MDBPanel.area_name.in_(contractor_areas),
                        MDBPanel.responsible_contractor_id == contractor.id
                    )
                )
        elif current_user.role != 'admin':
            # فلترة المستخدمين الآخرين حسب المناطق المخصصة
            assigned_areas = current_user.get_assigned_areas()
            if assigned_areas:  # إذا كانت هناك مناطق مخصصة
                panels_query = panels_query.filter(MDBPanel.area_name.in_(assigned_areas))

    panels = panels_query.all()

    # التحقق من وجود إحداثيات للوحات
    valid_panels = [panel for panel in panels if panel.x_coordinate is not None and panel.y_coordinate is not None]

    if len(valid_panels) == 0:
        flash('لا توجد لوحات بإحداثيات صالحة للعرض على الخارطة. يرجى التأكد من تحميل بيانات اللوحات مع الإحداثيات.', 'warning')

    # الحصول على القيم الفريدة للفلاتر والإحصائيات
    unique_areas = [area[0] for area in db.session.query(MDBPanel.area_name).distinct() if area[0]]
    unique_panel_types = [panel_type[0] for panel_type in db.session.query(MDBPanel.panel_type).distinct() if panel_type[0]]

    # الحصول على مناطق الخارطة النشطة
    map_areas = MapArea.query.filter_by(is_active=True).all()

    # الحصول على المقاولين
    contractors = Contractor.query.all()

    # الحصول على آخر القراءات لكل لوحة
    latest_readings = {}
    for panel in panels:
        reading = ElectricalReading.query.filter_by(panel_id=panel.id).order_by(ElectricalReading.timestamp.desc()).first()
        if reading:
            latest_readings[panel.id] = reading

    # الحصول على إعدادات الألوان والحدود
    normal_color = get_setting('normal_color', '#28a745')
    warning_color = get_setting('warning_color', '#ffc107')
    danger_color = get_setting('danger_color', '#dc3545')
    trip_color = get_setting('trip_color', '#6c757d')
    warning_threshold = float(get_setting('default_warning_threshold', '70'))
    danger_threshold = float(get_setting('default_danger_threshold', '80'))

    # الحصول على إعدادات مركز الخارطة
    map_center_str = get_setting('default_map_center', '21.3583, 39.9719')
    try:
        map_center = [float(coord.strip()) for coord in map_center_str.split(',')]
        if len(map_center) != 2:
            map_center = [21.3583, 39.9719]  # القيمة الافتراضية لعرفة، مكة المكرمة
    except (ValueError, IndexError):
        map_center = [21.3583, 39.9719]  # القيمة الافتراضية لعرفة، مكة المكرمة

    map_zoom = int(get_setting('default_map_zoom', '14'))

    # الحصول على إعدادات النقاط الساخنة
    enable_hotspots = get_setting('enable_hotspots', 'true') == 'true'
    hotspot_threshold = float(get_setting('hotspot_threshold', '75'))

    # إعداد بيانات اللوحات للخارطة
    panels_data = []
    for panel in valid_panels:
        panel_data = {
            'id': panel.id,
            'mdb': panel.mdb,
            'maximo_tag': str(panel.maximo_tag).zfill(7),
            'x_coordinate': panel.x_coordinate,
            'y_coordinate': panel.y_coordinate,
            'area_name': panel.area_name,
            'panel_type': panel.panel_type,
            'status': panel.status,
            'notes': panel.notes,
            'implementation_year': panel.implementation_year,
            'breaker_capacity': panel.breaker_capacity,
            'responsible_contractor_id': panel.responsible_contractor_id,
            'is_scada_connected': panel.is_scada_connected,  # حالة الارتباط بنظام سكادا
            # القيم الافتراضية للحالة الكهربائية
            'current_status': 'normal',
            'voltage_status': 'normal',
            'is_tripped': False,
            'current': None,
            'voltage': None,
            'power': None,
            'load_percentage': None,
            # معلومات الشركة والدولة والاتصال
            'company_name': None,
            'country_name': None,
            'contact_person': None,
            'company_phone': None,
            'camp_number': None,
            'square_number': None,
            'company_id': None,
            'country_id': None
        }

        # إضافة معلومات المخيم والشركة والدولة
        # أولاً: البحث عن المخيم المرتبط باللوحة
        if hasattr(panel, 'camp') and panel.camp:
            panel_data['camp_number'] = panel.camp.camp_number
            panel_data['square_number'] = panel.camp.square_number

            # إضافة معلومات الشركة من المخيم
            if hasattr(panel.camp, 'company') and panel.camp.company:
                panel_data['company_name'] = panel.camp.company.name
                panel_data['contact_person'] = panel.camp.company.contact_person
                panel_data['company_phone'] = panel.camp.company.phone
                panel_data['company_id'] = panel.camp.company.id

                # إضافة معلومات الدولة من الشركة
                if hasattr(panel.camp.company, 'country') and panel.camp.company.country:
                    panel_data['country_name'] = panel.camp.company.country.name
                    panel_data['country_id'] = panel.camp.company.country.id

        # ثانياً: إذا لم تكن اللوحة مرتبطة بمخيم، ابحث عن الشركة المرتبطة مباشرة
        elif hasattr(panel, 'company') and panel.company:
            panel_data['company_name'] = panel.company.name
            panel_data['contact_person'] = panel.company.contact_person
            panel_data['company_phone'] = panel.company.phone
            panel_data['company_id'] = panel.company.id

            if hasattr(panel.company, 'country') and panel.company.country:
                panel_data['country_name'] = panel.company.country.name
                panel_data['country_id'] = panel.company.country.id

        # ثالثاً: إذا لم تكن هناك شركة مرتبطة، ابحث عن الدولة المرتبطة مباشرة
        elif hasattr(panel, 'country') and panel.country:
            panel_data['country_name'] = panel.country.name
            panel_data['country_id'] = panel.country.id

        # إضافة معلومات القراءات الكهربائية إذا كانت متوفرة
        if panel.id in latest_readings:
            reading = latest_readings[panel.id]

            # تحديث البيانات الكهربائية
            panel_data.update({
                'current': reading.current,
                'voltage': reading.voltage,
                'power': reading.power,
                'current_status': reading.current_status,
                'voltage_status': reading.voltage_status,
                'last_reading': reading.timestamp.strftime('%Y-%m-%d %H:%M')
            })

            # تحديد حالة الفصل (Trip)
            is_tripped = False
            if reading.current is not None and reading.voltage is not None:
                if reading.current == 0 and reading.voltage > 0:
                    is_tripped = True
                elif reading.current is not None and reading.current < 0.1 and reading.voltage is not None and reading.voltage < 10:
                    is_tripped = True

            panel_data['is_tripped'] = is_tripped

            # حساب نسبة الحمل إذا كانت سعة القاطع متوفرة
            if panel.breaker_capacity and reading.current is not None and panel.breaker_capacity > 0:
                try:
                    panel_data['load_percentage'] = (reading.current / panel.breaker_capacity) * 100
                except (TypeError, ZeroDivisionError):
                    panel_data['load_percentage'] = None

        panels_data.append(panel_data)

    # الحصول على قوائم الدول والشركات للفلترة
    countries = Country.query.filter_by(is_active=True).order_by(Country.name).all()
    companies = Company.query.filter_by(is_active=True).order_by(Company.name).all()

    return render_template('map.html',
                           panels=panels,
                           panels_data=json.dumps(panels_data),
                           map_areas=map_areas,
                           unique_areas=unique_areas,
                           unique_panel_types=unique_panel_types,
                           contractors=contractors,
                           latest_readings=latest_readings,
                           normal_color=normal_color,
                           warning_color=warning_color,
                           danger_color=danger_color,
                           trip_color=trip_color,
                           warning_threshold=warning_threshold,
                           danger_threshold=danger_threshold,
                           map_center=map_center,
                           map_zoom=map_zoom,
                           enable_hotspots=enable_hotspots,
                           hotspot_threshold=hotspot_threshold,
                           countries=countries,
                           companies=companies,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# عرض تفاصيل اللوحة
@app.route('/panels/<int:panel_id>')
@login_required
def panel_details(panel_id):
    panel = MDBPanel.query.get_or_404(panel_id)

    # التحقق من صلاحية الوصول حسب نوع المستخدم
    has_access = False

    if current_user.role == 'admin':
        # المدير يملك صلاحية لرؤية جميع اللوحات
        has_access = True
    elif current_user.role in ['contractor', 'regular_contractor']:
        # المقاولين
        contractor = Contractor.query.get_or_404(current_user.contractor_id)
        contractor_areas = get_contractor_areas(contractor.id, current_user.is_manager)

        if panel.area_name in contractor_areas or panel.responsible_contractor_id == contractor.id:
            has_access = True
    elif current_user.role == 'readings_recorder':
        # مسجل القراءات يملك صلاحية محدودة
        # يمكنه رؤية اللوحة فقط إذا كان لديه مهام مرتبطة بها
        from models import UserGroupMembership
        user_groups = UserGroupMembership.query.filter_by(
            user_id=current_user.id,
            is_active=True
        ).all()
        user_group_ids = [membership.group_id for membership in user_groups]

        # فحص إذا كان لديه بلاغات أو طلبات فحص مرتبطة بهذه اللوحة
        related_issues = Issue.query.filter(
            Issue.panel_id == panel_id,
            db.or_(
                Issue.assignee_id == current_user.id,
                Issue.responsible_person == current_user.name,
                Issue.assigned_group_id.in_(user_group_ids) if user_group_ids else False
            )
        ).count()

        related_requests = InspectionRequest.query.filter(
            InspectionRequest.panel_id == panel_id,
            db.or_(
                InspectionRequest.assignee_id == current_user.id,
                InspectionRequest.assigned_to == current_user.id,
                InspectionRequest.responsible_person == current_user.name,
                InspectionRequest.assigned_group_id.in_(user_group_ids) if user_group_ids else False
            )
        ).count()

        if related_issues > 0 or related_requests > 0:
            has_access = True
    elif current_user.role == 'user':
        # المستخدم العادي يملك صلاحية حسب المناطق المخصصة له
        user_areas = current_user.get_assigned_areas()
        if not user_areas or panel.area_name in user_areas:
            has_access = True

    if not has_access:
        flash('ليس لديك صلاحية للوصول إلى هذه اللوحة', 'danger')
        return redirect(url_for('index'))

    contractors = Contractor.query.all()

    # الحصول على البلاغات المرتبطة باللوحة
    issues = Issue.query.filter_by(panel_id=panel_id).all()

    # الحصول على القراءات اليدوية المرتبطة باللوحة
    manual_readings = ManualReading.query.filter_by(panel_id=panel_id).order_by(ManualReading.timestamp.desc()).all()

    # إنشاء رابط Google Maps إذا لم يكن موجودًا
    if not panel.location_url and panel.x_coordinate and panel.y_coordinate:
        panel.location_url = create_google_maps_url(panel.x_coordinate, panel.y_coordinate)
        db.session.commit()

    return render_template('panel_details.html',
                           panel=panel,
                           issues=issues,
                           contractors=contractors,
                           manual_readings=manual_readings,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة تفاصيل القراءات الكهربائية للوحة (الدالة القديمة - سيتم حذفها)
@app.route('/panels/<int:panel_id>/readings-old')
def panel_readings_old(panel_id):
    panel = MDBPanel.query.get_or_404(panel_id)

    # التحقق مما إذا كان المستخدم مقاول أو مقاول عادي وليس لديه صلاحية للوصول إلى هذه اللوحة
    if current_user.is_authenticated and current_user.role in ['contractor', 'regular_contractor']:
        has_access = False
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # الحصول على مناطق المسؤولية للمقاول
        contractor_areas = get_contractor_areas(contractor.id, current_user.is_manager if hasattr(current_user, 'is_manager') else False)

        # التحقق مما إذا كانت اللوحة في منطقة مسؤولية المقاول
        if panel.area_name in contractor_areas or panel.responsible_contractor_id == contractor.id:
            has_access = True

        if not has_access:
            flash('ليس لديك صلاحية للوصول إلى هذه اللوحة', 'danger')
            return redirect(url_for('contractor_dashboard'))

    # الحصول على عدد الساعات للفلترة
    hours = request.args.get('hours', '24')
    try:
        hours = int(hours)
    except ValueError:
        hours = 24

    # حساب وقت البداية
    start_time = datetime.datetime.now() - datetime.timedelta(hours=hours)

    # الحصول على القراءات
    readings = ElectricalReading.query.filter(
        ElectricalReading.panel_id == panel_id,
        ElectricalReading.timestamp >= start_time
    ).order_by(ElectricalReading.timestamp.desc()).all()

    # الحصول على التنبيهات
    alerts = Alert.query.filter(
        Alert.panel_id == panel_id,
        Alert.timestamp >= start_time
    ).order_by(Alert.timestamp.desc()).all()

    # إعداد بيانات الرسوم البيانية
    timestamps = []
    current_values = []
    voltage_values = []
    power_values = []
    load_values = []

    if readings:
        for reading in reversed(readings):
            timestamps.append(reading.timestamp.strftime('%Y-%m-%d %H:%M'))
            current_values.append(reading.current)
            voltage_values.append(reading.voltage)
            power_values.append(reading.power)
            load_values.append(reading.load)

    # الحصول على المقاولين للعرض في الصفحة
    contractors = Contractor.query.all()

    # إضافة دالة get_setting إلى قالب Jinja
    def template_get_setting(key, default=None):
        return get_setting(key, default)

    return render_template('panel_readings.html',
                           panel=panel,
                           readings=readings,
                           alerts=alerts,
                           hours=hours,
                           timestamps=timestamps,
                           current_values=current_values,
                           voltage_values=voltage_values,
                           power_values=power_values,
                           load_values=load_values,
                           contractors=contractors,
                           get_setting=template_get_setting,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))



# تحديث حالة اللوحة ومعلومات القواطع
@app.route('/panels/<int:panel_id>/update-status', methods=['POST'])
def update_panel_status(panel_id):
    panel = MDBPanel.query.get_or_404(panel_id)

    # تحديث الحالة والملاحظات
    status = request.form.get('status')
    notes = request.form.get('notes')

    if status:
        panel.status = status

    if notes is not None:
        panel.notes = notes

    # تحديث معلومات القواطع والحدود
    breaker_capacity = request.form.get('breaker_capacity')
    min_voltage = request.form.get('min_voltage')
    max_voltage = request.form.get('max_voltage')
    warning_threshold = request.form.get('warning_threshold')
    danger_threshold = request.form.get('danger_threshold')
    location_url = request.form.get('location_url')
    responsible_contractor_id = request.form.get('responsible_contractor_id')

    if breaker_capacity:
        try:
            panel.breaker_capacity = float(breaker_capacity)
        except ValueError:
            pass

    if min_voltage:
        try:
            panel.min_voltage = float(min_voltage)
        except ValueError:
            pass

    if max_voltage:
        try:
            panel.max_voltage = float(max_voltage)
        except ValueError:
            pass

    if warning_threshold:
        try:
            panel.warning_threshold = float(warning_threshold)
        except ValueError:
            pass

    if danger_threshold:
        try:
            panel.danger_threshold = float(danger_threshold)
        except ValueError:
            pass

    # إذا لم يتم تحديد رابط موقع، قم بإنشائه من الإحداثيات
    if not location_url and panel.x_coordinate and panel.y_coordinate:
        location_url = create_google_maps_url(panel.x_coordinate, panel.y_coordinate)

    panel.location_url = location_url

    if responsible_contractor_id:
        panel.responsible_contractor_id = int(responsible_contractor_id)
    else:
        panel.responsible_contractor_id = None

    db.session.commit()

    flash('تم تحديث معلومات اللوحة بنجاح', 'success')
    return redirect(url_for('panel_details', panel_id=panel_id))

# صفحة إدارة المقاولين
@app.route('/contractors')
@login_required
def contractors():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    contractors_list = Contractor.query.all()

    # الحصول على المناطق الفريدة
    unique_areas = [area[0] for area in db.session.query(MDBPanel.area_name).distinct() if area[0]]

    # الحصول على المقاولين المديرين
    manager_contractors = Contractor.query.filter_by(is_manager=True).all()

    return render_template('contractors.html',
                           contractors=contractors_list,
                           unique_areas=unique_areas,
                           manager_contractors=manager_contractors,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة استيراد بيانات المقاولين
@app.route('/contractors/import', methods=['GET', 'POST'])
def import_contractors():
    if request.method == 'POST':
        # التحقق من وجود ملف
        if 'file' not in request.files:
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)

        file = request.files['file']

        # التحقق من اسم الملف
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)

        # التحقق من امتداد الملف
        if not allowed_file(file.filename, {'xlsx', 'xls'}):
            flash('امتداد الملف غير مسموح به. يرجى استخدام ملفات Excel (.xlsx, .xls)', 'danger')
            return redirect(request.url)

        try:
            # حفظ الملف مؤقتًا
            temp_dir = tempfile.gettempdir()
            temp_file_path = os.path.join(temp_dir, f"{uuid.uuid4()}.xlsx")
            file.save(temp_file_path)

            # قراءة الملف باستخدام pandas
            df = pd.read_excel(temp_file_path)

            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ['اسم المقاول', 'الشخص المسؤول', 'رقم الهاتف', 'البريد الإلكتروني', 'المناطق المسؤول عنها']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                flash(f'الملف لا يحتوي على الأعمدة المطلوبة: {", ".join(missing_columns)}', 'danger')
                os.remove(temp_file_path)
                return redirect(request.url)

            # عرض معاينة البيانات
            preview_data = df.head(5)

            return render_template('contractors_import.html',
                                  preview_data=preview_data,
                                  temp_file_path=temp_file_path,
                                  current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

        except Exception as e:
            flash(f'حدث خطأ أثناء قراءة الملف: {str(e)}', 'danger')
            return redirect(request.url)

    return render_template('contractors_import.html',
                          current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تأكيد استيراد بيانات المقاولين
@app.route('/contractors/import/confirm', methods=['POST'])
def confirm_import_contractors():
    file_path = request.form.get('file_path')

    if not file_path or not os.path.exists(file_path):
        flash('الملف غير موجود أو تم حذفه', 'danger')
        return redirect(url_for('import_contractors'))

    try:
        # قراءة الملف
        df = pd.read_excel(file_path)

        # إحصائيات الاستيراد
        import_results = {
            'added_contractors': 0,
            'updated_contractors': 0,
            'added_teams': 0,
            'updated_teams': 0,
            'added_members': 0
        }

        # معالجة بيانات المقاولين
        for _, row in df.iterrows():
            contractor_name = row['اسم المقاول']
            contact_person = row['الشخص المسؤول']
            phone = str(row['رقم الهاتف'])
            email = row['البريد الإلكتروني']
            area_responsibility = row['المناطق المسؤول عنها']

            # البحث عن المقاول في قاعدة البيانات
            contractor = Contractor.query.filter_by(name=contractor_name).first()

            if contractor:
                # تحديث بيانات المقاول الموجود
                contractor.contact_person = contact_person
                contractor.phone = phone
                contractor.email = email
                contractor.area_responsibility = area_responsibility
                import_results['updated_contractors'] += 1
            else:
                # إنشاء مقاول جديد
                contractor = Contractor(
                    name=contractor_name,
                    contact_person=contact_person,
                    phone=phone,
                    email=email,
                    area_responsibility=area_responsibility,
                    created_at=datetime.datetime.now()
                )
                db.session.add(contractor)
                import_results['added_contractors'] += 1

            # حفظ التغييرات لضمان وجود معرف للمقاول
            db.session.commit()

            # معالجة بيانات الفرق إذا كانت موجودة
            if 'اسم الفريق' in row and pd.notna(row['اسم الفريق']):
                team_name = row['اسم الفريق']
                team_description = row.get('وصف الفريق', '')
                team_areas = row.get('مناطق مسؤولية الفريق', area_responsibility)

                # البحث عن الفريق في قاعدة البيانات
                team = ContractorTeam.query.filter_by(contractor_id=contractor.id, name=team_name).first()

                if team:
                    # تحديث بيانات الفريق الموجود
                    team.description = team_description
                    team.area_responsibility = team_areas
                    import_results['updated_teams'] += 1
                else:
                    # إنشاء فريق جديد
                    team = ContractorTeam(
                        contractor_id=contractor.id,
                        name=team_name,
                        description=team_description,
                        area_responsibility=team_areas,
                        is_active=True,
                        created_at=datetime.datetime.now()
                    )
                    db.session.add(team)
                    import_results['added_teams'] += 1

                # حفظ التغييرات لضمان وجود معرف للفريق
                db.session.commit()

                # معالجة بيانات أعضاء الفريق إذا كانت موجودة
                if 'أعضاء الفريق' in row and pd.notna(row['أعضاء الفريق']):
                    members_data = str(row['أعضاء الفريق']).split(';')

                    for member_data in members_data:
                        if ':' in member_data:
                            member_parts = member_data.split(':')
                            member_name = member_parts[0].strip()
                            member_info = member_parts[1].strip() if len(member_parts) > 1 else ''

                            # استخراج معلومات العضو
                            member_position = ''
                            member_phone = ''
                            member_email = ''

                            if ',' in member_info:
                                info_parts = member_info.split(',')
                                member_position = info_parts[0].strip() if len(info_parts) > 0 else ''
                                member_phone = info_parts[1].strip() if len(info_parts) > 1 else ''
                                member_email = info_parts[2].strip() if len(info_parts) > 2 else ''

                            # إنشاء عضو جديد
                            member = ContractorTeamMember(
                                contractor_id=contractor.id,
                                team_id=team.id,
                                name=member_name,
                                position=member_position,
                                phone=member_phone,
                                email=member_email,
                                is_active=True,
                                created_at=datetime.datetime.now()
                            )
                            db.session.add(member)
                            import_results['added_members'] += 1

        # حفظ جميع التغييرات
        db.session.commit()

        # حذف الملف المؤقت
        os.remove(file_path)

        return render_template('contractors_import.html',
                              import_results=import_results,
                              current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

    except Exception as e:
        flash(f'حدث خطأ أثناء استيراد البيانات: {str(e)}', 'danger')
        return redirect(url_for('import_contractors'))

# تحميل قالب استيراد المقاولين
@app.route('/contractors/import/template')
def download_contractors_template():
    # إنشاء مصنف Excel جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات المقاولين"

    # تعريف الأعمدة
    columns = [
        'اسم المقاول',
        'الشخص المسؤول',
        'رقم الهاتف',
        'البريد الإلكتروني',
        'المناطق المسؤول عنها',
        'اسم الفريق',
        'وصف الفريق',
        'مناطق مسؤولية الفريق',
        'أعضاء الفريق'
    ]

    # إضافة الأعمدة
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="B89966", end_color="B89966", fill_type="solid")

    # إضافة بيانات مثال
    example_data = [
        [
            'شركةأزهر ',
            'ابو القاسم',
            '0555123456',
            'ahmed@example.com',
            'منى, عرفات',
            'فريق الصيانة',
            'مسؤول عن صيانة اللوحات الكهربائية',
            'منى',
            'محمد علي:مهندس كهرباء,0555111222,m.ali@example.com; خالد أحمد:فني,0555333444,khaled@example.com'
        ],
        [
            'كنترول تك',
            'وليد',
            '0555987654',
            'saeed@example.com',
            'مزدلفة, عرفات',
            'فريق الطوارئ',
            'مسؤول عن حالات الطوارئ',
            'مزدلفة',
            'عمر سعيد:مشرف,0555666777,omar@example.com; فهد محمد:فني,0555888999,fahad@example.com'
        ]
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='right')

    # ضبط عرض الأعمدة
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # إضافة ملاحظات توضيحية
    ws.cell(row=4, column=1).value = "ملاحظات:"
    ws.cell(row=4, column=1).font = Font(bold=True)

    notes = [
        "1. يجب ملء الأعمدة الخمسة الأولى لكل مقاول.",
        "2. يمكن إضافة فرق متعددة لنفس المقاول عن طريق تكرار بيانات المقاول مع تغيير اسم الفريق.",
        "3. صيغة أعضاء الفريق: اسم العضو:المنصب,رقم الهاتف,البريد الإلكتروني; اسم العضو الثاني:المنصب,رقم الهاتف,البريد الإلكتروني",
        "4. يمكن ترك حقول الفريق فارغة إذا لم يكن هناك فرق للمقاول."
    ]

    for i, note in enumerate(notes, 5):
        ws.cell(row=i, column=1).value = note
        ws.merge_cells(f'A{i}:I{i}')

    # حفظ الملف
    temp_file = os.path.join(tempfile.gettempdir(), "contractors_template.xlsx")
    wb.save(temp_file)

    return send_file(temp_file,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name="قالب_استيراد_المقاولين.xlsx")

# إضافة مقاول جديد
@app.route('/contractors/add', methods=['POST'])
def add_contractor():

    name = request.form.get('name')
    contact_person = request.form.get('contact_person')
    phone = request.form.get('phone')
    email = request.form.get('email')
    area_responsibility = request.form.get('area_responsibility')
    is_manager = 'is_manager' in request.form
    parent_contractor_id = request.form.get('parent_contractor_id')
    create_user = 'create_user' in request.form

    # تحويل parent_contractor_id إلى None إذا كان فارغًا
    if not parent_contractor_id:
        parent_contractor_id = None

    # التحقق من أن المقاول المدير لا يمكن أن يكون له مقاول أب
    if is_manager and parent_contractor_id:
        parent_contractor_id = None

    # إنشاء المقاول الجديد
    contractor = Contractor(
        name=name,
        contact_person=contact_person,
        phone=phone,
        email=email,
        area_responsibility=area_responsibility,
        is_manager=is_manager,
        parent_contractor_id=parent_contractor_id,
        issues_count=0
    )

    db.session.add(contractor)
    db.session.commit()

    # إنشاء حساب مستخدم للمقاول إذا تم اختيار ذلك
    if create_user:
        username = request.form.get('username')
        password = request.form.get('password')

        # إذا لم يتم تحديد اسم مستخدم، استخدم البريد الإلكتروني
        if not username and email:
            username = email.split('@')[0]  # استخدام الجزء الأول من البريد الإلكتروني
        elif not username:
            # إنشاء اسم مستخدم من الاسم
            username = name.lower().replace(' ', '_')

        # التحقق من عدم وجود مستخدم بنفس اسم المستخدم
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            # إضافة رقم عشوائي إلى اسم المستخدم
            import random
            username = f"{username}_{random.randint(100, 999)}"

        # إذا لم يتم تحديد كلمة مرور، استخدم اسم المستخدم
        if not password:
            password = username

        # إنشاء المستخدم الجديد
        user = User(
            username=username,
            name=contact_person if contact_person else name,
            email=email,
            role='contractor',
            contractor_id=contractor.id,
            is_active=True,
            is_manager=is_manager,
            created_at=datetime.datetime.now()
        )
        user.set_password(password)

        db.session.add(user)
        db.session.commit()

        flash(f'تم إنشاء حساب مستخدم للمقاول باسم {username}', 'success')

    flash('تم إضافة المقاول بنجاح', 'success')
    return redirect(url_for('users'))

# تحرير مقاول
@app.route('/contractors/<int:contractor_id>/edit', methods=['GET', 'POST'])
def edit_contractor(contractor_id):

    contractor = Contractor.query.get_or_404(contractor_id)

    if request.method == 'POST':
        name = request.form.get('name')
        contact_person = request.form.get('contact_person')
        phone = request.form.get('phone')
        email = request.form.get('email')
        area_responsibility = request.form.get('area_responsibility')
        is_manager = 'is_manager' in request.form
        parent_contractor_id = request.form.get('parent_contractor_id')

        # تحويل parent_contractor_id إلى None إذا كان فارغًا
        if not parent_contractor_id:
            parent_contractor_id = None

        # التحقق من أن المقاول المدير لا يمكن أن يكون له مقاول أب
        if is_manager and parent_contractor_id:
            parent_contractor_id = None

        contractor.name = name
        contractor.contact_person = contact_person
        contractor.phone = phone
        contractor.email = email
        contractor.area_responsibility = area_responsibility
        contractor.is_manager = is_manager
        contractor.parent_contractor_id = parent_contractor_id

        db.session.commit()

        flash('تم تحديث المقاول بنجاح', 'success')
        return redirect(url_for('contractors'))

    # الحصول على المناطق الفريدة
    unique_areas = [area[0] for area in db.session.query(MDBPanel.area_name).distinct() if area[0]]

    # الحصول على أعضاء مجموعة المقاول
    team_members = ContractorTeamMember.query.filter_by(contractor_id=contractor_id).all()

    # الحصول على فرق المقاول
    teams = ContractorTeam.query.filter_by(contractor_id=contractor_id).all()

    # الحصول على المقاولين المديرين
    manager_contractors = Contractor.query.filter_by(is_manager=True).all()

    return render_template('edit_contractor.html',
                           contractor=contractor,
                           unique_areas=unique_areas,
                           team_members=team_members,
                           teams=teams,
                           manager_contractors=manager_contractors,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة عضو جديد لمجموعة المقاول
@app.route('/contractors/<int:contractor_id>/team/<int:team_id>/add', methods=['POST'])
def add_team_member(contractor_id, team_id):
    contractor = Contractor.query.get_or_404(contractor_id)
    team = ContractorTeam.query.get_or_404(team_id)

    # التحقق مما إذا تم اختيار مستخدم موجود
    user_id = request.form.get('user_id')

    if user_id:
        # إضافة مستخدم موجود إلى الفريق
        user = User.query.get_or_404(int(user_id))

        # التحقق من أن المستخدم ينتمي للمقاول المحدد
        if user.contractor_id != contractor_id:
            flash('غير مسموح بإضافة هذا المستخدم للفريق', 'danger')
            return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

        # إنشاء عضو جديد من بيانات المستخدم
        member = ContractorTeamMember(
            contractor_id=contractor_id,
            team_id=team_id,
            name=user.name,
            position=request.form.get('position', ''),
            phone=request.form.get('phone', ''),
            email=user.email,
            is_active=True,
            created_at=datetime.datetime.now()
        )

        db.session.add(member)

        # تحديث بيانات المستخدم
        user.team_id = team_id

        db.session.commit()

        flash('تم إضافة المستخدم الموجود إلى الفريق بنجاح', 'success')
    else:
        # إضافة عضو جديد
        name = request.form.get('name')
        position = request.form.get('position', '')
        phone = request.form.get('phone', '')
        email = request.form.get('email', '')
        is_active = 'is_active' in request.form
        create_user = 'create_user' in request.form
        username = request.form.get('username', '')

        # إنشاء عضو جديد
        member = ContractorTeamMember(
            contractor_id=contractor_id,
            team_id=team_id,
            name=name,
            position=position,
            phone=phone,
            email=email,
            is_active=is_active,
            created_at=datetime.datetime.now()
        )

        db.session.add(member)
        db.session.commit()

        # إنشاء حساب مستخدم إذا تم اختيار ذلك
        if create_user:
            # إذا لم يتم تحديد اسم مستخدم، استخدم البريد الإلكتروني
            if not username and email:
                username = email.split('@')[0]  # استخدام الجزء الأول من البريد الإلكتروني
            elif not username:
                # إنشاء اسم مستخدم من الاسم
                username = name.lower().replace(' ', '_')

            # التحقق من عدم وجود مستخدم بنفس اسم المستخدم
            existing_user = User.query.filter_by(username=username).first()
            if existing_user:
                # إضافة رقم عشوائي إلى اسم المستخدم
                import random
                username = f"{username}_{random.randint(100, 999)}"

            # إنشاء المستخدم الجديد
            user = User(
                username=username,
                name=name,
                email=email,
                role='contractor',
                contractor_id=contractor_id,
                team_id=team_id,
                is_active=is_active,
                is_manager=False,
                created_at=datetime.datetime.now()
            )
            user.password_hash = username  # تعيين كلمة المرور نفس اسم المستخدم

            db.session.add(user)
            db.session.commit()

            flash(f'تم إنشاء حساب مستخدم باسم {username}', 'success')

        flash('تم إضافة عضو المجموعة بنجاح', 'success')

    return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

# تعديل عضو مجموعة المقاول
@app.route('/contractors/<int:contractor_id>/team/<int:team_id>/edit', methods=['POST'])
def edit_team_member(contractor_id, team_id):
    member_id = request.form.get('member_id')
    member = ContractorTeamMember.query.get_or_404(member_id)

    # التحقق من أن العضو ينتمي للمقاول المحدد
    if member.contractor_id != contractor_id or member.team_id != team_id:
        flash('غير مسموح بتعديل هذا العضو', 'danger')
        return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

    name = request.form.get('name')
    position = request.form.get('position', '')
    phone = request.form.get('phone', '')
    email = request.form.get('email', '')
    is_active = 'is_active' in request.form
    create_user = 'create_user' in request.form
    username = request.form.get('username', '')

    # تحديث بيانات العضو
    member.name = name
    member.position = position
    member.phone = phone
    member.email = email
    member.is_active = is_active

    db.session.commit()

    # البحث عن حساب مستخدم مرتبط بهذا العضو
    user = User.query.filter_by(name=member.name, contractor_id=contractor_id, team_id=team_id).first()

    # إنشاء حساب مستخدم إذا تم اختيار ذلك ولم يكن هناك حساب موجود
    if create_user and not user:
        # إذا لم يتم تحديد اسم مستخدم، استخدم البريد الإلكتروني
        if not username and email:
            username = email.split('@')[0]  # استخدام الجزء الأول من البريد الإلكتروني
        elif not username:
            # إنشاء اسم مستخدم من الاسم
            username = name.lower().replace(' ', '_')

        # التحقق من عدم وجود مستخدم بنفس اسم المستخدم
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            # إضافة رقم عشوائي إلى اسم المستخدم
            import random
            username = f"{username}_{random.randint(100, 999)}"

        # إنشاء المستخدم الجديد
        user = User(
            username=username,
            name=name,
            email=email,
            role='contractor',
            contractor_id=contractor_id,
            team_id=team_id,
            is_active=is_active,
            is_manager=False,
            created_at=datetime.datetime.now()
        )
        user.password_hash = username  # تعيين كلمة المرور نفس اسم المستخدم

        db.session.add(user)
        db.session.commit()

        flash(f'تم إنشاء حساب مستخدم باسم {username}', 'success')
    # تحديث بيانات المستخدم إذا كان موجودًا
    elif user:
        user.name = name
        user.email = email
        user.is_active = is_active
        db.session.commit()

    flash('تم تحديث بيانات عضو المجموعة بنجاح', 'success')
    return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

# حذف عضو مجموعة المقاول
@app.route('/contractors/<int:contractor_id>/team/<int:team_id>/member/<int:member_id>/delete')
def delete_team_member(contractor_id, team_id, member_id):
    member = ContractorTeamMember.query.get_or_404(member_id)

    # التحقق من أن العضو ينتمي للمقاول والفريق المحدد
    if member.contractor_id != contractor_id or member.team_id != team_id:
        flash('غير مسموح بحذف هذا العضو', 'danger')
        return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

    # البحث عن حساب مستخدم مرتبط بهذا العضو
    user = User.query.filter_by(name=member.name, contractor_id=contractor_id, team_id=team_id).first()

    # حذف المستخدم إذا كان موجودًا
    if user:
        db.session.delete(user)

    db.session.delete(member)
    db.session.commit()

    flash('تم حذف عضو المجموعة بنجاح', 'success')
    return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

# صفحة إدارة فرق المقاول
@app.route('/contractors/<int:contractor_id>/teams')
def contractor_teams(contractor_id):
    contractor = Contractor.query.get_or_404(contractor_id)
    teams = ContractorTeam.query.filter_by(contractor_id=contractor_id).all()

    # الحصول على المناطق الفريدة
    unique_areas = [area[0] for area in db.session.query(MDBPanel.area_name).distinct() if area[0]]

    return render_template('contractor_teams.html',
                           contractor=contractor,
                           teams=teams,
                           unique_areas=unique_areas,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة فريق جديد للمقاول
@app.route('/contractors/<int:contractor_id>/teams/add', methods=['POST'])
def add_team(contractor_id):
    contractor = Contractor.query.get_or_404(contractor_id)

    name = request.form.get('name')
    areas = request.form.getlist('area_responsibility')
    area_responsibility = ','.join(areas) if areas else ''
    description = request.form.get('description', '')
    is_active = 'is_active' in request.form
    parent_team_id = request.form.get('parent_team_id')

    # التحقق من صحة الفريق الأب
    if parent_team_id:
        parent_team = ContractorTeam.query.get(parent_team_id)
        if not parent_team or parent_team.contractor_id != contractor_id:
            parent_team_id = None

    # إنشاء فريق جديد
    team = ContractorTeam(
        contractor_id=contractor_id,
        name=name,
        area_responsibility=area_responsibility,
        description=description,
        is_active=is_active,
        parent_team_id=parent_team_id,
        created_at=datetime.datetime.now()
    )

    db.session.add(team)
    db.session.commit()

    flash('تم إضافة الفريق بنجاح', 'success')
    return redirect(url_for('contractor_teams', contractor_id=contractor_id))

# إضافة فريق جديد (من صفحة المستخدمين)
@app.route('/teams/add', methods=['POST'])
def add_team_from_users():
    contractor_id = request.form.get('contractor_id')
    contractor = Contractor.query.get_or_404(contractor_id)

    name = request.form.get('name')
    description = request.form.get('description', '')
    parent_team_id = request.form.get('parent_team_id')
    areas = request.form.getlist('area_responsibility')
    area_responsibility = json.dumps(areas) if areas else ''
    is_active = 'is_active' in request.form
    team_members = request.form.getlist('team_members')

    # تحويل parent_team_id إلى None إذا كان فارغًا
    if not parent_team_id:
        parent_team_id = None

    # إنشاء فريق جديد
    team = ContractorTeam(
        contractor_id=contractor_id,
        name=name,
        description=description,
        parent_team_id=parent_team_id,
        area_responsibility=area_responsibility,
        is_active=is_active,
        created_at=datetime.datetime.now()
    )

    db.session.add(team)
    db.session.commit()

    # إضافة أعضاء الفريق
    for member_id in team_members:
        user = User.query.get(member_id)
        if user and user.contractor_id == int(contractor_id):
            # تحديث المستخدم بمعرف الفريق
            user.team_id = team.id

            # إنشاء عضو فريق جديد
            team_member = ContractorTeamMember(
                contractor_id=contractor_id,
                team_id=team.id,
                name=user.name,
                email=user.email,
                is_active=True,
                created_at=datetime.datetime.now()
            )

            db.session.add(team_member)

    db.session.commit()

    flash(f'تم إضافة الفريق "{name}" بنجاح', 'success')
    return redirect(url_for('users'))

# API للحصول على فرق المقاول والمستخدمين
@app.route('/api/contractors/<int:contractor_id>/teams')
def get_contractor_teams_api(contractor_id):
    contractor = Contractor.query.get_or_404(contractor_id)

    # الحصول على فرق المقاول
    teams = ContractorTeam.query.filter_by(contractor_id=contractor_id).all()

    # إذا كان المقاول مديرًا، أضف أيضًا فرق المقاولين التابعين له
    if contractor.is_manager:
        # الحصول على المقاولين التابعين
        sub_contractors = Contractor.query.filter_by(parent_contractor_id=contractor_id).all()

        # إضافة فرق المقاولين التابعين
        for sub_contractor in sub_contractors:
            sub_teams = ContractorTeam.query.filter_by(contractor_id=sub_contractor.id).all()
            teams.extend(sub_teams)

    teams_data = [{'id': team.id, 'name': team.name} for team in teams]

    # الحصول على مستخدمي المقاول الذين ليس لديهم فريق
    users = User.query.filter_by(contractor_id=contractor_id, team_id=None).all()

    # إذا كان المقاول مديرًا، أضف أيضًا مستخدمي المقاولين التابعين له
    if contractor.is_manager:
        # الحصول على المقاولين التابعين
        sub_contractors = Contractor.query.filter_by(parent_contractor_id=contractor_id).all()

        # إضافة مستخدمي المقاولين التابعين
        for sub_contractor in sub_contractors:
            sub_users = User.query.filter_by(contractor_id=sub_contractor.id, team_id=None).all()
            users.extend(sub_users)

    users_data = [{'id': user.id, 'name': user.name, 'email': user.email} for user in users]

    return jsonify({'teams': teams_data, 'users': users_data})

# تعديل فريق المقاول
@app.route('/contractors/<int:contractor_id>/teams/edit', methods=['POST'])
def edit_team(contractor_id):
    team_id = request.form.get('team_id')
    team = ContractorTeam.query.get_or_404(team_id)

    # التحقق من أن الفريق ينتمي للمقاول المحدد
    if team.contractor_id != contractor_id:
        flash('غير مسموح بتعديل هذا الفريق', 'danger')
        return redirect(url_for('contractor_teams', contractor_id=contractor_id))

    name = request.form.get('name')
    areas = request.form.getlist('area_responsibility')
    area_responsibility = ','.join(areas) if areas else ''
    description = request.form.get('description', '')
    is_active = 'is_active' in request.form
    parent_team_id = request.form.get('parent_team_id')

    # التحقق من صحة الفريق الأب
    if parent_team_id:
        # التأكد من أن الفريق الأب ليس هو نفسه
        if parent_team_id == str(team_id):
            flash('لا يمكن جعل الفريق أب لنفسه', 'danger')
            return redirect(url_for('contractor_teams', contractor_id=contractor_id))

        parent_team = ContractorTeam.query.get(parent_team_id)
        if not parent_team or parent_team.contractor_id != contractor_id:
            parent_team_id = None

        # التحقق من عدم وجود دورة في العلاقات
        if parent_team and is_descendant(team_id, parent_team_id):
            flash('لا يمكن جعل الفريق الفرعي أب للفريق الأصلي', 'danger')
            return redirect(url_for('contractor_teams', contractor_id=contractor_id))

    # تحديث بيانات الفريق
    team.name = name
    team.area_responsibility = area_responsibility
    team.description = description
    team.is_active = is_active
    team.parent_team_id = parent_team_id

    db.session.commit()

    flash('تم تحديث بيانات الفريق بنجاح', 'success')
    return redirect(url_for('contractor_teams', contractor_id=contractor_id))

# حذف فريق المقاول
@app.route('/contractors/<int:contractor_id>/teams/<int:team_id>/delete')
def delete_team(contractor_id, team_id):
    team = ContractorTeam.query.get_or_404(team_id)

    # التحقق من أن الفريق ينتمي للمقاول المحدد
    if team.contractor_id != contractor_id:
        flash('غير مسموح بحذف هذا الفريق', 'danger')
        return redirect(url_for('contractor_teams', contractor_id=contractor_id))

    db.session.delete(team)
    db.session.commit()

    flash('تم حذف الفريق بنجاح', 'success')
    return redirect(url_for('contractor_teams', contractor_id=contractor_id))

# صفحة إدارة أعضاء الفريق (تم تعطيلها لتجنب التكرار)
# @app.route('/contractors/<int:contractor_id>/teams/<int:team_id>/members')
# def team_members_old(contractor_id, team_id):
#     contractor = Contractor.query.get_or_404(contractor_id)
#     team = ContractorTeam.query.get_or_404(team_id)
#
#     # التحقق من أن الفريق ينتمي للمقاول المحدد
#     if team.contractor_id != contractor_id:
#         flash('غير مسموح بعرض هذا الفريق', 'danger')
#         return redirect(url_for('contractor_teams', contractor_id=contractor_id))
#
#     members = ContractorTeamMember.query.filter_by(team_id=team_id).all()
#
#     # الحصول على المستخدمين المتاحين للإضافة
#     available_users = User.query.filter_by(role='contractor', contractor_id=contractor_id, team_id=None).all()
#
#     return render_template('team_members.html',
#                            contractor=contractor,
#                            team=team,
#                            members=members,
#                            available_users=available_users,
#                            current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة عضو جديد للفريق
@app.route('/contractors/<int:contractor_id>/teams/<int:team_id>/members/add', methods=['POST'])
def add_team_member_to_team(contractor_id, team_id):
    contractor = Contractor.query.get_or_404(contractor_id)
    team = ContractorTeam.query.get_or_404(team_id)

    # التحقق من أن الفريق ينتمي للمقاول المحدد
    if team.contractor_id != contractor_id:
        flash('غير مسموح بإضافة عضو لهذا الفريق', 'danger')
        return redirect(url_for('contractor_teams', contractor_id=contractor_id))

    name = request.form.get('name')
    position = request.form.get('position', '')
    phone = request.form.get('phone', '')
    email = request.form.get('email', '')
    is_active = 'is_active' in request.form

    # إنشاء عضو جديد
    member = ContractorTeamMember(
        contractor_id=contractor_id,
        team_id=team_id,
        name=name,
        position=position,
        phone=phone,
        email=email,
        is_active=is_active,
        created_at=datetime.datetime.now()
    )

    db.session.add(member)
    db.session.commit()

    flash('تم إضافة عضو الفريق بنجاح', 'success')
    return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

# تعديل عضو الفريق
@app.route('/contractors/<int:contractor_id>/teams/<int:team_id>/members/edit', methods=['POST'])
def edit_team_member_in_team(contractor_id, team_id):
    member_id = request.form.get('member_id')
    member = ContractorTeamMember.query.get_or_404(member_id)

    # التحقق من أن العضو ينتمي للفريق والمقاول المحددين
    if member.contractor_id != contractor_id or member.team_id != team_id:
        flash('غير مسموح بتعديل هذا العضو', 'danger')
        return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

    name = request.form.get('name')
    position = request.form.get('position', '')
    phone = request.form.get('phone', '')
    email = request.form.get('email', '')
    is_active = 'is_active' in request.form

    # تحديث بيانات العضو
    member.name = name
    member.position = position
    member.phone = phone
    member.email = email
    member.is_active = is_active

    db.session.commit()

    flash('تم تحديث بيانات عضو الفريق بنجاح', 'success')
    return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

# حذف عضو الفريق
@app.route('/contractors/<int:contractor_id>/teams/<int:team_id>/members/<int:member_id>/delete')
def delete_team_member_from_team(contractor_id, team_id, member_id):
    member = ContractorTeamMember.query.get_or_404(member_id)

    # التحقق من أن العضو ينتمي للفريق والمقاول المحددين
    if member.contractor_id != contractor_id or member.team_id != team_id:
        flash('غير مسموح بحذف هذا العضو', 'danger')
        return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

    db.session.delete(member)
    db.session.commit()

    flash('تم حذف عضو الفريق بنجاح', 'success')
    return redirect(url_for('team_members', contractor_id=contractor_id, team_id=team_id))

# لوحة تحكم المقاول
@app.route('/contractor/dashboard')
@login_required
def contractor_dashboard():
    # التحقق من أن المستخدم هو مقاول أو مقاول عادي
    if current_user.role not in ['contractor', 'regular_contractor'] or not current_user.contractor_id:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    contractor = Contractor.query.get_or_404(current_user.contractor_id)

    # الحصول على مناطق المسؤولية للمقاول
    contractor_areas = []
    if contractor.area_responsibility:
        try:
            # محاولة تحليل البيانات كـ JSON
            contractor_areas = json.loads(contractor.area_responsibility)
        except json.JSONDecodeError:
            # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
            contractor_areas = [area.strip() for area in contractor.area_responsibility.split(',')]

    # الحصول على اللوحات في مناطق المسؤولية
    panels_query = MDBPanel.query

    # إذا كان المستخدم مقاول مدير، يمكنه رؤية جميع اللوحات المرتبطة بالفرق التابعة له
    if hasattr(current_user, 'is_manager') and current_user.is_manager:
        # الحصول على جميع الفرق التابعة للمقاول
        teams = ContractorTeam.query.filter_by(contractor_id=contractor.id).all()

        # جمع جميع المناطق المسؤول عنها الفرق
        team_areas = []
        for team in teams:
            if team.area_responsibility:
                try:
                    # محاولة تحليل البيانات كـ JSON
                    team_areas_json = json.loads(team.area_responsibility)
                    if isinstance(team_areas_json, list):
                        team_areas.extend(team_areas_json)
                    else:
                        team_areas.append(team_areas_json)
                except json.JSONDecodeError:
                    # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                    team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

        # إضافة مناطق الفرق إلى مناطق المقاول
        contractor_areas.extend(team_areas)

        # إزالة التكرار
        contractor_areas = list(set(contractor_areas))

        # تصفية اللوحات حسب المقاول المسؤول أو المنطقة
        if contractor_areas:
            panels_query = panels_query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == contractor.id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            )
        else:
            # إذا لم تكن هناك مناطق محددة، استخدم فقط المقاول المسؤول
            panels_query = panels_query.filter(MDBPanel.responsible_contractor_id == contractor.id)
    else:
        # المقاول العادي يرى فقط اللوحات في مناطق مسؤوليته
        if contractor_areas:
            panels_query = panels_query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == contractor.id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            )
        else:
            # إذا لم تكن هناك مناطق محددة، استخدم فقط المقاول المسؤول
            panels_query = panels_query.filter(MDBPanel.responsible_contractor_id == contractor.id)

    panels = panels_query.all()

    # الحصول على آخر القراءات لكل لوحة
    latest_readings = {}
    for panel in panels:
        reading = ElectricalReading.query.filter_by(panel_id=panel.id).order_by(ElectricalReading.timestamp.desc()).first()
        if reading:
            latest_readings[panel.id] = reading

    # إحصائيات اللوحات
    working_panels_count = sum(1 for panel in panels if panel.status == 'عامل')

    # الحصول على البلاغات المفتوحة للمقاول
    open_issues = Issue.query.filter(
        Issue.contractor_id == contractor.id,
        Issue.status.in_(['مفتوح', 'قيد المعالجة'])
    ).all()

    open_issues_count = len(open_issues)

    # الحصول على اللوحات في حالة خطر
    danger_panels_count = 0
    warning_threshold = float(get_setting('default_warning_threshold', '70'))
    danger_threshold = float(get_setting('default_danger_threshold', '80'))

    # إعداد بيانات اللوحات للخارطة
    panels_data = []
    for panel in panels:
        panel_data = {
            'id': panel.id,
            'mdb': panel.mdb,
            'maximo_tag': str(panel.maximo_tag).zfill(7),
            'x_coordinate': panel.x_coordinate,
            'y_coordinate': panel.y_coordinate,
            'area_name': panel.area_name,
            'panel_type': panel.panel_type,
            'status': panel.status,
            'notes': panel.notes,
            'implementation_year': panel.implementation_year,
            'breaker_capacity': panel.breaker_capacity,
            'responsible_contractor_id': panel.responsible_contractor_id,
            # القيم الافتراضية للحالة الكهربائية
            'current_status': 'normal',
            'voltage_status': 'normal',
            'is_tripped': False,
            'current': None,
            'voltage': None,
            'power': None,
            'load_percentage': None,
            'last_reading': None
        }

        # إضافة معلومات القراءات الكهربائية إذا كانت متوفرة
        if panel.id in latest_readings:
            reading = latest_readings[panel.id]

            # تحديث البيانات الكهربائية
            panel_data.update({
                'current': reading.current,
                'voltage': reading.voltage,
                'power': reading.power,
                'current_status': reading.current_status,
                'voltage_status': reading.voltage_status,
                'last_reading': reading.timestamp.strftime('%Y-%m-%d %H:%M')
            })

            # تحديد حالة الفصل (Trip)
            is_tripped = False
            if reading.current is not None and reading.voltage is not None:
                if reading.current == 0 and reading.voltage > 0:
                    is_tripped = True
                elif reading.current is not None and reading.current < 0.1 and reading.voltage is not None and reading.voltage < 10:
                    is_tripped = True

            panel_data['is_tripped'] = is_tripped

            # حساب نسبة الحمل إذا كانت سعة القاطع متوفرة
            if panel.breaker_capacity and reading.current is not None and panel.breaker_capacity > 0:
                try:
                    load_percentage = (reading.current / panel.breaker_capacity) * 100
                    panel_data['load_percentage'] = load_percentage

                    # زيادة عدد اللوحات في حالة خطر
                    if load_percentage >= danger_threshold or reading.current_status == 'danger' or reading.voltage_status == 'danger' or is_tripped:
                        danger_panels_count += 1
                except (TypeError, ZeroDivisionError):
                    pass

        # إضافة اللوحة إلى البيانات إذا كانت لها إحداثيات
        if panel.x_coordinate is not None and panel.y_coordinate is not None:
            panels_data.append(panel_data)

    # الحصول على إعدادات مركز الخارطة
    map_center_str = get_setting('default_map_center', '21.3891, 39.8579')
    try:
        map_center = [float(coord.strip()) for coord in map_center_str.split(',')]
        if len(map_center) != 2:
            map_center = [21.3891, 39.8579]  # القيمة الافتراضية لعرفة، مكة المكرمة
    except (ValueError, IndexError):
        map_center = [21.3891, 39.8579]  # القيمة الافتراضية لعرفة، مكة المكرمة

    map_zoom = int(get_setting('default_map_zoom', '14'))

    # الحصول على إعدادات الألوان
    normal_color = get_setting('normal_color', '#28a745')
    warning_color = get_setting('warning_color', '#ffc107')
    danger_color = get_setting('danger_color', '#dc3545')
    trip_color = get_setting('trip_color', '#6c757d')

    # تنسيق التاريخ والوقت الحالي لنموذج إضافة القراءة
    current_datetime = datetime.datetime.now().strftime('%Y-%m-%dT%H:%M')

    # الحصول على آخر القراءات اليدوية
    manual_readings_query = ManualReading.query.join(MDBPanel)

    # تصفية القراءات اليدوية حسب مناطق المسؤولية
    if contractor_areas:
        manual_readings_query = manual_readings_query.filter(
            db.or_(
                MDBPanel.responsible_contractor_id == contractor.id,
                MDBPanel.area_name.in_(contractor_areas)
            )
        )
    else:
        manual_readings_query = manual_readings_query.filter(MDBPanel.responsible_contractor_id == contractor.id)

    # الحصول على آخر 5 قراءات يدوية
    manual_readings = manual_readings_query.order_by(ManualReading.timestamp.desc()).limit(5).all()

    # إحصائيات الفرق (للمقاول المدير فقط)
    teams = []
    if current_user.is_manager:
        # الحصول على فرق المقاول
        teams_query = ContractorTeam.query.filter_by(contractor_id=contractor.id)
        teams_list = teams_query.all()

        for team in teams_list:
            # الحصول على مناطق مسؤولية الفريق
            team_areas = []
            if team.area_responsibility:
                try:
                    # محاولة تحليل البيانات كـ JSON
                    team_areas_json = json.loads(team.area_responsibility)
                    if isinstance(team_areas_json, list):
                        team_areas.extend(team_areas_json)
                    else:
                        team_areas.append(team_areas_json)
                except json.JSONDecodeError:
                    # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                    team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

            # الحصول على اللوحات في مناطق مسؤولية الفريق
            team_panels = MDBPanel.query.filter(MDBPanel.area_name.in_(team_areas)).all()

            # الحصول على البلاغات المفتوحة للفريق
            team_open_issues = Issue.query.filter(
                Issue.team_id == team.id,
                Issue.status.in_(['مفتوح', 'قيد المعالجة'])
            ).all()

            # الحصول على إجمالي البلاغات للفريق
            team_all_issues = Issue.query.filter(Issue.team_id == team.id).all()

            # حساب نسبة الأعطال والمعالجة
            issues_percentage = 0
            resolution_percentage = 0

            if len(team_panels) > 0:
                issues_percentage = (len(team_open_issues) / len(team_panels)) * 100

            if len(team_all_issues) > 0:
                resolved_issues = sum(1 for issue in team_all_issues if issue.status == 'مغلق')
                resolution_percentage = (resolved_issues / len(team_all_issues)) * 100

            # إضافة إحصائيات الفريق
            team_stats = {
                'id': team.id,
                'name': team.name,
                'area_responsibility': team.area_responsibility,
                'panels_count': len(team_panels),
                'open_issues_count': len(team_open_issues),
                'issues_percentage': issues_percentage,
                'resolution_percentage': resolution_percentage
            }

            teams.append(team_stats)

    return render_template('contractor_dashboard.html',
                          contractor=contractor,
                          panels=panels,
                          panels_data=json.dumps(panels_data),
                          latest_readings=latest_readings,
                          working_panels_count=working_panels_count,
                          open_issues_count=open_issues_count,
                          danger_panels_count=danger_panels_count,
                          issues=open_issues,
                          manual_readings=manual_readings,
                          warning_threshold=warning_threshold,
                          danger_threshold=danger_threshold,
                          map_center=map_center,
                          map_zoom=map_zoom,
                          normal_color=normal_color,
                          warning_color=warning_color,
                          danger_color=danger_color,
                          trip_color=trip_color,
                          current_datetime=current_datetime,
                          teams=teams,
                          current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة قراءة يدوية
@app.route('/readings/add-manual', methods=['POST'])
@login_required
def add_manual_reading():
    # التحقق من أن المستخدم هو مقاول أو مقاول عادي
    if current_user.role not in ['contractor', 'regular_contractor'] or not current_user.contractor_id:
        flash('ليس لديك صلاحية للقيام بهذه العملية', 'danger')
        return redirect(url_for('index'))

    panel_id = request.form.get('panel_id')
    timestamp_str = request.form.get('timestamp')
    current = request.form.get('current')
    voltage = request.form.get('voltage')
    power = request.form.get('power')
    energy = request.form.get('energy')
    power_factor = request.form.get('power_factor')
    frequency = request.form.get('frequency')
    notes = request.form.get('notes')

    # التحقق من البيانات المطلوبة
    if not panel_id or not timestamp_str or not current or not voltage:
        flash('يرجى ملء جميع الحقول المطلوبة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # التحقق من أن اللوحة موجودة
    panel = MDBPanel.query.get_or_404(panel_id)

    # التحقق من أن اللوحة تنتمي لمناطق مسؤولية المقاول
    contractor = Contractor.query.get_or_404(current_user.contractor_id)
    contractor_areas = []

    if contractor.area_responsibility:
        try:
            # محاولة تحليل البيانات كـ JSON
            contractor_areas = json.loads(contractor.area_responsibility)
        except json.JSONDecodeError:
            # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
            contractor_areas = [area.strip() for area in contractor.area_responsibility.split(',')]

    if panel.responsible_contractor_id != contractor.id and panel.area_name not in contractor_areas:
        flash('ليس لديك صلاحية لإضافة قراءات لهذه اللوحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # تحويل البيانات إلى الأنواع المناسبة
    try:
        timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%dT%H:%M')
        current_value = float(current)
        voltage_value = float(voltage)
        power_value = float(power) if power else None
        energy_value = float(energy) if energy else None
        power_factor_value = float(power_factor) if power_factor else None
        frequency_value = float(frequency) if frequency else None
    except ValueError:
        flash('قيم غير صالحة. يرجى التأكد من إدخال أرقام صحيحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # حساب الحمل
    load_value = None
    if current_value is not None:
        load_value = current_value

    # تحديد حالة القراءات
    current_status = 'normal'
    voltage_status = 'normal'
    power_status = 'normal'

    # تحديد حالة التيار
    if panel.breaker_capacity and current_value is not None:
        warning_threshold = panel.warning_threshold or float(get_setting('default_warning_threshold', '70'))
        danger_threshold = panel.danger_threshold or float(get_setting('default_danger_threshold', '80'))

        load_percentage = (current_value / panel.breaker_capacity) * 100

        if load_percentage >= danger_threshold:
            current_status = 'danger'
        elif load_percentage >= warning_threshold:
            current_status = 'warning'

    # تحديد حالة الجهد
    if voltage_value is not None:
        min_voltage = panel.min_voltage or float(get_setting('default_min_voltage', '210'))
        max_voltage = panel.max_voltage or float(get_setting('default_max_voltage', '250'))

        if voltage_value < min_voltage or voltage_value > max_voltage:
            voltage_status = 'danger'
        elif voltage_value < min_voltage * 1.05 or voltage_value > max_voltage * 0.95:
            voltage_status = 'warning'

    # إنشاء قراءة جديدة
    reading = ElectricalReading(
        panel_id=panel_id,
        timestamp=timestamp,
        current=current_value,
        voltage=voltage_value,
        power=power_value,
        energy=energy_value,
        power_factor=power_factor_value,
        frequency=frequency_value,
        load=load_value,
        current_status=current_status,
        voltage_status=voltage_status,
        power_status=power_status
    )

    db.session.add(reading)

    # حفظ الصورة إذا تم تحميلها
    if 'image' in request.files and request.files['image'].filename:
        image_file = request.files['image']

        if allowed_file(image_file.filename, {'jpg', 'jpeg', 'png', 'gif'}):
            # إنشاء اسم فريد للملف
            filename = secure_filename(f"{panel.mdb}_{timestamp.strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}.{image_file.filename.rsplit('.', 1)[1].lower()}")

            # التأكد من وجود المجلد
            readings_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'readings')
            if not os.path.exists(readings_folder):
                os.makedirs(readings_folder)

            # حفظ الملف
            file_path = os.path.join(readings_folder, filename)
            image_file.save(file_path)

            # إضافة ملاحظة بوجود صورة
            if notes:
                notes += f" [تم إرفاق صورة: {filename}]"
            else:
                notes = f"تم إرفاق صورة: {filename}"

    # إضافة الملاحظات إذا وجدت
    if notes:
        # إنشاء بلاغ تلقائي إذا كانت هناك ملاحظات
        issue = Issue(
            panel_id=panel_id,
            contractor_id=contractor.id,
            title=f"ملاحظة على قراءة {panel.mdb}",
            issue_type="ملاحظة",
            description=notes,
            status="مفتوح",
            priority="متوسط",
            created_at=datetime.datetime.now()
        )
        db.session.add(issue)

        # تحديث عدد البلاغات للوحة والمقاول
        panel.issues_count = panel.issues_count + 1
        contractor.issues_count = contractor.issues_count + 1

    db.session.commit()

    flash('تم إضافة القراءة بنجاح', 'success')
    return redirect(url_for('contractor_dashboard'))

# عرض بلاغات المقاول
@app.route('/contractors/<int:contractor_id>/issues')
def contractor_issues(contractor_id):
    contractor = Contractor.query.get_or_404(contractor_id)

    # التحقق مما إذا كان المستخدم الحالي هو مقاول مدير
    is_manager = current_user.is_authenticated and current_user.role == 'contractor' and hasattr(current_user, 'is_manager') and current_user.is_manager

    # التحقق مما إذا كان المستخدم الحالي هو مقاول عادي
    is_regular_contractor = current_user.is_authenticated and current_user.role == 'regular_contractor'

    # الحصول على البلاغات
    if is_manager and current_user.contractor_id == contractor_id:
        # المقاول المدير يمكنه رؤية جميع البلاغات المرتبطة بالفرق التابعة له
        # الحصول على جميع الفرق التابعة للمقاول
        teams = ContractorTeam.query.filter_by(contractor_id=contractor_id).all()

        # جمع جميع المناطق المسؤول عنها الفرق
        team_areas = []
        for team in teams:
            if team.area_responsibility:
                try:
                    # محاولة تحليل البيانات كـ JSON
                    team_areas_json = json.loads(team.area_responsibility)
                    if isinstance(team_areas_json, list):
                        team_areas.extend(team_areas_json)
                    else:
                        team_areas.append(team_areas_json)
                except json.JSONDecodeError:
                    # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                    team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

        # إزالة التكرار
        team_areas = list(set(team_areas))

        # الحصول على اللوحات في مناطق المسؤولية
        panels_in_areas = MDBPanel.query.filter(MDBPanel.area_name.in_(team_areas)).all()
        panel_ids = [panel.id for panel in panels_in_areas]

        # الحصول على البلاغات المرتبطة بالمقاول أو اللوحات في مناطق المسؤولية
        issues_query = Issue.query.filter(
            db.or_(
                Issue.contractor_id == contractor_id,
                Issue.panel_id.in_(panel_ids)
            )
        )
    elif is_regular_contractor and current_user.contractor_id == contractor_id:
        # المقاول العادي يرى فقط البلاغات المرتبطة به مباشرة
        issues_query = Issue.query.filter_by(contractor_id=contractor_id)
    else:
        # المقاول العادي يرى فقط البلاغات المرتبطة به مباشرة
        issues_query = Issue.query.filter_by(contractor_id=contractor_id)

    # تطبيق الفلتر إذا تم تحديده
    status_filter = request.args.get('status')
    if status_filter and status_filter != 'all':
        issues_query = issues_query.filter(Issue.status == status_filter)

    issues = issues_query.order_by(Issue.created_at.desc()).all()

    # إحصائيات البلاغات
    total_issues = len(issues)
    open_issues = sum(1 for issue in issues if issue.status == 'مفتوح')
    in_progress_issues = sum(1 for issue in issues if issue.status == 'قيد المعالجة')
    closed_issues = sum(1 for issue in issues if issue.status == 'مغلق')

    return render_template('contractor_issues.html',
                           contractor=contractor,
                           issues=issues,
                           total_issues=total_issues,
                           open_issues=open_issues,
                           in_progress_issues=in_progress_issues,
                           closed_issues=closed_issues,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# عرض أعضاء الفريق
@app.route('/contractors/<int:contractor_id>/teams/<int:team_id>/members')
@login_required
def team_members(contractor_id, team_id):
    # التحقق من أن المستخدم هو مقاول أو مدير
    if current_user.role != 'admin' and (current_user.role != 'contractor' or current_user.contractor_id != contractor_id):
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    contractor = Contractor.query.get_or_404(contractor_id)
    team = ContractorTeam.query.get_or_404(team_id)

    # التحقق من أن الفريق ينتمي للمقاول المحدد
    if team.contractor_id != contractor_id:
        flash('غير مسموح بالوصول إلى هذا الفريق', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # الحصول على أعضاء الفريق
    team_members = ContractorTeamMember.query.filter_by(team_id=team_id).all()

    # الحصول على المستخدمين المتاحين للإضافة للفريق
    available_users = User.query.filter_by(contractor_id=contractor_id, team_id=None).all()

    return render_template('team_members.html',
                           contractor=contractor,
                           team=team,
                           team_members=team_members,
                           available_users=available_users,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# عرض بلاغات الفريق
@app.route('/contractors/<int:contractor_id>/teams/<int:team_id>/issues')
@login_required
def team_issues(contractor_id, team_id):
    # التحقق من أن المستخدم هو مقاول أو مدير
    if current_user.role != 'admin' and (current_user.role != 'contractor' or current_user.contractor_id != contractor_id):
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    contractor = Contractor.query.get_or_404(contractor_id)
    team = ContractorTeam.query.get_or_404(team_id)

    # التحقق من أن الفريق ينتمي للمقاول المحدد
    if team.contractor_id != contractor_id:
        flash('غير مسموح بالوصول إلى هذا الفريق', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # الحصول على البلاغات المرتبطة بالفريق
    issues = Issue.query.filter_by(team_id=team_id).order_by(Issue.created_at.desc()).all()

    # إحصائيات البلاغات
    total_issues = len(issues)
    open_issues = sum(1 for issue in issues if issue.status == 'مفتوح')
    in_progress_issues = sum(1 for issue in issues if issue.status == 'قيد المعالجة')
    closed_issues = sum(1 for issue in issues if issue.status == 'مغلق')

    return render_template('team_issues.html',
                           contractor=contractor,
                           team=team,
                           issues=issues,
                           total_issues=total_issues,
                           open_issues=open_issues,
                           in_progress_issues=in_progress_issues,
                           closed_issues=closed_issues,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة إدارة مناطق الخارطة
@app.route('/map-areas')
@login_required
def map_areas():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    map_areas_list = MapArea.query.all()
    panels = MDBPanel.query.all()

    return render_template('map_areas.html',
                           map_areas=map_areas_list,
                           panels=panels,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))


# صفحة تخصيص أسماء الأعمدة
@app.route('/column-mapping', methods=['GET', 'POST'])
@login_required
def column_mapping():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    if request.method == 'POST':
        # تحديث تخصيص أسماء الأعمدة
        column_mapping = {
            'column_name_mdb': request.form.get('column_name_mdb', 'MDB'),
            'column_name_timestamp': request.form.get('column_name_timestamp', 'وقت القراءة'),
            'column_name_current': request.form.get('column_name_current', 'التيار (أمبير)'),
            'column_name_voltage': request.form.get('column_name_voltage', 'الجهد (فولت)'),
            'column_name_power': request.form.get('column_name_power', 'القدرة (واط)'),
            'column_name_energy': request.form.get('column_name_energy', 'الطاقة (كيلوواط ساعة)'),
            'column_name_power_factor': request.form.get('column_name_power_factor', 'معامل القدرة'),
            'column_name_frequency': request.form.get('column_name_frequency', 'التردد (هرتز)'),
            'column_name_load': request.form.get('column_name_load', 'الحمل'),
            'column_name_breaker_capacity': request.form.get('column_name_breaker_capacity', 'سعة القاطع')
        }

        # حفظ الإعدادات
        for key, value in column_mapping.items():
            set_setting(key, value, f'اسم عمود {value}')

        flash('تم حفظ تخصيص أسماء الأعمدة بنجاح', 'success')
        return redirect(url_for('column_mapping'))

    # الحصول على تخصيص أسماء الأعمدة الحالي
    column_mapping = {
        'column_name_mdb': get_setting('column_name_mdb', 'MDB'),
        'column_name_timestamp': get_setting('column_name_timestamp', 'وقت القراءة'),
        'column_name_current': get_setting('column_name_current', 'التيار (أمبير)'),
        'column_name_voltage': get_setting('column_name_voltage', 'الجهد (فولت)'),
        'column_name_power': get_setting('column_name_power', 'القدرة (واط)'),
        'column_name_energy': get_setting('column_name_energy', 'الطاقة (كيلوواط ساعة)'),
        'column_name_power_factor': get_setting('column_name_power_factor', 'معامل القدرة'),
        'column_name_frequency': get_setting('column_name_frequency', 'التردد (هرتز)'),
        'column_name_load': get_setting('column_name_load', 'الحمل'),
        'column_name_breaker_capacity': get_setting('column_name_breaker_capacity', 'سعة القاطع')
    }

    return render_template('column_mapping.html',
                           column_mapping=column_mapping,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة الإعدادات
@app.route('/settings', methods=['GET', 'POST'])
def settings():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.is_authenticated and current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    if request.method == 'POST':
        # تحديث الإعدادات
        default_warning_threshold = request.form.get('default_warning_threshold')
        default_danger_threshold = request.form.get('default_danger_threshold')
        default_min_voltage = request.form.get('default_min_voltage')
        default_max_voltage = request.form.get('default_max_voltage')
        alert_check_interval = request.form.get('alert_check_interval')
        alert_method = request.form.get('alert_method')
        alert_calculation_period = request.form.get('alert_calculation_period')
        analysis_period = request.form.get('analysis_period')
        send_notifications = request.form.get('send_notifications', 'false')
        notification_method = request.form.get('notification_method')
        default_map_center = request.form.get('default_map_center')
        default_map_zoom = request.form.get('default_map_zoom')

        # إعدادات الألوان
        normal_color = request.form.get('normal_color', '#28a745')
        warning_color = request.form.get('warning_color', '#ffc107')
        danger_color = request.form.get('danger_color', '#dc3545')
        trip_color = request.form.get('trip_color', '#6c757d')

        # إعدادات النقاط الساخنة
        enable_hotspots = request.form.get('enable_hotspots', 'false')
        hotspot_threshold = request.form.get('hotspot_threshold', '75')

        # حفظ الإعدادات الأساسية
        set_setting('default_warning_threshold', default_warning_threshold, 'نسبة التحذير الافتراضية (%)')
        set_setting('default_danger_threshold', default_danger_threshold, 'نسبة الخطر الافتراضية (%)')
        set_setting('default_min_voltage', default_min_voltage, 'الحد الأدنى الافتراضي للجهد (فولت)')
        set_setting('default_max_voltage', default_max_voltage, 'الحد الأقصى الافتراضي للجهد (فولت)')
        set_setting('alert_check_interval', alert_check_interval, 'الفاصل الزمني للتحقق من التنبيهات (دقيقة)')
        set_setting('alert_method', alert_method, 'طريقة التنبيه (peak, average)')
        set_setting('alert_calculation_period', alert_calculation_period, 'فترة حساب التنبيهات (دقيقة)')
        set_setting('analysis_period', analysis_period, 'مدة التحليل الافتراضية (دقيقة)')
        set_setting('send_notifications', send_notifications, 'إرسال إشعارات للمقاولين')
        set_setting('notification_method', notification_method, 'طريقة إرسال الإشعارات (email, sms, both)')
        set_setting('default_map_center', default_map_center, 'مركز الخارطة الافتراضي (عرفة، مكة المكرمة)')
        set_setting('default_map_zoom', default_map_zoom, 'مستوى تكبير الخارطة الافتراضي')

        # حفظ إعدادات الألوان
        set_setting('normal_color', normal_color, 'لون الحالة الطبيعية')
        set_setting('warning_color', warning_color, 'لون حالة التحذير')
        set_setting('danger_color', danger_color, 'لون حالة الخطر')
        set_setting('trip_color', trip_color, 'لون حالة الفصل')

        # حفظ إعدادات النقاط الساخنة
        set_setting('enable_hotspots', enable_hotspots, 'تفعيل النقاط الساخنة على الخارطة')
        set_setting('hotspot_threshold', hotspot_threshold, 'حد النقاط الساخنة (%)')

        # حفظ إعدادات البريد الإلكتروني
        smtp_server = request.form.get('smtp_server', '')
        smtp_port = request.form.get('smtp_port', '587')
        smtp_username = request.form.get('smtp_username', '')
        smtp_password = request.form.get('smtp_password', '')
        sender_email = request.form.get('sender_email', '')

        set_setting('smtp_server', smtp_server, 'خادم SMTP')
        set_setting('smtp_port', smtp_port, 'منفذ SMTP')
        set_setting('smtp_username', smtp_username, 'اسم المستخدم SMTP')
        # حفظ كلمة المرور فقط إذا تم تغييرها (إذا كانت غير فارغة)
        if smtp_password:
            set_setting('smtp_password', smtp_password, 'كلمة المرور SMTP')
        set_setting('sender_email', sender_email, 'البريد الإلكتروني للمرسل')

        # حفظ إعدادات الرسائل النصية
        sms_provider = request.form.get('sms_provider', 'none')
        sms_api_key = request.form.get('sms_api_key', '')
        sms_api_secret = request.form.get('sms_api_secret', '')
        sms_from = request.form.get('sms_from', '')

        set_setting('sms_provider', sms_provider, 'مزود خدمة الرسائل النصية')
        set_setting('sms_api_key', sms_api_key, 'مفتاح API للرسائل النصية')
        # حفظ كلمة السر فقط إذا تم تغييرها (إذا كانت غير فارغة)
        if sms_api_secret:
            set_setting('sms_api_secret', sms_api_secret, 'كلمة سر API للرسائل النصية')
        set_setting('sms_from', sms_from, 'المرسل للرسائل النصية')

        flash('تم حفظ الإعدادات بنجاح', 'success')
        return redirect(url_for('settings'))

    # الحصول على الإعدادات الحالية
    settings_dict = {
        'default_warning_threshold': get_setting('default_warning_threshold', '70'),
        'default_danger_threshold': get_setting('default_danger_threshold', '80'),
        'default_min_voltage': get_setting('default_min_voltage', '210'),
        'default_max_voltage': get_setting('default_max_voltage', '250'),
        'alert_check_interval': get_setting('alert_check_interval', '15'),
        'alert_method': get_setting('alert_method', 'peak'),
        'alert_calculation_period': get_setting('alert_calculation_period', '60'),
        'analysis_period': get_setting('analysis_period', '60'),
        'send_notifications': get_setting('send_notifications', 'true'),
        'notification_method': get_setting('notification_method', 'email'),
        'default_map_center': get_setting('default_map_center', '21.3583, 39.9719'),
        'default_map_zoom': get_setting('default_map_zoom', '14'),
        'normal_color': get_setting('normal_color', '#28a745'),
        'warning_color': get_setting('warning_color', '#ffc107'),
        'danger_color': get_setting('danger_color', '#dc3545'),
        'trip_color': get_setting('trip_color', '#6c757d'),
        'enable_hotspots': get_setting('enable_hotspots', 'true'),
        'hotspot_threshold': get_setting('hotspot_threshold', '75'),

        # إعدادات البريد الإلكتروني
        'smtp_server': get_setting('smtp_server', ''),
        'smtp_port': get_setting('smtp_port', '587'),
        'smtp_username': get_setting('smtp_username', ''),
        'smtp_password': get_setting('smtp_password', ''),
        'sender_email': get_setting('sender_email', ''),

        # إعدادات الرسائل النصية
        'sms_provider': get_setting('sms_provider', 'none'),
        'sms_api_key': get_setting('sms_api_key', ''),
        'sms_api_secret': get_setting('sms_api_secret', ''),
        'sms_from': get_setting('sms_from', '')
    }

    return render_template('settings.html',
                           settings=settings_dict,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تحميل قالب استيراد مناطق الخارطة
@app.route('/map-areas/download-template')
@login_required
def download_map_areas_template():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # إنشاء DataFrame مع الأعمدة المطلوبة
    columns = ['الاسم', 'الوصف', 'الإحداثيات', 'اللون']
    df = pd.DataFrame(columns=columns)

    # إضافة مثال
    df.loc[0] = [
        'المنطقة الشرقية',
        'وصف المنطقة الشرقية',
        '[[24.7136, 46.6753], [24.7200, 46.6800], [24.7150, 46.6900], [24.7100, 46.6850]]',
        '#3388ff'
    ]

    # إنشاء ملف Excel في الذاكرة
    output = io.BytesIO()

    # استخدام openpyxl كمحرك بدلاً من xlsxwriter
    df.to_excel(output, index=False, sheet_name='Map Areas Template')
    output.seek(0)

    # إنشاء اسم الملف مع التاريخ والوقت
    now = datetime.datetime.now()
    filename = f"Map_Areas_Template_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)

# استيراد مناطق الخارطة
@app.route('/map-areas/import', methods=['POST'])
@login_required
def import_map_areas():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    if 'file' not in request.files:
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('map_areas'))

    file = request.files['file']

    if file.filename == '':
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('map_areas'))

    if file and allowed_file(file.filename, {'xlsx', 'xls'}):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        try:
            # قراءة ملف Excel
            df = pd.read_excel(file_path, header=0)

            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ['الاسم', 'الإحداثيات']
            for col in required_columns:
                if col not in df.columns:
                    flash(f'العمود {col} غير موجود في الملف', 'danger')
                    return redirect(url_for('map_areas'))

            # حذف المناطق الحالية إذا تم طلب ذلك
            if 'replace_existing' in request.form:
                MapArea.query.delete()

            # إضافة المناطق الجديدة
            for _, row in df.iterrows():
                try:
                    # التحقق من صحة الإحداثيات
                    coordinates = row['الإحداثيات']
                    if isinstance(coordinates, str):
                        # محاولة تحليل الإحداثيات من النص
                        coordinates_json = json.loads(coordinates.replace("'", '"'))
                    else:
                        flash(f'تنسيق الإحداثيات غير صحيح في الصف {_ + 1}', 'warning')
                        continue

                    # إنشاء منطقة جديدة
                    area = MapArea(
                        name=str(row['الاسم']),
                        description=str(row['الوصف']) if 'الوصف' in df.columns and pd.notna(row['الوصف']) else '',
                        coordinates=json.dumps(coordinates_json),
                        color=str(row['اللون']) if 'اللون' in df.columns and pd.notna(row['اللون']) else '#3388ff',
                        is_active=True,
                        created_at=datetime.datetime.now()
                    )

                    db.session.add(area)
                except Exception as e:
                    flash(f'حدث خطأ في الصف {_ + 1}: {str(e)}', 'warning')
                    continue

            db.session.commit()
            flash('تم استيراد مناطق الخارطة بنجاح', 'success')

        except Exception as e:
            flash(f'حدث خطأ أثناء استيراد البيانات: {str(e)}', 'danger')

        # حذف الملف بعد الاستيراد
        os.remove(file_path)

    else:
        flash('نوع الملف غير مسموح به. يرجى استخدام ملفات Excel (.xlsx, .xls)', 'danger')

    return redirect(url_for('map_areas'))

# تفعيل/تعطيل منطقة خارطة
@app.route('/map-areas/<int:area_id>/toggle')
@login_required
def toggle_map_area(area_id):
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    area = MapArea.query.get_or_404(area_id)

    area.is_active = not area.is_active
    db.session.commit()

    status = 'تفعيل' if area.is_active else 'تعطيل'
    flash(f'تم {status} منطقة الخارطة بنجاح', 'success')
    return redirect(url_for('map_areas'))

# تحرير منطقة خارطة
@app.route('/map-areas/<int:area_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_map_area(area_id):
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    area = MapArea.query.get_or_404(area_id)

    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        color = request.form.get('color')

        area.name = name
        area.description = description
        area.color = color

        db.session.commit()

        flash('تم تحديث منطقة الخارطة بنجاح', 'success')
        return redirect(url_for('map_areas'))

    return render_template('edit_map_area.html',
                           area=area,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة البلاغات
@app.route('/issues')
@login_required
def issues():
    # التحقق من صلاحية الوصول
    if current_user.role == 'readings_recorder':
        # مسجل القراءات يذهب إلى صفحة البلاغات المستلمة
        return redirect(url_for('received_issues'))
    elif current_user.role in ['contractor', 'regular_contractor']:
        # المقاول يذهب إلى صفحة بلاغاته الخاصة
        return redirect(url_for('contractor_issues', contractor_id=current_user.contractor_id))
    elif current_user.role not in ['admin', 'user']:
        # إذا لم يكن له صلاحية
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    # الحصول على معاملات الفلترة
    country_filter = request.args.get('country')
    company_filter = request.args.get('company')

    # الحصول على البلاغات (استخدام outerjoin لعرض البلاغات بدون لوحات أيضاً)
    issues_query = Issue.query.outerjoin(MDBPanel)

    # إضافة join للشركة والدولة إذا كانت اللوحة مرتبطة
    if country_filter or company_filter:
        issues_query = issues_query.outerjoin(Company, MDBPanel.company_id == Company.id)
        issues_query = issues_query.outerjoin(Country, MDBPanel.country_id == Country.id)

    # تطبيق فلترة الدولة
    if country_filter:
        issues_query = issues_query.filter(Country.id == int(country_filter))

    # تطبيق فلترة الشركة
    if company_filter:
        issues_query = issues_query.filter(Company.id == int(company_filter))

    # تطبيق فلتر الحالة
    status_filter = request.args.get('status')
    if status_filter and status_filter != 'all':
        issues_query = issues_query.filter(Issue.status == status_filter)

    # تطبيق فلتر البحث
    search_filter = request.args.get('search')
    if search_filter and search_filter.strip():
        search_term = f'%{search_filter.strip()}%'
        issues_query = issues_query.filter(
            db.or_(
                Issue.title.ilike(search_term),
                Issue.description.ilike(search_term),
                Issue.issue_type.ilike(search_term),
                Issue.responsible_person.ilike(search_term),
                MDBPanel.mdb.ilike(search_term),
                MDBPanel.maximo_tag.ilike(search_term),
                MDBPanel.area_name.ilike(search_term)
            )
        )

    # فلترة حسب المناطق المخصصة للمستخدم
    if current_user.is_authenticated and current_user.role != 'admin':
        user_areas = current_user.get_assigned_areas()
        if user_areas:
            issues_query = issues_query.filter(MDBPanel.area_name.in_(user_areas))

    issues_list = issues_query.order_by(Issue.created_at.desc()).all()

    # إحصائيات البلاغات (مع فلترة المناطق)
    stats_query = Issue.query.outerjoin(MDBPanel)

    # فلترة حسب المناطق المخصصة للمستخدم
    if current_user.is_authenticated and current_user.role != 'admin':
        user_areas = current_user.get_assigned_areas()
        if user_areas:
            stats_query = stats_query.filter(MDBPanel.area_name.in_(user_areas))

    total_issues = stats_query.count()
    open_issues = stats_query.filter(Issue.status == 'مفتوح').count()
    in_progress_issues = stats_query.filter(Issue.status == 'قيد المعالجة').count()
    closed_issues = stats_query.filter(Issue.status == 'مغلق').count()

    # الحصول على اللوحات والمقاولين والمستخدمين للنموذج
    panels_query = MDBPanel.query

    # فلترة اللوحات حسب المناطق المخصصة للمستخدم
    if current_user.is_authenticated and current_user.role != 'admin':
        user_areas = current_user.get_assigned_areas()
        if user_areas:
            panels_query = panels_query.filter(MDBPanel.area_name.in_(user_areas))

    panels = panels_query.all()
    contractors = Contractor.query.all()
    users = User.query.filter_by(is_active=True).all()

    # الحصول على أنواع البلاغات من قاعدة البيانات
    issue_types_dropdown = DropdownList.query.filter_by(name='issue_types', is_active=True).first()
    issue_types = []
    if issue_types_dropdown:
        issue_types = DropdownItem.query.filter_by(
            dropdown_id=issue_types_dropdown.id,
            is_active=True
        ).order_by(DropdownItem.order).all()

    # إحصائيات البلاغات حسب المنطقة
    areas_issues = db.session.query(
        MDBPanel.area_name,
        db.func.count(Issue.id)
    ).join(Issue, Issue.panel_id == MDBPanel.id).group_by(MDBPanel.area_name).all()

    # إحصائيات البلاغات حسب المقاول
    contractors_issues = db.session.query(
        Contractor.name,
        db.func.count(Issue.id)
    ).join(Issue, Issue.contractor_id == Contractor.id).group_by(Contractor.name).all()

    # إحصائيات البلاغات حسب الدول
    countries_issues = db.session.query(
        Country.name,
        db.func.count(Issue.id)
    ).join(Company, Country.id == Company.country_id)\
     .join(MDBPanel, Company.id == MDBPanel.company_id)\
     .join(Issue, MDBPanel.id == Issue.panel_id)\
     .group_by(Country.name)\
     .order_by(db.func.count(Issue.id).desc()).limit(10).all()

    # إحصائيات البلاغات حسب الشركات
    companies_issues = db.session.query(
        Company.name,
        db.func.count(Issue.id)
    ).join(MDBPanel, Company.id == MDBPanel.company_id)\
     .join(Issue, MDBPanel.id == Issue.panel_id)\
     .group_by(Company.name)\
     .order_by(db.func.count(Issue.id).desc()).limit(10).all()

    # إضافة قائمة المجموعات
    from models import UserGroup
    user_groups = UserGroup.query.filter_by(is_active=True).all()

    # الحصول على قوائم الدول والشركات للفلترة
    countries = Country.query.filter_by(is_active=True).order_by(Country.name).all()
    companies = Company.query.filter_by(is_active=True).order_by(Company.name).all()

    return render_template('issues.html',
                           issues=issues_list,
                           total_issues=total_issues,
                           open_issues=open_issues,
                           in_progress_issues=in_progress_issues,
                           closed_issues=closed_issues,
                           panels=panels,
                           contractors=contractors,
                           users=users,
                           user_groups=user_groups,
                           issue_types=issue_types,
                           areas_issues=areas_issues,
                           contractors_issues=contractors_issues,
                           countries_issues=countries_issues,
                           companies_issues=companies_issues,
                           countries=countries,
                           companies=companies,
                           country_filter=country_filter,
                           company_filter=company_filter,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة اختبار بسيطة
@app.route('/test-received-issues')
@login_required
def test_received_issues():
    try:
        # اختبار بسيط للبلاغات
        issues = Issue.query.filter_by(assignee_id=current_user.id).all()
        return f"عدد البلاغات للمستخدم {current_user.username}: {len(issues)}"
    except Exception as e:
        return f"خطأ: {str(e)}"

# صفحة البلاغات المستلمة (متاحة لجميع المستخدمين)
@app.route('/received-issues')
@login_required
def received_issues():
    # هذه الصفحة متاحة لجميع المستخدمين لرؤية البلاغات المخصصة لهم
    # لا حاجة لتقييد الوصول حسب نوع المستخدم

    # الحصول على البلاغات المخصصة للمستخدم الحالي
    # الحصول على مجموعات المستخدم
    from models import UserGroupMembership
    user_groups = UserGroupMembership.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).all()
    user_group_ids = [membership.group_id for membership in user_groups]

    # بناء شروط البحث
    search_conditions = [
        Issue.assignee_id == current_user.id,
        Issue.responsible_person == current_user.name,
        Issue.responsible_person == str(current_user.id)  # بحث بالID أيضاً
    ]

    # إضافة شرط المجموعات إذا وجدت
    if user_group_ids:
        search_conditions.append(Issue.assigned_group_id.in_(user_group_ids))

    # إضافة شرط التوزيعات الجديدة
    issue_assignments = IssueAssignment.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).all()
    if issue_assignments:
        assigned_issue_ids = [assignment.issue_id for assignment in issue_assignments]
        search_conditions.append(Issue.id.in_(assigned_issue_ids))

    issues_query = Issue.query.filter(db.or_(*search_conditions))

    # تطبيق فلترة المناطق للمستخدمين غير المديرين
    if current_user.role != 'admin':
        # الحصول على المناطق المخصصة للمستخدم
        assigned_areas = current_user.get_assigned_areas()
        print(f"مستخدم: {current_user.username}, المناطق المخصصة: {assigned_areas}")
        if assigned_areas:  # إذا كانت هناك مناطق مخصصة
            # فلترة البلاغات حسب المناطق المخصصة
            issues_query = issues_query.join(MDBPanel, Issue.panel_id == MDBPanel.id, isouter=True).filter(
                db.or_(
                    Issue.panel_id.is_(None),  # البلاغات العامة بدون لوحة
                    MDBPanel.area_name.in_(assigned_areas)  # البلاغات في المناطق المخصصة
                )
            )

    # تطبيق الفلتر إذا تم تحديده
    status_filter = request.args.get('status')
    if status_filter and status_filter != 'all':
        issues_query = issues_query.filter(Issue.status == status_filter)

    issues_list = issues_query.order_by(Issue.created_at.desc()).all()

    # إحصائيات البلاغات المستلمة
    total_issues = len(issues_list)
    open_issues = sum(1 for issue in issues_list if issue.status == 'مفتوح')
    in_progress_issues = sum(1 for issue in issues_list if issue.status == 'قيد المعالجة')
    closed_issues = sum(1 for issue in issues_list if issue.status == 'مغلق')

    return render_template('received_issues.html',
                           issues=issues_list,
                           total_issues=total_issues,
                           open_issues=open_issues,
                           in_progress_issues=in_progress_issues,
                           closed_issues=closed_issues,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة إدارة أنواع طلبات الفحص
@app.route('/inspection-request-types')
@login_required
def inspection_request_types():
    # التحقق من أن المستخدم مدير
    if current_user.role != 'admin':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    # الحصول على جميع أنواع طلبات الفحص
    types = InspectionRequestType.query.all()

    return render_template('inspection_request_types.html',
                           types=types,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة نوع طلب فحص جديد
@app.route('/inspection-request-types/add', methods=['POST'])
@login_required
def add_inspection_request_type():
    # التحقق من أن المستخدم مدير
    if current_user.role != 'admin':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    name = request.form.get('name')
    description = request.form.get('description')

    if not name:
        flash('يرجى إدخال اسم نوع الطلب', 'danger')
        return redirect(url_for('inspection_request_types'))

    # التحقق من عدم وجود نوع بنفس الاسم
    existing_type = InspectionRequestType.query.filter_by(name=name).first()
    if existing_type:
        flash('يوجد نوع طلب بنفس الاسم بالفعل', 'danger')
        return redirect(url_for('inspection_request_types'))

    # إنشاء نوع طلب جديد
    request_type = InspectionRequestType(
        name=name,
        description=description,
        is_active=True,
        created_at=datetime.datetime.now()
    )

    db.session.add(request_type)
    db.session.commit()

    flash('تم إضافة نوع الطلب بنجاح', 'success')
    return redirect(url_for('inspection_request_types'))

# تحديث حالة نوع طلب فحص
@app.route('/inspection-request-types/<int:type_id>/toggle-status')
@login_required
def toggle_inspection_request_type_status(type_id):
    # التحقق من أن المستخدم مدير
    if current_user.role != 'admin':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    request_type = InspectionRequestType.query.get_or_404(type_id)
    request_type.is_active = not request_type.is_active
    db.session.commit()

    status = 'تفعيل' if request_type.is_active else 'تعطيل'
    flash(f'تم {status} نوع الطلب بنجاح', 'success')
    return redirect(url_for('inspection_request_types'))

# صفحة طلبات الفحص
@app.route('/inspection-requests')
@login_required
def inspection_requests():
    # التحقق من صلاحية الوصول
    if current_user.role == 'readings_recorder':
        # مسجل القراءات يذهب إلى صفحة طلبات الفحص المستلمة
        return redirect(url_for('received_inspection_requests'))
    elif current_user.role == 'regular_contractor':
        # المقاول العادي لا يملك صلاحية لرؤية جميع طلبات الفحص
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    elif current_user.role not in ['admin', 'user', 'contractor']:
        # إذا لم يكن له صلاحية
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    # الحصول على معلمات التصفية
    status = request.args.get('status', 'all')
    priority = request.args.get('priority', 'all')
    panel_ids = request.args.getlist('panel_id')  # دعم أكثر من لوحة
    request_type_id = request.args.get('request_type_id', 'all')
    search_panel = request.args.get('search_panel', '')
    country_filter = request.args.get('country')
    company_filter = request.args.get('company')

    # إنشاء استعلام أساسي
    try:
        query = InspectionRequest.query.join(MDBPanel)

        # إضافة join للشركة والدولة إذا كانت اللوحة مرتبطة
        if country_filter or company_filter:
            query = query.outerjoin(Company, MDBPanel.company_id == Company.id)
            query = query.outerjoin(Country, MDBPanel.country_id == Country.id)

    except Exception as e:
        print(f'خطأ في join: {str(e)}')
        # في حالة فشل join، استخدم استعلام بسيط
        query = InspectionRequest.query

    # تطبيق التصفية
    if status != 'all':
        query = query.filter(InspectionRequest.status == status)
    if priority != 'all':
        query = query.filter(InspectionRequest.priority == priority)

    # تطبيق فلترة الدولة
    if country_filter:
        try:
            query = query.filter(Country.id == int(country_filter))
        except:
            pass

    # تطبيق فلترة الشركة
    if company_filter:
        try:
            query = query.filter(Company.id == int(company_filter))
        except:
            pass

    # تطبيق فلتر اللوحات (دعم أكثر من لوحة)
    if panel_ids and 'all' not in panel_ids:
        # تحويل القيم إلى أرقام صحيحة
        valid_panel_ids = []
        for pid in panel_ids:
            if pid.isdigit():
                valid_panel_ids.append(int(pid))

        if valid_panel_ids:
            query = query.filter(InspectionRequest.panel_id.in_(valid_panel_ids))

    if request_type_id != 'all' and request_type_id.isdigit():
        query = query.filter(InspectionRequest.request_type_id == int(request_type_id))

    # البحث عن اللوحة (برقم اللوحة أو رقم ماكسيمو أو المنطقة)
    if search_panel:
        search_term = f'%{search_panel}%'
        try:
            # محاولة البحث مع join
            query = query.filter(
                db.or_(
                    MDBPanel.mdb.ilike(search_term),
                    MDBPanel.maximo_tag.ilike(search_term),
                    MDBPanel.area_name.ilike(search_term),
                    InspectionRequest.description.ilike(search_term),
                    InspectionRequest.notes.ilike(search_term)
                )
            )
        except Exception as e:
            print(f'خطأ في البحث: {str(e)}')
            # بحث بسيط في حقول طلب الفحص فقط
            query = query.filter(
                db.or_(
                    InspectionRequest.description.ilike(search_term),
                    InspectionRequest.notes.ilike(search_term)
                )
            )

    # تصفية حسب المستخدم الحالي إذا كان مقاول أو مسجل قراءات
    if current_user.role in ['contractor', 'regular_contractor', 'readings_recorder']:
        # المقاول أو مسجل القراءات يرى فقط الطلبات المرتبطة به أو بفريقه
        query = query.filter(
            db.or_(
                InspectionRequest.contractor_id == current_user.contractor_id if hasattr(current_user, 'contractor_id') else False,
                InspectionRequest.assignee_id == current_user.id,
                InspectionRequest.requester_id == current_user.id
            )
        )

    # فلترة حسب المناطق المخصصة للمستخدم (لغير المدير)
    if current_user.role != 'admin':
        user_areas = current_user.get_assigned_areas()
        if user_areas:
            try:
                query = query.filter(MDBPanel.area_name.in_(user_areas))
            except Exception as e:
                print(f'خطأ في فلترة المناطق: {str(e)}')
                # فلترة باستخدام panel_id بدلاً من join
                panel_ids_in_areas = [p.id for p in MDBPanel.query.filter(MDBPanel.area_name.in_(user_areas)).all()]
                if panel_ids_in_areas:
                    query = query.filter(InspectionRequest.panel_id.in_(panel_ids_in_areas))

    # تطبيق فلتر البحث
    search_filter = request.args.get('search')
    if search_filter and search_filter.strip():
        search_term = f'%{search_filter.strip()}%'
        try:
            query = query.filter(
                db.or_(
                    InspectionRequest.title.ilike(search_term),
                    InspectionRequest.description.ilike(search_term),
                    InspectionRequest.request_number.ilike(search_term),
                    MDBPanel.mdb.ilike(search_term),
                    MDBPanel.maximo_tag.ilike(search_term),
                    MDBPanel.area_name.ilike(search_term)
                )
            )
        except Exception as e:
            print(f'خطأ في البحث العام: {str(e)}')
            # بحث بسيط في حقول طلب الفحص فقط
            query = query.filter(
                db.or_(
                    InspectionRequest.title.ilike(search_term),
                    InspectionRequest.description.ilike(search_term),
                    InspectionRequest.request_number.ilike(search_term)
                )
            )

    # الحصول على طلبات الفحص
    inspection_requests = query.order_by(InspectionRequest.created_at.desc()).all()

    # الحصول على اللوحات للتصفية مع فلترة المناطق
    panels_query = MDBPanel.query

    if current_user.role != 'admin':
        user_areas = current_user.get_assigned_areas()
        if user_areas:
            panels_query = panels_query.filter(MDBPanel.area_name.in_(user_areas))

    panels = panels_query.all()

    # الحصول على أنواع طلبات الفحص للتصفية
    request_types = InspectionRequestType.query.filter_by(is_active=True).all()

    # إضافة قائمة المستخدمين للتعيين
    users = User.query.filter(
        User.is_active == True
    ).all()

    # إضافة قائمة المجموعات
    from models import UserGroup
    user_groups = UserGroup.query.filter_by(is_active=True).all()

    # الحصول على قوائم الدول والشركات للفلترة
    countries = Country.query.filter_by(is_active=True).order_by(Country.name).all()
    companies = Company.query.filter_by(is_active=True).order_by(Company.name).all()

    # إحصائيات
    stats = {
        'total': len(inspection_requests),
        'new': sum(1 for r in inspection_requests if r.status == 'جديد'),
        'in_progress': sum(1 for r in inspection_requests if r.status == 'قيد التنفيذ'),
        'completed': sum(1 for r in inspection_requests if r.status == 'مكتمل'),
        'cancelled': sum(1 for r in inspection_requests if r.status == 'ملغي'),
        'high': sum(1 for r in inspection_requests if r.priority == 'عالي'),
        'medium': sum(1 for r in inspection_requests if r.priority == 'متوسط'),
        'low': sum(1 for r in inspection_requests if r.priority == 'منخفض')
    }

    return render_template('inspection_requests.html',
                           inspection_requests=inspection_requests,
                           panels=panels,
                           request_types=request_types,
                           users=users,
                           user_groups=user_groups,
                           stats=stats,
                           countries=countries,
                           companies=companies,
                           country_filter=country_filter,
                           company_filter=company_filter,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة إضافة طلب فحص جديد
@app.route('/inspection-requests/new')
@login_required
def new_inspection_request():
    # التحقق من أن المستخدم ليس مقاول عادي
    if current_user.role == 'regular_contractor':
        flash('ليس لديك صلاحية لإضافة طلبات فحص جديدة', 'danger')
        return redirect(url_for('inspection_requests'))

    # فلترة اللوحات حسب المناطق المخصصة للمستخدم
    panels_query = MDBPanel.query

    if current_user.role != 'admin':
        user_areas = current_user.get_assigned_areas()
        if user_areas:
            panels_query = panels_query.filter(MDBPanel.area_name.in_(user_areas))

    panels = panels_query.all()
    request_types = InspectionRequestType.query.filter_by(is_active=True).all()

    # إضافة قائمة المستخدمين للتعيين
    users = User.query.filter(
        User.is_active == True
    ).all()

    return render_template('new_inspection_request.html',
                           panels=panels,
                           request_types=request_types,
                           users=users,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة إضافة طلب فحص للوحة محددة
@app.route('/panels/<int:panel_id>/add-inspection-request')
@login_required
def add_inspection_request_form(panel_id):
    # المقاولون والمقاولون العاديون يمكنهم إضافة طلبات فحص

    panel = MDBPanel.query.get_or_404(panel_id)
    request_types = InspectionRequestType.query.filter_by(is_active=True).all()
    return render_template('add_inspection_request.html',
                           panel=panel,
                           request_types=request_types,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة طلب فحص جديد
@app.route('/inspection-requests/add', methods=['POST'])
@login_required
def add_inspection_request():
    # التحقق من أن المستخدم ليس مقاول عادي
    if current_user.role == 'regular_contractor':
        flash('ليس لديك صلاحية لإضافة طلبات فحص جديدة', 'danger')
        return redirect(url_for('inspection_requests'))

    panel_ids = request.form.getlist('panel_id')  # دعم أكثر من لوحة
    title = request.form.get('title') or 'طلب فحص عام'  # قيمة افتراضية
    description = request.form.get('description') or ''
    priority = request.form.get('priority') or 'متوسط'
    request_type_id = request.form.get('request_type_id') or None
    assignee_id = request.form.get('assignee_id') or None  # المستخدم المعين
    responsible_person_id = request.form.get('responsible_person') or None  # ID الشخص المسؤول
    assigned_group_id = request.form.get('assigned_group') or None  # المجموعة المسؤولة

    # الحصول على اسم الشخص المسؤول إذا تم تحديده
    responsible_person_name = ''
    if responsible_person_id:
        responsible_user = User.query.get(responsible_person_id)
        if responsible_user:
            responsible_person_name = responsible_user.name
    due_date_str = request.form.get('due_date')
    mutawif_name = request.form.get('mutawif_name') or ''
    pilgrim_guide_name = request.form.get('pilgrim_guide_name') or ''
    location_details = request.form.get('location_details') or ''

    # إذا لم يتم تحديد أي لوحة أو تم تحديد قيمة فارغة
    if not panel_ids or (len(panel_ids) == 1 and panel_ids[0] == ''):
        # إنشاء طلب فحص عام بدون لوحة محددة
        flash('يجب تحديد لوحة واحدة على الأقل لإنشاء طلب فحص', 'warning')
        return redirect(url_for('new_inspection_request'))

    # إنشاء طلب فحص لكل لوحة محددة
    created_requests = []
    for panel_id in panel_ids:
        if panel_id and panel_id != '':
            # التحقق من وجود اللوحة
            panel = MDBPanel.query.get(panel_id)
            if not panel:
                flash(f'اللوحة رقم {panel_id} غير موجودة', 'warning')
                continue

            # تحويل تاريخ الاستحقاق إلى كائن datetime
            due_date = None
            if due_date_str:
                try:
                    due_date = datetime.datetime.strptime(due_date_str, '%Y-%m-%d')
                except ValueError:
                    flash('صيغة تاريخ الاستحقاق غير صحيحة', 'warning')
                    continue

            # إنشاء رقم طلب فريد
            request_number = f"IR-{datetime.datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

            # تحديد المقاول المسؤول
            contractor_id = None
            team_id = None

            # إذا كان المستخدم مقاول، استخدم معرف المقاول الخاص به
            if current_user.role in ['contractor', 'regular_contractor']:
                contractor_id = current_user.contractor_id
                team_id = current_user.team_id

            # إذا كانت اللوحة مرتبطة بمقاول، استخدم معرف المقاول المرتبط
            elif panel.responsible_contractor_id:
                contractor_id = panel.responsible_contractor_id

            # إنشاء طلب فحص لهذه اللوحة
            inspection_request = InspectionRequest(
                request_number=request_number,
                panel_id=panel_id,
                requester_id=current_user.id,
                assignee_id=assignee_id,
                contractor_id=contractor_id,
                team_id=team_id,
                title=title,
                description=description,
                priority=priority,
                request_type_id=request_type_id if request_type_id and request_type_id.isdigit() else None,
                status='جديد',
                due_date=due_date,
                mutawif_name=mutawif_name,
                pilgrim_guide_name=pilgrim_guide_name,
                location_details=location_details,
                responsible_person=responsible_person_name,
                assigned_group_id=assigned_group_id,
                created_at=datetime.datetime.now()
            )

            db.session.add(inspection_request)
            created_requests.append(inspection_request)

    # حفظ الصورة إذا تم تحميلها
    if 'image' in request.files and request.files['image'].filename:
        image_file = request.files['image']

        if allowed_file(image_file.filename, {'jpg', 'jpeg', 'png', 'gif'}):
            # إنشاء اسم فريد للملف
            filename = secure_filename(f"inspection_{request_number}_{uuid.uuid4().hex}.{image_file.filename.rsplit('.', 1)[1].lower()}")

            # التأكد من وجود المجلد
            inspection_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'inspections')
            if not os.path.exists(inspection_folder):
                os.makedirs(inspection_folder)

            # حفظ الملف
            file_path = os.path.join(inspection_folder, filename)
            image_file.save(file_path)

            # تحديث مسار الصورة في طلب الفحص
            inspection_request.image_path = os.path.join('inspections', filename)

    db.session.add(inspection_request)
    db.session.commit()

    # معالجة التوزيع التلقائي إذا لم يتم تحديد مستخدم
    try:
        # إذا لم يتم تحديد مستخدم محدد، قم بالتوزيع التلقائي
        if not inspection_request.assignee_id and not inspection_request.assigned_to and not inspection_request.assigned_group_id:
            if inspection_request.panel and inspection_request.panel.area_name:
                auto_assign_to_area_users(inspection_request, inspection_request.panel.area_name)
    except Exception as e:
        print(f'خطأ في التوزيع التلقائي لطلب الفحص {inspection_request.id}: {str(e)}')

    # معالجة الإشعارات التلقائية
    try:
        from auto_notifications import process_new_inspection_request
        process_new_inspection_request(inspection_request)
    except Exception as e:
        print(f'خطأ في معالجة إشعارات طلب الفحص: {str(e)}')

    flash('تم إضافة طلب الفحص بنجاح', 'success')
    return redirect(url_for('inspection_requests'))

# تصدير طلبات الفحص
@app.route('/inspection-requests/export', methods=['POST'])
@login_required
def export_inspection_requests():
    # الحصول على معلمات التصفية
    status = request.form.get('status', 'all')
    priority = request.form.get('priority', 'all')
    panel_ids = request.form.getlist('panel_id')
    request_type_id = request.form.get('request_type_id', 'all')
    export_format = request.form.get('export_format', 'excel')

    # بناء الاستعلام
    query = InspectionRequest.query

    # تطبيق الفلاتر
    if status != 'all':
        query = query.filter(InspectionRequest.status == status)

    if priority != 'all':
        query = query.filter(InspectionRequest.priority == priority)

    if panel_ids and 'all' not in panel_ids:
        query = query.filter(InspectionRequest.panel_id.in_(panel_ids))

    if request_type_id != 'all':
        query = query.filter(InspectionRequest.request_type_id == request_type_id)

    # الحصول على البيانات
    requests = query.all()

    # إعداد بيانات التصدير
    data = []
    headers = ['رقم الطلب', 'العنوان', 'اللوحة', 'الحالة', 'الأولوية', 'نوع الطلب', 'مقدم الطلب', 'المعين', 'الشخص المسؤول', 'تاريخ الإنشاء', 'تاريخ الاستحقاق', 'وقت المعالجة (دقيقة)', 'وقت الإنجاز (دقيقة)']
    data.append(headers)

    for req in requests:
        panel_info = f"{req.panel.mdb} ({req.panel.maximo_tag})" if req.panel else 'غير محدد'
        request_type = req.request_type.name if req.request_type else 'غير محدد'
        requester = req.requester.name if req.requester else 'غير محدد'
        assignee = req.assignee.name if req.assignee else 'غير معين'
        responsible_person = req.responsible_person if req.responsible_person else 'غير محدد'

        # حساب وقت المعالجة (من الإنشاء حتى البدء)
        processing_time_minutes = ''
        if req.processing_time:
            processing_time_minutes = req.processing_time
        elif req.started_at and req.created_at:
            time_diff = req.started_at - req.created_at
            processing_time_minutes = int(time_diff.total_seconds() / 60)

        # حساب وقت الإنجاز (من البدء حتى الإكمال)
        completion_time_minutes = ''
        if req.completion_time:
            completion_time_minutes = req.completion_time
        elif req.completed_at and req.started_at:
            # حساب الفرق بين وقت البدء ووقت الإنجاز
            time_diff = req.completed_at - req.started_at
            completion_time_minutes = int(time_diff.total_seconds() / 60)
        elif req.status in ['مكتمل', 'مغلق'] and req.updated_at and req.started_at:
            # إذا لم يكن هناك completed_at، استخدم updated_at
            time_diff = req.updated_at - req.started_at
            completion_time_minutes = int(time_diff.total_seconds() / 60)

        row = [
            req.request_number,
            req.title,
            panel_info,
            req.status,
            req.priority,
            request_type,
            requester,
            assignee,
            responsible_person,
            req.created_at.strftime('%Y-%m-%d %H:%M') if req.created_at else '',
            req.due_date.strftime('%Y-%m-%d') if req.due_date else '',
            processing_time_minutes if processing_time_minutes != '' else '-',
            completion_time_minutes if completion_time_minutes != '' else '-'
        ]
        data.append(row)

    # تصدير حسب الصيغة
    now = datetime.datetime.now()

    if export_format == 'excel':
        # إنشاء DataFrame
        df = pd.DataFrame(data[1:], columns=data[0])

        # إنشاء ملف Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='طلبات الفحص')

            # تنسيق ورقة العمل
            workbook = writer.book
            worksheet = writer.sheets['طلبات الفحص']

            # تنسيق العناوين
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#B89966',
                'border': 1
            })

            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                worksheet.set_column(col_num, col_num, 20)

        output.seek(0)
        filename = f"Inspection_Requests_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(output,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name=filename)

    elif export_format == 'pdf':
        # إنشاء HTML للتحويل إلى PDF
        html_content = f"""
        <html dir="rtl">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; direction: rtl; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                th {{ background-color: #B89966; color: white; text-align: center; padding: 8px; border: 1px solid #ddd; }}
                td {{ padding: 8px; text-align: center; border: 1px solid #ddd; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                h1 {{ color: #333; text-align: center; }}
                .header {{ text-align: center; margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>تقرير طلبات الفحص</h1>
                <p>تاريخ التقرير: {now.strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>عدد الطلبات: {len(requests)}</p>
            </div>
            <table>
        """

        # إضافة العناوين
        html_content += "<tr>"
        for header in headers:
            html_content += f"<th>{header}</th>"
        html_content += "</tr>"

        # إضافة البيانات
        for row in data[1:]:
            html_content += "<tr>"
            for cell in row:
                html_content += f"<td>{cell}</td>"
            html_content += "</tr>"

        html_content += """
            </table>
        </body>
        </html>
        """

        # تحويل إلى PDF
        options = {
            'page-size': 'A4',
            'orientation': 'Landscape',
            'margin-top': '1cm',
            'margin-right': '1cm',
            'margin-bottom': '1cm',
            'margin-left': '1cm',
            'encoding': 'UTF-8'
        }

        # إنشاء ملف مؤقت
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as f:
            f.write(html_content.encode('utf-8'))
            html_path = f.name

        pdf_path = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False).name

        try:
            pdfkit.from_file(html_path, pdf_path, options=options)
            filename = f"Inspection_Requests_{now.strftime('%Y%m%d_%H%M%S')}.pdf"

            return send_file(pdf_path,
                            mimetype='application/pdf',
                            as_attachment=True,
                            download_name=filename)
        except Exception as e:
            flash(f'حدث خطأ في تصدير PDF: {str(e)}', 'danger')
            return redirect(url_for('inspection_requests'))
        finally:
            # حذف الملفات المؤقتة
            try:
                os.unlink(html_path)
                os.unlink(pdf_path)
            except:
                pass

# عرض تفاصيل طلب الفحص
@app.route('/inspection-requests/<int:request_id>')
@login_required
def inspection_request_details(request_id):
    inspection_request = InspectionRequest.query.get_or_404(request_id)
    return render_template('inspection_request_details.html',
                           inspection_request=inspection_request,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تحديث حالة طلب الفحص
@app.route('/inspection-requests/<int:request_id>/update-status/<status>')
@login_required
def update_inspection_request_status(request_id, status):
    inspection_request = InspectionRequest.query.get_or_404(request_id)

    # التحقق من أن الحالة صالحة
    valid_statuses = ['جديد', 'قيد التنفيذ', 'مكتمل', 'ملغي']
    if status not in valid_statuses:
        flash('حالة غير صالحة', 'danger')
        return redirect(url_for('inspection_request_details', request_id=request_id))

    # تحديث الحالة مع تتبع الأوقات
    current_time = datetime.datetime.now()
    old_status = inspection_request.status
    inspection_request.status = status

    # تتبع أوقات التنفيذ
    if old_status == 'جديد' and status == 'قيد التنفيذ':
        inspection_request.started_at = current_time
        if inspection_request.created_at:
            inspection_request.processing_time = int((current_time - inspection_request.created_at).total_seconds() / 60)

    # إذا كانت الحالة "مكتمل"، تحديث تاريخ الإكمال وحساب وقت الإنجاز
    if status == 'مكتمل':
        inspection_request.completed_at = current_time
        if inspection_request.created_at:
            inspection_request.completion_time = int((current_time - inspection_request.created_at).total_seconds() / 60)
        if current_user.is_authenticated:
            inspection_request.completed_by = current_user.id

    db.session.commit()

    flash(f'تم تحديث حالة طلب الفحص إلى "{status}" بنجاح', 'success')
    return redirect(url_for('inspection_request_details', request_id=request_id))

# صفحة إدارة مناطق المستخدمين
@app.route('/user-areas')
@login_required
def user_areas():
    # التحقق من أن المستخدم مدير
    if current_user.role != 'admin':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    users = User.query.filter_by(is_active=True).all()

    # الحصول على جميع المناطق المتاحة
    areas = db.session.query(MDBPanel.area_name).distinct().filter(
        MDBPanel.area_name.isnot(None),
        MDBPanel.area_name != ''
    ).order_by(MDBPanel.area_name).all()
    areas = [area[0] for area in areas]

    return render_template('user_areas.html', users=users, areas=areas)

# تحديث مناطق المستخدم
@app.route('/update-user-areas', methods=['POST'])
@login_required
def update_user_areas():
    # التحقق من أن المستخدم مدير
    if current_user.role != 'admin':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    user_id = request.form.get('user_id')
    selected_areas = request.form.getlist('areas')

    user = User.query.get_or_404(user_id)
    user.set_assigned_areas(selected_areas)

    db.session.commit()

    flash(f'تم تحديث مناطق المستخدم {user.name} بنجاح', 'success')
    return redirect(url_for('user_areas'))

# صفحة إدارة المجموعات والفرق
@app.route('/user-groups')
@login_required
def user_groups():
    # التحقق من أن المستخدم مدير
    if current_user.role != 'admin':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    from models import UserGroup
    groups = UserGroup.query.filter_by(is_active=True).all()
    users = User.query.filter_by(is_active=True).all()

    # الحصول على جميع المناطق المتاحة
    areas = db.session.query(MDBPanel.area_name).distinct().filter(
        MDBPanel.area_name.isnot(None),
        MDBPanel.area_name != ''
    ).order_by(MDBPanel.area_name).all()
    areas = [area[0] for area in areas]

    return render_template('user_groups.html', groups=groups, users=users, areas=areas)





# إنشاء مجموعة جديدة
@app.route('/create-user-group', methods=['POST'])
@login_required
def create_user_group():
    # التحقق من أن المستخدم مدير
    if current_user.role != 'admin':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('dashboard'))

    from models import UserGroup, UserGroupMembership

    name = request.form.get('name')
    description = request.form.get('description')
    leader_id = request.form.get('leader_id')
    selected_areas = request.form.getlist('areas')
    selected_members = request.form.getlist('members')

    # إنشاء المجموعة
    group = UserGroup(
        name=name,
        description=description,
        leader_id=leader_id if leader_id else None,
        created_by=current_user.id
    )
    group.set_assigned_areas(selected_areas)

    db.session.add(group)
    db.session.flush()  # للحصول على معرف المجموعة

    # إضافة الأعضاء
    for member_id in selected_members:
        if member_id:
            membership = UserGroupMembership(
                user_id=int(member_id),
                group_id=group.id,
                role_in_group='leader' if member_id == leader_id else 'member'
            )
            db.session.add(membership)

    db.session.commit()

    flash(f'تم إنشاء المجموعة "{name}" بنجاح', 'success')
    return redirect(url_for('users'))

# عرض تفاصيل المجموعة
@app.route('/user-groups/<int:group_id>')
@login_required
def group_details(group_id):
    from models import UserGroup
    group = UserGroup.query.get_or_404(group_id)

    # فحص الصلاحية
    if current_user.role != 'admin' and group.leader_id != current_user.id:
        flash('ليس لديك صلاحية للوصول إلى هذه المجموعة', 'danger')
        return redirect(url_for('user_groups'))

    return render_template('group_details.html', group=group)

# إدارة أعضاء المجموعة
@app.route('/user-groups/<int:group_id>/members')
@login_required
def group_members(group_id):
    from models import UserGroup
    group = UserGroup.query.get_or_404(group_id)

    # فحص الصلاحية
    if current_user.role != 'admin' and group.leader_id != current_user.id:
        return 'غير مصرح', 403

    users = User.query.filter_by(is_active=True).all()
    return render_template('group_members_partial.html', group=group, users=users)

# إدارة المجموعات الفرعية
@app.route('/user-groups/<int:group_id>/sub-groups')
@login_required
def group_sub_groups(group_id):
    from models import UserGroup
    group = UserGroup.query.get_or_404(group_id)

    # فحص الصلاحية
    if current_user.role != 'admin' and group.leader_id != current_user.id:
        return 'غير مصرح', 403

    # الحصول على جميع المناطق المتاحة
    areas = db.session.query(MDBPanel.area_name).distinct().filter(
        MDBPanel.area_name.isnot(None),
        MDBPanel.area_name != ''
    ).order_by(MDBPanel.area_name).all()
    areas = [area[0] for area in areas]

    return render_template('group_sub_groups_partial.html', group=group, areas=areas)

# صفحة إدارة أنواع البلاغات
@app.route('/issue-types')
@login_required
def issue_types():
    # التحقق من أن المستخدم مدير
    if current_user.role != 'admin':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('dashboard'))

    # الحصول على قائمة أنواع البلاغات
    issue_types_list = DropdownList.query.filter_by(name='issue_types').first()
    if not issue_types_list:
        # إنشاء قائمة أنواع البلاغات إذا لم تكن موجودة
        issue_types_list = DropdownList(
            name='issue_types',
            display_name='أنواع البلاغات',
            description='قائمة بأنواع البلاغات المختلفة',
            field_type='select',
            is_active=True,
            visibility='all'
        )
        db.session.add(issue_types_list)
        db.session.commit()

    # الحصول على عناصر القائمة
    issue_types_items = DropdownItem.query.filter_by(
        dropdown_id=issue_types_list.id
    ).order_by(DropdownItem.order).all()

    return render_template('issue_types_management.html',
                           issue_types=issue_types_items,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة نوع بلاغ جديد
@app.route('/issue-types/add', methods=['POST'])
@login_required
def add_issue_type():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'})

    # الحصول على قائمة أنواع البلاغات
    issue_types_list = DropdownList.query.filter_by(name='issue_types').first()
    if not issue_types_list:
        issue_types_list = DropdownList(
            name='issue_types',
            display_name='أنواع البلاغات',
            description='قائمة بأنواع البلاغات المختلفة',
            field_type='select',
            is_active=True,
            visibility='all'
        )
        db.session.add(issue_types_list)
        db.session.flush()

    value = request.form.get('value')
    display_text = request.form.get('display_text')
    description = request.form.get('description', '')
    order = int(request.form.get('order', 0))
    is_active = 'is_active' in request.form

    # التحقق من عدم وجود نوع بنفس القيمة
    existing_item = DropdownItem.query.filter_by(
        dropdown_id=issue_types_list.id,
        value=value
    ).first()

    if existing_item:
        flash(f'نوع البلاغ "{value}" موجود بالفعل', 'danger')
        return redirect(url_for('issue_types'))

    # إضافة العنصر الجديد
    new_item = DropdownItem(
        dropdown_id=issue_types_list.id,
        value=value,
        display_text=display_text,
        order=order,
        is_active=is_active
    )

    db.session.add(new_item)
    db.session.commit()

    flash('تم إضافة نوع البلاغ بنجاح', 'success')
    return redirect(url_for('issue_types'))

# تعديل نوع بلاغ
@app.route('/issue-types/edit', methods=['POST'])
@login_required
def edit_issue_type():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'})

    item_id = request.form.get('item_id')
    item = DropdownItem.query.get_or_404(item_id)

    item.value = request.form.get('value')
    item.display_text = request.form.get('display_text')
    item.order = int(request.form.get('order', 0))
    item.is_active = 'is_active' in request.form

    db.session.commit()

    flash('تم تعديل نوع البلاغ بنجاح', 'success')
    return redirect(url_for('issue_types'))

# تفعيل/تعطيل نوع بلاغ
@app.route('/issue-types/toggle', methods=['POST'])
@login_required
def toggle_issue_type():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'})

    item_id = request.form.get('item_id')
    item = DropdownItem.query.get_or_404(item_id)

    item.is_active = not item.is_active
    db.session.commit()

    status = 'تفعيل' if item.is_active else 'تعطيل'
    return jsonify({'success': True, 'message': f'تم {status} نوع البلاغ بنجاح'})

# حذف نوع بلاغ
@app.route('/issue-types/delete', methods=['POST'])
@login_required
def delete_issue_type():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'})

    item_id = request.form.get('item_id')
    item = DropdownItem.query.get_or_404(item_id)

    # التحقق من عدم وجود بلاغات تستخدم هذا النوع
    issues_count = Issue.query.filter_by(issue_type=item.value).count()
    if issues_count > 0:
        return jsonify({
            'success': False,
            'message': f'لا يمكن حذف هذا النوع لأنه مرتبط بـ {issues_count} بلاغ'
        })

    db.session.delete(item)
    db.session.commit()

    return jsonify({'success': True, 'message': 'تم حذف نوع البلاغ بنجاح'})

# إعادة ترتيب أنواع البلاغات
@app.route('/issue-types/reorder', methods=['POST'])
@login_required
def reorder_issue_types():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'})

    order_data = json.loads(request.form.get('order'))

    for item_data in order_data:
        item = DropdownItem.query.get(item_data['id'])
        if item:
            item.order = item_data['order']

    db.session.commit()

    return jsonify({'success': True, 'message': 'تم إعادة ترتيب أنواع البلاغات بنجاح'})

# API للبحث في اللوحات
@app.route('/api/search-panels')
@login_required
def search_panels():
    search_term = request.args.get('q', '').strip()

    if not search_term or len(search_term) < 2:
        return jsonify({'panels': []})

    # بحث في رقم اللوحة ورقم ماكسيمو
    panels_query = MDBPanel.query.filter(
        db.or_(
            MDBPanel.mdb.ilike(f'%{search_term}%'),
            MDBPanel.maximo_tag.ilike(f'%{search_term}%'),
            MDBPanel.area_name.ilike(f'%{search_term}%')
        )
    )

    # فلترة حسب المناطق المخصصة للمستخدم
    if current_user.role != 'admin':
        user_areas = current_user.get_assigned_areas()
        if user_areas:
            panels_query = panels_query.filter(MDBPanel.area_name.in_(user_areas))

    panels = panels_query.limit(20).all()

    panels_data = []
    for panel in panels:
        panels_data.append({
            'id': panel.id,
            'mdb': panel.mdb or '',
            'maximo_tag': panel.maximo_tag or '',
            'area_name': panel.area_name or '',
            'display_text': f"{panel.mdb or 'غير محدد'} ({panel.maximo_tag or 'غير محدد'}) - {panel.area_name or 'غير محدد'}"
        })

    return jsonify({'panels': panels_data})

# API للبحث في طلبات الفحص
@app.route('/api/search-inspection-requests')
@login_required
def api_search_inspection_requests():
    search_term = request.args.get('search_term', '')
    status_filter = request.args.get('status_filter', 'all')
    priority_filter = request.args.get('priority_filter', 'all')

    # إنشاء استعلام أساسي
    query = InspectionRequest.query.join(MDBPanel)

    # تطبيق فلتر البحث
    if search_term and len(search_term) >= 2:
        search_pattern = f'%{search_term}%'
        query = query.filter(
            db.or_(
                MDBPanel.mdb.ilike(search_pattern),
                MDBPanel.maximo_tag.ilike(search_pattern),
                MDBPanel.area_name.ilike(search_pattern),
                InspectionRequest.title.ilike(search_pattern),
                InspectionRequest.description.ilike(search_pattern),
                InspectionRequest.request_number.ilike(search_pattern)
            )
        )

    # تطبيق فلتر الحالة
    if status_filter != 'all':
        query = query.filter(InspectionRequest.status == status_filter)

    # تطبيق فلتر الأولوية
    if priority_filter != 'all':
        query = query.filter(InspectionRequest.priority == priority_filter)

    # فلترة حسب المناطق المخصصة للمستخدم
    if current_user.role != 'admin':
        user_areas = current_user.get_assigned_areas()
        if user_areas:
            query = query.filter(MDBPanel.area_name.in_(user_areas))

    # الحصول على النتائج مع ترتيب حسب تاريخ الإنشاء
    requests = query.order_by(InspectionRequest.created_at.desc()).limit(100).all()

    # تحويل النتائج إلى JSON
    requests_data = []
    for req in requests:
        requests_data.append({
            'id': req.id,
            'request_number': req.request_number,
            'title': req.title,
            'status': req.status,
            'priority': req.priority,
            'created_at': req.created_at.isoformat(),
            'panel': {
                'id': req.panel.id,
                'mdb': req.panel.mdb,
                'maximo_tag': req.panel.maximo_tag,
                'area_name': req.panel.area_name
            }
        })

    return jsonify({'requests': requests_data})

# صفحة إضافة بلاغ جديد
@app.route('/issues/new')
@login_required
def new_issue():
    # فلترة اللوحات حسب المناطق المخصصة للمستخدم
    if current_user.role == 'admin':
        panels = MDBPanel.query.all()
    else:
        user_areas = current_user.get_assigned_areas()
        if user_areas:
            panels = MDBPanel.query.filter(MDBPanel.area_name.in_(user_areas)).all()
        else:
            panels = MDBPanel.query.all()

    contractors = Contractor.query.all()
    users = User.query.filter_by(is_active=True).all()

    # الحصول على أنواع البلاغات من قاعدة البيانات
    issue_types_dropdown = DropdownList.query.filter_by(name='issue_types', is_active=True).first()
    issue_types = []
    if issue_types_dropdown:
        issue_types = DropdownItem.query.filter_by(
            dropdown_id=issue_types_dropdown.id,
            is_active=True
        ).order_by(DropdownItem.order).all()

    # الحصول على معرف اللوحة من المعاملات
    selected_panel_id = request.args.get('panel_id')
    selected_panel = None
    if selected_panel_id:
        selected_panel = MDBPanel.query.get(selected_panel_id)

    # الحصول على المجموعات النشطة
    from models import UserGroup
    groups = UserGroup.query.filter_by(is_active=True).all()

    # الحصول على أنواع البلاغات من قاعدة البيانات
    issue_types_list = DropdownList.query.filter_by(name='issue_types').first()
    if issue_types_list:
        issue_types_items = DropdownItem.query.filter_by(
            dropdown_id=issue_types_list.id,
            is_active=True
        ).order_by(DropdownItem.order).all()
        issue_types = [item.value for item in issue_types_items]
    else:
        # أنواع افتراضية إذا لم توجد في قاعدة البيانات
        issue_types = ['عطل كهربائي', 'تلف في المعدات', 'مشكلة في التوصيلات', 'أخرى']

    return render_template('add_issue_general.html',
                           panels=panels,
                           contractors=contractors,
                           users=users,
                           groups=groups,
                           issue_types=issue_types,
                           selected_panel=selected_panel,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# ==================== AI تحليل البلاغات ====================
@app.route('/api/issues/ai-query', methods=['POST'])
@csrf.exempt
def issues_ai_query():
    """استعلام ذكاء اصطناعي عن البلاغات"""
    try:
        data = request.get_json() or {}
        question = (data.get('question') or '').strip()
        if not question:
            return jsonify({'success': False, 'error': 'يرجى إدخال سؤال'})

        # جمع إحصائيات البلاغات من قاعدة البيانات
        total = Issue.query.count()
        open_c = Issue.query.filter_by(status='مفتوح').count()
        in_prog = Issue.query.filter_by(status='قيد المعالجة').count()
        closed = Issue.query.filter_by(status='مغلق').count()
        high_p = Issue.query.filter_by(priority='عالي').count()

        # أكثر أنواع البلاغات
        from sqlalchemy import func
        type_stats = db.session.query(Issue.issue_type, func.count(Issue.id)).group_by(Issue.issue_type).order_by(func.count(Issue.id).desc()).limit(5).all()
        type_text = '\n'.join([f'  - {t}: {c} بلاغ' for t, c in type_stats if t])

        # أكثر المقاولين بلاغات
        try:
            contr_stats = db.session.query(Contractor.name, func.count(Issue.id)).join(Issue, Issue.contractor_id == Contractor.id).group_by(Contractor.name).order_by(func.count(Issue.id).desc()).limit(5).all()
            contr_text = '\n'.join([f'  - {n}: {c} بلاغ' for n, c in contr_stats])
        except Exception:
            contr_text = 'غير متاح'

        # بناء السياق
        context = f"""أنت محلل بيانات ذكي لنظام إدارة البلاغات الكهربائية في مخيمات الحج.
البيانات الحالية:
- إجمالي البلاغات: {total}
- مفتوح: {open_c} | قيد المعالجة: {in_prog} | مغلق: {closed}
- عالي الأولوية: {high_p}
- أكثر أنواع البلاغات:
{type_text or '  لا توجد بيانات'}
- البلاغات حسب المقاولين:
{contr_text or '  لا توجد بيانات'}

السؤال: {question}

أجب بالعربية بشكل مختصر ومفيد ومنظم. استند فقط للبيانات المقدمة."""

        from advanced_ai_system import get_ai_response
        resp = get_ai_response(context, 'issue_analyzer')
        return jsonify({'success': True, 'answer': resp.get('response', ''), 'model': resp.get('model_used', '')})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# إضافة بلاغ جديد
@app.route('/issues/add', methods=['POST'])
def add_issue():

    panel_ids = request.form.getlist('panel_id')  # دعم أكثر من لوحة
    title = request.form.get('title') or 'بلاغ عام'  # قيمة افتراضية
    description = request.form.get('description') or ''
    issue_type = request.form.get('issue_type') or ''
    priority = request.form.get('priority') or 'متوسط'
    contractor_id = request.form.get('contractor_id') or None
    responsible_person_id = request.form.get('responsible_person') or None
    assigned_group_id = request.form.get('assigned_group') or None

    # الحصول على اسم الشخص المسؤول إذا تم تحديده
    responsible_person_name = ''
    if responsible_person_id:
        responsible_user = User.query.get(responsible_person_id)
        if responsible_user:
            responsible_person_name = responsible_user.name

    # معالجة التخصيص
    assignment_type = request.form.get('assignment_type')
    assignee_id = request.form.get('assignee_id') or None

    # إذا لم يتم تحديد أي لوحة أو تم تحديد قيمة فارغة
    if not panel_ids or (len(panel_ids) == 1 and panel_ids[0] == ''):
        # إنشاء بلاغ عام بدون لوحة محددة
        issue = Issue(
            panel_id=None,
            title=title,
            description=description,
            issue_type=issue_type,
            priority=priority,
            contractor_id=contractor_id,
            responsible_person=responsible_person_name,
            assignee_id=assignee_id,
            assigned_group_id=assigned_group_id,
            status='مفتوح',
            created_by=current_user.id,
            created_at=datetime.datetime.now()
        )

        db.session.add(issue)
        db.session.commit()

        flash('تم إضافة البلاغ العام بنجاح', 'success')
        return redirect(url_for('issues'))

    # إنشاء بلاغ لكل لوحة محددة
    created_issues = []
    for panel_id in panel_ids:
        if panel_id and panel_id != '':
            # التحقق من وجود اللوحة
            panel = MDBPanel.query.get(panel_id)
            if not panel:
                flash(f'اللوحة رقم {panel_id} غير موجودة', 'warning')
                continue

            # إذا كان المستخدم مقاول، استخدم معرف المقاول الخاص به
            current_contractor_id = contractor_id
            if current_user.is_authenticated and current_user.role == 'contractor':
                current_contractor_id = current_user.contractor_id

            # إضافة اسم المستخدم كشخص مسؤول إذا لم يتم تحديده
            current_responsible_person = responsible_person_name
            if current_user.is_authenticated and current_user.role == 'contractor' and not responsible_person_name:
                current_responsible_person = current_user.name

            # إنشاء بلاغ لهذه اللوحة
            issue = Issue(
                panel_id=panel_id,
                title=title,
                description=description,
                issue_type=issue_type,
                priority=priority,
                contractor_id=current_contractor_id,
                responsible_person=current_responsible_person,
                assignee_id=assignee_id,
                assigned_group_id=assigned_group_id,
                status='مفتوح',
                created_by=current_user.id,
                created_at=datetime.datetime.now()
            )

            db.session.add(issue)
            created_issues.append(issue)

            # تحديث عدد البلاغات للوحة
            panel.issues_count = panel.issues_count + 1

            # تحديث عدد البلاغات للمقاول إذا تم تحديده
            if current_contractor_id:
                contractor = Contractor.query.get(current_contractor_id)
                if contractor:
                    contractor.issues_count = contractor.issues_count + 1



    db.session.commit()

    # معالجة التوزيع التلقائي إذا لم يتم تحديد مستخدم
    try:
        # إذا لم يتم تحديد مستخدم محدد، قم بالتوزيع التلقائي
        if not issue.assignee_id and not issue.assigned_group_id and not issue.responsible_person:
            if issue.panel and issue.panel.area_name:
                auto_assign_to_area_users(issue, issue.panel.area_name)
    except Exception as e:
        print(f'خطأ في التوزيع التلقائي للبلاغ {issue.id}: {str(e)}')

    # معالجة الإشعارات التلقائية
    try:
        from auto_notifications import process_new_issue
        process_new_issue(issue)
    except Exception as e:
        print(f'خطأ في معالجة الإشعارات: {str(e)}')

    flash('تم إضافة البلاغ بنجاح', 'success')

    # إعادة التوجيه إلى الصفحة المناسبة
    if current_user.is_authenticated and current_user.role == 'contractor':
        return redirect(url_for('contractor_dashboard'))
    elif panel_id:
        return redirect(url_for('panel_details', panel_id=panel_id))
    else:
        return redirect(url_for('issues'))

# تصدير البلاغات
@app.route('/issues/export', methods=['POST'])
@login_required
def export_issues():
    # الحصول على معلمات التصفية
    status = request.form.get('status', 'all')
    priority = request.form.get('priority', 'all')
    panel_ids = request.form.getlist('panel_id')
    issue_type = request.form.get('issue_type', 'all')
    export_format = request.form.get('export_format', 'excel')

    # بناء الاستعلام
    query = Issue.query

    # تطبيق الفلاتر
    if status != 'all':
        query = query.filter(Issue.status == status)

    if priority != 'all':
        query = query.filter(Issue.priority == priority)

    if panel_ids and 'all' not in panel_ids:
        query = query.filter(Issue.panel_id.in_(panel_ids))

    if issue_type != 'all':
        query = query.filter(Issue.issue_type == issue_type)

    # الحصول على البيانات
    issues = query.all()

    # إعداد بيانات التصدير
    data = []
    headers = ['رقم البلاغ', 'العنوان', 'اللوحة', 'الحالة', 'الأولوية', 'نوع البلاغ', 'مقدم البلاغ', 'المعين', 'الشخص المسؤول', 'تاريخ الإنشاء', 'تاريخ الاستحقاق', 'وقت المعالجة (دقيقة)', 'وقت الإنجاز (دقيقة)']
    data.append(headers)

    for issue in issues:
        panel_info = f"{issue.panel.mdb} ({issue.panel.maximo_tag})" if issue.panel else 'غير محدد'
        reporter = issue.creator.name if issue.creator else 'غير محدد'
        assignee = issue.assignee.name if issue.assignee else 'غير معين'
        responsible_person = issue.responsible_person if issue.responsible_person else 'غير محدد'

        # حساب وقت المعالجة (من الإنشاء حتى البدء)
        processing_time_minutes = ''
        if hasattr(issue, 'processing_time') and issue.processing_time:
            processing_time_minutes = issue.processing_time
        elif hasattr(issue, 'started_at') and issue.started_at and issue.created_at:
            time_diff = issue.started_at - issue.created_at
            processing_time_minutes = int(time_diff.total_seconds() / 60)

        # حساب وقت الإنجاز (من البدء حتى الإغلاق)
        completion_time_minutes = ''
        if hasattr(issue, 'closure_time') and issue.closure_time:
            completion_time_minutes = issue.closure_time
        elif hasattr(issue, 'closed_at') and issue.closed_at and hasattr(issue, 'started_at') and issue.started_at:
            # حساب الفرق بين وقت البدء ووقت الإغلاق
            time_diff = issue.closed_at - issue.started_at
            completion_time_minutes = int(time_diff.total_seconds() / 60)
        elif issue.status in ['مغلق', 'تم الحل'] and hasattr(issue, 'updated_at') and issue.updated_at and hasattr(issue, 'started_at') and issue.started_at:
            # إذا لم يكن هناك closed_at، استخدم updated_at
            time_diff = issue.updated_at - issue.started_at
            completion_time_minutes = int(time_diff.total_seconds() / 60)

        row = [
            issue.issue_number if hasattr(issue, 'issue_number') else f'ISS-{issue.id}',
            issue.title,
            panel_info,
            issue.status,
            issue.priority,
            issue.issue_type,
            reporter,
            assignee,
            responsible_person,
            issue.created_at.strftime('%Y-%m-%d %H:%M') if issue.created_at else '',
            issue.due_date.strftime('%Y-%m-%d') if hasattr(issue, 'due_date') and issue.due_date else '',
            processing_time_minutes if processing_time_minutes != '' else '-',
            completion_time_minutes if completion_time_minutes != '' else '-'
        ]
        data.append(row)

    # تصدير حسب الصيغة
    now = datetime.datetime.now()

    if export_format == 'excel':
        # إنشاء DataFrame
        df = pd.DataFrame(data[1:], columns=data[0])

        # إنشاء ملف Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='البلاغات')

            # تنسيق ورقة العمل
            workbook = writer.book
            worksheet = writer.sheets['البلاغات']

            # تنسيق العناوين
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#B89966',
                'border': 1
            })

            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                worksheet.set_column(col_num, col_num, 20)

        output.seek(0)
        filename = f"Issues_Report_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(output,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name=filename)

    elif export_format == 'pdf':
        # إنشاء HTML للتحويل إلى PDF
        html_content = f"""
        <html dir="rtl">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; direction: rtl; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                th {{ background-color: #B89966; color: white; text-align: center; padding: 8px; border: 1px solid #ddd; }}
                td {{ padding: 8px; text-align: center; border: 1px solid #ddd; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                h1 {{ color: #333; text-align: center; }}
                .header {{ text-align: center; margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>تقرير البلاغات</h1>
                <p>تاريخ التقرير: {now.strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>عدد البلاغات: {len(issues)}</p>
            </div>
            <table>
        """

        # إضافة العناوين
        html_content += "<tr>"
        for header in headers:
            html_content += f"<th>{header}</th>"
        html_content += "</tr>"

        # إضافة البيانات
        for row in data[1:]:
            html_content += "<tr>"
            for cell in row:
                html_content += f"<td>{cell}</td>"
            html_content += "</tr>"

        html_content += """
            </table>
        </body>
        </html>
        """

        # تحويل إلى PDF
        options = {
            'page-size': 'A4',
            'orientation': 'Landscape',
            'margin-top': '1cm',
            'margin-right': '1cm',
            'margin-bottom': '1cm',
            'margin-left': '1cm',
            'encoding': 'UTF-8'
        }

        # إنشاء ملف مؤقت
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as f:
            f.write(html_content.encode('utf-8'))
            html_path = f.name

        pdf_path = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False).name

        try:
            pdfkit.from_file(html_path, pdf_path, options=options)
            filename = f"Issues_Report_{now.strftime('%Y%m%d_%H%M%S')}.pdf"

            return send_file(pdf_path,
                            mimetype='application/pdf',
                            as_attachment=True,
                            download_name=filename)
        except Exception as e:
            flash(f'حدث خطأ في تصدير PDF: {str(e)}', 'danger')
            return redirect(url_for('issues'))
        finally:
            # حذف الملفات المؤقتة
            try:
                os.unlink(html_path)
                os.unlink(pdf_path)
            except:
                pass

# صفحة إضافة بلاغ جديد للمقاول
@app.route('/panels/<int:panel_id>/add-issue')
@login_required
def add_issue_form(panel_id):
    # التحقق من أن المستخدم هو مقاول أو مقاول عادي
    if current_user.role not in ['contractor', 'regular_contractor'] or not current_user.contractor_id:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))

    # الحصول على اللوحة
    panel = MDBPanel.query.get_or_404(panel_id)

    # التحقق مما إذا كان المستخدم مقاول وليس لديه صلاحية للوصول إلى هذه اللوحة
    contractor = Contractor.query.get_or_404(current_user.contractor_id)

    # الحصول على مناطق المسؤولية للمقاول
    contractor_areas = get_contractor_areas(contractor.id, current_user.is_manager)

    # التحقق مما إذا كانت اللوحة في منطقة مسؤولية المقاول
    has_access = False
    if panel.area_name in contractor_areas or panel.responsible_contractor_id == contractor.id:
        has_access = True

    if not has_access:
        flash('ليس لديك صلاحية للوصول إلى هذه اللوحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    return render_template('add_issue.html',
                          panel=panel,
                          current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# عرض تفاصيل البلاغ
@app.route('/issues/<int:issue_id>')
def view_issue(issue_id):
    issue = Issue.query.get_or_404(issue_id)

    # التحقق مما إذا كان المستخدم مقاول أو مقاول عادي وليس لديه صلاحية للوصول إلى هذا البلاغ
    if current_user.is_authenticated and current_user.role in ['contractor', 'regular_contractor']:
        has_access = False
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # إذا كان البلاغ مرتبط بالمقاول مباشرة
        if issue.contractor_id == contractor.id:
            has_access = True
        else:
            # إذا كان المستخدم مقاول مدير، يمكنه رؤية جميع البلاغات المرتبطة بالفرق التابعة له
            if hasattr(current_user, 'is_manager') and current_user.is_manager:
                # الحصول على جميع الفرق التابعة للمقاول
                teams = ContractorTeam.query.filter_by(contractor_id=contractor.id).all()

                # جمع جميع المناطق المسؤول عنها الفرق
                team_areas = []
                for team in teams:
                    if team.area_responsibility:
                        try:
                            # محاولة تحليل البيانات كـ JSON
                            team_areas_json = json.loads(team.area_responsibility)
                            if isinstance(team_areas_json, list):
                                team_areas.extend(team_areas_json)
                            else:
                                team_areas.append(team_areas_json)
                        except json.JSONDecodeError:
                            # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                            team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

                # التحقق مما إذا كانت اللوحة المرتبطة بالبلاغ في منطقة مسؤولية أحد الفرق
                if issue.panel and issue.panel.area_name in team_areas:
                    has_access = True

            # التحقق مما إذا كانت اللوحة المرتبطة بالبلاغ في منطقة مسؤولية المقاول
            if not has_access and issue.panel:
                contractor_areas = []
                if contractor.area_responsibility:
                    try:
                        # محاولة تحليل البيانات كـ JSON
                        contractor_areas = json.loads(contractor.area_responsibility)
                    except json.JSONDecodeError:
                        # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                        contractor_areas = [area.strip() for area in contractor.area_responsibility.split(',')]

                if issue.panel.area_name in contractor_areas or issue.panel.responsible_contractor_id == contractor.id:
                    has_access = True

        # إذا لم يكن لديه صلاحية، إعادة توجيهه إلى لوحة تحكم المقاول
        if not has_access:
            flash('ليس لديك صلاحية للوصول إلى هذا البلاغ', 'danger')
            return redirect(url_for('contractor_dashboard'))

    return render_template('issue_details.html',
                           issue=issue,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تحديث حالة البلاغ
@app.route('/issues/<int:issue_id>/status/<status>')
def update_issue_status(issue_id, status):
    issue = Issue.query.get_or_404(issue_id)

    # التحقق من صحة الحالة
    if status not in ['مفتوح', 'قيد المعالجة', 'مغلق']:
        flash('حالة غير صالحة', 'danger')
        return redirect(url_for('issues'))

    # تحديث الحالة مع تتبع الأوقات
    current_time = datetime.datetime.now()
    old_status = issue.status
    issue.status = status

    # تتبع أوقات المعالجة
    if old_status == 'مفتوح' and status == 'قيد المعالجة':
        issue.started_at = current_time
        if issue.created_at:
            issue.processing_time = int((current_time - issue.created_at).total_seconds() / 60)

    # إذا تم إغلاق البلاغ، تحديث تاريخ الإغلاق وحساب وقت الإقفال
    if status == 'مغلق':
        issue.closed_at = current_time
        if issue.created_at:
            issue.closure_time = int((current_time - issue.created_at).total_seconds() / 60)
        if current_user.is_authenticated:
            issue.resolved_by = current_user.id

    # إضافة ملاحظة تلقائية إذا كان المستخدم مسجل الدخول
    if current_user.is_authenticated:
        note = f"تم تغيير الحالة إلى {status} بواسطة {current_user.name}"

        if issue.description:
            issue.description += f"\n\n[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}] {note}"
        else:
            issue.description = f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}] {note}"

    db.session.commit()

    flash('تم تحديث حالة البلاغ بنجاح', 'success')

    # إعادة التوجيه إلى الصفحة المناسبة حسب دور المستخدم
    if current_user.is_authenticated and current_user.role in ['contractor', 'regular_contractor']:
        return redirect(url_for('contractor_dashboard'))
    else:
        return redirect(url_for('issues'))

# تحديث حالة البلاغ (POST)
@app.route('/issues/<int:issue_id>/status', methods=['POST'])
def update_issue_status_post(issue_id):
    issue = Issue.query.get_or_404(issue_id)
    status = request.form.get('status')

    # التحقق من صحة الحالة
    if status not in ['مفتوح', 'قيد المعالجة', 'مغلق']:
        flash('حالة غير صالحة', 'danger')
        return redirect(url_for('issues'))

    # تحديث الحالة
    issue.status = status

    # إذا تم إغلاق البلاغ، تحديث تاريخ الإغلاق
    if status == 'مغلق':
        issue.closed_at = datetime.datetime.now()

    # إضافة ملاحظة تلقائية إذا كان المستخدم مسجل الدخول
    if current_user.is_authenticated:
        note = f"تم تغيير الحالة إلى {status} بواسطة {current_user.name}"

        if issue.description:
            issue.description += f"\n\n[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}] {note}"
        else:
            issue.description = f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}] {note}"

    db.session.commit()

    flash('تم تحديث حالة البلاغ بنجاح', 'success')

    # إعادة التوجيه إلى الصفحة المناسبة حسب دور المستخدم
    if current_user.is_authenticated and current_user.role in ['contractor', 'regular_contractor']:
        return redirect(url_for('contractor_dashboard'))
    else:
        return redirect(url_for('issues'))

# صفحة إدارة الأعمدة الديناميكية
@app.route('/dynamic-columns')
@login_required
def dynamic_columns():
    # التحقق من أن المستخدم ليس مقاول أو مقاول عادي
    if current_user.role in ['contractor', 'regular_contractor']:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    columns = DynamicColumn.query.order_by(DynamicColumn.created_at.desc()).all()
    return render_template('dynamic_columns.html',
                           dynamic_columns=columns,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة إعدادات النظام
@app.route('/system-settings', methods=['GET', 'POST'])
@login_required
def system_settings():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    if request.method == 'POST':
        # معالجة ملف الشعار إذا تم تحميله
        if 'setting_logo_file' in request.files:
            logo_file = request.files['setting_logo_file']
            if logo_file and logo_file.filename != '':
                # التحقق من امتداد الملف
                if allowed_file(logo_file.filename, {'png', 'jpg', 'jpeg', 'gif', 'svg'}):
                    # إنشاء مجلد الصور إذا لم يكن موجودًا
                    logo_folder = os.path.join(app.static_folder, 'images', 'logo')
                    if not os.path.exists(logo_folder):
                        os.makedirs(logo_folder)

                    # حفظ الملف بإسم آمن
                    filename = secure_filename(logo_file.filename)
                    # إضافة طابع زمني لتجنب مشاكل التخزين المؤقت
                    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                    filename = f"{timestamp}_{filename}"
                    file_path = os.path.join(logo_folder, filename)
                    logo_file.save(file_path)

                    # حفظ مسار الملف في الإعدادات (نسبي إلى مجلد static)
                    relative_path = os.path.join('images', 'logo', filename).replace('\\', '/')
                    set_setting('logo_path', relative_path, 'مسار شعار النظام')
                else:
                    flash('صيغة ملف الشعار غير مدعومة. الصيغ المدعومة هي: PNG, JPG, JPEG, GIF, SVG', 'danger')

        # معالجة إزالة الشعار
        if 'setting_remove_logo' in request.form and request.form.get('setting_remove_logo') == 'on':
            # الحصول على مسار الشعار الحالي
            logo_path = get_setting('logo_path')
            if logo_path:
                # حذف الملف إذا كان موجودًا
                full_path = os.path.join(app.static_folder, logo_path)
                if os.path.exists(full_path):
                    try:
                        os.remove(full_path)
                    except Exception as e:
                        print(f"خطأ في حذف ملف الشعار: {str(e)}")

                # إزالة الإعداد
                setting = SystemSettings.query.filter_by(key='logo_path').first()
                if setting:
                    db.session.delete(setting)
                    db.session.commit()

        # تحديث باقي الإعدادات
        for key in request.form:
            if key.startswith('setting_') and key not in ['setting_logo_file', 'setting_remove_logo']:
                setting_key = key.replace('setting_', '')
                set_setting(setting_key, request.form[key])

        flash('تم تحديث إعدادات النظام بنجاح', 'success')
        return redirect(url_for('system_settings'))

    # الحصول على جميع الإعدادات
    settings = SystemSettings.query.all()

    return render_template('system_settings.html',
                           settings=settings,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إضافة عمود ديناميكي جديد
@app.route('/dynamic-columns/add', methods=['POST'])
@login_required
def add_dynamic_column():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    name = request.form.get('name')
    display_name = request.form.get('display_name')
    data_type = request.form.get('data_type')
    options_str = request.form.get('options') if data_type == 'select' else None

    # التحقق من عدم وجود عمود بنفس الاسم
    existing_column = DynamicColumn.query.filter_by(name=name).first()
    if existing_column:
        flash('يوجد عمود بنفس الاسم بالفعل', 'danger')
        return redirect(url_for('dynamic_columns'))

    # إنشاء العمود الجديد
    column = DynamicColumn(
        name=name,
        display_name=display_name,
        data_type=data_type,
        options=options_str,
        is_active=True,
        created_at=datetime.datetime.now()
    )

    db.session.add(column)
    db.session.commit()

    flash('تم إضافة العمود بنجاح', 'success')
    return redirect(url_for('dynamic_columns'))

# تحرير عمود ديناميكي
@app.route('/dynamic-columns/<int:column_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_dynamic_column(column_id):
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    column = DynamicColumn.query.get_or_404(column_id)

    if request.method == 'POST':
        display_name = request.form.get('display_name')

        # تحديث اسم العرض فقط (لا يمكن تغيير اسم العمود أو نوع البيانات بعد الإنشاء)
        column.display_name = display_name
        db.session.commit()

        flash('تم تحديث العمود بنجاح', 'success')
        return redirect(url_for('dynamic_columns'))

    return render_template('edit_dynamic_column.html',
                           column=column,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تفعيل/تعطيل عمود ديناميكي
@app.route('/dynamic-columns/<int:column_id>/toggle')
@login_required
def toggle_dynamic_column(column_id):
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    column = DynamicColumn.query.get_or_404(column_id)

    # تبديل حالة التفعيل
    column.is_active = not column.is_active
    db.session.commit()

    status = 'تفعيل' if column.is_active else 'تعطيل'
    flash(f'تم {status} العمود بنجاح', 'success')
    return redirect(url_for('dynamic_columns'))

# تحميل قالب الاستيراد
@app.route('/download-template')
@login_required
def download_template():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # إنشاء DataFrame مع الأعمدة الأساسية
    columns = [
        'MDB', 'رقم تاج ماكسيمو', 'X', 'Y', 'الملاحظات', 'المرحلة',
        'سنة التنفيذ', 'كود المنطقة - تشغيل', 'نوع لوحة التوزيع', 'اسم المنطقة بالتنفيذ',
        'الحالة'
    ]

    # إضافة الأعمدة الديناميكية النشطة
    dynamic_columns = DynamicColumn.query.filter_by(is_active=True).all()
    for column in dynamic_columns:
        columns.append(column.display_name)

    # إنشاء DataFrame فارغ مع الأعمدة
    df = pd.DataFrame(columns=columns)

    # إنشاء ملف Excel في الذاكرة
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='MDB Panels Template')

        # تنسيق ورقة العمل
        workbook = writer.book
        worksheet = writer.sheets['MDB Panels Template']

        # تنسيق العناوين
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#94BCCB',
            'border': 1
        })

        # تطبيق التنسيق على الصف الأول
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            worksheet.set_column(col_num, col_num, 15)

    output.seek(0)

    # إنشاء اسم الملف مع التاريخ والوقت
    now = datetime.datetime.now()
    filename = f"MDB_Panels_Template_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)

# تصدير تقرير شامل للوحات
@app.route('/panels/export-report', methods=['POST'])
@login_required
def export_panels_report():
    # التحقق من وجود المستخدم
    if not current_user.is_authenticated:
        flash('يجب تسجيل الدخول للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('login'))

    try:
        # الحصول على معلمات التقرير
        report_scope = request.form.get('report_scope', 'all')
        report_format = request.form.get('report_format', 'excel')

        # الحقول المطلوبة
        include_basic_info = 'include_basic_info' in request.form
        include_location = 'include_location' in request.form
        include_scada_status = 'include_scada_status' in request.form
        include_readings = 'include_readings' in request.form
        include_issues = 'include_issues' in request.form

        # الحصول على اللوحات المفلترة إذا كان نطاق التقرير هو اللوحات المفلترة فقط
        filtered_panels_json = request.form.get('filtered_panels', '[]')
        filtered_panel_ids = json.loads(filtered_panels_json)

        # تحديد اللوحات المطلوبة
        if report_scope == 'filtered' and filtered_panel_ids:
            panels = MDBPanel.query.filter(MDBPanel.id.in_(filtered_panel_ids)).all()
        else:
            panels = MDBPanel.query.all()

        # إنشاء DataFrame للتقرير
        report_data = []

        for panel in panels:
            panel_data = {}

            # معلومات اللوحة الأساسية
            if include_basic_info:
                panel_data['MDB'] = panel.mdb
                panel_data['رقم تاج ماكسيمو'] = panel.maximo_tag
                panel_data['المنطقة'] = panel.area_name
                panel_data['نوع اللوحة'] = panel.panel_type
                panel_data['الحالة'] = panel.status
                panel_data['سنة التنفيذ'] = panel.implementation_year
                panel_data['سعة القاطع'] = panel.breaker_capacity
                panel_data['الملاحظات'] = panel.notes

            # معلومات الموقع
            if include_location:
                panel_data['إحداثي X'] = panel.x_coordinate
                panel_data['إحداثي Y'] = panel.y_coordinate
                if panel.x_coordinate and panel.y_coordinate:
                    panel_data['رابط الموقع'] = f"https://www.google.com/maps?q={panel.y_coordinate},{panel.x_coordinate}"
                else:
                    panel_data['رابط الموقع'] = ''

            # حالة الارتباط بنظام سكادا
            if include_scada_status:
                panel_data['مرتبط بنظام سكادا'] = 'نعم' if panel.is_scada_connected else 'لا'

            # القراءات الكهربائية
            if include_readings:
                # الحصول على آخر قراءة للوحة
                last_reading = ElectricalReading.query.filter_by(panel_id=panel.id).order_by(ElectricalReading.timestamp.desc()).first()

                if last_reading:
                    panel_data['التيار (أمبير)'] = last_reading.current
                    panel_data['الجهد (فولت)'] = last_reading.voltage
                    panel_data['القدرة (كيلو واط)'] = last_reading.power
                    panel_data['نسبة الحمل (%)'] = last_reading.load_percentage
                    panel_data['معامل القدرة'] = last_reading.power_factor
                    panel_data['التردد (هرتز)'] = last_reading.frequency
                    panel_data['تاريخ القراءة'] = last_reading.timestamp.strftime('%Y-%m-%d %H:%M:%S') if last_reading.timestamp else ''
                else:
                    panel_data['التيار (أمبير)'] = ''
                    panel_data['الجهد (فولت)'] = ''
                    panel_data['القدرة (كيلو واط)'] = ''
                    panel_data['نسبة الحمل (%)'] = ''
                    panel_data['معامل القدرة'] = ''
                    panel_data['التردد (هرتز)'] = ''
                    panel_data['تاريخ القراءة'] = ''

            # الأعطال والمشاكل
            if include_issues:
                # الحصول على عدد الأعطال النشطة للوحة
                active_issues_count = Issue.query.filter_by(panel_id=panel.id, status='مفتوح').count()
                panel_data['عدد الأعطال النشطة'] = active_issues_count

                # الحصول على آخر عطل للوحة
                last_issue = Issue.query.filter_by(panel_id=panel.id).order_by(Issue.created_at.desc()).first()

                if last_issue:
                    panel_data['آخر عطل'] = last_issue.issue_type
                    panel_data['حالة آخر عطل'] = last_issue.status
                    panel_data['تاريخ آخر عطل'] = last_issue.created_at.strftime('%Y-%m-%d %H:%M:%S') if last_issue.created_at else ''
                else:
                    panel_data['آخر عطل'] = ''
                    panel_data['حالة آخر عطل'] = ''
                    panel_data['تاريخ آخر عطل'] = ''

            report_data.append(panel_data)

        # إنشاء DataFrame
        df = pd.DataFrame(report_data)

        # إنشاء اسم الملف مع التاريخ والوقت
        now = datetime.datetime.now()
        filename_base = f"Panels_Report_{now.strftime('%Y%m%d_%H%M%S')}"

        # تصدير التقرير حسب التنسيق المطلوب
        if report_format == 'excel':
            # إنشاء ملف Excel في الذاكرة
            output = io.BytesIO()

            # استخدام openpyxl كمحرك
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Panels Report')

                # الحصول على ورقة العمل
                workbook = writer.book
                worksheet = writer.sheets['Panels Report']

                # تنسيق العناوين
                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="B89966", end_color="B89966", fill_type="solid")

                # ضبط عرض الأعمدة
                for i, column in enumerate(df.columns):
                    column_width = max(len(str(column)), df[column].astype(str).map(len).max())
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = column_width + 2

            output.seek(0)

            return send_file(output,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            download_name=f"{filename_base}.xlsx")

        elif report_format == 'pdf':
            # إنشاء ملف HTML مؤقت
            html = df.to_html(index=False, classes='table table-striped table-bordered', border=0)

            # إضافة تنسيق CSS
            html = f"""
            <html dir="rtl">
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; }}
                    table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                    th {{ background-color: #B89966; color: white; text-align: right; padding: 8px; }}
                    td {{ padding: 8px; text-align: right; }}
                    tr:nth-child(even) {{ background-color: #f2f2f2; }}
                    h1 {{ color: #333; text-align: center; }}
                </style>
            </head>
            <body>
                <h1>تقرير اللوحات</h1>
                <p>تاريخ التقرير: {now.strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>عدد اللوحات: {len(panels)}</p>
                {html}
            </body>
            </html>
            """

            # إنشاء ملف PDF في الذاكرة
            options = {
                'page-size': 'A4',
                'margin-top': '1cm',
                'margin-right': '1cm',
                'margin-bottom': '1cm',
                'margin-left': '1cm',
                'encoding': 'UTF-8',
                'no-outline': None,
                'enable-local-file-access': None
            }

            # إنشاء ملف مؤقت للـ HTML
            with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as f:
                f.write(html.encode('utf-8'))
                html_path = f.name

            # إنشاء ملف مؤقت للـ PDF
            pdf_path = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False).name

            # تحويل HTML إلى PDF
            pdfkit.from_file(html_path, pdf_path, options=options)

            # إرسال الملف للمستخدم
            return send_file(pdf_path,
                            mimetype='application/pdf',
                            as_attachment=True,
                            download_name=f"{filename_base}.pdf")

    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير التقرير: {str(e)}', 'danger')
        return redirect(url_for('map_view'))

# تحميل قالب استيراد لوحات سكادا
@app.route('/panels/import/scada-template')
@login_required
def download_scada_panels_template():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # إنشاء DataFrame مع الأعمدة المطلوبة
    columns = ['MDB', 'الحالة']
    df = pd.DataFrame(columns=columns)

    # إضافة مثال
    df.loc[0] = ['MDB-001', 'عامل']
    df.loc[1] = ['MDB-002', 'معطل']
    df.loc[2] = ['MDB-003', 'تحت الصيانة']

    # إنشاء ملف Excel في الذاكرة
    output = io.BytesIO()

    # استخدام openpyxl كمحرك بدلاً من xlsxwriter
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='SCADA Panels')

        # الحصول على ورقة العمل
        workbook = writer.book
        worksheet = writer.sheets['SCADA Panels']

        # إضافة تعليق توضيحي للحالة
        worksheet.cell(row=1, column=2).comment = openpyxl.comments.Comment(
            'القيم المسموح بها: عامل، معطل، تحت الصيانة', 'System'
        )

        # تنسيق العناوين
        for col in range(1, 3):
            cell = worksheet.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="B89966", end_color="B89966", fill_type="solid")

        # ضبط عرض الأعمدة
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15

    output.seek(0)

    # إنشاء اسم الملف مع التاريخ والوقت
    now = datetime.datetime.now()
    filename = f"SCADA_Panels_Template_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)

# استيراد اللوحات المربوطة بنظام سكادا
@app.route('/panels/import/scada', methods=['POST'])
@login_required
def import_scada_panels():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للقيام بهذه العملية', 'danger')
        return redirect(url_for('map_view'))

    # التحقق من وجود ملف
    if 'scada_file' not in request.files:
        flash('لم يتم تحديد ملف', 'danger')
        return redirect(url_for('map_view'))

    file = request.files['scada_file']
    if file.filename == '':
        flash('لم يتم تحديد ملف', 'danger')
        return redirect(url_for('map_view'))

    # التحقق من امتداد الملف
    if not allowed_file(file.filename, {'xlsx', 'xls', 'csv'}):
        flash('امتداد الملف غير مسموح به. يجب أن يكون الملف بامتداد xlsx أو xls أو csv', 'danger')
        return redirect(url_for('map_view'))

    # الحصول على اسم العمود الذي يحتوي على أرقام اللوحات
    column_name = request.form.get('column_name', 'MDB')

    # التحقق مما إذا كان يجب إعادة تعيين جميع اللوحات
    reset_scada = 'reset_scada' in request.form

    try:
        # قراءة الملف
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)

        # التحقق من وجود العمود المطلوب
        if column_name not in df.columns:
            flash(f'العمود {column_name} غير موجود في الملف', 'danger')
            return redirect(url_for('map_view'))

        # إعادة تعيين جميع اللوحات إذا تم اختيار ذلك
        if reset_scada:
            MDBPanel.query.update({MDBPanel.is_scada_connected: False})
            db.session.commit()

        # تحديث اللوحات المربوطة بنظام سكادا
        panel_count = 0
        status_column = 'الحالة'  # اسم عمود الحالة

        for _, row in df.iterrows():
            mdb = str(row[column_name])
            panel = MDBPanel.query.filter_by(mdb=mdb).first()
            if panel:
                panel.is_scada_connected = True

                # تحديث حالة اللوحة إذا كان العمود موجودًا
                if status_column in df.columns and pd.notna(row[status_column]):
                    status = str(row[status_column]).strip()
                    # التحقق من أن الحالة صالحة
                    if status in ['عامل', 'معطل', 'تحت الصيانة']:
                        panel.status = status

                panel_count += 1

        db.session.commit()

        flash(f'تم تحديث {panel_count} لوحة مربوطة بنظام سكادا بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء استيراد اللوحات: {str(e)}', 'danger')

    return redirect(url_for('map_view'))

# تحميل قالب استيراد قراءات الكهرباء
@app.route('/download-readings-template')
@login_required
def download_readings_template():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # الحصول على أسماء الأعمدة من الإعدادات
    column_mdb = get_setting('column_name_mdb', 'MDB')
    column_timestamp = get_setting('column_name_timestamp', 'وقت القراءة')
    column_current = get_setting('column_name_current', 'التيار (أمبير)')
    column_voltage = get_setting('column_name_voltage', 'الجهد (فولت)')
    column_power = get_setting('column_name_power', 'القدرة (واط)')
    column_energy = get_setting('column_name_energy', 'الطاقة (كيلوواط ساعة)')
    column_power_factor = get_setting('column_name_power_factor', 'معامل القدرة')
    column_frequency = get_setting('column_name_frequency', 'التردد (هرتز)')
    column_load = get_setting('column_name_load', 'الحمل')
    column_breaker_capacity = get_setting('column_name_breaker_capacity', 'سعة القاطع')

    # إنشاء DataFrame مع الأعمدة المطلوبة
    columns = [
        column_mdb, column_timestamp,
        # قراءات التيار الثلاثي
        'التيار L1 (أمبير)', 'التيار L2 (أمبير)', 'التيار L3 (أمبير)', column_current,
        # قراءات الجهد الثلاثي
        'الجهد L1-L2 (فولت)', 'الجهد L2-L3 (فولت)', 'الجهد L3-L1 (فولت)',
        'الجهد L1-N (فولت)', 'الجهد L2-N (فولت)', 'الجهد L3-N (فولت)', column_voltage,
        # قراءات القدرة
        'القدرة الفعالة L1 (واط)', 'القدرة الفعالة L2 (واط)', 'القدرة الفعالة L3 (واط)', 'القدرة الفعالة الكلية (واط)',
        'القدرة الظاهرية L1 (فولت أمبير)', 'القدرة الظاهرية L2 (فولت أمبير)', 'القدرة الظاهرية L3 (فولت أمبير)', 'القدرة الظاهرية الكلية (فولت أمبير)',
        'القدرة التفاعلية L1 (فار)', 'القدرة التفاعلية L2 (فار)', 'القدرة التفاعلية L3 (فار)', 'القدرة التفاعلية الكلية (فار)',
        column_power,
        # قراءات الطاقة
        'الطاقة الفعالة (كيلوواط ساعة)', 'الطاقة الظاهرية (كيلوفولت أمبير ساعة)', 'الطاقة التفاعلية (كيلوفار ساعة)',
        column_energy,
        # قراءات معامل القدرة
        'معامل القدرة L1', 'معامل القدرة L2', 'معامل القدرة L3', 'معامل القدرة الكلي',
        column_power_factor,
        # قراءات أخرى
        column_load, column_breaker_capacity, column_frequency,
        # نوع القراءة
        'نوع القراءة'
    ]

    df = pd.DataFrame(columns=columns)

    # إضافة مثال
    now = datetime.datetime.now()

    # مثال للقراءة الثلاثية
    example_row = {
        column_mdb: 'MDB-001',
        column_timestamp: now.strftime('%Y-%m-%d %H:%M:%S'),
        'التيار L1 (أمبير)': 25.0,
        'التيار L2 (أمبير)': 26.0,
        'التيار L3 (أمبير)': 25.5,
        column_current: 25.5,
        'الجهد L1-L2 (فولت)': 380.0,
        'الجهد L2-L3 (فولت)': 381.0,
        'الجهد L3-L1 (فولت)': 379.0,
        'الجهد L1-N (فولت)': 220.0,
        'الجهد L2-N (فولت)': 221.0,
        'الجهد L3-N (فولت)': 219.0,
        column_voltage: 220.0,
        'القدرة الفعالة L1 (واط)': 5500.0,
        'القدرة الفعالة L2 (واط)': 5600.0,
        'القدرة الفعالة L3 (واط)': 5550.0,
        'القدرة الفعالة الكلية (واط)': 16650.0,
        'القدرة الظاهرية L1 (فولت أمبير)': 5800.0,
        'القدرة الظاهرية L2 (فولت أمبير)': 5900.0,
        'القدرة الظاهرية L3 (فولت أمبير)': 5850.0,
        'القدرة الظاهرية الكلية (فولت أمبير)': 17550.0,
        'القدرة التفاعلية L1 (فار)': 1800.0,
        'القدرة التفاعلية L2 (فار)': 1850.0,
        'القدرة التفاعلية L3 (فار)': 1820.0,
        'القدرة التفاعلية الكلية (فار)': 5470.0,
        column_power: 16650.0,
        'الطاقة الفعالة (كيلوواط ساعة)': 120.5,
        'الطاقة الظاهرية (كيلوفولت أمبير ساعة)': 125.0,
        'الطاقة التفاعلية (كيلوفار ساعة)': 40.0,
        column_energy: 120.5,
        'معامل القدرة L1': 0.95,
        'معامل القدرة L2': 0.94,
        'معامل القدرة L3': 0.96,
        'معامل القدرة الكلي': 0.95,
        column_power_factor: 0.95,
        column_load: 25.5,
        column_breaker_capacity: 32.0,
        column_frequency: 50.0,
        'نوع القراءة': 'imported'
    }

    # إضافة الصف الأول
    df = pd.DataFrame([example_row])

    # إضافة الصف الثاني (قراءة سابقة)
    example_row2 = example_row.copy()
    example_row2[column_timestamp] = (now - datetime.timedelta(minutes=15)).strftime('%Y-%m-%d %H:%M:%S')
    example_row2['التيار L1 (أمبير)'] = 24.5
    example_row2['التيار L2 (أمبير)'] = 25.0
    example_row2['التيار L3 (أمبير)'] = 24.8
    example_row2[column_current] = 24.8
    example_row2[column_power] = 16000.0
    example_row2['نوع القراءة'] = 'manual'

    df = pd.concat([df, pd.DataFrame([example_row2])], ignore_index=True)

    # إنشاء ملف Excel في الذاكرة
    output = io.BytesIO()

    # استخدام openpyxl كمحرك بدلاً من xlsxwriter
    df.to_excel(output, index=False, sheet_name='Electrical Readings')

    output.seek(0)

    # إنشاء اسم الملف مع التاريخ والوقت
    filename = f"Electrical_Readings_Template_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)

# صفحة لوحة المعلومات الكهربائية
@app.route('/electrical-dashboard')
@login_required
def electrical_dashboard_view():
    # الحصول على اللوحات
    panels = MDBPanel.query.all()

    # الحصول على آخر القراءات لكل لوحة (تلقائية ويدوية)
    latest_readings = {}
    for panel in panels:
        # البحث عن آخر قراءة تلقائية
        auto_reading = ElectricalReading.query.filter_by(panel_id=panel.id).order_by(ElectricalReading.timestamp.desc()).first()

        # البحث عن آخر قراءة يدوية
        manual_reading = ManualReading.query.filter_by(panel_id=panel.id).order_by(ManualReading.timestamp.desc()).first()

        # تحديد أحدث قراءة (يدوية أو تلقائية)
        if auto_reading and manual_reading:
            # استخدام القراءة الأحدث
            if manual_reading.timestamp > auto_reading.timestamp:
                # تحويل القراءة اليدوية إلى قراءة كهربائية للتوافق
                # استخدام القيمة القصوى من الأطوار الثلاثة للتيار
                max_current = manual_reading.current
                if manual_reading.is_three_phase and manual_reading.current_l1 is not None and manual_reading.current_l2 is not None and manual_reading.current_l3 is not None:
                    max_current = max(manual_reading.current_l1, manual_reading.current_l2, manual_reading.current_l3)

                # حساب نسبة الحمل بناءً على سعة القاطع
                load_percentage = None
                current_status = 'normal'
                if max_current is not None and panel.breaker_capacity is not None and panel.breaker_capacity > 0:
                    load_percentage = (max_current / panel.breaker_capacity) * 100

                    # تحديد حالة التيار بناءً على نسبة الحمل
                    warning_threshold = float(get_setting('warning_threshold', '70'))
                    danger_threshold = float(get_setting('danger_threshold', '80'))

                    if load_percentage >= danger_threshold:
                        current_status = 'danger'
                    elif load_percentage >= warning_threshold:
                        current_status = 'warning'

                reading = ElectricalReading(
                    panel_id=panel.id,
                    timestamp=manual_reading.timestamp,
                    current=max_current,  # استخدام القيمة القصوى من الأطوار الثلاثة
                    voltage=manual_reading.voltage,
                    power=manual_reading.power,
                    energy=manual_reading.energy,
                    power_factor=manual_reading.power_factor,
                    frequency=manual_reading.frequency,
                    load=load_percentage,
                    current_status=current_status,
                    # إضافة قيم الأطوار الثلاثة (فقط الحقول الموجودة في ElectricalReading)
                    current_l1=manual_reading.current_l1,
                    current_l2=manual_reading.current_l2,
                    current_l3=manual_reading.current_l3,
                    voltage_l1_l2=manual_reading.voltage_l1_l2,
                    voltage_l2_l3=manual_reading.voltage_l2_l3,
                    voltage_l3_l1=manual_reading.voltage_l3_l1,
                    voltage_l1_n=manual_reading.voltage_l1_n,
                    voltage_l2_n=manual_reading.voltage_l2_n,
                    voltage_l3_n=manual_reading.voltage_l3_n,
                    active_power_l1=manual_reading.active_power_l1,
                    active_power_l2=manual_reading.active_power_l2,
                    active_power_l3=manual_reading.active_power_l3,
                    active_power_total=manual_reading.active_power_total,
                    reactive_power_l1=manual_reading.reactive_power_l1,
                    reactive_power_l2=manual_reading.reactive_power_l2,
                    reactive_power_l3=manual_reading.reactive_power_l3,
                    reactive_power_total=manual_reading.reactive_power_total,
                    apparent_power_l1=manual_reading.apparent_power_l1,
                    apparent_power_l2=manual_reading.apparent_power_l2,
                    apparent_power_l3=manual_reading.apparent_power_l3,
                    apparent_power_total=manual_reading.apparent_power_total,
                    power_factor_l1=manual_reading.power_factor_l1,
                    power_factor_l2=manual_reading.power_factor_l2,
                    power_factor_l3=manual_reading.power_factor_l3
                )
                # إضافة خاصية reading_type كخاصية ديناميكية
                reading.reading_type = 'manual'
                latest_readings[panel.id] = reading
            else:
                # إضافة خاصية reading_type كخاصية ديناميكية
                auto_reading.reading_type = 'auto'
                latest_readings[panel.id] = auto_reading
        elif manual_reading:
            # تحويل القراءة اليدوية إلى قراءة كهربائية للتوافق
            # استخدام القيمة القصوى من الأطوار الثلاثة للتيار
            max_current = manual_reading.current
            if manual_reading.is_three_phase and manual_reading.current_l1 is not None and manual_reading.current_l2 is not None and manual_reading.current_l3 is not None:
                max_current = max(manual_reading.current_l1, manual_reading.current_l2, manual_reading.current_l3)

            # حساب نسبة الحمل بناءً على سعة القاطع
            load_percentage = None
            current_status = 'normal'
            if max_current is not None and panel.breaker_capacity is not None and panel.breaker_capacity > 0:
                load_percentage = (max_current / panel.breaker_capacity) * 100

                # تحديد حالة التيار بناءً على نسبة الحمل
                warning_threshold = float(get_setting('warning_threshold', '70'))
                danger_threshold = float(get_setting('danger_threshold', '80'))

                if load_percentage >= danger_threshold:
                    current_status = 'danger'
                elif load_percentage >= warning_threshold:
                    current_status = 'warning'

            reading = ElectricalReading(
                panel_id=panel.id,
                timestamp=manual_reading.timestamp,
                current=max_current,  # استخدام القيمة القصوى من الأطوار الثلاثة
                voltage=manual_reading.voltage,
                power=manual_reading.power,
                energy=manual_reading.energy,
                power_factor=manual_reading.power_factor,
                frequency=manual_reading.frequency,
                load=load_percentage,
                current_status=current_status,
                # إضافة قيم الأطوار الثلاثة (فقط الحقول الموجودة في ElectricalReading)
                current_l1=manual_reading.current_l1,
                current_l2=manual_reading.current_l2,
                current_l3=manual_reading.current_l3,
                voltage_l1_l2=manual_reading.voltage_l1_l2,
                voltage_l2_l3=manual_reading.voltage_l2_l3,
                voltage_l3_l1=manual_reading.voltage_l3_l1,
                voltage_l1_n=manual_reading.voltage_l1_n,
                voltage_l2_n=manual_reading.voltage_l2_n,
                voltage_l3_n=manual_reading.voltage_l3_n,
                active_power_l1=manual_reading.active_power_l1,
                active_power_l2=manual_reading.active_power_l2,
                active_power_l3=manual_reading.active_power_l3,
                active_power_total=manual_reading.active_power_total,
                reactive_power_l1=manual_reading.reactive_power_l1,
                reactive_power_l2=manual_reading.reactive_power_l2,
                reactive_power_l3=manual_reading.reactive_power_l3,
                reactive_power_total=manual_reading.reactive_power_total,
                apparent_power_l1=manual_reading.apparent_power_l1,
                apparent_power_l2=manual_reading.apparent_power_l2,
                apparent_power_l3=manual_reading.apparent_power_l3,
                apparent_power_total=manual_reading.apparent_power_total,
                power_factor_l1=manual_reading.power_factor_l1,
                power_factor_l2=manual_reading.power_factor_l2,
                power_factor_l3=manual_reading.power_factor_l3
            )
            # إضافة خاصية reading_type كخاصية ديناميكية
            reading.reading_type = 'manual'
            latest_readings[panel.id] = reading
        elif auto_reading:
            # إضافة خاصية reading_type كخاصية ديناميكية
            auto_reading.reading_type = 'auto'
            latest_readings[panel.id] = auto_reading

    # الحصول على إعدادات الألوان
    normal_color = get_setting('normal_color', '#28a745')
    warning_color = get_setting('warning_color', '#ffc107')
    danger_color = get_setting('danger_color', '#dc3545')
    trip_color = get_setting('trip_color', '#6c757d')

    # الحصول على إعدادات الحدود
    warning_threshold = float(get_setting('default_warning_threshold', '70'))
    danger_threshold = float(get_setting('default_danger_threshold', '80'))

    # الحصول على التنبيهات غير المقروءة
    unread_alerts = Alert.query.filter_by(is_read=False).order_by(Alert.timestamp.desc()).all()

    # إحصائيات التنبيهات
    total_alerts = Alert.query.count()
    warning_alerts = Alert.query.filter_by(severity='warning').count()
    danger_alerts = Alert.query.filter_by(severity='danger').count()

    # الحصول على المقاولين
    contractors = Contractor.query.all()

    # تحديد اللوحات ذات الحمل العالي
    high_load_panels_list = []
    high_load_count = 0

    # بيانات للرسم البياني
    high_load_labels = []
    high_load_values = []
    high_load_colors = []
    high_load_borders = []

    # بيانات للخارطة الحرارية
    heat_map_data = []
    high_load_markers = ""

    for panel in panels:
        if panel.id in latest_readings and latest_readings[panel.id].current and panel.breaker_capacity:
            # حساب نسبة الحمل
            load_percentage = (latest_readings[panel.id].current / panel.breaker_capacity) * 100

            # تحديد اللوحات ذات الحمل العالي (أكبر من حد التحذير)
            if load_percentage >= warning_threshold:
                high_load_panels_list.append(panel)
                high_load_count += 1

                # إضافة بيانات للرسم البياني
                high_load_labels.append(panel.mdb)
                high_load_values.append(round(load_percentage, 1))

                # تحديد اللون حسب نسبة الحمل
                if load_percentage >= danger_threshold:
                    high_load_colors.append(danger_color)
                    high_load_borders.append(danger_color)
                else:
                    high_load_colors.append(warning_color)
                    high_load_borders.append(warning_color)

                # إضافة بيانات للخارطة الحرارية إذا كانت الإحداثيات متوفرة
                if panel.x_coordinate and panel.y_coordinate:
                    # إضافة نقطة للخارطة الحرارية مع وزن يعتمد على نسبة الحمل
                    intensity = min(1.0, load_percentage / 100)
                    heat_map_data.append([panel.y_coordinate, panel.x_coordinate, intensity])

                    # إضافة علامة للوحة على الخارطة
                    popup_content = f"""
                        <div style='direction: rtl; text-align: right;'>
                            <h5>{panel.mdb}</h5>
                            <p>المنطقة: {panel.area_name}</p>
                            <p>نسبة الحمل: {round(load_percentage, 1)}%</p>
                            <p>التيار: {round(latest_readings[panel.id].current, 1)} أمبير</p>
                            <p>سعة القاطع: {panel.breaker_capacity} أمبير</p>
                            <a href='/panels/{panel.id}/readings' class='btn btn-primary btn-sm' target='_blank'>عرض التفاصيل</a>
                        </div>
                    """

                    # تحديد لون العلامة حسب نسبة الحمل
                    marker_color = danger_color if load_percentage >= danger_threshold else warning_color

                    # إضافة العلامة إلى الخارطة
                    high_load_markers += f"""
                        L.marker([{panel.y_coordinate}, {panel.x_coordinate}], {{
                            icon: L.divIcon({{
                                className: 'custom-div-icon',
                                html: "<div style='background-color: {marker_color}; width: 10px; height: 10px; border-radius: 50%;'></div>",
                                iconSize: [10, 10],
                                iconAnchor: [5, 5]
                            }})
                        }}).addTo(map).bindPopup(`{popup_content}`);
                    """

    # ترتيب اللوحات ذات الحمل العالي حسب نسبة الحمل (تنازلياً)
    high_load_panels_list.sort(key=lambda panel: (latest_readings[panel.id].current / panel.breaker_capacity) * 100 if panel.id in latest_readings and latest_readings[panel.id].current and panel.breaker_capacity else 0, reverse=True)

    return render_template('electrical_dashboard.html',
                           panels=panels,
                           latest_readings=latest_readings,
                           unread_alerts=unread_alerts,
                           total_alerts=total_alerts,
                           warning_alerts=warning_alerts,
                           danger_alerts=danger_alerts,
                           contractors=contractors,
                           normal_color=normal_color,
                           warning_color=warning_color,
                           danger_color=danger_color,
                           trip_color=trip_color,
                           warning_threshold=warning_threshold,
                           danger_threshold=danger_threshold,
                           high_load_panels=high_load_count,
                           high_load_panels_list=high_load_panels_list,
                           high_load_labels=json.dumps(high_load_labels),
                           high_load_values=json.dumps(high_load_values),
                           high_load_colors=json.dumps(high_load_colors),
                           high_load_borders=json.dumps(high_load_borders),
                           heat_map_data=json.dumps(heat_map_data),
                           high_load_markers=high_load_markers,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة تفاصيل القراءات الكهربائية للوحة
@app.route('/panels/<int:panel_id>/readings')
@login_required
def panel_readings(panel_id):
    panel = MDBPanel.query.get_or_404(panel_id)

    # التحقق من الصلاحيات - المدير له صلاحية الوصول لجميع اللوحات
    if current_user.is_authenticated and current_user.role not in ['admin'] and current_user.role in ['contractor', 'regular_contractor']:
        has_access = False
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # الحصول على مناطق المسؤولية للمقاول
        contractor_areas = get_contractor_areas(contractor.id, current_user.is_manager if hasattr(current_user, 'is_manager') else False)

        # التحقق مما إذا كانت اللوحة في منطقة مسؤولية المقاول
        if panel.area_name in contractor_areas or panel.responsible_contractor_id == contractor.id:
            has_access = True

        if not has_access:
            flash('ليس لديك صلاحية للوصول إلى هذه اللوحة', 'danger')
            return redirect(url_for('contractor_dashboard'))

    # الحصول على عدد الساعات للفلترة
    hours = request.args.get('hours', '24')

    # التحقق من الفترة المخصصة
    start_date = None
    end_date = None

    if hours == 'custom':
        # استخدام الفترة المخصصة
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if start_date_str and end_date_str:
            try:
                start_date = datetime.datetime.fromisoformat(start_date_str)
                end_date = datetime.datetime.fromisoformat(end_date_str)
                start_time = start_date
            except ValueError:
                # إذا كان هناك خطأ في تنسيق التاريخ، استخدم الإعداد الافتراضي
                hours = get_setting('analysis_period', '24')
                try:
                    hours = float(hours)
                except ValueError:
                    hours = 24
                start_time = datetime.datetime.now() - datetime.timedelta(hours=hours)
        else:
            # إذا لم يتم تحديد تاريخ البداية أو النهاية، استخدم الإعداد الافتراضي
            hours = get_setting('analysis_period', '24')
            try:
                hours = float(hours)
            except ValueError:
                hours = 24
            start_time = datetime.datetime.now() - datetime.timedelta(hours=hours)
    else:
        # استخدام عدد الساعات المحدد
        try:
            hours = float(hours)
        except ValueError:
            hours = 24
        start_time = datetime.datetime.now() - datetime.timedelta(hours=hours)

    # الحصول على القراءات
    query = ElectricalReading.query.filter(ElectricalReading.panel_id == panel_id)

    if end_date:
        # إذا تم تحديد فترة مخصصة، استخدم تاريخ البداية والنهاية
        query = query.filter(
            ElectricalReading.timestamp >= start_date,
            ElectricalReading.timestamp <= end_date
        )
    else:
        # استخدام وقت البداية فقط
        query = query.filter(ElectricalReading.timestamp >= start_time)

    readings = query.order_by(ElectricalReading.timestamp.desc()).all()

    # إعادة حساب نسبة الحمل والحالة لكل قراءة
    warning_threshold = float(get_setting('warning_threshold', '70'))
    danger_threshold = float(get_setting('danger_threshold', '80'))

    for reading in readings:
        # حساب نسبة الحمل إذا لم تكن محسوبة أو كانت خاطئة
        if panel.breaker_capacity and reading.current is not None and panel.breaker_capacity > 0:
            # حساب نسبة الحمل الصحيحة
            load_percentage = (reading.current / panel.breaker_capacity) * 100
            reading.load = load_percentage

            # تحديد حالة التيار بناءً على نسبة الحمل
            if load_percentage >= danger_threshold:
                reading.current_status = 'danger'
            elif load_percentage >= warning_threshold:
                reading.current_status = 'warning'
            else:
                reading.current_status = 'normal'

        # تحديد حالة الجهد
        min_voltage = float(get_setting('min_voltage', '210'))
        max_voltage = float(get_setting('max_voltage', '250'))

        if reading.voltage is not None:
            if reading.voltage < min_voltage or reading.voltage > max_voltage:
                reading.voltage_status = 'danger'
            elif reading.voltage < (min_voltage + 5) or reading.voltage > (max_voltage - 5):
                reading.voltage_status = 'warning'
            else:
                reading.voltage_status = 'normal'

    # الحصول على التنبيهات
    alerts_query = Alert.query.filter(Alert.panel_id == panel_id)

    if end_date:
        # إذا تم تحديد فترة مخصصة، استخدم تاريخ البداية والنهاية
        alerts_query = alerts_query.filter(
            Alert.timestamp >= start_date,
            Alert.timestamp <= end_date
        )
    else:
        # استخدام وقت البداية فقط
        alerts_query = alerts_query.filter(Alert.timestamp >= start_time)

    alerts = alerts_query.order_by(Alert.timestamp.desc()).all()

    # إعداد بيانات الرسوم البيانية
    timestamps = []
    current_values = []
    voltage_values = []
    power_values = []

    if readings:
        for reading in reversed(readings):
            timestamps.append(reading.timestamp.strftime('%Y-%m-%d %H:%M'))
            current_values.append(reading.current)
            voltage_values.append(reading.voltage)
            power_values.append(reading.power)

    # الحصول على المقاولين للعرض في الصفحة
    contractors = Contractor.query.all()

    # إضافة دالة get_setting إلى قالب Jinja
    def template_get_setting(key, default=None):
        return get_setting(key, default)

    return render_template('panel_readings.html',
                           panel=panel,
                           readings=readings,
                           alerts=alerts,
                           hours=hours,
                           start_date=start_date.isoformat() if start_date else None,
                           end_date=end_date.isoformat() if end_date else None,
                           timestamps=timestamps,
                           current_values=current_values,
                           voltage_values=voltage_values,
                           power_values=power_values,
                           load_values=[reading.load for reading in reversed(readings) if reading.load is not None],
                           contractors=contractors,
                           get_setting=template_get_setting,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# صفحة التنبيهات
@app.route('/alerts')
@login_required
def alerts():
    # التحقق مما إذا كان المستخدم مقاول
    if current_user.is_authenticated and current_user.role == 'contractor':
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # الحصول على مناطق المسؤولية للمقاول
        contractor_areas = []
        if contractor.area_responsibility:
            try:
                # محاولة تحليل البيانات كـ JSON
                contractor_areas = json.loads(contractor.area_responsibility)
            except json.JSONDecodeError:
                # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                contractor_areas = [area.strip() for area in contractor.area_responsibility.split(',')]

        # إذا كان المستخدم مقاول مدير، يمكنه رؤية جميع التنبيهات المرتبطة بالفرق التابعة له
        if current_user.is_manager:
            # الحصول على جميع الفرق التابعة للمقاول
            teams = ContractorTeam.query.filter_by(contractor_id=contractor.id).all()

            # جمع جميع المناطق المسؤول عنها الفرق
            team_areas = []
            for team in teams:
                if team.area_responsibility:
                    try:
                        # محاولة تحليل البيانات كـ JSON
                        team_areas_json = json.loads(team.area_responsibility)
                        if isinstance(team_areas_json, list):
                            team_areas.extend(team_areas_json)
                        else:
                            team_areas.append(team_areas_json)
                    except json.JSONDecodeError:
                        # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                        team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

            # إضافة مناطق الفرق إلى مناطق المقاول
            contractor_areas.extend(team_areas)

            # إزالة التكرار
            contractor_areas = list(set(contractor_areas))

        # الحصول على اللوحات في مناطق المسؤولية
        if contractor_areas:
            panels = MDBPanel.query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == contractor.id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            ).all()
        else:
            # إذا لم تكن هناك مناطق محددة، استخدم فقط المقاول المسؤول
            panels = MDBPanel.query.filter(MDBPanel.responsible_contractor_id == contractor.id).all()

        # الحصول على معرفات اللوحات
        panel_ids = [panel.id for panel in panels]

        # تصفية التنبيهات حسب اللوحات المسؤول عنها المقاول
        alerts_query = Alert.query.filter(Alert.panel_id.in_(panel_ids))
    else:
        # الحصول على جميع التنبيهات للمستخدمين الآخرين
        alerts_query = Alert.query

    # تطبيق الفلترة
    severity = request.args.get('severity')
    if severity and severity != 'all':
        alerts_query = alerts_query.filter(Alert.severity == severity)

    alert_type = request.args.get('type')
    if alert_type and alert_type != 'all':
        alerts_query = alerts_query.filter(Alert.alert_type == alert_type)

    is_read = request.args.get('is_read')
    if is_read == 'read':
        alerts_query = alerts_query.filter(Alert.is_read == True)
    elif is_read == 'unread':
        alerts_query = alerts_query.filter(Alert.is_read == False)

    # الحصول على التنبيهات مرتبة حسب الوقت
    alerts_list = alerts_query.order_by(Alert.timestamp.desc()).all()

    # إحصائيات التنبيهات
    if current_user.is_authenticated and current_user.role == 'contractor':
        # إحصائيات التنبيهات للمقاول
        total_alerts = alerts_query.count()
        warning_alerts = alerts_query.filter(Alert.severity == 'warning').count()
        danger_alerts = alerts_query.filter(Alert.severity == 'danger').count()
        unread_alerts = alerts_query.filter(Alert.is_read == False).count()
    else:
        # إحصائيات التنبيهات لجميع المستخدمين
        total_alerts = Alert.query.count()
        warning_alerts = Alert.query.filter_by(severity='warning').count()
        danger_alerts = Alert.query.filter_by(severity='danger').count()
        unread_alerts = Alert.query.filter_by(is_read=False).count()

    return render_template('alerts.html',
                           alerts=alerts_list,
                           total_alerts=total_alerts,
                           warning_alerts=warning_alerts,
                           danger_alerts=danger_alerts,
                           unread_alerts=unread_alerts,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تحديث حالة التنبيه (مقروء/غير مقروء)
@app.route('/alerts/<int:alert_id>/toggle-read')
def toggle_alert_read(alert_id):
    alert = Alert.query.get_or_404(alert_id)

    # التحقق مما إذا كان المستخدم مقاول وليس لديه صلاحية للوصول إلى هذا التنبيه
    if current_user.is_authenticated and current_user.role == 'contractor':
        has_access = False
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # الحصول على مناطق المسؤولية للمقاول
        contractor_areas = []
        if contractor.area_responsibility:
            try:
                # محاولة تحليل البيانات كـ JSON
                contractor_areas = json.loads(contractor.area_responsibility)
            except json.JSONDecodeError:
                # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                contractor_areas = [area.strip() for area in contractor.area_responsibility.split(',')]

        # إذا كان المستخدم مقاول مدير، يمكنه رؤية جميع التنبيهات المرتبطة بالفرق التابعة له
        if current_user.is_manager:
            # الحصول على جميع الفرق التابعة للمقاول
            teams = ContractorTeam.query.filter_by(contractor_id=contractor.id).all()

            # جمع جميع المناطق المسؤول عنها الفرق
            team_areas = []
            for team in teams:
                if team.area_responsibility:
                    try:
                        # محاولة تحليل البيانات كـ JSON
                        team_areas_json = json.loads(team.area_responsibility)
                        if isinstance(team_areas_json, list):
                            team_areas.extend(team_areas_json)
                        else:
                            team_areas.append(team_areas_json)
                    except json.JSONDecodeError:
                        # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                        team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

            # إضافة مناطق الفرق إلى مناطق المقاول
            contractor_areas.extend(team_areas)

            # إزالة التكرار
            contractor_areas = list(set(contractor_areas))

        # التحقق مما إذا كانت اللوحة المرتبطة بالتنبيه في منطقة مسؤولية المقاول
        panel = MDBPanel.query.get(alert.panel_id)
        if panel:
            if panel.responsible_contractor_id == contractor.id or panel.area_name in contractor_areas:
                has_access = True

        # إذا لم يكن لديه صلاحية، إعادة توجيهه إلى لوحة تحكم المقاول
        if not has_access:
            flash('ليس لديك صلاحية للوصول إلى هذا التنبيه', 'danger')
            return redirect(url_for('contractor_dashboard'))

    alert.is_read = not alert.is_read
    db.session.commit()

    status = 'مقروء' if alert.is_read else 'غير مقروء'
    flash(f'تم تحديث حالة التنبيه إلى {status}', 'success')

    return redirect(url_for('alerts'))

# تحديث حالة التنبيه (تم الحل/لم يتم الحل)
@app.route('/alerts/<int:alert_id>/toggle-resolved')
def toggle_alert_resolved(alert_id):
    alert = Alert.query.get_or_404(alert_id)

    # التحقق مما إذا كان المستخدم مقاول وليس لديه صلاحية للوصول إلى هذا التنبيه
    if current_user.is_authenticated and current_user.role == 'contractor':
        has_access = False
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # الحصول على مناطق المسؤولية للمقاول
        contractor_areas = []
        if contractor.area_responsibility:
            try:
                # محاولة تحليل البيانات كـ JSON
                contractor_areas = json.loads(contractor.area_responsibility)
            except json.JSONDecodeError:
                # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                contractor_areas = [area.strip() for area in contractor.area_responsibility.split(',')]

        # إذا كان المستخدم مقاول مدير، يمكنه رؤية جميع التنبيهات المرتبطة بالفرق التابعة له
        if current_user.is_manager:
            # الحصول على جميع الفرق التابعة للمقاول
            teams = ContractorTeam.query.filter_by(contractor_id=contractor.id).all()

            # جمع جميع المناطق المسؤول عنها الفرق
            team_areas = []
            for team in teams:
                if team.area_responsibility:
                    try:
                        # محاولة تحليل البيانات كـ JSON
                        team_areas_json = json.loads(team.area_responsibility)
                        if isinstance(team_areas_json, list):
                            team_areas.extend(team_areas_json)
                        else:
                            team_areas.append(team_areas_json)
                    except json.JSONDecodeError:
                        # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                        team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

            # إضافة مناطق الفرق إلى مناطق المقاول
            contractor_areas.extend(team_areas)

            # إزالة التكرار
            contractor_areas = list(set(contractor_areas))

        # التحقق مما إذا كانت اللوحة المرتبطة بالتنبيه في منطقة مسؤولية المقاول
        panel = MDBPanel.query.get(alert.panel_id)
        if panel:
            if panel.responsible_contractor_id == contractor.id or panel.area_name in contractor_areas:
                has_access = True

        # إذا لم يكن لديه صلاحية، إعادة توجيهه إلى لوحة تحكم المقاول
        if not has_access:
            flash('ليس لديك صلاحية للوصول إلى هذا التنبيه', 'danger')
            return redirect(url_for('contractor_dashboard'))

    alert.is_resolved = not alert.is_resolved
    if alert.is_resolved:
        alert.resolved_at = datetime.datetime.now()
    else:
        alert.resolved_at = None

    db.session.commit()

    status = 'تم الحل' if alert.is_resolved else 'لم يتم الحل'
    flash(f'تم تحديث حالة التنبيه إلى {status}', 'success')

    return redirect(url_for('alerts'))

# استيراد قراءات الكهرباء
@app.route('/import-readings', methods=['POST'])
@login_required
def import_readings():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    if 'file' not in request.files:
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('electrical_dashboard_view'))

    file = request.files['file']

    if file.filename == '':
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('electrical_dashboard_view'))

    if file and allowed_file(file.filename, {'xlsx', 'xls', 'csv'}):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        try:
            # قراءة الملف حسب نوعه
            if filename.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                try:
                    df = pd.read_excel(file_path, engine='openpyxl')
                except Exception as e:
                    # محاولة استخدام محرك آخر إذا فشل المحرك الأول
                    try:
                        df = pd.read_excel(file_path, engine='xlrd')
                    except Exception as e2:
                        raise Exception(f"فشل قراءة الملف: {str(e)} | {str(e2)}")

            # الحصول على تخصيص أسماء الأعمدة من الإعدادات
            column_mapping = {
                'mdb_column': get_setting('column_name_mdb', 'MDB'),
                'timestamp_column': get_setting('column_name_timestamp', 'وقت القراءة'),
                'current_column': get_setting('column_name_current', 'التيار (أمبير)'),
                'voltage_column': get_setting('column_name_voltage', 'الجهد (فولت)'),
                'power_column': get_setting('column_name_power', 'القدرة (واط)'),
                'energy_column': get_setting('column_name_energy', 'الطاقة (كيلوواط ساعة)'),
                'power_factor_column': get_setting('column_name_power_factor', 'معامل القدرة'),
                'frequency_column': get_setting('column_name_frequency', 'التردد (هرتز)'),
                'load_column': get_setting('column_name_load', 'الحمل'),
                'breaker_capacity_column': get_setting('column_name_breaker_capacity', 'سعة القاطع')
            }

            # التحقق من وجود الأعمدة المطلوبة
            required_columns = [column_mapping['mdb_column'], column_mapping['timestamp_column'],
                               column_mapping['current_column'], column_mapping['voltage_column']]

            for col in required_columns:
                if col not in df.columns:
                    flash(f'العمود {col} غير موجود في الملف', 'danger')
                    return redirect(url_for('electrical_dashboard_view'))

            # عدد القراءات التي تمت إضافتها
            readings_added = 0
            alerts_generated = 0

            # إضافة القراءات
            for _, row in df.iterrows():
                # البحث عن اللوحة بواسطة MDB
                panel = MDBPanel.query.filter_by(mdb=row[column_mapping['mdb_column']]).first()
                if not panel:
                    continue

                # تحويل وقت القراءة إلى كائن datetime
                try:
                    # التحقق من وجود قيمة صالحة للوقت
                    if pd.isna(row[column_mapping['timestamp_column']]):
                        continue

                    timestamp = pd.to_datetime(row[column_mapping['timestamp_column']])

                    # التحقق من أن التاريخ ليس في المستقبل
                    if timestamp > datetime.datetime.now() + datetime.timedelta(hours=1):  # سماح بفارق ساعة للتوقيت
                        print(f"تم تجاهل قراءة بتاريخ مستقبلي: {timestamp}")
                        continue
                except Exception as e:
                    print(f"خطأ في تحويل التاريخ: {str(e)}")
                    continue

                # تحديث سعة القاطع إذا كانت موجودة في الملف
                if column_mapping['breaker_capacity_column'] in df.columns and pd.notna(row[column_mapping['breaker_capacity_column']]):
                    try:
                        # التحقق من أن القيمة رقمية
                        breaker_capacity_value = row[column_mapping['breaker_capacity_column']]
                        if isinstance(breaker_capacity_value, (int, float)) or (isinstance(breaker_capacity_value, str) and breaker_capacity_value.replace('.', '', 1).isdigit()):
                            breaker_capacity = float(breaker_capacity_value)
                            if breaker_capacity > 0:
                                panel.breaker_capacity = breaker_capacity
                                db.session.add(panel)
                                print(f"تم تحديث سعة القاطع للوحة {panel.mdb} إلى {breaker_capacity}")
                    except (ValueError, TypeError) as e:
                        print(f"خطأ في تحويل سعة القاطع: {str(e)}")

                # التحقق من عدم وجود قراءة بنفس الوقت لنفس اللوحة
                existing_reading = ElectricalReading.query.filter_by(
                    panel_id=panel.id,
                    timestamp=timestamp
                ).first()

                if existing_reading:
                    continue

                # استخراج القيم مع التعامل مع القيم غير الرقمية
                try:
                    current = float(row[column_mapping['current_column']]) if pd.notna(row[column_mapping['current_column']]) else None
                except (ValueError, TypeError):
                    current = None

                try:
                    voltage = float(row[column_mapping['voltage_column']]) if pd.notna(row[column_mapping['voltage_column']]) else None
                except (ValueError, TypeError):
                    voltage = None

                try:
                    power = float(row[column_mapping['power_column']]) if column_mapping['power_column'] in row and pd.notna(row[column_mapping['power_column']]) else None
                except (ValueError, TypeError):
                    power = None

                try:
                    energy = float(row[column_mapping['energy_column']]) if column_mapping['energy_column'] in row and pd.notna(row[column_mapping['energy_column']]) else None
                except (ValueError, TypeError):
                    energy = None

                try:
                    power_factor = float(row[column_mapping['power_factor_column']]) if column_mapping['power_factor_column'] in row and pd.notna(row[column_mapping['power_factor_column']]) else None
                except (ValueError, TypeError):
                    power_factor = None

                try:
                    frequency = float(row[column_mapping['frequency_column']]) if column_mapping['frequency_column'] in row and pd.notna(row[column_mapping['frequency_column']]) else None
                except (ValueError, TypeError):
                    frequency = None

                # استخراج قيمة الحمل
                load = None
                if column_mapping['load_column'] in row and pd.notna(row[column_mapping['load_column']]):
                    try:
                        load = float(row[column_mapping['load_column']])
                    except (ValueError, TypeError):
                        load = None
                elif current is not None:
                    # إذا لم يكن الحمل موجودًا، نستخدم قيمة التيار كبديل
                    load = current

                # تحديد حالة القراءات
                current_status = 'normal'
                voltage_status = 'normal'
                power_status = 'normal'
                is_tripped = False

                # التحقق من حالة اللوحة (مفصولة أم لا)
                if current is not None and current == 0 and voltage is not None and voltage > 0:
                    # اللوحة مفصولة (trip) - يوجد جهد ولكن لا يوجد تيار
                    is_tripped = True
                    current_status = 'danger'
                elif current is not None and current < 0.1 and voltage is not None and voltage < 10:
                    # اللوحة مفصولة تماماً (لا يوجد جهد ولا تيار)
                    is_tripped = True
                    current_status = 'danger'
                    voltage_status = 'danger'

                # التحقق من حالة التيار والحمل (إذا لم تكن اللوحة مفصولة)
                if not is_tripped and panel.breaker_capacity is not None and panel.breaker_capacity > 0:
                    try:
                        # حساب نسبة التيار
                        current_percentage = None
                        if current is not None and isinstance(current, (int, float)):
                            current_percentage = (current / panel.breaker_capacity) * 100

                        # حساب نسبة الحمل
                        load_percentage = None
                        if load is not None and isinstance(load, (int, float)):
                            load_percentage = (load / panel.breaker_capacity) * 100

                        # استخدام النسبة الأعلى للتنبيهات (إما التيار أو الحمل)
                        percentage_to_use = None
                        if current_percentage is not None and load_percentage is not None:
                            percentage_to_use = max(current_percentage, load_percentage)
                        elif current_percentage is not None:
                            percentage_to_use = current_percentage
                        elif load_percentage is not None:
                            percentage_to_use = load_percentage
                    except (ValueError, TypeError, ZeroDivisionError) as e:
                        print(f"خطأ في حساب نسبة الحمل: {str(e)}")
                        percentage_to_use = None

                    # استخدام قيم الحدود من اللوحة أو القيم الافتراضية من الإعدادات
                    warning_threshold = panel.warning_threshold if panel.warning_threshold is not None else float(get_setting('default_warning_threshold', '70'))
                    danger_threshold = panel.danger_threshold if panel.danger_threshold is not None else float(get_setting('default_danger_threshold', '80'))

                    # تحديد حالة التيار بناءً على النسبة المستخدمة
                    if percentage_to_use is not None:
                        # التحقق من طريقة التنبيه (القيمة القصوى أو المتوسط)
                        alert_method = get_setting('alert_method', 'peak')

                        # إذا كانت طريقة التنبيه هي القيمة القصوى
                        if alert_method == 'peak':
                            if percentage_to_use >= danger_threshold:
                                current_status = 'danger'
                            elif percentage_to_use >= warning_threshold:
                                current_status = 'warning'
                    # إذا كانت طريقة التنبيه هي المتوسط، سنقوم بحساب المتوسط لاحقاً
                    if percentage_to_use is not None and alert_method == 'average':
                        # الحصول على فترة حساب المتوسط من الإعدادات
                        calculation_period = int(get_setting('alert_calculation_period', '60'))

                        # حساب وقت بداية الفترة
                        start_time = timestamp - datetime.timedelta(minutes=calculation_period)

                        # الحصول على القراءات السابقة خلال الفترة المحددة
                        previous_readings = ElectricalReading.query.filter(
                            ElectricalReading.panel_id == panel.id,
                            ElectricalReading.timestamp >= start_time,
                            ElectricalReading.timestamp <= timestamp
                        ).all()

                        # حساب متوسط التيار والحمل إذا كانت هناك قراءات سابقة
                        if previous_readings:
                            # حساب متوسط التيار
                            avg_current = None
                            current_readings = [r.current for r in previous_readings if r.current is not None]
                            if current_readings:
                                avg_current = sum(current_readings) / len(current_readings)

                            # حساب متوسط الحمل
                            avg_load = None
                            load_readings = [r.load for r in previous_readings if r.load is not None]
                            if load_readings:
                                avg_load = sum(load_readings) / len(load_readings)

                            # حساب النسب المئوية
                            avg_current_percentage = None
                            if avg_current is not None:
                                avg_current_percentage = (avg_current / panel.breaker_capacity) * 100

                            avg_load_percentage = None
                            if avg_load is not None:
                                avg_load_percentage = (avg_load / panel.breaker_capacity) * 100

                            # استخدام النسبة الأعلى
                            avg_percentage = None
                            if avg_current_percentage is not None and avg_load_percentage is not None:
                                avg_percentage = max(avg_current_percentage, avg_load_percentage)
                            elif avg_current_percentage is not None:
                                avg_percentage = avg_current_percentage
                            elif avg_load_percentage is not None:
                                avg_percentage = avg_load_percentage

                            # تحديد الحالة بناءً على النسبة
                            if avg_percentage is not None:
                                if avg_percentage >= danger_threshold:
                                    current_status = 'danger'
                                elif avg_percentage >= warning_threshold:
                                    current_status = 'warning'

                # التحقق من حالة الجهد
                if not is_tripped and voltage is not None:
                    # استخدام قيم الحدود من اللوحة أو القيم الافتراضية من الإعدادات
                    min_voltage = panel.min_voltage if panel.min_voltage is not None else float(get_setting('default_min_voltage', '210'))
                    max_voltage = panel.max_voltage if panel.max_voltage is not None else float(get_setting('default_max_voltage', '250'))

                    if voltage > max_voltage:
                        voltage_status = 'danger'
                    elif voltage < min_voltage:
                        voltage_status = 'danger'

                # إنشاء قراءة جديدة
                reading = ElectricalReading(
                    panel_id=panel.id,
                    timestamp=timestamp,
                    current=current,
                    voltage=voltage,
                    power=power,
                    energy=energy,
                    power_factor=power_factor,
                    frequency=frequency,
                    load=load,
                    current_status=current_status,
                    voltage_status=voltage_status,
                    power_status=power_status
                )

                db.session.add(reading)
                readings_added += 1

                # إنشاء تنبيهات إذا لزم الأمر
                if is_tripped:
                    # تنبيه للوحة المفصولة
                    if current is not None and current == 0 and voltage is not None and voltage > 0:
                        message = f'اللوحة مفصولة (Trip): يوجد جهد ({voltage:.1f} فولت) ولكن لا يوجد تيار'
                    else:
                        message = f'اللوحة مفصولة تماماً: لا يوجد جهد ولا تيار'

                    alert = Alert(
                        panel_id=panel.id,
                        reading_id=reading.id,
                        alert_type='trip',
                        severity='danger',
                        message=message,
                        timestamp=timestamp
                    )
                    db.session.add(alert)
                    alerts_generated += 1

                    # إرسال إشعار للمقاول المسؤول
                    if panel.responsible_contractor_id and get_setting('send_notifications', 'true') == 'true':
                        # الحصول على معلومات المقاول
                        contractor = Contractor.query.get(panel.responsible_contractor_id)
                        if contractor:
                            # إنشاء رابط الموقع إذا كان متاحًا
                            location_link = ""
                            if panel.location_url:
                                location_link = f"<a href='{panel.location_url}' target='_blank'>رابط الموقع على الخارطة</a>"
                            elif panel.x_coordinate and panel.y_coordinate:
                                maps_url = create_google_maps_url(panel.x_coordinate, panel.y_coordinate)
                                if maps_url:
                                    location_link = f"<a href='{maps_url}' target='_blank'>رابط الموقع على الخارطة</a>"

                            # إنشاء رسالة التنبيه
                            alert_message = f"""
                            <h3>تنبيه: فصل لوحة كهربائية (Trip)</h3>
                            <p><strong>اللوحة:</strong> {panel.mdb}</p>
                            <p><strong>رقم تاج ماكسيمو:</strong> {panel.maximo_tag}</p>
                            <p><strong>المنطقة:</strong> {panel.area_name}</p>
                            <p><strong>وقت الفصل:</strong> {timestamp.strftime('%Y-%m-%d %H:%M:%S')}</p>
                            <p><strong>نوع التنبيه:</strong> فصل (Trip)</p>
                            {location_link}
                            <p>يرجى التوجه للموقع للكشف على اللوحة وإصلاح المشكلة في أقرب وقت ممكن.</p>
                            """

                            # إنشاء نص عادي للرسائل النصية
                            sms_message = f"تنبيه: فصل لوحة كهربائية (Trip) - اللوحة: {panel.mdb} - المنطقة: {panel.area_name} - يرجى التوجه للموقع للكشف على اللوحة"

                            # إرسال الإشعار للمقاول
                            subject = f"تنبيه: فصل لوحة كهربائية {panel.mdb}"
                            notify_contractor(contractor, subject, alert_message, sms_message)

                            print(f"تم إرسال تنبيه فصل للمقاول {contractor.name}")

                elif current_status in ['warning', 'danger']:
                    # استخدام قيم الحدود من اللوحة أو القيم الافتراضية من الإعدادات
                    warning_threshold = panel.warning_threshold if panel.warning_threshold is not None else float(get_setting('default_warning_threshold', '70'))
                    danger_threshold = panel.danger_threshold if panel.danger_threshold is not None else float(get_setting('default_danger_threshold', '80'))

                    # حساب النسب المئوية
                    current_percentage = None
                    if current is not None:
                        current_percentage = (current / panel.breaker_capacity) * 100

                    load_percentage = None
                    if load is not None:
                        load_percentage = (load / panel.breaker_capacity) * 100

                    # تحديد النسبة المستخدمة للتنبيه
                    percentage_to_use = None
                    value_to_use = None
                    value_type = None

                    if current_percentage is not None and load_percentage is not None:
                        if current_percentage >= load_percentage:
                            percentage_to_use = current_percentage
                            value_to_use = current
                            value_type = "التيار"
                        else:
                            percentage_to_use = load_percentage
                            value_to_use = load
                            value_type = "الحمل"
                    elif current_percentage is not None:
                        percentage_to_use = current_percentage
                        value_to_use = current
                        value_type = "التيار"
                    elif load_percentage is not None:
                        percentage_to_use = load_percentage
                        value_to_use = load
                        value_type = "الحمل"

                    # تحديد نوع التنبيه (متوسط أو قيمة قصوى)
                    alert_method = get_setting('alert_method', 'peak')
                    calculation_period = int(get_setting('alert_calculation_period', '60'))

                    if alert_method == 'average':
                        message = f'تجاوز متوسط {value_type} خلال {calculation_period} دقيقة النسبة المسموحة: {percentage_to_use:.1f}% من سعة القاطع ({value_to_use:.1f} أمبير)'
                    else:
                        message = f'تجاوز {value_type} النسبة المسموحة: {percentage_to_use:.1f}% من سعة القاطع ({value_to_use:.1f} أمبير)'

                    # إضافة معلومات الحدود إلى الرسالة
                    if current_status == 'danger':
                        message += f' (تجاوز حد الخطر: {danger_threshold}%)'
                    else:
                        message += f' (تجاوز حد التحذير: {warning_threshold}%)'

                    alert = Alert(
                        panel_id=panel.id,
                        reading_id=reading.id,
                        alert_type='current',
                        severity=current_status,
                        message=message,
                        timestamp=timestamp
                    )
                    db.session.add(alert)
                    alerts_generated += 1

                    # إرسال إشعار للمقاول المسؤول إذا كانت الحالة خطرة
                    if current_status == 'danger' and panel.responsible_contractor_id and get_setting('send_notifications', 'true') == 'true':
                        # الحصول على معلومات المقاول
                        contractor = Contractor.query.get(panel.responsible_contractor_id)
                        if contractor:
                            # إنشاء رابط الموقع إذا كان متاحًا
                            location_link = ""
                            if panel.location_url:
                                location_link = f"<a href='{panel.location_url}' target='_blank'>رابط الموقع على الخارطة</a>"
                            elif panel.x_coordinate and panel.y_coordinate:
                                maps_url = create_google_maps_url(panel.x_coordinate, panel.y_coordinate)
                                if maps_url:
                                    location_link = f"<a href='{maps_url}' target='_blank'>رابط الموقع على الخارطة</a>"

                            # إنشاء رسالة التنبيه
                            alert_message = f"""
                            <h3>تنبيه: تجاوز الحد الخطر للتيار</h3>
                            <p><strong>اللوحة:</strong> {panel.mdb}</p>
                            <p><strong>رقم تاج ماكسيمو:</strong> {panel.maximo_tag}</p>
                            <p><strong>المنطقة:</strong> {panel.area_name}</p>
                            <p><strong>وقت التنبيه:</strong> {timestamp.strftime('%Y-%m-%d %H:%M:%S')}</p>
                            <p><strong>نسبة التحميل:</strong> {percentage_to_use:.1f}%</p>
                            <p><strong>نوع التنبيه:</strong> تجاوز الحد الخطر ({danger_threshold}%)</p>
                            {location_link}
                            <p>يرجى التحقق من اللوحة وتخفيف الحمل لتجنب انقطاع التيار.</p>
                            """

                            # إنشاء نص عادي للرسائل النصية
                            sms_message = f"تنبيه: تجاوز الحد الخطر للتيار - اللوحة: {panel.mdb} - المنطقة: {panel.area_name} - نسبة التحميل: {percentage_to_use:.1f}% - يرجى التحقق من اللوحة"

                            # إرسال الإشعار للمقاول
                            subject = f"تنبيه: تجاوز الحد الخطر للتيار في اللوحة {panel.mdb}"
                            notify_contractor(contractor, subject, alert_message, sms_message)

                            print(f"تم إرسال تنبيه تجاوز الحد الخطر للمقاول {contractor.name}")

                if voltage_status == 'danger' and not is_tripped:
                    # استخدام قيم الحدود من اللوحة أو القيم الافتراضية من الإعدادات
                    min_voltage = panel.min_voltage if panel.min_voltage is not None else float(get_setting('default_min_voltage', '210'))
                    max_voltage = panel.max_voltage if panel.max_voltage is not None else float(get_setting('default_max_voltage', '250'))

                    if voltage > max_voltage:
                        message = f'تجاوز الجهد الحد الأقصى: {voltage:.1f} فولت (الحد الأقصى: {max_voltage} فولت)'
                    else:
                        message = f'انخفاض الجهد عن الحد الأدنى: {voltage:.1f} فولت (الحد الأدنى: {min_voltage} فولت)'

                    alert = Alert(
                        panel_id=panel.id,
                        reading_id=reading.id,
                        alert_type='voltage',
                        severity='danger',
                        message=message,
                        timestamp=timestamp
                    )
                    db.session.add(alert)
                    alerts_generated += 1

                    # إرسال إشعار للمقاول المسؤول
                    if panel.responsible_contractor_id and get_setting('send_notifications', 'true') == 'true':
                        # الحصول على معلومات المقاول
                        contractor = Contractor.query.get(panel.responsible_contractor_id)
                        if contractor:
                            # إنشاء رابط الموقع إذا كان متاحًا
                            location_link = ""
                            if panel.location_url:
                                location_link = f"<a href='{panel.location_url}' target='_blank'>رابط الموقع على الخارطة</a>"
                            elif panel.x_coordinate and panel.y_coordinate:
                                maps_url = create_google_maps_url(panel.x_coordinate, panel.y_coordinate)
                                if maps_url:
                                    location_link = f"<a href='{maps_url}' target='_blank'>رابط الموقع على الخارطة</a>"

                            # تحديد نوع مشكلة الجهد
                            voltage_issue_type = "تجاوز الحد الأقصى" if voltage > max_voltage else "انخفاض عن الحد الأدنى"
                            voltage_value = f"{voltage:.1f} فولت"
                            voltage_limit = f"{max_voltage} فولت" if voltage > max_voltage else f"{min_voltage} فولت"

                            # إنشاء رسالة التنبيه
                            alert_message = f"""
                            <h3>تنبيه: مشكلة في الجهد الكهربائي</h3>
                            <p><strong>اللوحة:</strong> {panel.mdb}</p>
                            <p><strong>رقم تاج ماكسيمو:</strong> {panel.maximo_tag}</p>
                            <p><strong>المنطقة:</strong> {panel.area_name}</p>
                            <p><strong>وقت التنبيه:</strong> {timestamp.strftime('%Y-%m-%d %H:%M:%S')}</p>
                            <p><strong>نوع المشكلة:</strong> {voltage_issue_type}</p>
                            <p><strong>قيمة الجهد:</strong> {voltage_value}</p>
                            <p><strong>الحد المسموح:</strong> {voltage_limit}</p>
                            {location_link}
                            <p>يرجى التحقق من اللوحة ومصدر التغذية الكهربائية لتجنب تلف الأجهزة.</p>
                            """

                            # إنشاء نص عادي للرسائل النصية
                            sms_message = f"تنبيه: مشكلة في الجهد الكهربائي - اللوحة: {panel.mdb} - المنطقة: {panel.area_name} - نوع المشكلة: {voltage_issue_type} - قيمة الجهد: {voltage_value}"

                            # إرسال الإشعار للمقاول
                            subject = f"تنبيه: مشكلة في الجهد الكهربائي في اللوحة {panel.mdb}"
                            notify_contractor(contractor, subject, alert_message, sms_message)

                            print(f"تم إرسال تنبيه مشكلة جهد للمقاول {contractor.name}")

            db.session.commit()

            flash(f'تم استيراد {readings_added} قراءة بنجاح وإنشاء {alerts_generated} تنبيه', 'success')

        except Exception as e:
            flash(f'حدث خطأ أثناء استيراد البيانات: {str(e)}', 'danger')

        # حذف الملف بعد الاستيراد
        os.remove(file_path)

    else:
        flash('نوع الملف غير مسموح به. يرجى استخدام ملفات Excel (.xlsx, .xls) أو CSV', 'danger')

    return redirect(url_for('electrical_dashboard_view'))

# استيراد البيانات من ملف Excel
@app.route('/import', methods=['POST'])
@login_required
def import_data():
    # التحقق من أن المستخدم ليس مقاول
    if current_user.role == 'contractor':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    if 'file' not in request.files:
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('index'))

    file = request.files['file']

    if file.filename == '':
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('index'))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        try:
            # قراءة ملف Excel
            try:
                df = pd.read_excel(file_path, header=0, engine='openpyxl')  # تحديد الصف الأول كعناوين
            except Exception as e:
                # محاولة استخدام محرك آخر إذا فشل المحرك الأول
                try:
                    df = pd.read_excel(file_path, header=0, engine='xlrd')
                except Exception as e2:
                    raise Exception(f"فشل قراءة الملف: {str(e)} | {str(e2)}")

            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ['MDB', 'رقم تاج ماكسيمو', 'X', 'Y', 'الملاحظات', 'المرحلة',
                               'سنة التنفيذ', 'كود المنطقة - تشغيل', 'نوع لوحة التوزيع', 'اسم المنطقة بالتنفيذ']

            # التحقق من وجود عمود سعة القاطع (اختياري)
            has_breaker_capacity = 'سعة القاطع' in df.columns

            for col in required_columns:
                if col not in df.columns:
                    flash(f'العمود {col} غير موجود في الملف', 'danger')
                    return redirect(url_for('index'))

            # حذف البيانات الحالية (اختياري)
            MDBPanel.query.delete()

            # تجاهل الصف الأول إذا كان يحتوي على عناوين الأعمدة مرة أخرى
            if df.iloc[0]['X'] == 'X' and df.iloc[0]['Y'] == 'Y':
                df = df.iloc[1:]

            # إعادة تعيين الفهرس
            df = df.reset_index(drop=True)

            # إضافة البيانات الجديدة
            for _, row in df.iterrows():
                try:
                    # محاولة تحويل الإحداثيات إلى أرقام
                    x_coord = None
                    if pd.notna(row['X']):
                        try:
                            x_coord = float(row['X'])
                        except (ValueError, TypeError):
                            x_coord = None

                    y_coord = None
                    if pd.notna(row['Y']):
                        try:
                            y_coord = float(row['Y'])
                        except (ValueError, TypeError):
                            y_coord = None

                    # محاولة تحويل سنة التنفيذ إلى رقم
                    impl_year = None
                    if pd.notna(row['سنة التنفيذ']):
                        try:
                            impl_year = int(row['سنة التنفيذ'])
                        except (ValueError, TypeError):
                            impl_year = None

                    # إعداد بيانات اللوحة
                    panel_data = {
                        'mdb': str(row['MDB']) if pd.notna(row['MDB']) else '',
                        'maximo_tag': str(row['رقم تاج ماكسيمو']) if pd.notna(row['رقم تاج ماكسيمو']) else '',
                        'x_coordinate': x_coord,
                        'y_coordinate': y_coord,
                        'notes': str(row['الملاحظات']) if pd.notna(row['الملاحظات']) else '',
                        'phase': str(row['المرحلة']) if pd.notna(row['المرحلة']) else '',
                        'implementation_year': impl_year,
                        'area_code': str(row['كود المنطقة - تشغيل']) if pd.notna(row['كود المنطقة - تشغيل']) else '',
                        'panel_type': str(row['نوع لوحة التوزيع']) if pd.notna(row['نوع لوحة التوزيع']) else '',
                        'area_name': str(row['اسم المنطقة بالتنفيذ']) if pd.notna(row['اسم المنطقة بالتنفيذ']) else '',
                        'status': 'عامل'  # القيمة الافتراضية
                    }

                    # إضافة سعة القاطع إذا كانت موجودة في الملف
                    if has_breaker_capacity and pd.notna(row['سعة القاطع']):
                        try:
                            breaker_capacity = float(row['سعة القاطع'])
                            panel_data['breaker_capacity'] = breaker_capacity
                        except (ValueError, TypeError):
                            # تجاهل القيمة إذا لم تكن رقمية
                            pass

                    # إنشاء اللوحة
                    panel = MDBPanel(**panel_data)
                    db.session.add(panel)
                except Exception as e:
                    flash(f'حدث خطأ في الصف {_ + 1}: {str(e)}', 'warning')
                    continue

            db.session.commit()
            flash('تم استيراد البيانات بنجاح', 'success')

        except Exception as e:
            flash(f'حدث خطأ أثناء استيراد البيانات: {str(e)}', 'danger')

        # حذف الملف بعد الاستيراد
        os.remove(file_path)

    else:
        flash('نوع الملف غير مسموح به. يرجى استخدام ملفات Excel (.xlsx, .xls)', 'danger')

    return redirect(url_for('index'))

# تصفية البيانات
@app.route('/filter', methods=['POST'])
def filter_data():
    area = request.form.get('area')
    panel_type = request.form.get('panel_type')
    year = request.form.get('year')

    query = MDBPanel.query

    if area and area != 'all':
        query = query.filter(MDBPanel.area_name == area)

    if panel_type and panel_type != 'all':
        query = query.filter(MDBPanel.panel_type == panel_type)

    if year and year != 'all':
        query = query.filter(MDBPanel.implementation_year == int(year))

    panels = query.all()
    total_filtered = len(panels)

    # إحصائيات للوحة المعلومات المفلترة
    panel_types = {}
    years = {}
    areas = {}

    for panel in panels:
        # إحصائيات أنواع اللوحات
        if panel.panel_type in panel_types:
            panel_types[panel.panel_type] += 1
        else:
            panel_types[panel.panel_type] = 1

        # إحصائيات السنوات
        if panel.implementation_year in years:
            years[panel.implementation_year] += 1
        else:
            years[panel.implementation_year] = 1

        # إحصائيات المناطق
        if panel.area_name in areas:
            areas[panel.area_name] += 1
        else:
            areas[panel.area_name] = 1

    # تحويل القواميس إلى قوائم من الأزواج
    panel_types_list = [(k, v) for k, v in panel_types.items()]
    years_list = [(k, v) for k, v in years.items()]
    areas_list = [(k, v) for k, v in areas.items()]

    return jsonify({
        'panels': [
            {
                'id': panel.id,
                'mdb': panel.mdb,
                'maximo_tag': str(panel.maximo_tag).zfill(7),
                'x_coordinate': panel.x_coordinate,
                'y_coordinate': panel.y_coordinate,
                'notes': panel.notes,
                'phase': panel.phase,
                'implementation_year': panel.implementation_year,
                'area_code': panel.area_code,
                'panel_type': panel.panel_type,
                'area_name': panel.area_name
            } for panel in panels
        ],
        'total_filtered': total_filtered,
        'panel_types': panel_types_list,
        'years': years_list,
        'areas': areas_list
    })

# تصدير البيانات إلى Excel
@app.route('/export/excel', methods=['POST'])
@app.route('/export-excel', methods=['POST'])
@login_required
def export_excel():
    # التحقق من أن المستخدم ليس مقاول أو أنه مقاول ولكن يطلب تصدير البيانات الكهربائية فقط
    if current_user.role == 'contractor' and request.form.get('export_type') != 'electrical':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # الحصول على معايير التصفية
    area = request.form.get('area')
    panel_type = request.form.get('panel_type')
    year = request.form.get('year')
    export_type = request.form.get('export_type', 'general')  # نوع التصدير (عام أو كهربائي)

    query = MDBPanel.query

    # إذا كان المستخدم مقاول، قم بتصفية اللوحات حسب مناطق المسؤولية
    if current_user.is_authenticated and current_user.role == 'contractor':
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # الحصول على مناطق المسؤولية للمقاول
        contractor_areas = []
        if contractor.area_responsibility:
            try:
                # محاولة تحليل البيانات كـ JSON
                contractor_areas = json.loads(contractor.area_responsibility)
            except json.JSONDecodeError:
                # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                contractor_areas = [area.strip() for area in contractor.area_responsibility.split(',')]

        # إذا كان المستخدم مقاول مدير، يمكنه رؤية جميع اللوحات المرتبطة بالفرق التابعة له
        if current_user.is_manager:
            # الحصول على جميع الفرق التابعة للمقاول
            teams = ContractorTeam.query.filter_by(contractor_id=contractor.id).all()

            # جمع جميع المناطق المسؤول عنها الفرق
            team_areas = []
            for team in teams:
                if team.area_responsibility:
                    try:
                        # محاولة تحليل البيانات كـ JSON
                        team_areas_json = json.loads(team.area_responsibility)
                        if isinstance(team_areas_json, list):
                            team_areas.extend(team_areas_json)
                        else:
                            team_areas.append(team_areas_json)
                    except json.JSONDecodeError:
                        # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                        team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

            # إضافة مناطق الفرق إلى مناطق المقاول
            contractor_areas.extend(team_areas)

            # إزالة التكرار
            contractor_areas = list(set(contractor_areas))

        # تصفية اللوحات حسب المقاول المسؤول أو المنطقة
        if contractor_areas:
            query = query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == contractor.id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            )
        else:
            # إذا لم تكن هناك مناطق محددة، استخدم فقط المقاول المسؤول
            query = query.filter(MDBPanel.responsible_contractor_id == contractor.id)

    if area and area != 'all':
        query = query.filter(MDBPanel.area_name == area)

    if panel_type and panel_type != 'all':
        query = query.filter(MDBPanel.panel_type == panel_type)

    if year and year != 'all':
        query = query.filter(MDBPanel.implementation_year == int(year))

    panels = query.all()

    # إنشاء DataFrame
    data = []

    if export_type == 'electrical':
        # تصدير البيانات الكهربائية مع آخر القراءات
        for panel in panels:
            panel_data = {
                'MDB': panel.mdb,
                'رقم تاج ماكسيمو': panel.maximo_tag,
                'المنطقة': panel.area_name,
                'نوع اللوحة': panel.panel_type,
                'سعة القاطع (أمبير)': panel.breaker_capacity
            }

            # إضافة بيانات آخر قراءة إذا كانت متوفرة
            reading = ElectricalReading.query.filter_by(panel_id=panel.id).order_by(ElectricalReading.timestamp.desc()).first()
            if reading:
                panel_data.update({
                    'آخر قراءة': reading.timestamp.strftime('%Y-%m-%d %H:%M'),
                    'التيار (أمبير)': reading.current,
                    'الجهد (فولت)': reading.voltage,
                    'القدرة (واط)': reading.power,
                    'الطاقة (كيلوواط ساعة)': reading.energy,
                    'معامل القدرة': reading.power_factor,
                    'التردد (هرتز)': reading.frequency
                })

                # حساب نسبة الحمل إذا كانت سعة القاطع متوفرة
                if panel.breaker_capacity and reading.current:
                    panel_data['نسبة الحمل (%)'] = (reading.current / panel.breaker_capacity) * 100
            else:
                panel_data.update({
                    'آخر قراءة': 'غير متوفر',
                    'التيار (أمبير)': None,
                    'الجهد (فولت)': None,
                    'القدرة (واط)': None,
                    'الطاقة (كيلوواط ساعة)': None,
                    'معامل القدرة': None,
                    'التردد (هرتز)': None,
                    'نسبة الحمل (%)': None
                })

            data.append(panel_data)
    else:
        # تصدير البيانات العامة
        for panel in panels:
            data.append({
                'MDB': panel.mdb,
                'رقم تاج ماكسيمو': panel.maximo_tag,
                'X': panel.x_coordinate,
                'Y': panel.y_coordinate,
                'الملاحظات': panel.notes,
                'المرحلة': panel.phase,
                'سنة التنفيذ': panel.implementation_year,
                'كود المنطقة - تشغيل': panel.area_code,
                'نوع لوحة التوزيع': panel.panel_type,
                'اسم المنطقة بالتنفيذ': panel.area_name
            })

    df = pd.DataFrame(data)

    # إنشاء ملف Excel في الذاكرة
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        sheet_name = 'Electrical Data' if export_type == 'electrical' else 'MDB Panels'
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        # تنسيق ورقة العمل
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # تنسيق العناوين
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#94BCCB',
            'border': 1
        })

        # تطبيق التنسيق على الصف الأول
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            worksheet.set_column(col_num, col_num, 15)

    output.seek(0)

    # إنشاء اسم الملف مع التاريخ والوقت
    now = datetime.datetime.now()
    prefix = "Electrical" if export_type == 'electrical' else "MDB_Panels"
    filename = f"{prefix}_Export_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)

# تصدير التنبيهات
@app.route('/export-alerts', methods=['POST'])
@login_required
def export_alerts():
    # الحصول على معايير التصفية
    severity = request.form.get('severity', 'all')
    alert_type = request.form.get('type', 'all')
    is_read = request.form.get('is_read', 'all')
    export_format = request.form.get('export_format', 'excel')
    date_range = request.form.get('date_range', 'all')

    # الحصول على الأعمدة المطلوبة
    include_panel = 'include_panel' in request.form
    include_type = 'include_type' in request.form
    include_message = 'include_message' in request.form
    include_severity = 'include_severity' in request.form
    include_timestamp = 'include_timestamp' in request.form
    include_status = 'include_status' in request.form

    # بناء استعلام التنبيهات
    query = Alert.query

    # تطبيق معايير التصفية
    if severity != 'all':
        query = query.filter(Alert.severity == severity)

    if alert_type != 'all':
        query = query.filter(Alert.alert_type == alert_type)

    if is_read != 'all':
        is_read_bool = (is_read == 'read')
        query = query.filter(Alert.is_read == is_read_bool)

    # تطبيق نطاق التاريخ
    now = datetime.datetime.now()
    if date_range == 'today':
        start_date = datetime.datetime.combine(now.date(), datetime.time.min)
        query = query.filter(Alert.timestamp >= start_date)
    elif date_range == 'yesterday':
        yesterday = now - datetime.timedelta(days=1)
        start_date = datetime.datetime.combine(yesterday.date(), datetime.time.min)
        end_date = datetime.datetime.combine(yesterday.date(), datetime.time.max)
        query = query.filter(Alert.timestamp >= start_date, Alert.timestamp <= end_date)
    elif date_range == 'last_week':
        start_date = now - datetime.timedelta(days=7)
        query = query.filter(Alert.timestamp >= start_date)
    elif date_range == 'last_month':
        start_date = now - datetime.timedelta(days=30)
        query = query.filter(Alert.timestamp >= start_date)
    elif date_range == 'custom':
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')

        if start_date_str:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = datetime.datetime.combine(start_date, datetime.time.min)
            query = query.filter(Alert.timestamp >= start_date)

        if end_date_str:
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = datetime.datetime.combine(end_date, datetime.time.max)
            query = query.filter(Alert.timestamp <= end_date)

    # إذا كان المستخدم مقاول، قم بتصفية التنبيهات حسب مناطق المسؤولية
    if current_user.is_authenticated and current_user.role == 'contractor':
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # الحصول على مناطق المسؤولية للمقاول
        contractor_areas = get_contractor_areas(contractor, current_user)

        # الحصول على اللوحات في مناطق المسؤولية
        panels_in_areas = MDBPanel.query.filter(
            db.or_(
                MDBPanel.responsible_contractor_id == contractor.id,
                MDBPanel.area_name.in_(contractor_areas)
            )
        ).all()

        panel_ids = [panel.id for panel in panels_in_areas]

        # تصفية التنبيهات حسب اللوحات
        query = query.filter(Alert.panel_id.in_(panel_ids))

    # الحصول على التنبيهات
    alerts = query.order_by(Alert.timestamp.desc()).all()

    # إنشاء البيانات للتصدير
    data = []

    # إضافة عناوين الأعمدة
    headers = []
    if include_panel:
        headers.append('اللوحة')
    if include_type:
        headers.append('نوع التنبيه')
    if include_message:
        headers.append('الرسالة')
    if include_severity:
        headers.append('الخطورة')
    if include_timestamp:
        headers.append('الوقت')
    if include_status:
        headers.append('حالة القراءة')
        headers.append('حالة الحل')

    data.append(headers)

    # إضافة بيانات التنبيهات
    for alert in alerts:
        row = []

        if include_panel:
            panel = MDBPanel.query.get(alert.panel_id)
            row.append(panel.mdb if panel else '')

        if include_type:
            alert_type_text = ''
            if alert.alert_type == 'current':
                alert_type_text = 'التيار'
            elif alert.alert_type == 'voltage':
                alert_type_text = 'الجهد'
            elif alert.alert_type == 'power':
                alert_type_text = 'القدرة'
            elif alert.alert_type == 'trip':
                alert_type_text = 'فصل'
            row.append(alert_type_text)

        if include_message:
            row.append(alert.message)

        if include_severity:
            severity_text = 'خطر' if alert.severity == 'danger' else 'تحذير'
            row.append(severity_text)

        if include_timestamp:
            row.append(alert.timestamp.strftime('%Y-%m-%d %H:%M'))

        if include_status:
            row.append('مقروء' if alert.is_read else 'غير مقروء')
            row.append('تم الحل' if alert.is_resolved else 'لم يتم الحل')

        data.append(row)

    # تصدير البيانات حسب الصيغة المطلوبة
    if export_format == 'excel':
        # إنشاء ملف Excel في الذاكرة
        output = io.BytesIO()

        # إنشاء DataFrame من البيانات
        df = pd.DataFrame(data[1:], columns=data[0])

        # تصدير إلى Excel
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Alerts')

            # تنسيق ورقة العمل
            workbook = writer.book
            worksheet = writer.sheets['Alerts']

            # تنسيق العناوين
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#94BCCB',
                'border': 1
            })

            # تطبيق التنسيق على الصف الأول
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                worksheet.set_column(col_num, col_num, 20)

        output.seek(0)

        # إنشاء اسم الملف مع التاريخ والوقت
        filename = f"Alerts_Report_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=filename)
    else:
        # إنشاء ملف PDF في الذاكرة
        buffer = io.BytesIO()

        # إنشاء مستند PDF مع دعم اللغة العربية
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=60,  # زيادة الهامش العلوي لإضافة الشعار
            bottomMargin=30,
            encoding='UTF-8'
        )

        # قائمة العناصر التي سيتم إضافتها إلى المستند
        elements = []

        # إضافة الشعار
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'img', 'logo.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path)
            logo.drawHeight = 40
            logo.drawWidth = 120
            elements.append(logo)
            elements.append(Spacer(1, 10))

        # إضافة العنوان والتاريخ
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontName='NotoSansArabic-Bold',
            alignment=1,  # وسط
            textColor=colors.HexColor('#B89966'),
            spaceAfter=12
        )

        elements.append(Paragraph("تقرير التنبيهات", title_style))

        # إضافة التاريخ والوقت
        date_style = ParagraphStyle(
            'Date',
            parent=styles['Normal'],
            fontName='NotoSansArabic',
            alignment=1,  # وسط
            textColor=colors.HexColor('#94BCCB'),
            fontSize=10,
            spaceAfter=12
        )
        elements.append(Paragraph(f"تاريخ التصدير: {now.strftime('%Y-%m-%d %H:%M:%S')}", date_style))

        # إضافة معلومات التصفية
        filter_info = "معايير التصفية: "
        if severity != 'all':
            filter_info += f"الخطورة: {severity}, "
        if alert_type != 'all':
            filter_info += f"نوع التنبيه: {alert_type}, "
        if is_read != 'all':
            filter_info += f"حالة القراءة: {is_read}, "
        if date_range != 'all':
            filter_info += f"نطاق التاريخ: {date_range}, "

        if filter_info == "معايير التصفية: ":
            filter_info += "جميع البيانات"
        else:
            filter_info = filter_info[:-2]  # إزالة الفاصلة والمسافة الأخيرة

        elements.append(Paragraph(filter_info, date_style))
        elements.append(Spacer(1, 12))

        # إنشاء جدول البيانات
        table = Table(data, repeatRows=1)

        # تنسيق الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#94BCCB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'NotoSansArabic-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'NotoSansArabic'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])

        # إضافة تنسيق متناوب للصفوف
        for i in range(1, len(data), 2):
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F5F5'))

        table.setStyle(table_style)

        # إضافة الجدول إلى المستند
        elements.append(table)

        # إضافة معلومات إجمالية
        elements.append(Spacer(1, 12))
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontName='NotoSansArabic',
            alignment=0,  # يمين
            textColor=colors.HexColor('#B89966'),
            fontSize=10,
            spaceAfter=6
        )
        elements.append(Paragraph(f"إجمالي عدد التنبيهات: {len(alerts)}", summary_style))

        # بناء المستند
        doc.build(elements)

        buffer.seek(0)

        # إنشاء اسم الملف مع التاريخ والوقت
        filename = f"Alerts_Report_{now.strftime('%Y%m%d_%H%M%S')}.pdf"

        return send_file(buffer,
                         mimetype='application/pdf',
                         as_attachment=True,
                         download_name=filename)

# تصدير البيانات إلى PDF
@app.route('/export/pdf', methods=['POST'])
@app.route('/export-pdf', methods=['POST'])
@login_required
def export_pdf():
    # التحقق من أن المستخدم ليس مقاول أو أنه مقاول ولكن يطلب تصدير البيانات الكهربائية فقط
    if current_user.role == 'contractor' and request.form.get('export_type') != 'electrical':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('contractor_dashboard'))

    # الحصول على معايير التصفية
    area = request.form.get('area')
    panel_type = request.form.get('panel_type')
    year = request.form.get('year')
    export_type = request.form.get('export_type', 'general')  # نوع التصدير (عام أو كهربائي)

    query = MDBPanel.query

    # إذا كان المستخدم مقاول، قم بتصفية اللوحات حسب مناطق المسؤولية
    if current_user.is_authenticated and current_user.role == 'contractor':
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # الحصول على مناطق المسؤولية للمقاول
        contractor_areas = []
        if contractor.area_responsibility:
            try:
                # محاولة تحليل البيانات كـ JSON
                contractor_areas = json.loads(contractor.area_responsibility)
            except json.JSONDecodeError:
                # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                contractor_areas = [area.strip() for area in contractor.area_responsibility.split(',')]

        # إذا كان المستخدم مقاول مدير، يمكنه رؤية جميع اللوحات المرتبطة بالفرق التابعة له
        if current_user.is_manager:
            # الحصول على جميع الفرق التابعة للمقاول
            teams = ContractorTeam.query.filter_by(contractor_id=contractor.id).all()

            # جمع جميع المناطق المسؤول عنها الفرق
            team_areas = []
            for team in teams:
                if team.area_responsibility:
                    try:
                        # محاولة تحليل البيانات كـ JSON
                        team_areas_json = json.loads(team.area_responsibility)
                        if isinstance(team_areas_json, list):
                            team_areas.extend(team_areas_json)
                        else:
                            team_areas.append(team_areas_json)
                    except json.JSONDecodeError:
                        # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                        team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

            # إضافة مناطق الفرق إلى مناطق المقاول
            contractor_areas.extend(team_areas)

            # إزالة التكرار
            contractor_areas = list(set(contractor_areas))

        # تصفية اللوحات حسب المقاول المسؤول أو المنطقة
        if contractor_areas:
            query = query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == contractor.id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            )
        else:
            # إذا لم تكن هناك مناطق محددة، استخدم فقط المقاول المسؤول
            query = query.filter(MDBPanel.responsible_contractor_id == contractor.id)

    if area and area != 'all':
        query = query.filter(MDBPanel.area_name == area)

    if panel_type and panel_type != 'all':
        query = query.filter(MDBPanel.panel_type == panel_type)

    if year and year != 'all':
        query = query.filter(MDBPanel.implementation_year == int(year))

    panels = query.all()

    # إنشاء ملف PDF في الذاكرة
    buffer = io.BytesIO()

    # إنشاء مستند PDF مع دعم اللغة العربية
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=60,  # زيادة الهامش العلوي لإضافة الشعار
        bottomMargin=30,
        encoding='UTF-8'
    )

    # قائمة العناصر التي سيتم إضافتها إلى المستند
    elements = []

    # إضافة الشعار
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'img', 'logo.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path)
        logo.drawHeight = 40
        logo.drawWidth = 120
        elements.append(logo)
        elements.append(Spacer(1, 10))

    # إضافة العنوان والتاريخ
    styles = getSampleStyleSheet()

    # استخدام الخطوط العربية المسجلة مسبقاً
    arabic_font_bold_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'fonts', 'NotoSansArabic-Bold.ttf')

    # التأكد من وجود مجلد الخطوط
    fonts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'fonts')
    if not os.path.exists(fonts_dir):
        os.makedirs(fonts_dir)

    # تنزيل الخط العربي الغامق إذا لم يكن موجوداً
    if not os.path.exists(arabic_font_bold_path):
        import urllib.request
        try:
            urllib.request.urlretrieve('https://github.com/googlefonts/noto-fonts/raw/main/hinted/ttf/NotoSansArabic/NotoSansArabic-Bold.ttf', arabic_font_bold_path)
            print("تم تنزيل الخط العربي الغامق")
        except Exception as e:
            print(f"فشل تنزيل الخط العربي الغامق: {str(e)}")

    # تسجيل الخط العربي الغامق
    try:
        pdfmetrics.registerFont(TTFont('NotoSansArabic-Bold', arabic_font_bold_path))
    except Exception as e:
        print(f"فشل تسجيل الخط العربي الغامق: {str(e)}")

    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontName='NotoSansArabic-Bold',
        alignment=1,  # وسط
        textColor=colors.HexColor('#B89966'),
        spaceAfter=12
    )

    # إضافة العنوان حسب نوع التصدير
    if export_type == 'electrical':
        elements.append(Paragraph("تقرير البيانات الكهربائية للوحات", title_style))
    else:
        elements.append(Paragraph("تقرير لوحات التوزيع الكهربائية", title_style))

    # إضافة التاريخ والوقت
    now = datetime.datetime.now()
    date_style = ParagraphStyle(
        'Date',
        parent=styles['Normal'],
        fontName='NotoSansArabic',
        alignment=1,  # وسط
        textColor=colors.HexColor('#94BCCB'),
        fontSize=10,
        spaceAfter=12
    )
    elements.append(Paragraph(f"تاريخ التصدير: {now.strftime('%Y-%m-%d %H:%M:%S')}", date_style))

    # إضافة معلومات التصفية
    filter_info = "معايير التصفية: "
    if area and area != 'all':
        filter_info += f"المنطقة: {area}, "
    if panel_type and panel_type != 'all':
        filter_info += f"نوع اللوحة: {panel_type}, "
    if year and year != 'all':
        filter_info += f"سنة التنفيذ: {year}, "

    if filter_info == "معايير التصفية: ":
        filter_info += "جميع البيانات"
    else:
        filter_info = filter_info[:-2]  # إزالة الفاصلة والمسافة الأخيرة

    elements.append(Paragraph(filter_info, date_style))
    elements.append(Spacer(1, 12))

    # إنشاء جدول البيانات حسب نوع التصدير
    if export_type == 'electrical':
        # تصدير البيانات الكهربائية
        data = [['MDB', 'رقم تاج ماكسيمو', 'المنطقة', 'نوع اللوحة', 'سعة القاطع', 'آخر قراءة', 'التيار (أمبير)', 'الجهد (فولت)', 'القدرة (واط)', 'نسبة الحمل (%)']]
        col_widths = [80, 100, 100, 80, 60, 80, 70, 70, 70, 70]

        for panel in panels:
            row_data = [
                panel.mdb,
                panel.maximo_tag,
                panel.area_name,
                panel.panel_type,
                str(panel.breaker_capacity) if panel.breaker_capacity else '-'
            ]

            # إضافة بيانات آخر قراءة إذا كانت متوفرة
            reading = ElectricalReading.query.filter_by(panel_id=panel.id).order_by(ElectricalReading.timestamp.desc()).first()
            if reading:
                row_data.extend([
                    reading.timestamp.strftime('%Y-%m-%d %H:%M'),
                    str(round(reading.current, 1)) if reading.current is not None else '-',
                    str(round(reading.voltage, 1)) if reading.voltage is not None else '-',
                    str(round(reading.power, 1)) if reading.power is not None else '-'
                ])

                # حساب نسبة الحمل إذا كانت سعة القاطع متوفرة
                if panel.breaker_capacity and reading.current:
                    load_percentage = (reading.current / panel.breaker_capacity) * 100
                    row_data.append(str(round(load_percentage, 1)) + '%')
                else:
                    row_data.append('-')
            else:
                row_data.extend(['-', '-', '-', '-', '-'])

            data.append(row_data)
    else:
        # تصدير البيانات العامة
        data = [['MDB', 'رقم تاج ماكسيمو', 'X', 'Y', 'الملاحظات', 'المرحلة', 'سنة التنفيذ', 'كود المنطقة', 'نوع اللوحة', 'اسم المنطقة']]
        col_widths = [80, 100, 50, 50, 150, 80, 80, 80, 80, 100]

        for panel in panels:
            data.append([
                panel.mdb,
                panel.maximo_tag,
                str(panel.x_coordinate) if panel.x_coordinate else '',
                str(panel.y_coordinate) if panel.y_coordinate else '',
                panel.notes,
                panel.phase,
                str(panel.implementation_year) if panel.implementation_year else '',
                panel.area_code,
                panel.panel_type,
                panel.area_name
            ])

    # إنشاء الجدول مع تحديد عرض الأعمدة
    table = Table(data, repeatRows=1, colWidths=col_widths)

    # تنسيق الجدول
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#94BCCB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'NotoSansArabic-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'NotoSansArabic'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])

    # إضافة تنسيق متناوب للصفوف
    for i in range(1, len(data), 2):
        table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F5F5'))

    table.setStyle(table_style)

    # إضافة الجدول إلى المستند
    elements.append(table)

    # إضافة معلومات إجمالية
    elements.append(Spacer(1, 12))
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontName='NotoSansArabic',
        alignment=0,  # يمين
        textColor=colors.HexColor('#B89966'),
        fontSize=10,
        spaceAfter=6
    )
    elements.append(Paragraph(f"إجمالي عدد اللوحات: {len(panels)}", summary_style))

    # بناء المستند
    doc.build(elements)

    buffer.seek(0)

    # إنشاء اسم الملف مع التاريخ والوقت
    prefix = "Electrical" if export_type == 'electrical' else "MDB_Panels"
    filename = f"{prefix}_Report_{now.strftime('%Y%m%d_%H%M%S')}.pdf"

    return send_file(buffer,
                     mimetype='application/pdf',
                     as_attachment=True,
                     download_name=filename)

# صفحة البحث العام
@app.route('/search')
def search():
    # الحصول على مصطلح البحث
    query = request.args.get('q', '')

    if not query:
        flash('يرجى إدخال مصطلح البحث', 'warning')
        return redirect(url_for('index'))

    # البحث في اللوحات
    panels_query = MDBPanel.query.filter(
        db.or_(
            MDBPanel.mdb.like(f'%{query}%'),
            MDBPanel.maximo_tag.like(f'%{query}%'),
            MDBPanel.area_name.like(f'%{query}%'),
            MDBPanel.notes.like(f'%{query}%')
        )
    )

    # إذا كان المستخدم مقاول، قم بتصفية اللوحات حسب مناطق المسؤولية
    if current_user.is_authenticated and current_user.role == 'contractor':
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # الحصول على مناطق المسؤولية للمقاول
        contractor_areas = []
        if contractor.area_responsibility:
            try:
                # محاولة تحليل البيانات كـ JSON
                contractor_areas = json.loads(contractor.area_responsibility)
            except json.JSONDecodeError:
                # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                contractor_areas = [area.strip() for area in contractor.area_responsibility.split(',')]

        # إذا كان المستخدم مقاول مدير، يمكنه رؤية جميع اللوحات المرتبطة بالفرق التابعة له
        if current_user.is_manager:
            # الحصول على جميع الفرق التابعة للمقاول
            teams = ContractorTeam.query.filter_by(contractor_id=contractor.id).all()

            # جمع جميع المناطق المسؤول عنها الفرق
            team_areas = []
            for team in teams:
                if team.area_responsibility:
                    try:
                        # محاولة تحليل البيانات كـ JSON
                        team_areas_json = json.loads(team.area_responsibility)
                        if isinstance(team_areas_json, list):
                            team_areas.extend(team_areas_json)
                        else:
                            team_areas.append(team_areas_json)
                    except json.JSONDecodeError:
                        # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                        team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

            # إضافة مناطق الفرق إلى مناطق المقاول
            contractor_areas.extend(team_areas)

            # إزالة التكرار
            contractor_areas = list(set(contractor_areas))

        # تصفية اللوحات حسب المقاول المسؤول أو المنطقة
        if contractor_areas:
            panels_query = panels_query.filter(
                db.or_(
                    MDBPanel.responsible_contractor_id == contractor.id,
                    MDBPanel.area_name.in_(contractor_areas)
                )
            )
        else:
            # إذا لم تكن هناك مناطق محددة، استخدم فقط المقاول المسؤول
            panels_query = panels_query.filter(MDBPanel.responsible_contractor_id == contractor.id)

    panels = panels_query.all()

    # البحث في البلاغات
    issues_query = Issue.query.join(MDBPanel).filter(
        db.or_(
            Issue.title.like(f'%{query}%'),
            Issue.description.like(f'%{query}%'),
            MDBPanel.mdb.like(f'%{query}%'),
            MDBPanel.maximo_tag.like(f'%{query}%')
        )
    )

    # إذا كان المستخدم مقاول، قم بتصفية البلاغات حسب المقاول
    if current_user.is_authenticated and current_user.role == 'contractor':
        contractor = Contractor.query.get_or_404(current_user.contractor_id)

        # إذا كان المستخدم مقاول مدير، يمكنه رؤية جميع البلاغات المرتبطة بالفرق التابعة له
        if current_user.is_manager:
            # الحصول على جميع الفرق التابعة للمقاول
            teams = ContractorTeam.query.filter_by(contractor_id=contractor.id).all()

            # جمع جميع المناطق المسؤول عنها الفرق
            team_areas = []
            for team in teams:
                if team.area_responsibility:
                    try:
                        # محاولة تحليل البيانات كـ JSON
                        team_areas_json = json.loads(team.area_responsibility)
                        if isinstance(team_areas_json, list):
                            team_areas.extend(team_areas_json)
                        else:
                            team_areas.append(team_areas_json)
                    except json.JSONDecodeError:
                        # إذا لم يكن JSON، نفترض أنه نص مفصول بفواصل
                        team_areas.extend([area.strip() for area in team.area_responsibility.split(',')])

            # الحصول على اللوحات في مناطق المسؤولية
            panels_in_areas = MDBPanel.query.filter(MDBPanel.area_name.in_(team_areas)).all()
            panel_ids = [panel.id for panel in panels_in_areas]

            # الحصول على البلاغات المرتبطة بالمقاول أو اللوحات في مناطق المسؤولية
            issues_query = issues_query.filter(
                db.or_(
                    Issue.contractor_id == contractor.id,
                    Issue.panel_id.in_(panel_ids)
                )
            )
        else:
            # المقاول العادي يرى فقط البلاغات المرتبطة به مباشرة
            issues_query = issues_query.filter_by(contractor_id=contractor.id)

    issues = issues_query.all()

    # الحصول على آخر القراءات للوحات
    latest_readings = {}
    for panel in panels:
        reading = ElectricalReading.query.filter_by(panel_id=panel.id).order_by(ElectricalReading.timestamp.desc()).first()
        if reading:
            latest_readings[panel.id] = reading

    # الحصول على إعدادات الألوان والحدود
    normal_color = get_setting('normal_color', '#28a745')
    warning_color = get_setting('warning_color', '#ffc107')
    danger_color = get_setting('danger_color', '#dc3545')
    trip_color = get_setting('trip_color', '#6c757d')
    warning_threshold = float(get_setting('default_warning_threshold', '70'))
    danger_threshold = float(get_setting('default_danger_threshold', '80'))

    return render_template('search_results.html',
                           query=query,
                           panels=panels,
                           issues=issues,
                           latest_readings=latest_readings,
                           normal_color=normal_color,
                           warning_color=warning_color,
                           danger_color=danger_color,
                           trip_color=trip_color,
                           warning_threshold=warning_threshold,
                           danger_threshold=danger_threshold,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# تشغيل التطبيق
if __name__ == '__main__':
    print("بدء تشغيل التطبيق...")

    # إنشاء قاعدة البيانات والجداول
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mdb_dashboard.db')
    
    # إذا كان الملف موجود لكن فاضي (0 بايت)، احذفه
    if os.path.exists(db_path) and os.path.getsize(db_path) == 0:
        print(f"⚠️ قاعدة البيانات فاضية - سيتم حذفها وإعادة إنشائها")
        os.remove(db_path)
    
    # إنشاء الجداول
    print("إنشاء الجداول المفقودة...")
    with app.app_context():
        db.create_all()  # ✅ إنشاء الجداول المفقودة فقط
        print("تم إنشاء جميع الجداول بنجاح!")
        
        # التحقق من نجاح الإنشاء
        import sqlite3
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = [t[0] for t in cursor.fetchall()]
        
        # إضافة حقل panorama_id إلى جدول mdb_panel (مهم جداً قبل إضافة البيانات)
        if 'mdb_panel' in tables:
            cursor.execute("PRAGMA table_info(mdb_panel)")
            columns = [column[1] for column in cursor.fetchall()]
            if 'panorama_id' not in columns:
                print("⚠️ إضافة حقل panorama_id إلى جدول mdb_panel...")
                cursor.execute("ALTER TABLE mdb_panel ADD COLUMN panorama_id INTEGER")
                conn.commit()
                print("✅ تم إضافة حقل panorama_id بنجاح!")
        
        conn.close()
        print(f"✅ تم إنشاء {len(tables)} جدول")
    
    # إضافة البيانات الافتراضية فقط إذا كان RECREATE_DB مفعل
    if RECREATE_DB:
        print("إنشاء البيانات الافتراضية...")
        with app.app_context():

            # إضافة بيانات افتراضية
            print("بدء إضافة البيانات الافتراضية...")

            # إضافة مستخدم افتراضي (مدير)
            print("إضافة مستخدم افتراضي...")
            existing_admin = User.query.filter_by(username="admin").first()
            if not existing_admin:
                admin_user = User(
                    username="admin",
                    name="مدير النظام",
                    email="admin@example.com",
                    role="admin",
                    is_active=True,
                    created_at=datetime.datetime.now()
                )
                admin_user.set_password("admin123")
                db.session.add(admin_user)
                db.session.commit()
                print("تم إضافة المستخدم الافتراضي بنجاح!")
            else:
                print("المستخدم الافتراضي موجود بالفعل - تم التخطي")

            # إضافة بيانات المقاولين
            print("إضافة بيانات المقاولين الافتراضية...")
            contractors = [
                Contractor(name="أزهر", contact_person="ابو القاسم", phone="0555555555", email="ahmed@example.com", area_responsibility="المنطقة الشرقية"),
                Contractor(name="كنترول تك", contact_person="وليد", phone="0566666666", email="khalid@example.com", area_responsibility="المنطقة الغربية"),
               ## Contractor(name="شركة التقنية المتطورة", contact_person="سعيد علي", phone="0577777777", email="saeed@example.com", area_responsibility="المنطقة الشمالية")
            ]
            db.session.add_all(contractors)
            db.session.commit()
            print(f"تم إضافة {len(contractors)} مقاولين بنجاح!")

            # إضافة مستخدمين للمقاولين
            print("إضافة مستخدمين للمقاولين...")
            contractor_usernames = ["contractor1", "contractor2", "contractor3"]
            contractor_users_to_add = []
            
            for i, username in enumerate(contractor_usernames, 1):
                if not User.query.filter_by(username=username).first():
                    user = User(
                        username=username,
                        name=f"مستخدم المقاول {i}",
                        email=f"{username}@example.com",
                        role="contractor",
                        contractor_id=i,
                        is_active=True,
                        created_at=datetime.datetime.now()
                    )
                    user.set_password("contractor123")
                    contractor_users_to_add.append(user)
            
            if contractor_users_to_add:
                db.session.add_all(contractor_users_to_add)
                db.session.commit()
                print(f"تم إضافة {len(contractor_users_to_add)} مستخدمين للمقاولين بنجاح!")
            else:
                print("مستخدمو المقاولين موجودون بالفعل - تم التخطي")

            # إضافة أعمدة ديناميكية افتراضية
            print("إضافة الأعمدة الديناميكية الافتراضية...")
            columns = [
                DynamicColumn(name="maintenance_company", display_name="شركة الصيانة", data_type="text", is_active=True),
                DynamicColumn(name="installation_date", display_name="تاريخ التركيب", data_type="date", is_active=True),
                DynamicColumn(name="power_capacity", display_name="السعة الكهربائية", data_type="number", is_active=True)
            ]
            db.session.add_all(columns)
            db.session.commit()
            print(f"تم إضافة {len(columns)} أعمدة ديناميكية بنجاح!")

            # إضافة بعض بيانات اللوحات للعرض
            print("إضافة بيانات اللوحات الافتراضية...")
            panels = [
                MDBPanel(
                    mdb="MDB-001",
                    maximo_tag="MAX-001",
                    x_coordinate=46.6753,
                    y_coordinate=24.7136,
                    notes="لوحة رئيسية للمنطقة الشرقية",
                    phase="المرحلة الأولى",
                    implementation_year=2020,
                    area_code="E-001",
                    panel_type="رئيسية",
                    area_name="المنطقة الشرقية",
                    status="عامل"
                ),
                MDBPanel(
                    mdb="MDB-002",
                    maximo_tag="MAX-002",
                    x_coordinate=46.7000,
                    y_coordinate=24.7200,
                    notes="لوحة فرعية للمنطقة الشرقية",
                    phase="المرحلة الأولى",
                    implementation_year=2020,
                    area_code="E-002",
                    panel_type="فرعية",
                    area_name="المنطقة الشرقية",
                    status="عامل"
                ),
                MDBPanel(
                    mdb="MDB-003",
                    maximo_tag="MAX-003",
                    x_coordinate=46.6900,
                    y_coordinate=24.7000,
                    notes="لوحة رئيسية للمنطقة الغربية",
                    phase="المرحلة الثانية",
                    implementation_year=2021,
                    area_code="W-001",
                    panel_type="رئيسية",
                    area_name="المنطقة الغربية",
                    status="تحت الصيانة"
                ),
                MDBPanel(
                    mdb="MDB-004",
                    maximo_tag="MAX-004",
                    x_coordinate=46.6800,
                    y_coordinate=24.6900,
                    notes="لوحة فرعية للمنطقة الغربية",
                    phase="المرحلة الثانية",
                    implementation_year=2021,
                    area_code="W-002",
                    panel_type="فرعية",
                    area_name="المنطقة الغربية",
                    status="معطل"
                ),
                MDBPanel(
                    mdb="MDB-005",
                    maximo_tag="MAX-005",
                    x_coordinate=46.7100,
                    y_coordinate=24.7300,
                    notes="لوحة رئيسية للمنطقة الشمالية",
                    phase="المرحلة الثالثة",
                    implementation_year=2022,
                    area_code="N-001",
                    panel_type="رئيسية",
                    area_name="المنطقة الشمالية",
                    status="عامل"
                )
            ]
            db.session.add_all(panels)
            db.session.commit()
            print(f"تم إضافة {len(panels)} لوحات بنجاح!")

            # إضافة مستخدم مسجل قراءات للاختبار
            print("إضافة مستخدم مسجل قراءات...")
            if not User.query.filter_by(username="recorder1").first():
                readings_recorder = User(
                    username="recorder1",
                    name="مسجل القراءات الأول",
                    email="recorder@example.com",
                    role="readings_recorder",
                    is_active=True,
                    created_at=datetime.datetime.now()
                )
                readings_recorder.set_password("123456")
                # تخصيص مناطق للمستخدم
                readings_recorder.set_assigned_areas(["عرفة", "منى"])
                db.session.add(readings_recorder)
                db.session.commit()
                print("تم إضافة مستخدم مسجل القراءات بنجاح!")
            else:
                print("مستخدم مسجل القراءات موجود بالفعل - تم التخطي")

            # إضافة بعض البلاغات للعرض
            print("إضافة بلاغات افتراضية...")
            # الحصول على مستخدم مسجل القراءات
            readings_recorder = User.query.filter_by(username="recorder1").first()
            
            issues = [
                Issue(
                    panel_id=4,  # اللوحة المعطلة
                    contractor_id=2,
                    title="عطل في اللوحة الكهربائية",
                    issue_type="عطل كهربائي",
                    description="انقطاع التيار الكهربائي بشكل متكرر",
                    status="مفتوح",
                    priority="عالي",
                    assignee_id=readings_recorder.id if readings_recorder else None,
                    responsible_person=readings_recorder.name if readings_recorder else "غير محدد",
                    created_at=datetime.datetime.now() - datetime.timedelta(days=2)
                ),
                Issue(
                    panel_id=3,  # اللوحة تحت الصيانة
                    contractor_id=1,
                    title="تلف في المعدات الداخلية",
                    issue_type="تلف في المعدات",
                    description="تلف في بعض المكونات الداخلية",
                    status="قيد المعالجة",
                    priority="متوسط",
                    assignee_id=readings_recorder.id,  # تخصيص البلاغ لمسجل القراءات
                    responsible_person=readings_recorder.name,
                    created_at=datetime.datetime.now() - datetime.timedelta(days=5)
                ),
                Issue(
                    panel_id=2,
                    contractor_id=3,
                    title="مشكلة في التوصيلات",
                    issue_type="مشكلة في التوصيلات",
                    description="ضعف في التوصيلات الخارجية",
                    status="مغلق",
                    priority="منخفض",
                    created_at=datetime.datetime.now() - datetime.timedelta(days=10),
                    closed_at=datetime.datetime.now() - datetime.timedelta(days=8)
                ),
                # بلاغ عام بدون لوحة محددة
                Issue(
                    panel_id=None,
                    contractor_id=None,
                    title="بلاغ عام للصيانة",
                    issue_type="صيانة عامة",
                    description="طلب صيانة عامة للمنطقة",
                    status="مفتوح",
                    priority="متوسط",
                    assignee_id=readings_recorder.id,  # تخصيص البلاغ لمسجل القراءات
                    responsible_person=readings_recorder.name,
                    created_at=datetime.datetime.now() - datetime.timedelta(days=1)
                )
            ]
            db.session.add_all(issues)
            db.session.commit()
            print(f"تم إضافة {len(issues)} بلاغات بنجاح!")

            # إضافة مناطق الخارطة الافتراضية
            print("إضافة مناطق الخارطة الافتراضية...")
            map_areas = [
                MapArea(
                    name="منطقة عرفة",
                    description="منطقة عرفة في مكة المكرمة",
                    coordinates=json.dumps([[21.3583, 39.9719], [21.3650, 39.9750], [21.3630, 39.9800], [21.3550, 39.9780]]),
                    color="#33ff33",
                    is_active=True,
                    created_at=datetime.datetime.now()
                )
            ]
            db.session.add_all(map_areas)
            db.session.commit()
            print(f"تم إضافة {len(map_areas)} مناطق خارطة بنجاح!")

    # تهيئة الإعدادات الافتراضية
    with app.app_context():
        # دالة لتعيين قيمة إعداد
        def set_setting(key, value, description=None):
            """تعيين قيمة إعداد في قاعدة البيانات"""
            setting = SystemSettings.query.filter_by(key=key).first()
            if setting:
                setting.value = value
                if description:
                    setting.description = description
            else:
                setting = SystemSettings(key=key, value=value, description=description)
                db.session.add(setting)
            db.session.commit()

        # تهيئة الإعدادات الافتراضية
        def init_default_settings():
            """تهيئة الإعدادات الافتراضية للنظام"""
            default_settings = {
                'default_warning_threshold': ('70', 'نسبة التحذير الافتراضية (%)'),
                'default_danger_threshold': ('80', 'نسبة الخطر الافتراضية (%)'),
                'default_min_voltage': ('210', 'الحد الأدنى الافتراضي للجهد (فولت)'),
                'default_max_voltage': ('250', 'الحد الأقصى الافتراضي للجهد (فولت)'),
                'alert_check_interval': ('15', 'الفاصل الزمني للتحقق من التنبيهات (دقيقة)'),
                'alert_method': ('peak', 'طريقة التنبيه (peak, average)'),
                'alert_calculation_period': ('60', 'فترة حساب التنبيهات (دقيقة)'),
                'analysis_period': ('60', 'مدة التحليل الافتراضية (دقيقة)'),
                'send_notifications': ('true', 'إرسال إشعارات للمقاولين'),
                'notification_method': ('email', 'طريقة إرسال الإشعارات (email, sms, both)'),
                'default_map_center': ('21.3583, 39.9719', 'مركز الخارطة الافتراضي (عرفة، مكة المكرمة)'),
                'default_map_zoom': ('14', 'مستوى تكبير الخارطة الافتراضي'),
                'column_name_mdb': ('MDB', 'اسم عمود MDB'),
                'column_name_timestamp': ('وقت القراءة', 'اسم عمود وقت القراءة'),
                'column_name_current': ('التيار (أمبير)', 'اسم عمود التيار'),
                'column_name_voltage': ('الجهد (فولت)', 'اسم عمود الجهد'),
                'column_name_power': ('القدرة (واط)', 'اسم عمود القدرة'),
                'column_name_energy': ('الطاقة (كيلوواط ساعة)', 'اسم عمود الطاقة'),
                'column_name_power_factor': ('معامل القدرة', 'اسم عمود معامل القدرة'),
                'column_name_frequency': ('التردد (هرتز)', 'اسم عمود التردد'),
                'column_name_load': ('الحمل', 'اسم عمود الحمل'),
                'column_name_breaker_capacity': ('سعة القاطع', 'اسم عمود سعة القاطع'),
                'normal_color': ('#28a745', 'لون الحالة الطبيعية'),
                'warning_color': ('#ffc107', 'لون حالة التحذير'),
                'danger_color': ('#dc3545', 'لون حالة الخطر'),
                'trip_color': ('#6c757d', 'لون حالة الفصل'),
                'enable_hotspots': ('true', 'تفعيل النقاط الساخنة على الخارطة'),
                'hotspot_threshold': ('75', 'حد النقاط الساخنة (%)')
            }

            for key, (value, description) in default_settings.items():
                set_setting(key, value, description)

        # تنفيذ تهيئة الإعدادات الافتراضية
        init_default_settings()

        # تشغيل ترحيل قاعدة البيانات للجداول الجديدة
        try:
            from migrations.add_assignment_tables import run_migration
            run_migration()
        except Exception as e:
            print(f"خطأ في ترحيل قاعدة البيانات: {str(e)}")

# عرض طلبات الفحص المستلمة (متاحة لجميع المستخدمين)
@app.route('/received-inspection-requests')
@login_required
def received_inspection_requests():
    # هذه الصفحة متاحة لجميع المستخدمين لرؤية طلبات الفحص المخصصة لهم
    # لا حاجة لتقييد الوصول حسب نوع المستخدم

    # الحصول على طلبات الفحص المخصصة للمستخدم الحالي
    # الحصول على مجموعات المستخدم
    from models import UserGroup, UserGroupMembership
    user_groups = UserGroupMembership.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).all()
    user_group_ids = [membership.group_id for membership in user_groups]

    # بناء شروط البحث
    search_conditions = [
        InspectionRequest.assignee_id == current_user.id,
        InspectionRequest.assigned_to == current_user.id,
        InspectionRequest.responsible_person == current_user.name,
        InspectionRequest.responsible_person == str(current_user.id)  # بحث بالID أيضاً
    ]

    # إضافة شرط المجموعات إذا وجدت
    if user_group_ids:
        search_conditions.append(InspectionRequest.assigned_group_id.in_(user_group_ids))

    # إضافة شرط التوزيعات الجديدة
    inspection_assignments = InspectionRequestAssignment.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).all()
    if inspection_assignments:
        assigned_request_ids = [assignment.inspection_request_id for assignment in inspection_assignments]
        search_conditions.append(InspectionRequest.id.in_(assigned_request_ids))

    requests_query = InspectionRequest.query.filter(db.or_(*search_conditions))

    # تطبيق الفلتر إذا تم تحديده
    status_filter = request.args.get('status')
    if status_filter and status_filter != 'all':
        requests_query = requests_query.filter(InspectionRequest.status == status_filter)

    requests = requests_query.order_by(InspectionRequest.created_at.desc()).all()

    # الحصول على أنواع طلبات الفحص
    request_types = InspectionRequestType.query.all()

    return render_template('received_inspection_requests.html',
                           requests=requests,
                           request_types=request_types,
                           current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# إكمال طلب فحص
@app.route('/complete-inspection-request/<int:request_id>', methods=['POST'])
@login_required
def complete_inspection_request(request_id):
    # التحقق من أن المستخدم هو مسجل قراءات
    if current_user.role != 'readings_recorder':
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية للقيام بهذه العملية'})

    # الحصول على طلب الفحص
    inspection_request = InspectionRequest.query.get_or_404(request_id)

    # التحقق من أن الطلب مخصص للمستخدم الحالي (مباشرة أو عبر التوزيع)
    is_directly_assigned = (
        inspection_request.assignee_id == current_user.id or
        inspection_request.assigned_to == current_user.id or
        inspection_request.responsible_person == current_user.name or
        inspection_request.responsible_person == str(current_user.id)
    )

    # التحقق من التخصيص عبر المجموعات
    from models import UserGroupMembership, SubGroupMembership
    user_groups = UserGroupMembership.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).all()
    user_group_ids = [membership.group_id for membership in user_groups]

    is_group_assigned = False
    if user_group_ids and inspection_request.assigned_group_id in user_group_ids:
        is_group_assigned = True

    user_sub_groups = SubGroupMembership.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).all()
    user_sub_group_ids = [membership.sub_group_id for membership in user_sub_groups]

    is_sub_group_assigned = False
    if user_sub_group_ids and inspection_request.assigned_sub_group_id in user_sub_group_ids:
        is_sub_group_assigned = True

    # التحقق من التخصيص عبر جدول التوزيعات
    assignment = InspectionRequestAssignment.query.filter_by(
        inspection_request_id=request_id,
        user_id=current_user.id,
        is_active=True
    ).first()

    # إذا لم يكن المستخدم مخصص بأي طريقة
    if not (is_directly_assigned or is_group_assigned or is_sub_group_assigned or assignment):
        return jsonify({'success': False, 'message': 'هذا الطلب غير مخصص لك'})

    # تحديث حالة الطلب
    inspection_request.status = 'مكتمل'
    inspection_request.completed_at = datetime.datetime.now()
    inspection_request.completed_by = current_user.id

    # حساب وقت الإنجاز (من البدء حتى الإكمال)
    if inspection_request.started_at:
        completion_time_minutes = int((inspection_request.completed_at - inspection_request.started_at).total_seconds() / 60)
        inspection_request.completion_time = completion_time_minutes

    # إضافة ملاحظات الإكمال
    notes = request.form.get('notes', '')
    if notes:
        inspection_request.notes = notes

    try:
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# حل بلاغ
@app.route('/resolve-issue/<int:issue_id>', methods=['POST'])
@login_required
def resolve_issue(issue_id):
    # التحقق من أن المستخدم هو مسجل قراءات
    if current_user.role != 'readings_recorder':
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية للقيام بهذه العملية'})

    # الحصول على البلاغ
    issue = Issue.query.get_or_404(issue_id)

    # التحقق من أن البلاغ مخصص للمستخدم الحالي (مباشرة أو عبر التوزيع)
    is_directly_assigned = (
        issue.assignee_id == current_user.id or
        issue.responsible_person == current_user.name or
        issue.responsible_person == str(current_user.id)
    )

    # التحقق من التخصيص عبر المجموعات
    from models import UserGroupMembership, SubGroupMembership
    user_groups = UserGroupMembership.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).all()
    user_group_ids = [membership.group_id for membership in user_groups]

    is_group_assigned = False
    if user_group_ids and issue.assigned_group_id in user_group_ids:
        is_group_assigned = True

    user_sub_groups = SubGroupMembership.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).all()
    user_sub_group_ids = [membership.sub_group_id for membership in user_sub_groups]

    is_sub_group_assigned = False
    if user_sub_group_ids and issue.assigned_sub_group_id in user_sub_group_ids:
        is_sub_group_assigned = True

    # التحقق من التخصيص عبر جدول التوزيعات
    assignment = IssueAssignment.query.filter_by(
        issue_id=issue_id,
        user_id=current_user.id,
        is_active=True
    ).first()

    # إذا لم يكن المستخدم مخصص بأي طريقة
    if not (is_directly_assigned or is_group_assigned or is_sub_group_assigned or assignment):
        return jsonify({'success': False, 'message': 'هذا البلاغ غير مخصص لك'})

    # تحديث حالة البلاغ
    issue.status = 'resolved'
    issue.closed_at = datetime.datetime.now()
    issue.resolved_by = current_user.id

    # حساب وقت الإنجاز (من البدء حتى الحل)
    if issue.started_at:
        closure_time_minutes = int((issue.closed_at - issue.started_at).total_seconds() / 60)
        issue.closure_time = closure_time_minutes

    # إضافة ملاحظات الحل
    notes = request.form.get('notes', '')
    if notes:
        issue.resolution_notes = notes
        if issue.description:
            issue.description += f"\n\n[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}] تم الحل: {notes}"
        else:
            issue.description = f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}] تم الحل: {notes}"

    try:
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# قبول طلب فحص
@app.route('/accept-inspection-request/<int:request_id>', methods=['POST'])
@login_required
def accept_inspection_request(request_id):
    try:
        # الحصول على طلب الفحص
        inspection_request = InspectionRequest.query.get_or_404(request_id)

        # التحقق من أن المستخدم مخصص لهذا الطلب (مباشرة أو عبر التوزيع)
        is_directly_assigned = (
            inspection_request.assignee_id == current_user.id or
            inspection_request.assigned_to == current_user.id or
            inspection_request.responsible_person == current_user.name or
            inspection_request.responsible_person == str(current_user.id)
        )

        # التحقق من التخصيص عبر المجموعات
        from models import UserGroupMembership, SubGroupMembership
        user_groups = UserGroupMembership.query.filter_by(
            user_id=current_user.id,
            is_active=True
        ).all()
        user_group_ids = [membership.group_id for membership in user_groups]

        is_group_assigned = False
        if user_group_ids and inspection_request.assigned_group_id in user_group_ids:
            is_group_assigned = True

        user_sub_groups = SubGroupMembership.query.filter_by(
            user_id=current_user.id,
            is_active=True
        ).all()
        user_sub_group_ids = [membership.sub_group_id for membership in user_sub_groups]

        is_sub_group_assigned = False
        if user_sub_group_ids and inspection_request.assigned_sub_group_id in user_sub_group_ids:
            is_sub_group_assigned = True

        # التحقق من التخصيص عبر جدول التوزيعات
        assignment = InspectionRequestAssignment.query.filter_by(
            inspection_request_id=request_id,
            user_id=current_user.id,
            is_active=True
        ).first()

        # إذا لم يكن المستخدم مخصص بأي طريقة
        if not (is_directly_assigned or is_group_assigned or is_sub_group_assigned or assignment):
            return jsonify({'success': False, 'message': 'هذا الطلب غير مخصص لك'})

        # إذا كان هناك تخصيص في جدول التوزيعات، قم بتحديثه
        if assignment:
            assignment.is_accepted = True
            assignment.accepted_at = datetime.datetime.now()

            # تعطيل جميع التخصيصات الأخرى لهذا الطلب
            other_assignments = InspectionRequestAssignment.query.filter(
                InspectionRequestAssignment.inspection_request_id == request_id,
                InspectionRequestAssignment.user_id != current_user.id,
                InspectionRequestAssignment.is_active == True
            ).all()

            for other_assignment in other_assignments:
                other_assignment.is_active = False

        # تحديث طلب الفحص
        inspection_request.assignee_id = current_user.id
        inspection_request.status = 'قيد التنفيذ'
        inspection_request.started_at = datetime.datetime.now()

        # حساب وقت المعالجة (من الإنشاء حتى البدء)
        processing_time_minutes = int((inspection_request.started_at - inspection_request.created_at).total_seconds() / 60)
        inspection_request.processing_time = processing_time_minutes

        db.session.commit()

        return jsonify({'success': True, 'message': 'تم قبول طلب الفحص بنجاح'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# قبول بلاغ
@app.route('/accept-issue/<int:issue_id>', methods=['POST'])
@login_required
def accept_issue(issue_id):
    try:
        # الحصول على البلاغ
        issue = Issue.query.get_or_404(issue_id)

        # التحقق من أن المستخدم مخصص لهذا البلاغ (مباشرة أو عبر التوزيع)
        is_directly_assigned = (
            issue.assignee_id == current_user.id or
            issue.responsible_person == current_user.name or
            issue.responsible_person == str(current_user.id)
        )

        # التحقق من التخصيص عبر المجموعات
        from models import UserGroupMembership, SubGroupMembership
        user_groups = UserGroupMembership.query.filter_by(
            user_id=current_user.id,
            is_active=True
        ).all()
        user_group_ids = [membership.group_id for membership in user_groups]

        is_group_assigned = False
        if user_group_ids and issue.assigned_group_id in user_group_ids:
            is_group_assigned = True

        user_sub_groups = SubGroupMembership.query.filter_by(
            user_id=current_user.id,
            is_active=True
        ).all()
        user_sub_group_ids = [membership.sub_group_id for membership in user_sub_groups]

        is_sub_group_assigned = False
        if user_sub_group_ids and issue.assigned_sub_group_id in user_sub_group_ids:
            is_sub_group_assigned = True

        # التحقق من التخصيص عبر جدول التوزيعات
        assignment = IssueAssignment.query.filter_by(
            issue_id=issue_id,
            user_id=current_user.id,
            is_active=True
        ).first()

        # إذا لم يكن المستخدم مخصص بأي طريقة
        if not (is_directly_assigned or is_group_assigned or is_sub_group_assigned or assignment):
            return jsonify({'success': False, 'message': 'هذا البلاغ غير مخصص لك'})

        # إذا كان هناك تخصيص في جدول التوزيعات، قم بتحديثه
        if assignment:
            assignment.is_accepted = True
            assignment.accepted_at = datetime.datetime.now()

            # تعطيل جميع التخصيصات الأخرى لهذا البلاغ
            other_assignments = IssueAssignment.query.filter(
                IssueAssignment.issue_id == issue_id,
                IssueAssignment.user_id != current_user.id,
                IssueAssignment.is_active == True
            ).all()

            for other_assignment in other_assignments:
                other_assignment.is_active = False

        # تحديث البلاغ
        issue.assignee_id = current_user.id
        issue.status = 'in_progress'
        issue.started_at = datetime.datetime.now()

        # حساب وقت المعالجة (من الإنشاء حتى البدء)
        processing_time_minutes = int((issue.started_at - issue.created_at).total_seconds() / 60)
        issue.processing_time = processing_time_minutes

        db.session.commit()

        return jsonify({'success': True, 'message': 'تم قبول البلاغ بنجاح'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ===== دوال مساعدة للوحة معلومات المخيمات =====

def get_countries_list():
    """الحصول على قائمة الدول النشطة"""
    try:
        return Country.query.filter_by(is_active=True).order_by(Country.name).all()
    except Exception as e:
        print(f"خطأ في الحصول على قائمة الدول: {e}")
        return []

def get_companies_list(country_filter=None):
    """الحصول على قائمة الشركات النشطة"""
    try:
        query = Company.query.filter_by(is_active=True)
        if country_filter:
            query = query.filter_by(country_id=int(country_filter))
        return query.order_by(Company.name).all()
    except Exception as e:
        print(f"خطأ في الحصول على قائمة الشركات: {e}")
        return []

def get_filtered_camps(filters):
    """الحصول على المخيمات مع تطبيق الفلاتر"""
    try:
        # بناء الاستعلام الأساسي
        query = Camp.query.join(Company).join(Country).filter(Camp.is_active == True)

        # تطبيق فلتر الدولة
        if filters['country']:
            query = query.filter(Country.id == int(filters['country']))

        # تطبيق فلتر الشركة
        if filters['company']:
            query = query.filter(Company.id == int(filters['company']))

        # تطبيق فلتر البحث
        if filters['search']:
            search_term = f"%{filters['search']}%"
            query = query.filter(
                db.or_(
                    Camp.camp_number.like(search_term),
                    Camp.square_number.like(search_term),
                    Company.name.like(search_term)
                )
            )

        # الحصول على المخيمات الأولية
        camps = query.order_by(Camp.camp_number, Camp.square_number).all()

        # تطبيق الفلاتر المتقدمة وإضافة عدد اللوحات
        filtered_camps = []
        for camp in camps:
            if should_include_camp(camp, filters):
                # إضافة عدد اللوحات للمخيم
                panels_count = PanelCampAssignment.query.filter_by(
                    camp_id=camp.id, is_active=True
                ).count()
                camp.panels_count = panels_count
                filtered_camps.append(camp)

        return filtered_camps

    except Exception as e:
        print(f"خطأ في فلترة المخيمات: {e}")
        return []

def should_include_camp(camp, filters):
    """تحديد ما إذا كان يجب تضمين المخيم بناءً على الفلاتر"""
    try:
        # فلتر اللوحات
        if filters['panels_filter']:
            panels_count = PanelCampAssignment.query.filter_by(
                camp_id=camp.id, is_active=True
            ).count()

            if filters['panels_filter'] == 'with_panels' and panels_count == 0:
                return False
            elif filters['panels_filter'] == 'without_panels' and panels_count > 0:
                return False

        # فلتر البلاغات
        if filters['issues_filter']:
            issues_count = get_camp_issues_count(camp.id)

            if filters['issues_filter'] == 'with_issues' and issues_count == 0:
                return False
            elif filters['issues_filter'] == 'without_issues' and issues_count > 0:
                return False

        # فلتر طلبات الفحص
        if filters['inspections_filter']:
            inspections_count = get_camp_inspections_count(camp.id)

            if filters['inspections_filter'] == 'with_inspections' and inspections_count == 0:
                return False
            elif filters['inspections_filter'] == 'without_inspections' and inspections_count > 0:
                return False

        return True

    except Exception as e:
        print(f"خطأ في فحص المخيم {camp.id}: {e}")
        return True  # في حالة الخطأ، نضمن المخيم

def get_camp_issues_count(camp_id):
    """حساب عدد البلاغات للمخيم"""
    try:
        panel_ids = db.session.query(PanelCampAssignment.panel_id).filter_by(
            camp_id=camp_id, is_active=True
        ).subquery()

        return Issue.query.filter(
            Issue.panel_id.in_(panel_ids),
            Issue.status.in_(['مفتوح', 'قيد المعالجة'])
        ).count()
    except Exception as e:
        print(f"خطأ في حساب البلاغات للمخيم {camp_id}: {e}")
        return 0

def get_camp_inspections_count(camp_id):
    """حساب عدد طلبات الفحص للمخيم"""
    try:
        panel_ids = db.session.query(PanelCampAssignment.panel_id).filter_by(
            camp_id=camp_id, is_active=True
        ).subquery()

        return InspectionRequest.query.filter(
            InspectionRequest.panel_id.in_(panel_ids),
            InspectionRequest.status.in_(['pending', 'in_progress'])
        ).count()
    except Exception as e:
        print(f"خطأ في حساب طلبات الفحص للمخيم {camp_id}: {e}")
        return 0

def calculate_camps_statistics(camps, filters):
    """حساب الإحصائيات العامة للمخيمات"""
    try:
        stats = {
            'total_camps': len(camps),
            'total_pilgrims': sum(camp.pilgrims_count or 0 for camp in camps),
            'total_companies': len(set(camp.company_id for camp in camps)) if camps else 0,
            'total_countries': len(set(camp.company.country_id for camp in camps)) if camps else 0,
            'total_area': sum(camp.total_area or 0 for camp in camps),
            'total_panels': 0,
            'total_issues': 0,
            'total_inspections': 0,
            'avg_load_percentage': 0,
            'load_per_pilgrim': 0,
            'total_consumption': 0,
            'consumption_per_pilgrim': 0
        }

        if not camps:
            return stats

        # حساب اللوحات والقراءات
        camp_ids = [camp.id for camp in camps]
        panel_assignments = PanelCampAssignment.query.filter(
            PanelCampAssignment.camp_id.in_(camp_ids),
            PanelCampAssignment.is_active == True
        ).all()

        stats['total_panels'] = len(panel_assignments)

        # حساب البلاغات وطلبات الفحص
        panel_ids = [assignment.panel_id for assignment in panel_assignments]
        if panel_ids:
            stats['total_issues'] = Issue.query.filter(Issue.panel_id.in_(panel_ids)).count()
            stats['total_inspections'] = InspectionRequest.query.filter(
                InspectionRequest.panel_id.in_(panel_ids)
            ).count()

        # حساب الأحمال والاستهلاك
        total_load = 0
        total_consumption = 0
        panels_with_readings = 0

        for assignment in panel_assignments:
            latest_reading = ElectricalReading.query.filter_by(
                panel_id=assignment.panel_id
            ).order_by(ElectricalReading.timestamp.desc()).first()

            if latest_reading:
                if latest_reading.load_percentage is not None:
                    total_load += latest_reading.load_percentage
                    panels_with_readings += 1
                if latest_reading.power is not None:
                    total_consumption += latest_reading.power

        # حساب المتوسطات
        if panels_with_readings > 0:
            stats['avg_load_percentage'] = total_load / panels_with_readings

        stats['total_consumption'] = total_consumption

        if stats['total_pilgrims'] > 0:
            stats['load_per_pilgrim'] = stats['avg_load_percentage'] / stats['total_pilgrims']
            stats['consumption_per_pilgrim'] = stats['total_consumption'] / stats['total_pilgrims']

        return stats

    except Exception as e:
        print(f"خطأ في حساب الإحصائيات: {e}")
        return {
            'total_camps': 0, 'total_pilgrims': 0, 'total_companies': 0,
            'total_countries': 0, 'total_area': 0, 'total_panels': 0,
            'total_issues': 0, 'total_inspections': 0, 'avg_load_percentage': 0,
            'load_per_pilgrim': 0, 'total_consumption': 0, 'consumption_per_pilgrim': 0
        }

def prepare_map_data(camps):
    """إعداد بيانات الخرائط"""
    try:
        camps_data = []
        print(f"🗺️ إعداد بيانات الخارطة لـ {len(camps)} مخيم...")

        for camp in camps:
            try:
                # محاولة الحصول على الإحداثيات
                center_coords = None
                polygon_coords = None

                # طريقة 1: استخدام get_center_coordinates و get_coordinates_list
                if hasattr(camp, 'get_center_coordinates') and hasattr(camp, 'get_coordinates_list'):
                    try:
                        center_coords = camp.get_center_coordinates()
                        polygon_coords = camp.get_coordinates_list()

                        # تشخيص الإحداثيات (فقط للمخيمات الأولى لتجنب الإزعاج)
                        if len(camps_data) < 3:
                            print(f"🏕️ المخيم {camp.camp_number}:")
                            print(f"   المركز: {center_coords}")
                            print(f"   عدد نقاط المضلع: {len(polygon_coords) if polygon_coords else 0}")

                    except Exception as e:
                        if len(camps_data) < 3:
                            print(f"   ⚠️ خطأ في استخراج الإحداثيات: {e}")
                        center_coords = None
                        polygon_coords = None

                # طريقة 2: تحليل coordinates مباشرة إذا فشلت الطريقة الأولى
                if not center_coords and camp.coordinates:
                    try:
                        coords_str = camp.coordinates.strip()
                        print(f"   📍 إحداثيات خام: {coords_str[:100]}...")

                        if ',' in coords_str:
                            # إذا كانت نقطة واحدة: lat,lon
                            if ' ' not in coords_str:
                                parts = coords_str.split(',')
                                if len(parts) >= 2:
                                    lat = float(parts[0].strip())
                                    lon = float(parts[1].strip())
                                    center_coords = [lat, lon]
                            else:
                                # إذا كانت نقاط متعددة: lat1,lon1,z1 lat2,lon2,z2
                                coord_pairs = coords_str.split(' ')
                                points = []
                                for pair in coord_pairs:
                                    if ',' in pair:
                                        parts = pair.split(',')
                                        if len(parts) >= 2:
                                            try:
                                                lat = float(parts[0].strip())
                                                lon = float(parts[1].strip())
                                                # التحقق من صحة الإحداثيات
                                                if 20.0 <= lat <= 22.0 and 39.0 <= lon <= 41.0:
                                                    points.append([lat, lon])
                                            except ValueError:
                                                continue

                                if points:
                                    polygon_coords = points
                                    # حساب المركز
                                    avg_lat = sum(point[0] for point in points) / len(points)
                                    avg_lon = sum(point[1] for point in points) / len(points)
                                    center_coords = [avg_lat, avg_lon]
                                    print(f"   ✅ تم استخراج {len(points)} نقطة، المركز: {center_coords}")
                    except Exception as e:
                        print(f"   ❌ خطأ في تحليل الإحداثيات: {e}")
                        center_coords = None

                # طريقة 3: إحداثيات افتراضية لمنى
                if not center_coords:
                    print(f"   🔄 استخدام إحداثيات افتراضية للمخيم {camp.camp_number}")
                    center_coords = [21.3891, 39.8579]

                # التأكد من صحة الإحداثيات
                if center_coords and len(center_coords) == 2:
                    lat, lon = center_coords
                    if -90 <= lat <= 90 and -180 <= lon <= 180:
                        panels_count = PanelCampAssignment.query.filter_by(
                            camp_id=camp.id, is_active=True
                        ).count()

                        camp_data = {
                            'id': camp.id,
                            'camp_number': camp.camp_number,
                            'square_number': camp.square_number,
                            'company_name': camp.company.name if camp.company else 'غير محدد',
                            'country_name': camp.company.country.name if camp.company and camp.company.country else 'غير محدد',
                            'pilgrims_count': camp.pilgrims_count or 0,
                            'panels_count': panels_count,
                            'coordinates': center_coords,
                            'polygon_coordinates': polygon_coords,  # إضافة إحداثيات المضلع
                            'total_area': camp.total_area or 0
                        }
                        camps_data.append(camp_data)
                        print(f"   ✅ تم إضافة المخيم {camp.camp_number} للخارطة")

            except Exception as e:
                print(f"❌ خطأ في معالجة المخيم {camp.id}: {e}")
                continue

        print(f"🎉 تم إعداد بيانات {len(camps_data)} مخيم للخارطة")
        return {'camps': camps_data}

    except Exception as e:
        print(f"❌ خطأ في إعداد بيانات الخرائط: {e}")
        return {'camps': []}

def prepare_chart_data(camps, stats):
    """إعداد بيانات الرسوم البيانية"""
    try:
        # بيانات الحجاج حسب الدولة
        country_pilgrims = {}
        for camp in camps:
            country = camp.company.country.name
            if country not in country_pilgrims:
                country_pilgrims[country] = 0
            country_pilgrims[country] += camp.pilgrims_count or 0

        countries_labels = list(country_pilgrims.keys())
        pilgrims_by_country = list(country_pilgrims.values())

        # بيانات المخيمات حسب الشركة (أكبر 10 شركات)
        company_camps = {}
        for camp in camps:
            company = camp.company.name
            if company not in company_camps:
                company_camps[company] = 0
            company_camps[company] += 1

        # ترتيب الشركات حسب عدد المخيمات
        sorted_companies = sorted(company_camps.items(), key=lambda x: x[1], reverse=True)[:10]
        companies_labels = [item[0] for item in sorted_companies]
        camps_by_company = [item[1] for item in sorted_companies]

        return {
            'countries_labels': countries_labels,
            'pilgrims_by_country': pilgrims_by_country,
            'companies_labels': companies_labels,
            'camps_by_company': camps_by_company
        }

    except Exception as e:
        print(f"خطأ في إعداد بيانات الرسوم البيانية: {e}")
        return {
            'countries_labels': [],
            'pilgrims_by_country': [],
            'companies_labels': [],
            'camps_by_company': []
        }

def calculate_country_statistics(camps):
    """حساب إحصائيات الدول"""
    try:
        country_stats = {}

        for camp in camps:
            country_name = camp.company.country.name
            country_id = camp.company.country.id

            if country_name not in country_stats:
                country_stats[country_name] = {
                    'country_id': country_id,
                    'camps_count': 0,
                    'pilgrims_count': 0,
                    'companies_count': set(),
                    'panels_count': 0
                }

            country_stats[country_name]['camps_count'] += 1
            country_stats[country_name]['pilgrims_count'] += camp.pilgrims_count or 0
            country_stats[country_name]['companies_count'].add(camp.company_id)

            # حساب اللوحات
            panels_count = PanelCampAssignment.query.filter_by(
                camp_id=camp.id, is_active=True
            ).count()
            country_stats[country_name]['panels_count'] += panels_count

        # تحويل sets إلى أعداد
        for country in country_stats:
            country_stats[country]['companies_count'] = len(country_stats[country]['companies_count'])

        return country_stats

    except Exception as e:
        print(f"خطأ في حساب إحصائيات الدول: {e}")
        return {}

def calculate_company_statistics(camps):
    """حساب إحصائيات الشركات"""
    try:
        company_stats = {}

        for camp in camps:
            company_name = camp.company.name

            if company_name not in company_stats:
                company_stats[company_name] = {
                    'company_id': camp.company.id,
                    'camps_count': 0,
                    'pilgrims_count': 0,
                    'country': camp.company.country.name,
                    'country_id': camp.company.country.id,
                    'panels_count': 0
                }

            company_stats[company_name]['camps_count'] += 1
            company_stats[company_name]['pilgrims_count'] += camp.pilgrims_count or 0

            # حساب اللوحات
            panels_count = PanelCampAssignment.query.filter_by(
                camp_id=camp.id, is_active=True
            ).count()
            company_stats[company_name]['panels_count'] += panels_count

        return company_stats

    except Exception as e:
        print(f"خطأ في حساب إحصائيات الشركات: {e}")
        return {}

# ===== لوحة معلومات المخيمات الرئيسية =====

# لوحة معلومات المخيمات والشركات - إعادة هيكلة كاملة
@app.route('/camps-dashboard')
@login_required
def camps_dashboard():
    """لوحة معلومات المخيمات والشركات - نسخة محسنة ومعاد هيكلتها"""
    try:
        # 1. التحقق من الصلاحيات
        if not current_user or current_user.role not in ['admin', 'user']:
            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
            return redirect(url_for('login'))

        # 2. الحصول على معاملات الفلترة مع قيم افتراضية آمنة
        filters = {
            'country': request.args.get('country', ''),
            'company': request.args.get('company', ''),
            'search': request.args.get('search', ''),
            'panels_filter': request.args.get('panels_filter', ''),
            'issues_filter': request.args.get('issues_filter', ''),
            'inspections_filter': request.args.get('inspections_filter', '')
        }

        # 3. الحصول على البيانات الأساسية
        countries = get_countries_list()
        companies = get_companies_list(filters['country'])
        camps = get_filtered_camps(filters)

        # 4. حساب الإحصائيات
        stats = calculate_camps_statistics(camps, filters)

        # 5. إعداد بيانات الخرائط والرسوم البيانية
        map_data = prepare_map_data(camps)
        chart_data = prepare_chart_data(camps, stats)

        # 6. إعداد إحصائيات الدول والشركات
        country_stats = calculate_country_statistics(camps)
        company_stats = calculate_company_statistics(camps)

        # 7. إرجاع البيانات للقالب
        return render_template('camps_dashboard.html',
                             # البيانات الأساسية
                             camps=camps,
                             countries=countries,
                             companies=companies,

                             # الفلاتر
                             country_filter=filters['country'],
                             company_filter=filters['company'],
                             search=filters['search'],
                             panels_filter=filters['panels_filter'],
                             issues_filter=filters['issues_filter'],
                             inspections_filter=filters['inspections_filter'],

                             # الإحصائيات العامة
                             total_camps=stats['total_camps'],
                             total_pilgrims=stats['total_pilgrims'],
                             total_companies=stats['total_companies'],
                             total_countries=stats['total_countries'],
                             total_panels=stats['total_panels'],
                             total_area=stats['total_area'],
                             total_issues=stats['total_issues'],
                             total_inspections=stats['total_inspections'],
                             avg_load_percentage=stats['avg_load_percentage'],
                             load_per_pilgrim=stats['load_per_pilgrim'],
                             total_consumption=stats['total_consumption'],
                             consumption_per_pilgrim=stats['consumption_per_pilgrim'],

                             # إحصائيات الدول والشركات
                             country_stats=country_stats,
                             company_stats=company_stats,

                             # بيانات الخرائط والرسوم البيانية
                             camps_map_data=json.dumps(map_data['camps']),
                             pilgrims_by_country=json.dumps(chart_data['pilgrims_by_country']),
                             countries_labels=json.dumps(chart_data['countries_labels']),
                             camps_by_company=json.dumps(chart_data['camps_by_company']),
                             companies_labels=json.dumps(chart_data['companies_labels']),

                             # معلومات إضافية
                             current_date=datetime.datetime.now().strftime("%Y-%m-%d"),
                             panel_assignments=[])

    except Exception as e:
        # معالجة شاملة للأخطاء
        error_msg = f"خطأ في لوحة معلومات المخيمات: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()

        flash('حدث خطأ أثناء تحميل لوحة المعلومات. يرجى المحاولة مرة أخرى.', 'danger')

        # إرجاع صفحة فارغة في حالة الخطأ
        return render_template('camps_dashboard.html',
                             camps=[], countries=[], companies=[],
                             country_filter='', company_filter='', search='',
                             panels_filter='', issues_filter='', inspections_filter='',
                             total_camps=0, total_pilgrims=0, total_companies=0,
                             total_countries=0, total_panels=0, total_area=0,
                             total_issues=0, total_inspections=0, avg_load_percentage=0,
                             load_per_pilgrim=0, total_consumption=0, consumption_per_pilgrim=0,
                             country_stats={}, company_stats={},
                             camps_map_data='[]', pilgrims_by_country='[]',
                             countries_labels='[]', camps_by_company='[]',
                             companies_labels='[]', current_date=datetime.datetime.now().strftime("%Y-%m-%d"),
                             panel_assignments=[])

# استيراد بيانات المخيمات من Excel
@app.route('/import-camps', methods=['POST'])
@login_required
def import_camps():
    """استيراد بيانات المخيمات من Excel"""
    # التحقق من الصلاحيات
    if current_user.role not in ['admin']:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('camps_dashboard'))

    if 'file' not in request.files:
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('camps_dashboard'))

    file = request.files['file']

    if file.filename == '':
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('camps_dashboard'))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        try:
            # قراءة الملف
            if file.filename.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path, engine='openpyxl')

            # التحقق من وجود الأعمدة الأساسية المطلوبة
            basic_required_columns = [
                'رقم المربع', 'رقم المخيم', 'عدد الحجاج',
                'اسم الشركة', 'الدولة'
            ]

            missing_basic_columns = [col for col in basic_required_columns if col not in df.columns]
            if missing_basic_columns:
                flash(f'الأعمدة الأساسية التالية مفقودة: {", ".join(missing_basic_columns)}', 'danger')
                return redirect(url_for('camps_dashboard'))

            # تعريف أعمدة بديلة للحقول الاختيارية
            area_columns = ['المساحة الإجمالية', 'المساحة', 'Total Area', 'Area']
            coord_columns = ['الإحداثيات', 'Coordinates', 'إحداثيات المضلع', 'Polygon Coordinates']
            zone_columns = ['Zone/Style', 'Zone', 'Style', 'المنطقة', 'النمط']
            person_columns = ['الشخص المسؤول', 'المسؤول', 'Responsible Person', 'Contact Person']
            contact_columns = ['رقم الاتصال', 'الهاتف', 'Contact Number', 'Phone']

            # معالجة البيانات
            camps_added = 0
            countries_added = 0
            companies_added = 0

            for index, row in df.iterrows():
                try:
                    # معالجة الشخص المسؤول - البحث في أعمدة متعددة
                    contact_person = ''
                    for col in person_columns:
                        if col in row and pd.notna(row[col]):
                            contact_person = str(row[col]).strip()
                            if contact_person and contact_person != 'nan':
                                break

                    # معالجة رقم الاتصال - البحث في أعمدة متعددة
                    phone = ''
                    for col in contact_columns:
                        if col in row and pd.notna(row[col]):
                            phone = str(row[col]).strip()
                            if phone and phone != 'nan':
                                break

                    # معالجة الدولة
                    country_name = str(row['الدولة']).strip()
                    country = Country.query.filter_by(name=country_name).first()
                    if not country:
                        country = Country(name=country_name, is_active=True)
                        db.session.add(country)
                        db.session.flush()  # للحصول على ID
                        countries_added += 1

                    # معالجة الشركة
                    company_name = str(row['اسم الشركة']).strip()
                    company = Company.query.filter_by(
                        name=company_name,
                        country_id=country.id
                    ).first()
                    if not company:
                        company = Company(
                            name=company_name,
                            country_id=country.id,
                            contact_person=contact_person if contact_person else None,
                            phone=phone if phone else None,
                            is_active=True
                        )
                        db.session.add(company)
                        db.session.flush()  # للحصول على ID
                        companies_added += 1
                    else:
                        # تحديث معلومات الاتصال إذا كانت متوفرة
                        updated = False
                        if contact_person and not company.contact_person:
                            company.contact_person = contact_person
                            updated = True
                        if phone and not company.phone:
                            company.phone = phone
                            updated = True
                        if updated:
                            db.session.flush()

                    # معالجة المخيم
                    camp_number = str(row['رقم المخيم']).strip()
                    square_number = str(row['رقم المربع']).strip()

                    # التحقق من عدم وجود المخيم مسبقاً
                    existing_camp = Camp.query.filter_by(
                        camp_number=camp_number,
                        square_number=square_number,
                        company_id=company.id
                    ).first()

                    if not existing_camp:
                        # معالجة البيانات الرقمية
                        pilgrims_count = 0
                        try:
                            pilgrims_count = int(float(row['عدد الحجاج'])) if pd.notna(row['عدد الحجاج']) else 0
                        except (ValueError, TypeError):
                            pilgrims_count = 0

                        # معالجة المساحة - البحث في أعمدة متعددة
                        total_area = None
                        for col in area_columns:
                            if col in row and pd.notna(row[col]):
                                try:
                                    total_area = float(row[col])
                                    break
                                except (ValueError, TypeError):
                                    continue

                        # معالجة الإحداثيات - البحث في أعمدة متعددة
                        coordinates = ''
                        for col in coord_columns:
                            if col in row and pd.notna(row[col]):
                                coordinates = str(row[col]).strip()
                                if coordinates and coordinates != 'nan':
                                    break

                        # معالجة Zone/Style - البحث في أعمدة متعددة
                        zone_style = ''
                        for col in zone_columns:
                            if col in row and pd.notna(row[col]):
                                zone_style = str(row[col]).strip()
                                if zone_style and zone_style != 'nan':
                                    break

                        # إنشاء المخيم
                        camp = Camp(
                            camp_number=camp_number,
                            square_number=square_number,
                            company_id=company.id,
                            pilgrims_count=pilgrims_count,
                            total_area=total_area,
                            zone_style=zone_style,
                            coordinates=coordinates,
                            is_active=True
                        )

                        db.session.add(camp)
                        camps_added += 1

                except Exception as e:
                    print(f"خطأ في معالجة الصف {index + 1}: {str(e)}")
                    continue

            # حفظ التغييرات
            db.session.commit()

            # ربط اللوحات بالمخيمات
            try:
                from camp_management import assign_panels_to_camps
                assignments_count = assign_panels_to_camps()
                if assignments_count > 0:
                    flash(f'تم استيراد {camps_added} مخيم، {companies_added} شركة، {countries_added} دولة، وربط {assignments_count} لوحة بالمخيمات بنجاح', 'success')
                else:
                    flash(f'تم استيراد {camps_added} مخيم، {companies_added} شركة، {countries_added} دولة. لم يتم ربط أي لوحة (تحقق من الإحداثيات)', 'warning')
            except Exception as e:
                print(f"خطأ في ربط اللوحات بالمخيمات: {str(e)}")
                flash(f'تم استيراد {camps_added} مخيم، {companies_added} شركة، {countries_added} دولة. خطأ في ربط اللوحات: {str(e)}', 'warning')

        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء استيراد البيانات: {str(e)}', 'danger')

        # حذف الملف بعد الاستيراد
        os.remove(file_path)

    else:
        flash('نوع الملف غير مسموح به. يرجى استخدام ملفات Excel (.xlsx, .xls) أو CSV', 'danger')

    return redirect(url_for('camps_dashboard'))

# صفحة ربط اللوحات بالمخيمات يدوياً
@app.route('/link-panels-to-camps')
@login_required
def link_panels_to_camps():
    """صفحة ربط اللوحات بالمخيمات يدوياً"""
    if current_user.role not in ['admin']:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('camps_dashboard'))

    # الحصول على اللوحات غير المرتبطة
    unlinked_panels = MDBPanel.query.filter(
        MDBPanel.camp_id.is_(None)
    ).order_by(MDBPanel.mdb).limit(50).all()

    # الحصول على المخيمات
    camps = Camp.query.join(Company).join(Country).order_by(
        Country.name, Company.name, Camp.camp_number
    ).all()

    # الحصول على الدول والشركات للفلترة
    countries = Country.query.filter_by(is_active=True).order_by(Country.name).all()
    companies = Company.query.filter_by(is_active=True).order_by(Company.name).all()

    return render_template('link_panels_to_camps.html',
                         unlinked_panels=unlinked_panels,
                         camps=camps,
                         countries=countries,
                         companies=companies,
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# ربط لوحة بمخيم يدوياً
@app.route('/link-panel-to-camp', methods=['POST'])
@login_required
def link_panel_to_camp():
    """ربط لوحة بمخيم يدوياً"""
    if current_user.role not in ['admin']:
        flash('ليس لديك صلاحية لتنفيذ هذا الإجراء', 'danger')
        return redirect(url_for('camps_dashboard'))

    try:
        panel_id = request.form.get('panel_id')
        camp_id = request.form.get('camp_id')

        if not panel_id or not camp_id:
            flash('يرجى اختيار اللوحة والمخيم', 'danger')
            return redirect(url_for('link_panels_to_camps'))

        # الحصول على اللوحة والمخيم
        panel = MDBPanel.query.get_or_404(panel_id)
        camp = Camp.query.get_or_404(camp_id)

        # التحقق من عدم وجود ربط سابق
        existing_assignment = PanelCampAssignment.query.filter_by(
            panel_id=panel.id,
            camp_id=camp.id
        ).first()

        if existing_assignment:
            flash(f'اللوحة {panel.mdb} مرتبطة بالفعل بالمخيم {camp.camp_number}', 'warning')
            return redirect(url_for('link_panels_to_camps'))

        # حذف أي ربط سابق للوحة
        PanelCampAssignment.query.filter_by(panel_id=panel.id).delete()

        # إنشاء ربط جديد
        assignment = PanelCampAssignment(
            panel_id=panel.id,
            camp_id=camp.id,
            is_active=True
        )
        db.session.add(assignment)

        # تحديث معلومات اللوحة
        panel.camp_id = camp.id
        panel.company_id = camp.company_id
        panel.country_id = camp.company.country_id

        db.session.commit()

        flash(f'تم ربط اللوحة {panel.mdb} بالمخيم {camp.camp_number} ({camp.company.name}) بنجاح', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء الربط: {str(e)}', 'danger')

    return redirect(url_for('link_panels_to_camps'))

# API للحصول على الشركات حسب الدولة
@app.route('/api/companies/<int:country_id>')
@login_required
def get_companies_by_country(country_id):
    """API للحصول على الشركات حسب الدولة"""
    companies = Company.query.filter_by(country_id=country_id, is_active=True).order_by(Company.name).all()
    return jsonify([{
        'id': company.id,
        'name': company.name
    } for company in companies])

# ربط اللوحات بالمخيمات يدوياً
@app.route('/assign-panels-to-camps')
@login_required
def assign_panels_to_camps_route():
    """ربط اللوحات بالمخيمات يدوياً"""
    if current_user.role not in ['admin']:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('camps_dashboard'))

    try:
        from camp_management import assign_panels_to_camps
        assignments_count = assign_panels_to_camps()
        flash(f'تم ربط {assignments_count} لوحة بالمخيمات بنجاح', 'success')
    except Exception as e:
        print(f"خطأ في ربط اللوحات بالمخيمات: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'حدث خطأ أثناء ربط اللوحات: {str(e)}', 'danger')

    return redirect(url_for('camps_dashboard'))

# تحميل قالب بيانات المخيمات
@app.route('/download-camps-template')
@login_required
def download_camps_template():
    """تحميل قالب Excel لاستيراد بيانات المخيمات"""
    # التحقق من الصلاحيات
    if current_user.role not in ['admin']:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('camps_dashboard'))

    try:
        # إنشاء ملف Excel جديد
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "قالب بيانات المخيمات"

        # تعيين العناوين
        headers = [
            'رقم المربع',
            'رقم المخيم',
            'عدد الحجاج',
            'اسم الشركة',
            'الدولة',
            'المساحة الإجمالية',
            'إحداثيات المضلع (JSON)',
            'Zone/Style',
            'الشخص المسؤول',
            'رقم الاتصال'
        ]

        # كتابة العناوين
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # إضافة بيانات تجريبية كمثال
        sample_data = [
            [
                "A1",
                "001",
                "500",
                "شركة آراء الخير لخدمات الحجاج",
                "إندونيسيا",
                "2500",
                '[[21.3891, 39.8579], [21.3892, 39.8580], [21.3893, 39.8581], [21.3891, 39.8579]]',
                "منطقة أ - نمط 1",
                "أحمد محمد",
                "+966501234567"
            ],
            [
                "B2",
                "002",
                "750",
                "شركة إقامة لخدمات الحجاج",
                "ماليزيا",
                "3000",
                '[[21.3894, 39.8582], [21.3895, 39.8583], [21.3896, 39.8584], [21.3894, 39.8582]]',
                "منطقة ب - نمط 2",
                "فاطمة علي",
                "+966502345678"
            ]
        ]

        # كتابة البيانات التجريبية
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # تعديل عرض الأعمدة
        column_widths = [15, 15, 15, 30, 15, 15, 50, 20, 20, 20]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = width

        # إضافة ورقة تعليمات
        instructions_sheet = workbook.create_sheet("تعليمات الاستيراد")
        instructions = [
            "تعليمات استيراد بيانات المخيمات:",
            "",
            "1. رقم المربع: رقم المربع الذي يحتوي على المخيم (مطلوب)",
            "2. رقم المخيم: رقم المخيم داخل المربع (مطلوب)",
            "3. عدد الحجاج: عدد الحجاج في المخيم (رقم)",
            "4. اسم الشركة: اسم الشركة المسؤولة عن المخيم (مطلوب)",
            "5. الدولة: اسم الدولة التابع لها المخيم (مطلوب)",
            "6. المساحة الإجمالية: المساحة بالمتر المربع (رقم)",
            "7. إحداثيات المضلع: إحداثيات حدود المخيم بصيغة JSON (اختياري)",
            "8. Zone/Style: منطقة أو نمط المخيم (اختياري)",
            "9. الشخص المسؤول: اسم الشخص المسؤول عن المخيم (اختياري)",
            "10. رقم الاتصال: رقم هاتف الشخص المسؤول (اختياري)",
            "",
            "ملاحظات مهمة:",
            "- يجب ملء الحقول المطلوبة (رقم المربع، رقم المخيم، اسم الشركة، الدولة)",
            "- إحداثيات المضلع يجب أن تكون بصيغة JSON صحيحة",
            "- مثال على إحداثيات المضلع: [[21.3891, 39.8579], [21.3892, 39.8580]]",
            "- سيتم إنشاء الشركات والدول الجديدة تلقائياً إذا لم تكن موجودة",
            "- تأكد من صحة البيانات قبل الاستيراد"
        ]

        for row_idx, instruction in enumerate(instructions, 1):
            cell = instructions_sheet.cell(row=row_idx, column=1, value=instruction)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif instruction.startswith(("ملاحظات مهمة:", "تعليمات")):
                cell.font = Font(bold=True, size=12)

        instructions_sheet.column_dimensions['A'].width = 80

        # حفظ الملف في ذاكرة مؤقتة
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # إرسال الملف للتحميل
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'قالب_بيانات_المخيمات_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        flash(f'حدث خطأ أثناء إنشاء القالب: {str(e)}', 'danger')
        return redirect(url_for('camps_dashboard'))

# تصدير تقرير شامل للمخيمات والشركات
@app.route('/export-camps-report', methods=['POST'])
@login_required
def export_camps_report():
    """تصدير تقرير شامل للمخيمات والشركات مع حساب الاستهلاك"""
    try:
        # الحصول على المعاملات من النموذج
        country_filter = request.form.get('country', '')
        company_filter = request.form.get('company', '')
        report_format = request.form.get('report_format', 'excel')

        # البيانات المراد تضمينها
        include_camps = 'include_camps' in request.form
        include_panels = 'include_panels' in request.form
        include_readings = 'include_readings' in request.form
        include_issues = 'include_issues' in request.form
        include_inspections = 'include_inspections' in request.form
        include_consumption = 'include_consumption' in request.form

        # بناء الاستعلام للمخيمات
        camps_query = Camp.query.join(Company).join(Country)

        # تطبيق الفلاتر
        if country_filter:
            camps_query = camps_query.filter(Country.id == country_filter)
        if company_filter:
            camps_query = camps_query.filter(Company.id == company_filter)

        camps = camps_query.order_by(Camp.camp_number, Camp.square_number).all()

        # إنشاء ملف Excel
        workbook = Workbook()

        # حذف الورقة الافتراضية
        workbook.remove(workbook.active)

        # إعداد الألوان والتنسيق
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # ورقة بيانات المخيمات
        if include_camps:
            camps_sheet = workbook.create_sheet("بيانات المخيمات")
            camps_headers = [
                'رقم المربع', 'رقم المخيم', 'عدد الحجاج', 'اسم الشركة', 'الدولة',
                'المساحة الإجمالية', 'Zone/Style', 'الشخص المسؤول', 'رقم الاتصال',
                'عدد اللوحات المرتبطة', 'تاريخ الإنشاء'
            ]

            # كتابة العناوين
            for col, header in enumerate(camps_headers, 1):
                cell = camps_sheet.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # كتابة بيانات المخيمات
            for row_idx, camp in enumerate(camps, 2):
                panels_count = PanelCampAssignment.query.filter_by(
                    camp_id=camp.id, is_active=True
                ).count()

                camps_sheet.cell(row=row_idx, column=1, value=camp.square_number)
                camps_sheet.cell(row=row_idx, column=2, value=camp.camp_number)
                camps_sheet.cell(row=row_idx, column=3, value=camp.pilgrims_count or 0)
                camps_sheet.cell(row=row_idx, column=4, value=camp.company.name)
                camps_sheet.cell(row=row_idx, column=5, value=camp.company.country.name)
                camps_sheet.cell(row=row_idx, column=6, value=camp.total_area or 0)
                camps_sheet.cell(row=row_idx, column=7, value=camp.zone_style or '')
                camps_sheet.cell(row=row_idx, column=8, value=camp.responsible_person or '')
                camps_sheet.cell(row=row_idx, column=9, value=camp.contact_number or '')
                camps_sheet.cell(row=row_idx, column=10, value=panels_count)
                camps_sheet.cell(row=row_idx, column=11, value=camp.created_at.strftime('%Y-%m-%d %H:%M') if camp.created_at else '')

            # تعديل عرض الأعمدة
            for col in range(1, len(camps_headers) + 1):
                camps_sheet.column_dimensions[camps_sheet.cell(row=1, column=col).column_letter].width = 15

        # ورقة اللوحات الكهربائية
        if include_panels:
            panels_sheet = workbook.create_sheet("اللوحات الكهربائية")

            # الحصول على اللوحات المرتبطة بالمخيمات
            panel_assignments = PanelCampAssignment.query.join(MDBPanel).join(Camp).join(Company).join(Country)

            if country_filter:
                panel_assignments = panel_assignments.filter(Country.id == country_filter)
            if company_filter:
                panel_assignments = panel_assignments.filter(Company.id == company_filter)

            panel_assignments = panel_assignments.filter(PanelCampAssignment.is_active == True).all()

            panels_headers = [
                'رقم اللوحة', 'تاج ماكسيمو', 'المنطقة', 'نوع اللوحة', 'الحالة',
                'سعة القاطع', 'SCADA', 'رقم المربع', 'رقم المخيم', 'الشركة', 'الدولة',
                'آخر قراءة تيار', 'آخر قراءة جهد', 'نسبة الحمولة', 'عدد الأعطال', 'عدد التنبيهات'
            ]

            # كتابة العناوين
            for col, header in enumerate(panels_headers, 1):
                cell = panels_sheet.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # كتابة بيانات اللوحات
            for row_idx, assignment in enumerate(panel_assignments, 2):
                panel = assignment.panel
                camp = assignment.camp

                # الحصول على آخر قراءة
                latest_reading = ElectricalReading.query.filter_by(
                    panel_id=panel.id
                ).order_by(ElectricalReading.timestamp.desc()).first()

                # حساب عدد الأعطال والتنبيهات
                issues_count = Issue.query.filter_by(panel_id=panel.id, status='open').count()
                alerts_count = Alert.query.filter_by(panel_id=panel.id, is_resolved=False).count()

                panels_sheet.cell(row=row_idx, column=1, value=panel.mdb)
                panels_sheet.cell(row=row_idx, column=2, value=panel.maximo_tag or '')
                panels_sheet.cell(row=row_idx, column=3, value=panel.area_name or '')
                panels_sheet.cell(row=row_idx, column=4, value=panel.panel_type or '')
                panels_sheet.cell(row=row_idx, column=5, value=panel.status or '')
                panels_sheet.cell(row=row_idx, column=6, value=panel.breaker_capacity or 0)
                panels_sheet.cell(row=row_idx, column=7, value='متصل' if panel.is_scada_connected else 'غير متصل')
                panels_sheet.cell(row=row_idx, column=8, value=camp.square_number)
                panels_sheet.cell(row=row_idx, column=9, value=camp.camp_number)
                panels_sheet.cell(row=row_idx, column=10, value=camp.company.name)
                panels_sheet.cell(row=row_idx, column=11, value=camp.company.country.name)

                if latest_reading:
                    panels_sheet.cell(row=row_idx, column=12, value=f"{latest_reading.current_l1 or 0:.2f}")
                    panels_sheet.cell(row=row_idx, column=13, value=f"{latest_reading.voltage_l1 or 0:.2f}")
                    panels_sheet.cell(row=row_idx, column=14, value=f"{latest_reading.load_percentage or 0:.1f}%")
                else:
                    panels_sheet.cell(row=row_idx, column=12, value="لا توجد قراءات")
                    panels_sheet.cell(row=row_idx, column=13, value="لا توجد قراءات")
                    panels_sheet.cell(row=row_idx, column=14, value="0%")

                panels_sheet.cell(row=row_idx, column=15, value=issues_count)
                panels_sheet.cell(row=row_idx, column=16, value=alerts_count)

            # تعديل عرض الأعمدة
            for col in range(1, len(panels_headers) + 1):
                panels_sheet.column_dimensions[panels_sheet.cell(row=1, column=col).column_letter].width = 15

        # ورقة حساب الاستهلاك
        if include_consumption:
            consumption_sheet = workbook.create_sheet("حساب الاستهلاك")

            consumption_headers = [
                'الدولة', 'الشركة', 'عدد المخيمات', 'عدد الحجاج', 'عدد اللوحات',
                'إجمالي الاستهلاك (كيلو واط)', 'متوسط الاستهلاك لكل حاج', 'متوسط الاستهلاك لكل مخيم'
            ]

            # كتابة العناوين
            for col, header in enumerate(consumption_headers, 1):
                cell = consumption_sheet.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # حساب الاستهلاك حسب الشركة
            consumption_data = {}
            for camp in camps:
                company_key = f"{camp.company.country.name}_{camp.company.name}"

                if company_key not in consumption_data:
                    consumption_data[company_key] = {
                        'country': camp.company.country.name,
                        'company': camp.company.name,
                        'camps_count': 0,
                        'pilgrims_count': 0,
                        'panels_count': 0,
                        'total_consumption': 0
                    }

                consumption_data[company_key]['camps_count'] += 1
                consumption_data[company_key]['pilgrims_count'] += camp.pilgrims_count or 0

                # حساب استهلاك اللوحات المرتبطة بالمخيم
                panel_assignments = PanelCampAssignment.query.filter_by(
                    camp_id=camp.id, is_active=True
                ).all()

                consumption_data[company_key]['panels_count'] += len(panel_assignments)

                for assignment in panel_assignments:
                    # الحصول على آخر قراءة للوحة
                    latest_reading = ElectricalReading.query.filter_by(
                        panel_id=assignment.panel_id
                    ).order_by(ElectricalReading.timestamp.desc()).first()

                    if latest_reading and latest_reading.power:
                        consumption_data[company_key]['total_consumption'] += latest_reading.power

            # كتابة بيانات الاستهلاك
            row_idx = 2
            for data in consumption_data.values():
                consumption_sheet.cell(row=row_idx, column=1, value=data['country'])
                consumption_sheet.cell(row=row_idx, column=2, value=data['company'])
                consumption_sheet.cell(row=row_idx, column=3, value=data['camps_count'])
                consumption_sheet.cell(row=row_idx, column=4, value=data['pilgrims_count'])
                consumption_sheet.cell(row=row_idx, column=5, value=data['panels_count'])
                consumption_sheet.cell(row=row_idx, column=6, value=f"{data['total_consumption']:.2f}")

                # متوسط الاستهلاك لكل حاج
                avg_per_pilgrim = data['total_consumption'] / data['pilgrims_count'] if data['pilgrims_count'] > 0 else 0
                consumption_sheet.cell(row=row_idx, column=7, value=f"{avg_per_pilgrim:.2f}")

                # متوسط الاستهلاك لكل مخيم
                avg_per_camp = data['total_consumption'] / data['camps_count'] if data['camps_count'] > 0 else 0
                consumption_sheet.cell(row=row_idx, column=8, value=f"{avg_per_camp:.2f}")

                row_idx += 1

            # تعديل عرض الأعمدة
            for col in range(1, len(consumption_headers) + 1):
                consumption_sheet.column_dimensions[consumption_sheet.cell(row=1, column=col).column_letter].width = 20

        # إضافة ورقة ملخص
        summary_sheet = workbook.create_sheet("ملخص التقرير")
        workbook.active = summary_sheet  # جعل ورقة الملخص هي الورقة النشطة

        summary_data = [
            ['تقرير شامل للمخيمات والشركات', ''],
            ['تاريخ التقرير', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['المستخدم', current_user.name],
            ['', ''],
            ['إجمالي المخيمات', len(camps)],
            ['إجمالي الحجاج', sum(camp.pilgrims_count or 0 for camp in camps)],
            ['إجمالي الشركات', len(set(camp.company_id for camp in camps))],
            ['إجمالي الدول', len(set(camp.company.country_id for camp in camps))],
        ]

        for row_idx, (label, value) in enumerate(summary_data, 1):
            summary_sheet.cell(row=row_idx, column=1, value=label)
            summary_sheet.cell(row=row_idx, column=2, value=value)

            if row_idx == 1:  # العنوان الرئيسي
                summary_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=16)
                summary_sheet.merge_cells(f'A{row_idx}:B{row_idx}')
            elif label and value:  # البيانات
                summary_sheet.cell(row=row_idx, column=1).font = Font(bold=True)

        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 25

        # حفظ الملف
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # تحديد اسم الملف
        filter_text = ""
        if country_filter:
            country = Country.query.get(country_filter)
            filter_text += f"_{country.name}" if country else ""
        if company_filter:
            company = Company.query.get(company_filter)
            filter_text += f"_{company.name}" if company else ""

        filename = f'تقرير_المخيمات_والشركات{filter_text}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير التقرير: {str(e)}', 'danger')
        return redirect(url_for('camps_dashboard'))



# صفحة التقرير المفصل
@app.route('/detailed-report/<entity_type>/<int:entity_id>')
@login_required
def detailed_report(entity_type, entity_id):
    """عرض تقرير مفصل للدولة أو الشركة"""
    try:
        if entity_type == 'country':
            # الحصول على بيانات الدولة
            country = Country.query.get_or_404(entity_id)
            entity_name = country.name

            # الحصول على المخيمات
            camps = Camp.query.join(Company).filter(Company.country_id == country.id).all()

        elif entity_type == 'company':
            # الحصول على بيانات الشركة
            company = Company.query.get_or_404(entity_id)
            entity_name = company.name

            # الحصول على المخيمات
            camps = Camp.query.filter_by(company_id=company.id).all()

        else:
            flash('نوع التقرير غير صحيح', 'error')
            return redirect(url_for('camps_dashboard'))

        # حساب الإحصائيات الأساسية
        total_camps = len(camps)
        total_pilgrims = sum(camp.pilgrims_count or 0 for camp in camps)
        total_area = sum(camp.total_area or 0 for camp in camps)

        # إحصائيات مجمعة
        stats = {
            'camps_count': total_camps,
            'pilgrims_count': total_pilgrims,
            'panels_count': 0,
            'total_area': total_area,
            'issues_count': 0,
            'inspections_count': 0,
            'avg_load_percentage': 0,
            'total_consumption': 0
        }

        return render_template('detailed_report.html',
                             entity_type=entity_type,
                             entity_id=entity_id,
                             entity_name=entity_name,
                             camps=camps or [],
                             panels=[],
                             stats=stats,
                             camps_map_data=json.dumps([]),
                             panels_map_data=json.dumps([]),
                             pilgrims_chart_data=json.dumps({'labels': [], 'data': []}),
                             panels_chart_data=json.dumps({'labels': [], 'data': []}),
                             report_date=datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                             current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

    except Exception as e:
        print(f"خطأ في التقرير المفصل: {str(e)}")
        flash(f'حدث خطأ أثناء تحميل التقرير: {str(e)}', 'error')
        return redirect(url_for('camps_dashboard'))

# تصدير التقرير المفصل
@app.route('/export-detailed-report', methods=['POST'])
@login_required
def export_detailed_report():
    """تصدير التقرير المفصل"""
    try:
        entity_type = request.form.get('entity_type')
        entity_id = int(request.form.get('entity_id'))
        export_format = request.form.get('format', 'excel')

        flash('تم تصدير التقرير بنجاح', 'success')
        return redirect(url_for('detailed_report', entity_type=entity_type, entity_id=entity_id))

    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير التقرير: {str(e)}', 'error')
        return redirect(url_for('camps_dashboard'))

# تصدير بيانات لوحة معلومات المخيمات
@app.route('/export-camps-dashboard', methods=['POST'])
@login_required
def export_camps_dashboard():
    """تصدير بيانات لوحة معلومات المخيمات"""
    try:
        # الحصول على الفلاتر
        filters = {
            'country': request.form.get('country', ''),
            'company': request.form.get('company', ''),
            'search': request.form.get('search', ''),
            'panels_filter': request.form.get('panels_filter', ''),
            'issues_filter': request.form.get('issues_filter', ''),
            'inspections_filter': request.form.get('inspections_filter', '')
        }

        export_format = request.form.get('format', 'excel')

        # الحصول على البيانات المفلترة
        camps = get_filtered_camps(filters)
        stats = calculate_camps_statistics(camps, filters)
        country_stats = calculate_country_statistics(camps)
        company_stats = calculate_company_statistics(camps)

        if export_format == 'excel':
            return export_camps_excel(camps, stats, country_stats, company_stats, filters)
        elif export_format == 'pdf':
            return export_camps_pdf(camps, stats, country_stats, company_stats, filters)
        else:
            flash('نوع التصدير غير مدعوم', 'error')
            return redirect(url_for('camps_dashboard'))

    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'error')
        return redirect(url_for('camps_dashboard'))

def export_camps_excel(camps, stats, country_stats, company_stats, filters):
    """تصدير البيانات إلى Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        from flask import send_file

        workbook = Workbook()
        workbook.remove(workbook.active)

        # إعداد التنسيق
        header_fill = PatternFill(start_color="B89966", end_color="B89966", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # ورقة الملخص
        summary_sheet = workbook.create_sheet("ملخص عام")
        workbook.active = summary_sheet

        summary_data = [
            ['تقرير لوحة معلومات المخيمات', ''],
            ['تاريخ التقرير', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['المستخدم', current_user.name],
            ['', ''],
            ['إجمالي المخيمات', stats['total_camps']],
            ['إجمالي الحجاج', stats['total_pilgrims']],
            ['إجمالي الشركات', stats['total_companies']],
            ['إجمالي الدول', stats['total_countries']],
            ['إجمالي اللوحات', stats['total_panels']],
            ['إجمالي المساحة (م²)', stats['total_area']],
            ['إجمالي البلاغات', stats['total_issues']],
            ['إجمالي طلبات الفحص', stats['total_inspections']],
            ['متوسط نسبة الحمل (%)', f"{stats['avg_load_percentage']:.1f}"],
            ['إجمالي الاستهلاك (كيلوواط)', f"{stats['total_consumption']:.1f}"]
        ]

        for row_idx, (label, value) in enumerate(summary_data, 1):
            summary_sheet.cell(row=row_idx, column=1, value=label)
            summary_sheet.cell(row=row_idx, column=2, value=value)

            if row_idx == 1:
                summary_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=16)
                summary_sheet.merge_cells(f'A{row_idx}:B{row_idx}')
            elif label and value:
                summary_sheet.cell(row=row_idx, column=1).font = Font(bold=True)

        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 25

        # ورقة المخيمات
        camps_sheet = workbook.create_sheet("المخيمات")
        camps_headers = [
            'رقم المربع', 'رقم المخيم', 'الشركة', 'الدولة', 'عدد الحجاج',
            'المساحة (م²)', 'Zone/Style', 'الشخص المسؤول', 'رقم الاتصال'
        ]

        for col, header in enumerate(camps_headers, 1):
            cell = camps_sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row_idx, camp in enumerate(camps, 2):
            camps_sheet.cell(row=row_idx, column=1, value=camp.square_number)
            camps_sheet.cell(row=row_idx, column=2, value=camp.camp_number)
            camps_sheet.cell(row=row_idx, column=3, value=camp.company.name)
            camps_sheet.cell(row=row_idx, column=4, value=camp.company.country.name)
            camps_sheet.cell(row=row_idx, column=5, value=camp.pilgrims_count or 0)
            camps_sheet.cell(row=row_idx, column=6, value=camp.total_area or 0)
            camps_sheet.cell(row=row_idx, column=7, value=camp.zone_style or '')
            camps_sheet.cell(row=row_idx, column=8, value=camp.responsible_person or '')
            camps_sheet.cell(row=row_idx, column=9, value=camp.contact_number or '')

        for col in range(1, len(camps_headers) + 1):
            camps_sheet.column_dimensions[camps_sheet.cell(row=1, column=col).column_letter].width = 15

        # ورقة إحصائيات الدول
        countries_sheet = workbook.create_sheet("إحصائيات الدول")
        countries_headers = ['الدولة', 'عدد المخيمات', 'عدد الحجاج', 'عدد الشركات', 'عدد اللوحات']

        for col, header in enumerate(countries_headers, 1):
            cell = countries_sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row_idx, (country, stats_data) in enumerate(country_stats.items(), 2):
            countries_sheet.cell(row=row_idx, column=1, value=country)
            countries_sheet.cell(row=row_idx, column=2, value=stats_data['camps_count'])
            countries_sheet.cell(row=row_idx, column=3, value=stats_data['pilgrims_count'])
            countries_sheet.cell(row=row_idx, column=4, value=stats_data['companies_count'])
            countries_sheet.cell(row=row_idx, column=5, value=stats_data['panels_count'])

        for col in range(1, len(countries_headers) + 1):
            countries_sheet.column_dimensions[countries_sheet.cell(row=1, column=col).column_letter].width = 20

        # ورقة إحصائيات الشركات
        companies_sheet = workbook.create_sheet("إحصائيات الشركات")
        companies_headers = ['الشركة', 'الدولة', 'عدد المخيمات', 'عدد الحجاج', 'عدد اللوحات']

        for col, header in enumerate(companies_headers, 1):
            cell = companies_sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row_idx, (company, stats_data) in enumerate(company_stats.items(), 2):
            companies_sheet.cell(row=row_idx, column=1, value=company)
            companies_sheet.cell(row=row_idx, column=2, value=stats_data['country'])
            companies_sheet.cell(row=row_idx, column=3, value=stats_data['camps_count'])
            companies_sheet.cell(row=row_idx, column=4, value=stats_data['pilgrims_count'])
            companies_sheet.cell(row=row_idx, column=5, value=stats_data['panels_count'])

        for col in range(1, len(companies_headers) + 1):
            companies_sheet.column_dimensions[companies_sheet.cell(row=1, column=col).column_letter].width = 20

        # حفظ الملف
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f'تقرير_المخيمات_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"خطأ في تصدير Excel: {str(e)}")
        flash(f'حدث خطأ أثناء تصدير Excel: {str(e)}', 'error')
        return redirect(url_for('camps_dashboard'))

def export_camps_pdf(camps, stats, country_stats, company_stats, filters):
    """تصدير البيانات إلى PDF"""
    try:
        flash('تصدير PDF قيد التطوير', 'info')
        return redirect(url_for('camps_dashboard'))

    except Exception as e:
        print(f"خطأ في تصدير PDF: {str(e)}")
        flash(f'حدث خطأ أثناء تصدير PDF: {str(e)}', 'error')
        return redirect(url_for('camps_dashboard'))

def auto_link_panels_to_camps():
    """ربط اللوحات بالمخيمات تلقائياً بناءً على الإحداثيات"""
    try:
        import math

        # الحصول على جميع المخيمات التي لها إحداثيات
        camps = Camp.query.filter(
            Camp.coordinates.isnot(None),
            Camp.is_active == True
        ).all()

        # الحصول على جميع اللوحات التي لها إحداثيات وليست مرتبطة بمخيم
        panels = MDBPanel.query.filter(
            MDBPanel.x_coordinate.isnot(None),
            MDBPanel.y_coordinate.isnot(None),
            MDBPanel.camp_id.is_(None)  # فقط اللوحات غير المرتبطة
        ).all()

        print(f"🔍 بدء ربط اللوحات: {len(panels)} لوحة غير مرتبطة، {len(camps)} مخيم متاح")

        linked_count = 0

        def calculate_distance_haversine(lat1, lon1, lat2, lon2):
            """حساب المسافة بين نقطتين بالكيلومتر باستخدام صيغة Haversine"""
            R = 6371  # نصف قطر الأرض بالكيلومترات

            lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
            dlat = lat2 - lat1
            dlon = lon2 - lon1

            a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
            c = 2 * math.asin(math.sqrt(a))

            return R * c

        def point_in_polygon(lat, lon, polygon_points):
            """فحص إذا كانت النقطة داخل المضلع"""
            if not polygon_points or len(polygon_points) < 3:
                return False

            x, y = lat, lon
            n = len(polygon_points)
            inside = False

            p1x, p1y = polygon_points[0]
            for i in range(1, n + 1):
                p2x, p2y = polygon_points[i % n]
                if y > min(p1y, p2y):
                    if y <= max(p1y, p2y):
                        if x <= max(p1x, p2x):
                            if p1y != p2y:
                                xinters = (y - p1y) * (p2x - p1x) / (p2y - p1y) + p1x
                            if p1x == p2x or x <= xinters:
                                inside = not inside
                p1x, p1y = p2x, p2y

            return inside

        for panel in panels:
            closest_camp = None
            min_distance = float('inf')
            found_inside_polygon = False

            panel_lat = panel.y_coordinate
            panel_lon = panel.x_coordinate

            print(f"🔍 معالجة اللوحة {panel.mdb} في ({panel_lat}, {panel_lon})")

            for camp in camps:
                try:
                    # الحصول على إحداثيات المخيم
                    camp_coords = camp.get_coordinates_list()

                    if not camp_coords:
                        continue

                    # إذا كان هناك أكثر من نقطة واحدة، فحص إذا كانت اللوحة داخل المضلع
                    if len(camp_coords) > 2:
                        if point_in_polygon(panel_lat, panel_lon, camp_coords):
                            closest_camp = camp
                            min_distance = 0
                            found_inside_polygon = True
                            print(f"   ✅ اللوحة داخل مضلع المخيم {camp.camp_number}")
                            break

                    # حساب المسافة إلى مركز المخيم
                    center_coords = camp.get_center_coordinates()
                    if center_coords and len(center_coords) == 2:
                        camp_lat, camp_lon = center_coords
                        distance = calculate_distance_haversine(panel_lat, panel_lon, camp_lat, camp_lon)

                        if distance < min_distance:
                            min_distance = distance
                            closest_camp = camp

                except Exception as e:
                    print(f"   ⚠️ خطأ في معالجة المخيم {camp.camp_number}: {e}")
                    continue

            # ربط اللوحة بأقرب مخيم إذا وُجد
            if closest_camp:
                # إذا كانت اللوحة داخل المضلع أو ضمن نطاق 10 كم
                if found_inside_polygon or min_distance <= 10.0:
                    try:
                        # تحديث اللوحة مع التحقق من وجود الشركة والدولة
                        panel.camp_id = closest_camp.id

                        if closest_camp.company:
                            panel.company_id = closest_camp.company_id
                            if closest_camp.company.country:
                                panel.country_id = closest_camp.company.country_id
                            else:
                                print(f"   ⚠️ المخيم {closest_camp.camp_number} - الشركة بدون دولة")
                        else:
                            print(f"   ⚠️ المخيم {closest_camp.camp_number} بدون شركة")

                        # إنشاء سجل ربط
                        assignment = PanelCampAssignment(
                            panel_id=panel.id,
                            camp_id=closest_camp.id,
                            assigned_at=datetime.datetime.now(),
                            is_active=True
                        )
                        db.session.add(assignment)
                        linked_count += 1

                        if found_inside_polygon:
                            print(f"   ✅ ربط مباشر: {panel.mdb} -> مخيم {closest_camp.camp_number}")
                        else:
                            print(f"   📏 ربط بالمسافة: {panel.mdb} -> مخيم {closest_camp.camp_number} (مسافة: {min_distance:.2f} كم)")

                    except Exception as e:
                        print(f"   ❌ خطأ في ربط اللوحة {panel.mdb}: {e}")
                        continue
                else:
                    print(f"   ❌ أقرب مخيم للوحة {panel.mdb} هو {closest_camp.camp_number} (مسافة: {min_distance:.2f} كم - بعيد جداً)")
            else:
                print(f"   ❌ لم يتم العثور على مخيم مناسب للوحة {panel.mdb}")

        # حفظ التغييرات
        if linked_count > 0:
            db.session.commit()
            print(f"🎉 تم ربط {linked_count} لوحة بالمخيمات بنجاح!")
        else:
            print("⚠️ لم يتم ربط أي لوحة")

        return linked_count

    except Exception as e:
        print(f"❌ خطأ في ربط اللوحات بالمخيمات: {str(e)}")
        db.session.rollback()
        return 0

def update_existing_panel_links():
    """تحديث الروابط الموجودة للوحات المرتبطة بالمخيمات"""
    try:
        print("🔄 تحديث الروابط الموجودة للوحات...")

        # الحصول على اللوحات المرتبطة بمخيمات لكن بدون company_id أو country_id
        panels_to_update = MDBPanel.query.filter(
            MDBPanel.camp_id.isnot(None),
            db.or_(
                MDBPanel.company_id.is_(None),
                MDBPanel.country_id.is_(None)
            )
        ).all()

        updated_count = 0

        for panel in panels_to_update:
            try:
                if panel.camp and panel.camp.company:
                    # تحديث company_id
                    if not panel.company_id:
                        panel.company_id = panel.camp.company_id

                    # تحديث country_id
                    if not panel.country_id and panel.camp.company.country:
                        panel.country_id = panel.camp.company.country_id

                    updated_count += 1
                    print(f"   ✅ تم تحديث اللوحة {panel.mdb}")

            except Exception as e:
                print(f"   ❌ خطأ في تحديث اللوحة {panel.mdb}: {e}")
                continue

        if updated_count > 0:
            db.session.commit()
            print(f"🎉 تم تحديث {updated_count} لوحة بنجاح!")
        else:
            print("ℹ️ لا توجد لوحات تحتاج تحديث")

        return updated_count

    except Exception as e:
        print(f"❌ خطأ في تحديث الروابط: {str(e)}")
        db.session.rollback()
        return 0

def calculate_distance(lat1, lon1, lat2, lon2):
    """حساب المسافة بين نقطتين بالكيلومتر باستخدام صيغة Haversine"""
    try:
        import math
        # تحويل الدرجات إلى راديان
        lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])

        # صيغة Haversine
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))

        # نصف قطر الأرض بالكيلومتر
        r = 6371

        return c * r

    except Exception as e:
        print(f"خطأ في حساب المسافة: {str(e)}")
        return float('inf')

# تحميل قالب استيراد المخيمات للوحة المعلومات
@app.route('/camps-dashboard/download-template')
@login_required
def download_camps_dashboard_template():
    """تحميل قالب Excel لاستيراد بيانات المخيمات"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        from flask import send_file

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "قالب المخيمات"

        # إعداد التنسيق
        header_fill = PatternFill(start_color="B89966", end_color="B89966", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # رؤوس الأعمدة
        headers = [
            'رقم المربع',
            'رقم المخيم',
            'عدد الحجاج',
            'اسم الشركة',
            'الدولة',
            'المساحة الإجمالية',
            'الإحداثيات',
            'Zone/Style',
            'الشخص المسؤول',
            'رقم الاتصال'
        ]

        # إضافة رؤوس الأعمدة
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # إضافة بيانات تجريبية
        sample_data = [
            ['A1', '001', 100, 'شركة الحج المتميزة', 'السعودية', 1000.5, '21.3891,39.8579', 'منطقة أ', 'أحمد محمد', '966501234567'],
            ['A2', '002', 150, 'شركة الحج المتميزة', 'السعودية', 1200.0, '21.3892,39.8580', 'منطقة أ', 'محمد أحمد', '966501234568'],
            ['B1', '003', 120, 'شركة الخدمات الذهبية', 'مصر', 1100.0, '21.3893,39.8581', 'منطقة ب', 'علي حسن', '966501234569']
        ]

        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # تعديل عرض الأعمدة
        column_widths = [15, 15, 12, 25, 15, 18, 20, 15, 20, 18]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[sheet.cell(row=1, column=col).column_letter].width = width

        # إضافة ورقة التعليمات
        instructions_sheet = workbook.create_sheet("تعليمات الاستيراد")

        instructions = [
            ['تعليمات استيراد بيانات المخيمات', ''],
            ['', ''],
            ['الأعمدة المطلوبة:', ''],
            ['رقم المربع', 'رقم المربع الذي يحتوي على المخيم (مثل: A1, B2)'],
            ['رقم المخيم', 'رقم المخيم الفريد (مثل: 001, 002)'],
            ['عدد الحجاج', 'عدد الحجاج في المخيم (رقم صحيح)'],
            ['اسم الشركة', 'اسم الشركة المسؤولة عن المخيم'],
            ['الدولة', 'اسم الدولة التي تنتمي إليها الشركة'],
            ['المساحة الإجمالية', 'مساحة المخيم بالمتر المربع (رقم عشري)'],
            ['الإحداثيات', 'إحداثيات المخيم بصيغة: خط العرض,خط الطول'],
            ['Zone/Style', 'منطقة أو نمط المخيم (اختياري)'],
            ['الشخص المسؤول', 'اسم الشخص المسؤول عن المخيم (اختياري)'],
            ['رقم الاتصال', 'رقم هاتف الشخص المسؤول (اختياري)'],
            ['', ''],
            ['ملاحظات مهمة:', ''],
            ['• تأكد من صحة أسماء الدول والشركات', ''],
            ['• الإحداثيات يجب أن تكون بصيغة: خط العرض,خط الطول', ''],
            ['• الأعمدة الاختيارية يمكن تركها فارغة', ''],
            ['• تأكد من عدم وجود مخيمات مكررة', ''],
            ['• احفظ الملف بصيغة Excel (.xlsx)', '']
        ]

        for row_idx, (col1, col2) in enumerate(instructions, 1):
            instructions_sheet.cell(row=row_idx, column=1, value=col1)
            instructions_sheet.cell(row=row_idx, column=2, value=col2)

            if row_idx == 1:
                instructions_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=16)
                instructions_sheet.merge_cells(f'A{row_idx}:B{row_idx}')
            elif col1 and not col2:
                instructions_sheet.cell(row=row_idx, column=1).font = Font(bold=True)

        instructions_sheet.column_dimensions['A'].width = 25
        instructions_sheet.column_dimensions['B'].width = 50

        # حفظ الملف
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f'قالب_استيراد_المخيمات_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"خطأ في تحميل القالب: {str(e)}")
        flash(f'حدث خطأ أثناء تحميل القالب: {str(e)}', 'error')
        return redirect(url_for('camps_dashboard'))

# استيراد بيانات المخيمات للوحة المعلومات
@app.route('/camps-dashboard/import-data', methods=['POST'])
@login_required
def import_camps_dashboard_data():
    """استيراد بيانات المخيمات من ملف Excel"""
    try:
        # التحقق من الصلاحيات
        if current_user.role not in ['admin']:
            flash('ليس لديك صلاحية لاستيراد البيانات', 'error')
            return redirect(url_for('camps_dashboard'))

        # التحقق من وجود الملف
        if 'camps_file' not in request.files:
            flash('لم يتم اختيار ملف للاستيراد', 'error')
            return redirect(url_for('camps_dashboard'))

        file = request.files['camps_file']
        import_mode = request.form.get('import_mode', 'add')

        if file.filename == '':
            flash('لم يتم اختيار ملف للاستيراد', 'error')
            return redirect(url_for('camps_dashboard'))

        # التحقق من نوع الملف
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('يرجى اختيار ملف Excel صحيح (.xlsx أو .xls)', 'error')
            return redirect(url_for('camps_dashboard'))

        # قراءة الملف
        import pandas as pd

        try:
            df = pd.read_excel(file)
        except Exception as e:
            flash(f'خطأ في قراءة ملف Excel: {str(e)}', 'error')
            return redirect(url_for('camps_dashboard'))

        # التحقق من وجود الأعمدة المطلوبة
        required_columns = ['رقم المربع', 'رقم المخيم', 'عدد الحجاج', 'اسم الشركة', 'الدولة']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            flash(f'الأعمدة التالية مفقودة في الملف: {", ".join(missing_columns)}', 'error')
            return redirect(url_for('camps_dashboard'))

        # معالجة البيانات
        success_count = 0
        error_count = 0
        errors = []

        # في حالة الاستبدال، حذف البيانات الموجودة
        if import_mode == 'replace':
            try:
                Camp.query.delete()
                db.session.commit()
                flash('تم حذف البيانات الموجودة', 'info')
            except Exception as e:
                db.session.rollback()
                flash(f'خطأ في حذف البيانات الموجودة: {str(e)}', 'error')
                return redirect(url_for('camps_dashboard'))

        # معالجة كل صف
        for index, row in df.iterrows():
            try:
                # التحقق من البيانات الأساسية
                if pd.isna(row['رقم المربع']) or pd.isna(row['رقم المخيم']):
                    errors.append(f'الصف {index + 2}: رقم المربع أو رقم المخيم فارغ')
                    error_count += 1
                    continue

                # البحث عن الدولة أو إنشاؤها
                country_name = str(row['الدولة']).strip()
                country = Country.query.filter_by(name=country_name).first()
                if not country:
                    country = Country(name=country_name, is_active=True)
                    db.session.add(country)
                    db.session.flush()

                # البحث عن الشركة أو إنشاؤها
                company_name = str(row['اسم الشركة']).strip()
                company = Company.query.filter_by(name=company_name, country_id=country.id).first()
                if not company:
                    company = Company(name=company_name, country_id=country.id, is_active=True)
                    db.session.add(company)
                    db.session.flush()

                # التحقق من وجود المخيم
                square_number = str(row['رقم المربع']).strip()
                camp_number = str(row['رقم المخيم']).strip()

                existing_camp = Camp.query.filter_by(
                    square_number=square_number,
                    camp_number=camp_number
                ).first()

                if existing_camp and import_mode == 'add':
                    errors.append(f'الصف {index + 2}: المخيم {camp_number} في المربع {square_number} موجود مسبقاً')
                    error_count += 1
                    continue

                # إنشاء أو تحديث المخيم
                if existing_camp and import_mode == 'update':
                    camp = existing_camp
                else:
                    camp = Camp()

                # تعيين البيانات
                camp.square_number = square_number
                camp.camp_number = camp_number
                camp.company_id = company.id
                camp.pilgrims_count = int(row['عدد الحجاج']) if not pd.isna(row['عدد الحجاج']) else 0

                # معالجة المساحة الإجمالية مع أسماء أعمدة متعددة محتملة
                total_area = 0
                area_columns = ['المساحة الإجمالية', 'المساحة', 'Total Area', 'Area']
                for col in area_columns:
                    if col in row and not pd.isna(row[col]):
                        try:
                            total_area = float(row[col])
                            break
                        except (ValueError, TypeError):
                            continue
                camp.total_area = total_area

                # معالجة الإحداثيات مع أسماء أعمدة متعددة محتملة
                coordinates = None
                coord_columns = ['الإحداثيات', 'Coordinates', 'إحداثيات المضلع', 'Polygon Coordinates']
                for col in coord_columns:
                    if col in row and not pd.isna(row[col]):
                        coordinates = str(row[col]).strip()
                        if coordinates and coordinates != 'nan':
                            break
                camp.coordinates = coordinates

                # معالجة Zone/Style مع أسماء أعمدة متعددة محتملة
                zone_style = None
                zone_columns = ['Zone/Style', 'Zone', 'Style', 'المنطقة', 'النمط']
                for col in zone_columns:
                    if col in row and not pd.isna(row[col]):
                        zone_style = str(row[col]).strip()
                        if zone_style and zone_style != 'nan':
                            break
                camp.zone_style = zone_style

                # معالجة الشخص المسؤول مع أسماء أعمدة متعددة محتملة
                responsible_person = None
                person_columns = ['الشخص المسؤول', 'المسؤول', 'Responsible Person', 'Contact Person']
                for col in person_columns:
                    if col in row and not pd.isna(row[col]):
                        responsible_person = str(row[col]).strip()
                        if responsible_person and responsible_person != 'nan':
                            break
                camp.responsible_person = responsible_person

                # معالجة رقم الاتصال مع أسماء أعمدة متعددة محتملة
                contact_number = None
                contact_columns = ['رقم الاتصال', 'الهاتف', 'Contact Number', 'Phone']
                for col in contact_columns:
                    if col in row and not pd.isna(row[col]):
                        contact_number = str(row[col]).strip()
                        if contact_number and contact_number != 'nan':
                            break
                camp.contact_number = contact_number

                camp.is_active = True

                if not existing_camp or import_mode != 'update':
                    db.session.add(camp)

                success_count += 1

            except Exception as e:
                errors.append(f'الصف {index + 2}: {str(e)}')
                error_count += 1
                continue

        # حفظ التغييرات
        try:
            db.session.commit()

            # ربط اللوحات بالمخيمات تلقائياً بناءً على الإحداثيات
            linked_panels = 0
            if success_count > 0:
                try:
                    linked_panels = auto_link_panels_to_camps()
                    if linked_panels > 0:
                        flash(f'تم ربط {linked_panels} لوحة بالمخيمات تلقائياً', 'info')
                except Exception as e:
                    print(f"خطأ في ربط اللوحات: {str(e)}")

            # رسائل النتائج
            if success_count > 0:
                flash(f'تم استيراد {success_count} مخيم بنجاح', 'success')

            if error_count > 0:
                flash(f'فشل في استيراد {error_count} مخيم', 'warning')

                # عرض أول 5 أخطاء
                if errors:
                    error_details = '<br>'.join(errors[:5])
                    if len(errors) > 5:
                        error_details += f'<br>... و {len(errors) - 5} أخطاء أخرى'
                    flash(f'تفاصيل الأخطاء:<br>{error_details}', 'warning')

        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في حفظ البيانات: {str(e)}', 'error')

        return redirect(url_for('camps_dashboard'))

    except Exception as e:
        print(f"خطأ في استيراد البيانات: {str(e)}")
        flash(f'حدث خطأ أثناء استيراد البيانات: {str(e)}', 'error')
        return redirect(url_for('camps_dashboard'))

# ربط اللوحات بالمخيمات تلقائياً - API
@app.route('/camps-dashboard/link-panels', methods=['POST'])
@login_required
def auto_link_panels_api():
    """ربط اللوحات بالمخيمات تلقائياً عبر API"""
    try:
        # التحقق من الصلاحيات
        if current_user.role not in ['admin']:
            return jsonify({
                'success': False,
                'message': 'ليس لديك صلاحية لربط اللوحات'
            }), 403

        print("🚀 بدء عملية الربط التلقائي للوحات بالمخيمات عبر API...")

        # تنفيذ عملية الربط
        linked_count = auto_link_panels_to_camps()

        if linked_count > 0:
            return jsonify({
                'success': True,
                'linked_count': linked_count,
                'message': f'تم ربط {linked_count} لوحة بالمخيمات بنجاح! يمكنك الآن رؤية معلومات المخيم والشركة والدولة في النوافذ المنبثقة.'
            })
        else:
            return jsonify({
                'success': True,
                'linked_count': 0,
                'message': 'لا توجد لوحات جديدة للربط. جميع اللوحات مرتبطة بالفعل أو لا تحتوي على إحداثيات صحيحة.'
            })

    except Exception as e:
        print(f"❌ خطأ في API الربط التلقائي: {e}")
        return jsonify({
            'success': False,
            'message': f'حدث خطأ أثناء ربط اللوحات: {str(e)}'
        }), 500

# تحديث الروابط الموجودة - API
@app.route('/camps-dashboard/update-links', methods=['POST'])
@login_required
def update_panel_links_api():
    """تحديث الروابط الموجودة للوحات عبر API"""
    try:
        # التحقق من الصلاحيات
        if current_user.role not in ['admin']:
            return jsonify({
                'success': False,
                'message': 'ليس لديك صلاحية لتحديث الروابط'
            }), 403

        print("🔄 بدء تحديث الروابط الموجودة عبر API...")

        # تنفيذ عملية التحديث
        updated_count = update_existing_panel_links()

        if updated_count > 0:
            return jsonify({
                'success': True,
                'updated_count': updated_count,
                'message': f'تم تحديث {updated_count} لوحة بنجاح! الآن ستظهر معلومات الشركة والدولة في النوافذ المنبثقة.'
            })
        else:
            return jsonify({
                'success': True,
                'updated_count': 0,
                'message': 'جميع اللوحات محدثة بالفعل. لا توجد روابط تحتاج تحديث.'
            })

    except Exception as e:
        print(f"❌ خطأ في API تحديث الروابط: {e}")
        return jsonify({
            'success': False,
            'message': f'حدث خطأ أثناء تحديث الروابط: {str(e)}'
        }), 500

# إضافة routes مؤقتة لحل مشكلة pilgrimage_360
@app.route('/pilgrimage-360/')
def pilgrimage_index():
    """الصفحة الرئيسية لتجربة الحجاج والسياح 360° - مؤقت"""
    return render_template('pilgrimage_360/index.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/enhanced-interactive-map')
def pilgrimage_enhanced_map():
    """الخريطة التفاعلية المحسنة - مؤقت"""
    return render_template('pilgrimage_360/enhanced_interactive_map.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/map')
def pilgrimage_map_360():
    """خريطة 360 درجة - مؤقت"""
    return render_template('pilgrimage_360/map_360.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/interactive-map')
def pilgrimage_interactive_map():
    """الخريطة التفاعلية العادية - مؤقت"""
    return render_template('pilgrimage_360/interactive_map.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/virtual-tour')
def pilgrimage_virtual_tour():
    """الجولة الافتراضية - مؤقت"""
    return render_template('pilgrimage_360/virtual_tour.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/report-issue')
def pilgrimage_report_issue():
    """تقديم بلاغ - مؤقت"""
    return render_template('pilgrimage_360/report_issue.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/viewer-3d')
def pilgrimage_viewer_3d():
    """عارض ثلاثي الأبعاد - مؤقت"""
    return render_template('pilgrimage_360/3d_viewer.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/help-center')
def pilgrimage_help_center():
    """مركز المساعدة - مؤقت"""
    return render_template('pilgrimage_360/help_center.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/admin-locations')
def pilgrimage_admin_locations():
    """إدارة المواقع - مؤقت"""
    return render_template('pilgrimage_360/admin_locations.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/location/<int:location_id>')
def pilgrimage_location_detail(location_id):
    """تفاصيل الموقع - مؤقت"""
    return render_template('pilgrimage_360/location_detail.html',
                         location_id=location_id,
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

# API endpoints للخريطة 360°
@app.route('/pilgrimage-360/api/locations')
def pilgrimage_api_locations():
    """API للحصول على المواقع - مؤقت"""
    try:
        # إنشاء بيانات تجريبية
        locations = [
            {
                'id': 1,
                'name': 'المسجد الحرام',
                'latitude': 21.4225,
                'longitude': 39.8262,
                'category': {'name': 'مساجد'},
                'description': 'أقدس مكان في الإسلام',
                'panorama_360_images': '["/static/360_images/kaaba_360.jpg"]'
            },
            {
                'id': 2,
                'name': 'مطعم الحرمين',
                'latitude': 21.4200,
                'longitude': 39.8280,
                'category': {'name': 'مطاعم'},
                'description': 'مطعم يقدم الأطعمة الحلال',
                'panorama_360_images': '["/static/360_images/restaurant_360.jpg"]'
            },
            {
                'id': 3,
                'name': 'مستشفى الحرم',
                'latitude': 21.4250,
                'longitude': 39.8240,
                'category': {'name': 'خدمات طبية'},
                'description': 'خدمات طبية للحجاج',
                'panorama_360_images': None
            }
        ]
        return jsonify({'locations': locations})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/pilgrimage-360/upload-panorama', methods=['POST'])
def pilgrimage_upload_panorama():
    """رفع صورة بانورامية 360° - مؤقت"""
    try:
        print("تم استلام طلب رفع صورة")
        print(f"Files: {list(request.files.keys())}")
        print(f"Form data: {dict(request.form)}")

        location_id = request.form.get('location_id')
        if not location_id:
            print("خطأ: معرف الموقع مفقود")
            return jsonify({'success': False, 'message': 'معرف الموقع مطلوب'}), 400

        if 'panorama_image' not in request.files:
            print("خطأ: لم يتم رفع أي صورة")
            return jsonify({'success': False, 'message': 'لم يتم رفع أي صورة'}), 400

        file = request.files['panorama_image']
        if file.filename == '':
            print("خطأ: لم يتم اختيار ملف")
            return jsonify({'success': False, 'message': 'لم يتم اختيار ملف'}), 400

        print(f"اسم الملف: {file.filename}")

        if file:
            # إنشاء مجلد الحفظ
            upload_folder = os.path.join('static', '360_images')
            os.makedirs(upload_folder, exist_ok=True)
            print(f"مجلد الحفظ: {upload_folder}")

            # حفظ الملف
            filename = f"location_{location_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            file_path = os.path.join(upload_folder, filename)
            file.save(file_path)
            print(f"تم حفظ الملف في: {file_path}")

            image_url = f'/static/360_images/{filename}'
            print(f"رابط الصورة: {image_url}")

            return jsonify({
                'success': True,
                'message': 'تم رفع الصورة بنجاح',
                'image_url': image_url
            })
        else:
            return jsonify({'success': False, 'message': 'نوع الملف غير مدعوم'}), 400

    except Exception as e:
        print(f"خطأ في رفع الصورة: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'خطأ في رفع الصورة: {str(e)}'
        }), 500

@app.route('/pilgrimage-360/add-quick-location', methods=['POST'])
def pilgrimage_add_quick_location():
    """إضافة موقع سريع - مؤقت"""
    try:
        data = request.get_json()

        name = data.get('name')
        description = data.get('description', '')
        latitude = float(data.get('latitude'))
        longitude = float(data.get('longitude'))

        if not name:
            return jsonify({'success': False, 'message': 'اسم الموقع مطلوب'}), 400

        # محاكاة إضافة الموقع
        new_id = datetime.datetime.now().microsecond  # ID مؤقت

        return jsonify({
            'success': True,
            'message': 'تم إضافة الموقع بنجاح',
            'location_id': new_id
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في إضافة الموقع: {str(e)}'
        }), 500

@app.route('/pilgrimage-360/add-location')
def pilgrimage_add_location():
    """صفحة إضافة موقع جديد - مؤقت"""
    return render_template('pilgrimage_360/add_location.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/report-status')
def pilgrimage_report_status():
    """حالة البلاغات - مؤقت"""
    return render_template('pilgrimage_360/report_status.html',
                         current_date=datetime.datetime.now().strftime("%Y-%m-%d"))

@app.route('/pilgrimage-360/admin/add-location', methods=['POST'])
def pilgrimage_admin_add_location():
    """API لإضافة موقع من لوحة الإدارة - مؤقت"""
    try:
        data = request.get_json()
        # محاكاة إضافة الموقع
        return jsonify({
            'success': True,
            'message': 'تم إضافة الموقع بنجاح',
            'location_id': datetime.datetime.now().microsecond
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/pilgrimage-360/admin/delete-location/<int:location_id>', methods=['DELETE'])
def pilgrimage_admin_delete_location(location_id):
    """API لحذف موقع من لوحة الإدارة - مؤقت"""
    try:
        # محاكاة حذف الموقع
        return jsonify({
            'success': True,
            'message': 'تم حذف الموقع بنجاح'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== APIs للبانوراما ====================


@app.route('/api/panoramas')
@login_required
def get_panoramas():
    """الحصول على قائمة البانوراما"""
    panoramas = Panorama360.query.filter_by(is_active=True).all()
    return jsonify([{
        'id': p.id,
        'title': p.title,
        'location_name': p.location_name,
        'latitude': p.latitude,
        'longitude': p.longitude
    } for p in panoramas])

@app.route('/api/panoramas/<int:panorama_id>')
@login_required
def get_panorama(panorama_id):
    """الحصول على بانوراما محددة"""
    panorama = Panorama360.query.get_or_404(panorama_id)
    return jsonify({
        'id': panorama.id,
        'title': panorama.title,
        'location_name': panorama.location_name,
        'latitude': panorama.latitude,
        'longitude': panorama.longitude
    })

@app.route('/api/panels/<int:panel_id>/link-panorama', methods=['POST'])
@login_required
def link_panel_to_panorama(panel_id):
    """ربط لوحة ببانوراما"""
    try:
        data = request.get_json()
        panorama_id = data.get('panorama_id')
        
        panel = MDBPanel.query.get_or_404(panel_id)
        panel.panorama_id = panorama_id
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'تم الربط بنجاح'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/panels/<int:panel_id>/unlink-panorama', methods=['POST'])
@login_required
def unlink_panel_from_panorama(panel_id):
    """إلغاء ربط لوحة من البانوراما"""
    try:
        panel = MDBPanel.query.get_or_404(panel_id)
        panel.panorama_id = None
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'تم إلغاء الربط بنجاح'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ==========================================
# API المفضلة (بسيطة - مخزنة في الجلسة)
# ==========================================

@app.route('/api/favorites/add', methods=['POST'])
@csrf.exempt
@login_required
def api_favorites_add():
    """إضافة عنصر للمفضلة"""
    try:
        data     = request.get_json() or {}
        fav_type = data.get('type', 'panorama')
        fav_id   = data.get('id')
        if not fav_id:
            return jsonify({'success': False, 'message': 'المعرف مطلوب'}), 400

        key = f'favorites_{fav_type}'
        favs = session.get(key, [])
        if fav_id not in favs:
            favs.append(fav_id)
            session[key] = favs
            session.modified = True

        return jsonify({'success': True, 'message': 'تم الحفظ في المفضلة', 'count': len(favs)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/favorites/remove', methods=['POST'])
@csrf.exempt
@login_required
def api_favorites_remove():
    """إزالة عنصر من المفضلة"""
    try:
        data     = request.get_json() or {}
        fav_type = data.get('type', 'panorama')
        fav_id   = data.get('id')
        key      = f'favorites_{fav_type}'
        favs     = session.get(key, [])
        if fav_id in favs:
            favs.remove(fav_id)
            session[key] = favs
            session.modified = True
        return jsonify({'success': True, 'message': 'تم الحذف من المفضلة'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/favorites/list')
@csrf.exempt
@login_required
def api_favorites_list():
    """قائمة المفضلة"""
    try:
        fav_type   = request.args.get('type', 'panorama')
        key        = f'favorites_{fav_type}'
        fav_ids    = session.get(key, [])
        items      = []
        if fav_type == 'panorama' and fav_ids:
            pans = Panorama360.query.filter(Panorama360.id.in_(fav_ids)).all()
            items = [{'id': p.id, 'title': p.title, 'location': p.location_name or p.area} for p in pans]
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/favorites')
@login_required
def favorites_page():
    """صفحة المفضلة"""
    fav_ids  = session.get('favorites_panorama', [])
    panoramas = []
    if fav_ids:
        panoramas = Panorama360.query.filter(Panorama360.id.in_(fav_ids)).all()
    return render_template('favorites.html', panoramas=panoramas)



if __name__ == '__main__':
    print("بدء تشغيل الخادم على المنفذ 5000...")
    # تعديل إعدادات الخادم لتحسين الاستقرار
    app.run(
        debug=True,  # تفعيل وضع التصحيح لرؤية الأخطاء
        host='0.0.0.0',
        port=5000,
        threaded=True,  # تفعيل وضع الخيوط المتعددة
        use_reloader=True  # تم تفعيله لضمان تحديث الكود تلقائياً
    )
